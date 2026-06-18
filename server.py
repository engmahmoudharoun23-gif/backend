from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Query, Request, WebSocket, WebSocketDisconnect, BackgroundTasks, Body, Header, Response
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from fastapi.responses import StreamingResponse
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict
import uuid
import asyncio
import re
from concurrent.futures import ThreadPoolExecutor

# ⚡ Thread pool لمعالجة الصور والعمليات الثقيلة
thread_pool = ThreadPoolExecutor(max_workers=20)

def compress_image(content: bytes, target_kb: int = 300) -> bytes:
    """ضغط الصورة ليصل حجمها إلى target_kb كحد أقصى (افتراضي 300KB)"""
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(content))
        # تحويل إلى RGB إذا كانت RGBA
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        # تقليل الحجم إذا كانت كبيرة جداً
        max_size = (1920, 1920)
        img.thumbnail(max_size, Image.Resampling.LANCZOS)
        
        target_bytes = target_kb * 1024
        # محاولة ضغط تدريجي: خفض الجودة حتى الحجم المطلوب
        for quality in (85, 75, 65, 55, 45, 35, 25):
            output = io.BytesIO()
            img.save(output, format='JPEG', quality=quality, optimize=True)
            data = output.getvalue()
            if len(data) <= target_bytes:
                return data
        
        # لو ظلّ أكبر: تقليل الأبعاد تدريجياً
        for scale in (0.75, 0.6, 0.5, 0.4):
            w, h = img.size
            resized = img.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
            output = io.BytesIO()
            resized.save(output, format='JPEG', quality=55, optimize=True)
            data = output.getvalue()
            if len(data) <= target_bytes:
                return data
        
        return data  # أفضل ما استطعنا الوصول إليه
    except Exception:
        return content  # إرجاع الأصلي إذا فشل الضغط
from uuid import uuid4
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import base64
from emergentintegrations.llm.chat import LlmChat, UserMessage
from PIL import Image
from io import BytesIO
import time
import random
import resend
import arabic_reshaper
from bidi.algorithm import get_display
import html

# ============= CACHING SYSTEM =============
# كاش بسيط وسريع للأداء العالي
from functools import lru_cache
import time as time_module

class SimpleCache:
    """كاش بسيط مع TTL"""
    def __init__(self):
        self._cache = {}
        self._timestamps = {}
    
    def get(self, key, ttl=10):
        """جلب من الكاش إذا لم تنتهي الصلاحية"""
        if key in self._cache:
            if time_module.time() - self._timestamps.get(key, 0) < ttl:
                return self._cache[key]
        return None
    
    def set(self, key, value):
        """حفظ في الكاش"""
        self._cache[key] = value
        self._timestamps[key] = time_module.time()
    
    def clear(self, prefix=None):
        """مسح الكاش"""
        if prefix:
            keys_to_delete = [k for k in self._cache if k.startswith(prefix)]
            for k in keys_to_delete:
                del self._cache[k]
                del self._timestamps[k]
        else:
            self._cache.clear()
            self._timestamps.clear()

# كاش عام
cache = SimpleCache()
stats_cache = {}
CACHE_TTL = 10  # 10 ثواني

# ======================================================
# دالة تطبيع حالة الأسفلت - توحيد كل الصيغ إلى قيمة واحدة
# تضمن عدم تكرار مشكلة عدم العد الصحيح مهما كان المشروع
# ======================================================
_ASPHALT_CANONICAL = "تم الإصلاح-ومتبقي الأسفلت"
import re as _re
_ASPHALT_PATTERN = _re.compile(r"متبقيـ*[اأإ]سفلت", _re.IGNORECASE)

def normalize_asphalt_status(status_val: str) -> str:
    """
    يُوحّد كل صيغ حالة 'بانتظار الأسفلت' و'تم الإصلاح - ومتبقي الأسفلت'
    إلى قيم قياسية ثابتة لضمان حساب صحيح في كل الاستعلامات.
    يعمل مع أي مشروع حالي أو مستقبلي.
    """
    if not status_val:
        return status_val
    s = status_val.strip()
    # توحيد حالة 'تم الإصلاح - ومتبقي الأسفلت' بكل صيغها
    if _ASPHALT_PATTERN.search(s):
        return _ASPHALT_CANONICAL
    return s



ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection - محسن للأداء العالي
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(
    mongo_url,
    maxPoolSize=100,        # عدد الاتصالات المتزامنة
    minPoolSize=10,         # الحد الأدنى للاتصالات
    maxIdleTimeMS=30000,    # وقت الخمول 30 ثانية
    connectTimeoutMS=5000,  # وقت الاتصال 5 ثواني
    serverSelectionTimeoutMS=5000
)
db = client[os.environ['DB_NAME']]

# mongo_url = os.environ.get('MONGO_URL')
# db_name = os.environ.get('DB_NAME')

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto", bcrypt__rounds=12)

# JWT settings
SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 525600  # 365 يوم (سنة كاملة) - الجلسة مفتوحة حتى تسجيل الخروج

# Security
security = HTTPBearer()

# LLM Integration
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY')

# Resend Email
resend.api_key = os.environ.get('RESEND_API_KEY')

# Store for password reset codes (in production, use Redis)
password_reset_codes = {}

# Create the main app
app = FastAPI(
    title="نظام إدارة البلاغات",
    description="API للتعامل مع البلاغات والمستخدمين",
    version="2.0",
    # ⚡ تحسينات الأداء
    openapi_url=None,  # تعطيل OpenAPI في الإنتاج للسرعة
)

os.makedirs("uploads", exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

print("SERVER RELOADING - WFM TOGGLE ACTIVE")

def arabic_text(text):
    """دالة عالمية لتحويل النص العربي ليدعم الـ PDF (Reshaping + Bidi)"""
    if not text:
        return ''
    if not isinstance(text, str):
        text = str(text)
    try:
        # فك تشفير أي نصوص مشفرة أو Unicode Escapes (مثل \u0627)
        if '\\u' in text:
            try:
                text = text.encode('utf-8').decode('unicode-escape')
            except:
                pass
        
        # فك تشفير HTML Entities
        text = html.unescape(text)
        
        # معالجة العربية
        reshaped = arabic_reshaper.reshape(text)
        bidi_text = get_display(reshaped)
        return bidi_text
    except Exception:
        return text

def normalize_arabic(text: str) -> str:
    """توحيد النص العربي لضمان تطابق الأسماء (المشاريع/المحافظات)"""
    if not text:
        return ""
    # توحيد المسافات والهمزات والتاء المربوطة والألف المقصورة
    res = str(text).strip()
    res = re.sub(r'\s+', ' ', res)
    res = res.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')
    res = res.replace('ة', 'ه')
    res = res.replace('ى', 'ي')
    return res

def normalize_arabic_regex(s: str) -> str:
    """توحيد النص العربي وبناء تعبير نمطي (Regex) للبحث المرن دون استبدال تكراري"""
    if not s:
        return ""
    mapping = {
        '\u0623': '[\u0623\u0627]',  # أ
        '\u0625': '[\u0625\u0627]',  # إ
        '\u0627': '[\u0627\u0623\u0625]',  # ا
        '\u0629': '[\u0647\u0629]',  # ة
        '\u0647': '[\u0647\u0629]',  # ه
        '\u064a': '[\u064a\u0649]',  # ي
        '\u0649': '[\u064a\u0649]',  # ى
        ' ': '.*'
    }
    return "".join(mapping.get(c, c) for c in s)

# ⚡ Semaphore للتحكم في عدد العمليات المتزامنة الثقيلة
import asyncio
upload_semaphore = asyncio.Semaphore(50)  # 50 رفع صور متزامن
query_semaphore = asyncio.Semaphore(100)  # 100 استعلام متزامن

# WebSocket Connection Manager للدردشة الفورية
class ConnectionManager:
    def __init__(self):
        self.active_connections: dict[str, WebSocket] = {}

    async def connect(self, websocket: WebSocket, user_id: str):
        await websocket.accept()
        self.active_connections[user_id] = websocket

    def disconnect(self, user_id: str):
        if user_id in self.active_connections:
            del self.active_connections[user_id]

    async def send_personal_message(self, message: dict, user_id: str):
        if user_id in self.active_connections:
            try:
                await self.active_connections[user_id].send_json(message)
            except Exception:
                self.disconnect(user_id)

manager = ConnectionManager()

# ============= OBJECT STORAGE =============
import requests as http_requests

# ============= OBJECT STORAGE =============
from storage import upload_image as _upload_image, get_object as _get_object, guess_ext as _guess_ext

def put_object(path: str, data: bytes, content_type: str) -> dict:
    """رفع الملف إلى Cloudinary بدلاً من التخزين القديم"""
    try:
        url = _upload_image(data, category="uploads", content_type=content_type)
        return {"path": url}
    except Exception as e:
        logging.error(f"Cloudinary put_object failed: {e}")
        raise

async def process_images_for_storage(images: List[str], category: str = "general") -> List[str]:
    """مساعد: تحويل أي صور base64 في القائمة إلى روابط Cloudinary"""
    if not images:
        return []
    processed = []
    for img in images:
        if isinstance(img, str) and img.startswith("data:"):
            # هذا base64، قم برفعه إلى Cloudinary
            try:
                header, encoded = img.split(",", 1)
                data = base64.b64decode(encoded)
                content_type = header.split(";")[0].split(":")[1]
                url = _upload_image(data, category=category, content_type=content_type)
                processed.append(url)
            except Exception as e:
                logging.error(f"Failed to upload base64 to Cloudinary: {e}")
                processed.append(img)
        else:
            # رابط جاهز بالفعل
            processed.append(img)
    return processed

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ============= نظام تخزين الصور (Cloudinary) =============
# تم نقل الإعدادات إلى storage.py


# ============= MODELS =============

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: Optional[EmailStr] = None
    full_name: str
    title: Optional[str] = None  # اللقب مثل "المهندس /"
    profile_picture: Optional[str] = None  # الصورة الشخصية (base64 أو URL)
    role: str = "user"  # admin or user
    governorates: List[str] = []  # قائمة المحافظات المسموح بها (فارغة = كل المحافظات للـ admin)
    projects: List[str] = []  # قائمة المشاريع المخصصة (فارغة = كل المشاريع للـ admin)
    hashed_password: str
    created_by: Optional[str] = None  # ID المستخدم الذي أنشأ هذا الحساب
    can_create_subusers: bool = True  # هل يمكن للمستخدم إنشاء مستخدمين فرعيين
    permissions: List[str] = []  # صلاحيات المستخدم (عامة - غير مرتبطة بمشروع)
    allowed_chat_users: List[str] = []  # قائمة معرفات المستخدمين المسموح بالتحدث معهم بشكل استثنائي
    connection_permissions: Optional[dict] = None  # صلاحيات التوصيلات حسب المشروع (legacy)
    project_permissions: Optional[Dict[str, List[str]]] = None  # صلاحيات لكل مشروع على حدة {"project_name": ["perm1", "perm2"]}
    personal_theme: Optional[str] = None  # الثيم الشخصي للمستخدم
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_active: bool = True

def get_flexible_project_query(project_name):
    """
    بناء استعلام Mongo مرن جداً للمشاريع يدعم:
    1. تنوع الهمزات (أ، إ، آ، ا)
    2. الهاء والتاء المربوطة (ه، ة)
    3. الياء والألف المقصورة (ي، ى)
    4. تجاهل الكلمات العامة (مشروع، إصلاح، أعمال، إلخ) لزيادة دقة التطابق
    5. التعامل مع المسافات والشرطات
    """
    if not project_name:
        return None
    
    # ⚡ توحيد الاسم الخاص بكشف التسربات لتسهيل المطابقة بغض النظر عن طريقة كتابته
    norm = normalize_arabic(project_name)
    if "كشف" in norm and "تسرب" in norm:
        # هذا هو مشروع كشف التسربات وإصلاحها
        # نرجّع تعبيراً نمطياً يطابق أي صيغة تحتوي على "كشف" و "تسرب"
        full_regex = r"^(مشروع\s+)?(ال)?كشف[\s\-_]*(ال)?تسرب([اأإآ]ت)?([\s\-_]*(ال)?و[اأإآ]صل[اأإآ]ح[هة][اأإآ])?$"
        return {"$regex": full_regex, "$options": "i"}
    
    # تنظيف النص وتقسيمه لكلمات بشكل عدواني (تجاهل المسافات والشرطات)
    cleaned_name = re.sub(r'[\s\-_]+', ' ', project_name).strip()
    keywords = [k.strip() for k in cleaned_name.split() if len(k.strip()) > 1]
    
    # الكلمات التي نفضل تجاهلها إذا كان هناك كلمات أخرى أكثر تحديداً
    # تم إزالة (الرياض، القطاع، الأوسط) لأنها كلمات مفصلية للتمييز بين المشاريع
    generic_words = ['مشروع', 'إصلاح', 'أعمال', 'بناء', 'عمليات', 'منطقة', 'بلدية', 'نظام']
    important_keywords = [k for k in keywords if k not in generic_words]
    
    # إذا كان الاسم يتكون فقط من كلمات عامة، نستخدمها جميعاً، وإلا نستخدم الكلمات الهامة فقط
    search_keywords = important_keywords if important_keywords else keywords
    
    if not search_keywords:
        return {"$regex": re.escape(project_name), "$options": "i"}

    regex_parts = []
    for k in search_keywords:
        pattern = ""
        # جعل الـ (الـ) اختيارية في بداية الكلمة
        if k.startswith('ال'):
            pattern += '(ال)?'
            k_no_al = k[2:]
        else:
            pattern += '(ال)?'
            k_no_al = k
            
        for char in k_no_al:
            if char in 'اأإآ': pattern += '[اأإآ]'
            elif char in 'هة': pattern += '[هة]'
            elif char in 'يى': pattern += '[يى]'
            else: pattern += re.escape(char)
        regex_parts.append(pattern)
    
    # الربط بين الكلمات بـ .* بدلاً من [\s\-_]* لضمان عدم فشل البحث إذا حذفت كلمة مثل "إصلاح" وكانت موجودة في النص الأصلي
    # وإضافة التطابق الحرفي الدقيق كخيار أول في الـ Regex
    exact_pattern = re.escape(project_name)
    norm_pattern = re.escape(normalize_arabic(project_name))
    flexible_pattern = r"(مشروع\s+)?.*" + r".*".join(regex_parts) + r".*"
    
    full_regex = f"^({exact_pattern})$|^({norm_pattern})$|^({flexible_pattern})$"
    return {"$regex": full_regex, "$options": "i"}


def get_flexible_in_query(items: List[str], field_name: str = "project") -> dict:
    """
    بناء استعلام Mongo مرن لقائمة من العناصر (المشاريع أو المحافظات)
    يدعم التنوع في الحروف والمسافات.
    """
    if not items:
        return {}
    
    # تحويل "الكل" أو "جميع المشاريع" إلى استعلام فارغ (لا يفلتر)
    all_keywords = ["الكل", "جميع المحافظات", "كل المحافظات", "جميع المشاريع", "كل المشاريع"]
    if any(item in all_keywords for item in items):
        return {}

    clauses = []
    for item in items:
        if not item or item in all_keywords:
            continue
        clean_item = str(item).strip()
        norm_item = normalize_arabic(clean_item)
        q = get_flexible_project_query(clean_item)
        clauses.append({field_name: q})
        clauses.append({field_name: clean_item})
        if norm_item != clean_item:
            clauses.append({field_name: norm_item})
            
    if not clauses:
        return {}
    if len(clauses) == 1:
        return clauses[0]
    return {"$or": clauses}


def get_loose_project_query(project_name: str) -> dict:
    """
    بناء استعلام Mongo مرن جداً للمشاريع يدعم المطابقة الجزئية
    للتعامل مع الفروقات بين اسم المشروع الكامل والاسم المختصر.
    """
    # التقسيم بناءً على الفواصل الشائعة مثل الشرطة أو الشرطة المائلة
    parts = re.split(r'[-/|()]', project_name)
    main_part = parts[0].strip() if parts else project_name
    
    # تقسيم الجزء الرئيسي إلى كلمات
    keywords = [k for k in re.split(r'\s+', main_part) if k]
    generic_words = ['مشروع', 'إصلاح', 'أعمال', 'بناء', 'عمليات', 'منطقة', 'بلدية', 'نظام']
    important_keywords = [k for k in keywords if k not in generic_words]
    search_keywords = important_keywords if important_keywords else keywords
    
    if not search_keywords:
        return {"$regex": re.escape(project_name), "$options": "i"}
        
    regex_parts = []
    for k in search_keywords:
        pattern = "(ال)?"
        k_no_al = k[2:] if k.startswith('ال') else k
        for char in k_no_al:
            if char in 'اأإآ': pattern += '[اأإآ]'
            elif char in 'هة': pattern += '[هة]'
            elif char in 'يى': pattern += '[يى]'
            else: pattern += re.escape(char)
        regex_parts.append(pattern)
        
    # البحث الجزئي غير المقيد ببدء ونهاية السلسلة
    full_regex = r".*" + r".*".join(regex_parts) + r".*"
    return {"$regex": full_regex, "$options": "i"}


def get_loose_in_query(items: List[str], field_name: str = "project") -> dict:
    """
    بناء استعلام Mongo مرن لقائمة من المشاريع للسماح بالمطابقة الجزئية.
    """
    if not items:
        return {}
    all_keywords = ["الكل", "جميع المحافظات", "كل المحافظات", "جميع المشاريع", "كل المشاريع"]
    if any(item in all_keywords for item in items):
        return {}
        
    clauses = []
    for item in items:
        if not item or item in all_keywords:
            continue
        clean_item = str(item).strip()
        norm_item = normalize_arabic(clean_item)
        q = get_loose_project_query(clean_item)
        clauses.append({field_name: q})
        clauses.append({field_name: clean_item})
        if norm_item != clean_item:
            clauses.append({field_name: norm_item})
            
    if not clauses:
        return {}
    if len(clauses) == 1:
        return clauses[0]
    return {"$or": clauses}


# قائمة جميع الصلاحيات المتاحة
ALL_PERMISSIONS = [
    {"key": "dashboard", "label": "لوحة التحكم وتحليل البيانات", "group": "عام"},
    {"key": "reports_view", "label": "عرض البلاغات", "group": "البلاغات"},
    {"key": "reports_add", "label": "إضافة بلاغ", "group": "البلاغات"},
    {"key": "reports_edit", "label": "تعديل بلاغ", "group": "البلاغات"},
    {"key": "reports_delete", "label": "حذف بلاغ", "group": "البلاغات"},
    {"key": "reports_review", "label": "مراجعة البلاغات", "group": "البلاغات"},
    {"key": "reports_import", "label": "استيراد بلاغات من Excel", "group": "البلاغات"},
    {"key": "reports_notifications", "label": "إشعارات البلاغات الجديدة", "group": "البلاغات"},
    {"key": "consultant_notes", "label": "ملاحظات الاستشاري", "group": "البلاغات"},
    {"key": "report_notes", "label": "ملاحظات البلاغات", "group": "البلاغات"},
    {"key": "owner_notes", "label": "ملاحظات المالك", "group": "البلاغات"},
    {"key": "water_connections", "label": "توصيلات المياه", "group": "التوصيلات"},
    {"key": "water_connections_import", "label": "استيراد توصيلات المياه من Excel", "group": "التوصيلات"},
    {"key": "sewage_connections", "label": "توصيلات الصرف الصحي", "group": "التوصيلات"},
    {"key": "sewage_connections_import", "label": "استيراد توصيلات الصرف من Excel", "group": "التوصيلات"},
    {"key": "contractors", "label": "المقاولين", "group": "الإدارة"},
    {"key": "projects", "label": "المشاريع", "group": "الإدارة"},
    {"key": "users_manage", "label": "إدارة المستخدمين", "group": "الإدارة"},
    {"key": "team", "label": "فريق العمل", "group": "الإدارة"},
    {"key": "project_settings", "label": "حالات البلاغ", "group": "الإدارة"},
    {"key": "cars", "label": "عرض سيارات المشاريع", "group": "السيارات"},
    {"key": "cars_manage", "label": "تسليم وإدارة السيارات", "group": "السيارات"},
    {"key": "fleet_maintenance", "label": "صيانة الأسطول", "group": "السيارات"},
    {"key": "invoices", "label": "الفواتير والعهدة", "group": "المالية"},
    {"key": "extracts", "label": "المستخلصات", "group": "المالية"},
    {"key": "view_all_invoices", "label": "عرض جميع الفواتير", "group": "المالية"},
    {"key": "review_invoices", "label": "اعتماد من المدير المباشر", "group": "المالية"},
    {"key": "review_invoices_3", "label": "اعتماد نهائي", "group": "المالية"},
    {"key": "view_extracts_all", "label": "عرض جميع المستخلصات (الوارد والمسجل)", "group": "المالية"},
    {"key": "employee_requests", "label": "طلبات الموظفين", "group": "الموارد البشرية"},
    {"key": "review_employee_requests", "label": "مراجعة واعتماد طلبات الموظفين", "group": "الموارد البشرية"},
    {"key": "view_all_employee_requests", "label": "عرض جميع طلبات الموظفين", "group": "الموارد البشرية"},
    {"key": "hr_management", "label": "شؤون الموظفين", "group": "الموارد البشرية"},
    {"key": "support_messages", "label": "رسائل الدعم", "group": "الدعم"},
    {"key": "settings", "label": "الإعدادات", "group": "النظام"},
    {"key": "trash", "label": "سجل المحذوفات", "group": "النظام"},
    {"key": "safety_reports", "label": "تقارير السلامة", "group": "التقارير"},
    {"key": "quality_reports", "label": "تقارير الجودة", "group": "التقارير"},
    {"key": "business_reports", "label": "تقارير الأعمال", "group": "التقارير"},
    {"key": "safety_reports_edit", "label": "تعديل تقرير السلامة", "group": "التقارير"},
    {"key": "safety_reports_delete", "label": "حذف تقرير السلامة", "group": "التقارير"},
    {"key": "quality_reports_edit", "label": "تعديل تقرير الجودة", "group": "التقارير"},
    {"key": "quality_reports_delete", "label": "حذف تقرير الجودة", "group": "التقارير"},
    {"key": "business_reports_edit", "label": "تعديل تقرير الأعمال", "group": "التقارير"},
    {"key": "business_reports_delete", "label": "حذف تقرير الأعمال", "group": "التقارير"},
    {"key": "business_reports_review", "label": "مراجعة تقارير الأعمال", "group": "التقارير"},
    {"key": "consultant_close", "label": "إغلاق الرخصة بواسطة الاستشاري", "group": "البلاغات"},
    {"key": "violations", "label": "المخالفات الميدانية", "group": "التقارير"},
    {"key": "work_permits", "label": "تصاريح العمل", "group": "التقارير"},
    {"key": "work_permits_edit", "label": "تعديل تصاريح العمل", "group": "التقارير"},
    {"key": "work_permits_delete", "label": "حذف تصاريح العمل", "group": "التقارير"},
    {"key": "view_governorate_data", "label": "رؤية إجمالي بيانات المحافظة (تجاوز منشئ البلاغ)", "group": "النظام"},
    {"key": "meetings", "label": "الاجتماعات", "group": "الإدارة"},
    {"key": "meetings_add", "label": "إضافة اجتماع", "group": "الإدارة"},
]


# الصلاحيات المرتبطة بمشروع (يمكن منحها لكل مشروع على حدة)
PROJECT_SCOPED_PERMISSIONS = {
    "reports_view", "reports_add", "reports_edit", "reports_delete",
    "reports_review", "reports_import", "reports_notifications", "consultant_notes", "report_notes", "owner_notes",
    "water_connections", "water_connections_import",
    "sewage_connections", "sewage_connections_import",
    "invoices", "review_invoices", "review_invoices_3", "view_all_invoices",
    "extracts", "view_extracts_all",
    "employee_requests", "review_employee_requests", "view_all_employee_requests",
    "contractors", "projects", "users_manage", "team", "project_settings",
    "cars", "cars_manage", "fleet_maintenance", "hr_management",
    "dashboard", "trash", "settings", "support_messages",
    "safety_reports", "quality_reports", "business_reports", "safety_reports_edit", "safety_reports_delete", "quality_reports_edit", "quality_reports_delete", "business_reports_edit", "business_reports_delete", "business_reports_review", "consultant_close",
    "violations", "work_permits", "work_permits_edit", "work_permits_delete", "meetings", "meetings_add"
}


def get_user_permissions_for_project(user_doc_or_obj, project: Optional[str]) -> set:
    """
    يعيد مجموعة الصلاحيات المرتبطة بمشروع المتاحة للمستخدم في سياق مشروع محدد.
    - إذا كان project_permissions[project] محدد (غير فارغ): يستخدمها حصرياً (تتجاوز العامة)
    - إذا لم يكن محدداً: يستخدم الصلاحيات العامة (backward compat)
    """
    if isinstance(user_doc_or_obj, dict):
        global_perms = user_doc_or_obj.get("permissions") or []
        pp = user_doc_or_obj.get("project_permissions") or {}
    else:
        global_perms = getattr(user_doc_or_obj, "permissions", None) or []
        pp = getattr(user_doc_or_obj, "project_permissions", None) or {}
    
    if project and isinstance(pp, dict):
        project_specific = pp.get(project) or []
        if project_specific:
            # المشروع يملك صلاحيات مخصصة → استخدمها حصرياً (override) للمرتبطة بمشروع
            # لكن ندمج مع الصلاحيات العامة غير المرتبطة بمشروع
            non_scoped_global = [p for p in global_perms if p not in PROJECT_SCOPED_PERMISSIONS]
            return set(non_scoped_global) | set(project_specific)
    
    # fallback: الصلاحيات العامة كاملة
    return set(global_perms)


def has_project_permission(user_doc_or_obj, project: Optional[str], perm_key: str) -> bool:
    """
    يفحص إذا كان المستخدم يملك صلاحية محددة لمشروع محدد.
    - Admin: دائماً True
    - الصلاحيات غير المرتبطة بمشروع: تُفحص من القائمة العامة فقط
    - الصلاحيات المرتبطة بمشروع:
      • إذا كان للمشروع project_permissions محدد (غير فارغ): تُسخدم حصرياً
      • غير ذلك: القائمة العامة
    """
    role = user_doc_or_obj.get("role") if isinstance(user_doc_or_obj, dict) else getattr(user_doc_or_obj, "role", None)
    if role == "admin":
        return True
    
    connection_fields = {
        "connections_full_form", "connections_show_phone", "connections_show_request_number",
        "connections_show_restriction_number", "connections_show_account_number",
        "connections_show_ccb_number", "connections_show_ccp_number", "connections_show_dates",
        "connections_show_measurements", "connections_show_meter", "connections_show_location"
    }
    if perm_key in connection_fields:
        return (
            has_project_permission(user_doc_or_obj, project, "water_connections") or
            has_project_permission(user_doc_or_obj, project, "sewage_connections")
        )
    
    global_perms = user_doc_or_obj.get("permissions") if isinstance(user_doc_or_obj, dict) else getattr(user_doc_or_obj, "permissions", None)
    global_perms = global_perms or []
    
    # الصلاحيات غير المرتبطة بمشروع
    if perm_key not in PROJECT_SCOPED_PERMISSIONS:
        return perm_key in global_perms
    
    # الصلاحيات المرتبطة بمشروع
    pp = user_doc_or_obj.get("project_permissions") if isinstance(user_doc_or_obj, dict) else getattr(user_doc_or_obj, "project_permissions", None)
    pp = pp or {}
    if project and isinstance(pp, dict):
        project_specific = pp.get(project) or []
        if project_specific:
            # override exclusive
            return perm_key in project_specific
    
    # fallback
    return perm_key in global_perms


def user_has_any_project_permission(user_doc_or_obj, perm_key: str) -> bool:
    """يعيد True إذا كان المستخدم يملك هذه الصلاحية في أي مشروع من مشاريعه أو عامة."""
    role = user_doc_or_obj.get("role") if isinstance(user_doc_or_obj, dict) else getattr(user_doc_or_obj, "role", None)
    if role == "admin":
        return True
    
    connection_fields = {
        "connections_full_form", "connections_show_phone", "connections_show_request_number",
        "connections_show_restriction_number", "connections_show_account_number",
        "connections_show_ccb_number", "connections_show_ccp_number", "connections_show_dates",
        "connections_show_measurements", "connections_show_meter", "connections_show_location"
    }
    if perm_key in connection_fields:
        return (
            user_has_any_project_permission(user_doc_or_obj, "water_connections") or
            user_has_any_project_permission(user_doc_or_obj, "sewage_connections")
        )
    
    global_perms = user_doc_or_obj.get("permissions") if isinstance(user_doc_or_obj, dict) else getattr(user_doc_or_obj, "permissions", None)
    if perm_key in (global_perms or []):
        return True
    
    pp = user_doc_or_obj.get("project_permissions") if isinstance(user_doc_or_obj, dict) else getattr(user_doc_or_obj, "project_permissions", None)
    if isinstance(pp, dict):
        for perms in pp.values():
            if perm_key in (perms or []):
                return True
    return False


def get_projects_with_permission(user_doc_or_obj, perm_key: str) -> List[str]:
    """يعيد قائمة المشاريع التي يملك فيها المستخدم الصلاحية المحددة (مع منطق override الخاص بكل مشروع)."""
    role = user_doc_or_obj.get("role") if isinstance(user_doc_or_obj, dict) else getattr(user_doc_or_obj, "role", None)
    if isinstance(user_doc_or_obj, dict):
        user_projects = user_doc_or_obj.get("projects") or []
        global_perms = user_doc_or_obj.get("permissions") or []
        pp = user_doc_or_obj.get("project_permissions") or {}
    else:
        user_projects = getattr(user_doc_or_obj, "projects", None) or []
        global_perms = getattr(user_doc_or_obj, "permissions", None) or []
        pp = getattr(user_doc_or_obj, "project_permissions", None) or {}
    
    if role == "admin":
        return list(user_projects)
    
    connection_fields = {
        "connections_full_form", "connections_show_phone", "connections_show_request_number",
        "connections_show_restriction_number", "connections_show_account_number",
        "connections_show_ccb_number", "connections_show_ccp_number", "connections_show_dates",
        "connections_show_measurements", "connections_show_meter", "connections_show_location"
    }
    if perm_key in connection_fields:
        water_projects = get_projects_with_permission(user_doc_or_obj, "water_connections")
        sewage_projects = get_projects_with_permission(user_doc_or_obj, "sewage_connections")
        return list(set(water_projects) | set(sewage_projects))
    
    # الصلاحيات غير المرتبطة بمشروع
    if perm_key not in PROJECT_SCOPED_PERMISSIONS:
        return list(user_projects) if perm_key in (global_perms or []) else []
    
    result = []
    for proj in user_projects:
        proj_specific = (pp or {}).get(proj) or []
        if proj_specific:
            # override exclusive
            if perm_key in proj_specific:
                result.append(proj)
        else:
            # fallback to global
            if perm_key in (global_perms or []):
                result.append(proj)
    return result


# ============ نماذج توصيلات المياه ============
class WaterConnectionCreate(BaseModel):
    project: str
    governorate: Optional[str] = ""
    contractor: str
    account_number: Optional[str] = ""
    request_number: Optional[str] = ""
    restriction_number: Optional[str] = ""
    ccb_report_number: Optional[str] = ""
    customer_name: Optional[str] = ""
    phone_number: Optional[str] = ""
    area: Optional[str] = ""
    work_order_date: Optional[str] = ""
    diameter: Optional[str] = ""
    connection_length: Optional[str] = ""
    notes: Optional[str] = ""
    latitude: Optional[str] = ""
    longitude: Optional[str] = ""
    commissioning_date: Optional[str] = ""
    permit_number: Optional[str] = ""
    publication_date: Optional[str] = ""
    issue_date: Optional[str] = ""
    expected_execution_date: Optional[str] = ""
    connection_type: Optional[str] = ""  # نوع التوصيلة: توصيله مفرده، شجريه، صندوق وعداد...
    connections_count: Optional[str] = ""
    connection_length_without_extra: Optional[str] = ""
    connections_length_without_main: Optional[str] = ""
    network_diameter_63: Optional[str] = ""
    network_line_length: Optional[str] = ""
    network_diameter_16: Optional[str] = ""
    meter_number: Optional[str] = ""
    meter_type: Optional[str] = ""
    meter_removal_installation: Optional[str] = ""  # إزالة عداد وتركيب عداد جديد
    execution_date: Optional[str] = ""
    system_closing_date: Optional[str] = ""
    request_status: Optional[str] = "جديد"
    cancellation_date: Optional[str] = ""
    cancellation_reason: Optional[str] = ""
    images: Optional[List[str]] = []  # صور التوصيلة


# ============ نماذج توصيلات الصرف الصحي ============
class SewageConnectionCreate(BaseModel):
    project: str
    governorate: Optional[str] = ""
    contractors: List[str] = []  # مقاولين متعددين
    request_number: Optional[str] = ""
    account_number: Optional[str] = ""
    restriction_number: Optional[str] = ""
    ccb_report_number: Optional[str] = ""
    customer_name: Optional[str] = ""
    customer_number: Optional[str] = ""
    area: Optional[str] = ""
    work_order_date: Optional[str] = ""
    diameter: Optional[str] = ""
    meter_number: Optional[str] = ""
    latitude: Optional[str] = ""
    longitude: Optional[str] = ""
    commissioning_date: Optional[str] = ""
    permit: Optional[str] = ""
    publication_date: Optional[str] = ""
    issue_date: Optional[str] = ""
    expected_execution_date: Optional[str] = ""
    connection_type: Optional[str] = ""
    ventilation_installation: Optional[bool] = False
    inspection_room_installation: Optional[bool] = False
    back_drop: Optional[bool] = False
    actual_length: Optional[str] = ""
    network_line_length: Optional[str] = ""
    cesspool_breaking: Optional[bool] = False
    attack: Optional[bool] = False
    connection_removal: Optional[bool] = False
    execution_date: Optional[str] = ""
    system_closing_date: Optional[str] = ""
    request_status: Optional[str] = "جديد"
    cancellation_reason: Optional[str] = ""
    phone_number: Optional[str] = ""
    notes: Optional[str] = ""
    images: Optional[List[str]] = []  # صور التوصيلة


# ============ نماذج السيارات ============
class CarCreate(BaseModel):
    project: str
    assigned_user_id: str
    assigned_user_name: str
    car_type: str
    plate_number: str
    model: str
    authorization_start: str
    authorization_end: str
    color: str
    notes: Optional[str] = ""
    kilometers: Optional[str] = ""

class CarUpdate(BaseModel):
    project: Optional[str] = None
    assigned_user_id: Optional[str] = None
    assigned_user_name: Optional[str] = None
    car_type: Optional[str] = None
    plate_number: Optional[str] = None
    model: Optional[str] = None
    authorization_start: Optional[str] = None
    authorization_end: Optional[str] = None
    color: Optional[str] = None
    notes: Optional[str] = None
    kilometers: Optional[str] = None


# ============= نماذج سجل السيارات والصيانة =============
class FleetCarCreate(BaseModel):
    """سيارة في أسطول الشركة"""
    plate_number: str  # رقم اللوحة
    car_type: str  # نوع السيارة
    model: str  # الموديل
    year: Optional[str] = ""  # سنة الصنع
    color: Optional[str] = ""  # اللون
    owner_name: str  # اسم المالك
    company: Optional[str] = ""  # الشركة التابع لها
    project_name: Optional[str] = ""  # اسم المشروع
    current_user_name: Optional[str] = ""  # اسم المستخدم الحالي للسيارة
    registration_start: Optional[str] = ""  # بداية الاستمارة
    registration_end: Optional[str] = ""  # نهاية الاستمارة
    inspection_start: Optional[str] = ""  # بداية الفحص الدوري
    inspection_end: Optional[str] = ""  # نهاية الفحص الدوري
    authorization_start: Optional[str] = ""  # بداية التفويض
    authorization_end: Optional[str] = ""  # نهاية التفويض
    notes: Optional[str] = ""

class FleetCarUpdate(BaseModel):
    plate_number: Optional[str] = None
    car_type: Optional[str] = None
    model: Optional[str] = None
    year: Optional[str] = None
    color: Optional[str] = None
    owner_name: Optional[str] = None
    company: Optional[str] = None
    project_name: Optional[str] = None
    current_user_name: Optional[str] = None
    registration_start: Optional[str] = None
    registration_end: Optional[str] = None
    inspection_start: Optional[str] = None
    inspection_end: Optional[str] = None
    authorization_start: Optional[str] = None
    authorization_end: Optional[str] = None
    notes: Optional[str] = None

class CarUserHistoryCreate(BaseModel):
    """سجل تسليم السيارة لمستخدم"""
    fleet_car_id: str  # معرف السيارة
    user_name: str  # اسم المستخدم
    assigned_date: str  # تاريخ التسليم
    returned_date: Optional[str] = None  # تاريخ الإرجاع (فارغ = مستلم حالياً)
    notes: Optional[str] = ""

class MaintenanceRecordCreate(BaseModel):
    """سجل صيانة لسيارة"""
    fleet_car_id: str  # معرف السيارة
    maintenance_type: str  # نوع الصيانة (صيانة دورية، إصلاح، صدمة، إطارات، زيت، إلخ)
    description: str  # وصف الصيانة
    cost: Optional[float] = 0  # التكلفة
    workshop: Optional[str] = ""  # ورشة الصيانة
    date: str  # تاريخ الصيانة
    notes: Optional[str] = ""


class UserCreate(BaseModel):
    username: str
    email: Optional[EmailStr] = None
    full_name: str
    title: Optional[str] = None  # اللقب مثل "المهندس /"
    password: str
    role: str = "user"
    governorates: List[str] = []
    projects: List[str] = []
    permissions: List[str] = []  # صلاحيات المستخدم
    allowed_chat_users: List[str] = []  # صلاحيات التحدث الاستثنائية

class UserLogin(BaseModel):
    username: str
    password: str


class UserResponse(BaseModel):
    id: str
    username: str
    email: Optional[str] = None
    full_name: str
    title: Optional[str] = None
    profile_picture: Optional[str] = None  # الصورة الشخصية (base64 أو URL)
    role: str
    governorates: List[str]
    projects: List[str]
    created_by: Optional[str] = None
    can_create_subusers: bool = True
    has_sub_users: bool = False  # هل لديه فعلاً موظفين تحت إدارته (يحدد إن كان مديراً حقيقياً)
    permissions: List[str] = []  # صلاحيات المستخدم (عامة)
    allowed_chat_users: List[str] = []  # صلاحيات التحدث الاستثنائية
    connection_permissions: Optional[dict] = None  # صلاحيات التوصيلات حسب المشروع (legacy)
    project_permissions: Optional[Dict[str, List[str]]] = None  # صلاحيات لكل مشروع
    personal_theme: Optional[str] = None  # الثيم الشخصي للمستخدم
    is_active: bool
    created_at: datetime


class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse


class ProfileUpdate(BaseModel):
    full_name: Optional[str] = None
    current_password: Optional[str] = None
    new_password: Optional[str] = None


class SupportMessage(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid4()))
    name: str
    email: str
    message: str
    username: Optional[str] = None  # اسم المستخدم إذا كان مسجل دخول
    status: str = "جديدة"  # جديدة، قيد المعالجة، تم الحل
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    resolved_by: Optional[str] = None  # من قام بحل المشكلة
    resolved_at: Optional[datetime] = None


class SupportMessageCreate(BaseModel):
    name: str
    email: str
    message: str


class SupportMessageResponse(BaseModel):
    id: str
    name: str
    email: str
    message: str
    username: Optional[str] = None
    status: str
    created_at: datetime
    resolved_by: Optional[str] = None
    resolved_at: Optional[datetime] = None


# ============= نموذج فواتير العهدة =============
class Invoice(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid4()))
    invoice_number: str  # رقم الفاتورة
    amount: float  # المبلغ
    description: str  # الوصف
    image: Optional[str] = None  # صورة الفاتورة (base64)
    uploaded_by: str  # ID المستخدم الذي رفع الفاتورة
    uploaded_by_name: str  # اسم المستخدم
    project: str  # المشروع
    governorate: Optional[str] = None  # المحافظة
    status: str = "pending"  # pending, approved_by_manager, approved_by_admin, rejected
    manager_id: Optional[str] = None  # ID المدير المباشر (Level 2)
    manager_name: Optional[str] = None
    reviewed_by_manager: Optional[str] = None  # من راجعها من المستوى 2
    reviewed_by_manager_at: Optional[str] = None
    manager_notes: Optional[str] = None
    reviewed_by_admin: Optional[str] = None  # من راجعها من الأدمن
    reviewed_by_admin_at: Optional[str] = None
    admin_notes: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class InvoiceCreate(BaseModel):
    invoice_number: str
    amount: float
    description: str
    images: Optional[List[str]] = []  # صور متعددة
    image: Optional[str] = None  # للتوافق مع الكود القديم
    project: str
    governorate: Optional[str] = None
    invoice_date: Optional[str] = None  # تاريخ الفاتورة


class Report(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    report_number: str
    report_date: Optional[datetime] = None
    license_number: str
    report_type: str
    status: str
    governorate: str
    project: str  # المشروع المرتبط بالبلاغ
    depth_meters: float
    diameter_mm: float
    contractor: str
    latitude: Optional[str] = None  # خط العرض
    longitude: Optional[str] = None  # خط الطول
    asphalt_license_issued: bool = False  # هل تم إصدار رخصة الأسفلت
    notes: Optional[str] = None  # ملاحظات على البلاغ
    images: List[str] = []
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # تاريخ استلام البلاغ (قد يكون قديم)
    start_date: Optional[datetime] = None  # تاريخ مباشرة البلاغ
    added_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # تاريخ إضافة البلاغ للنظام (فعلياً)
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    closed_at: Optional[datetime] = None
    is_deleted: bool = False
    deleted_at: Optional[datetime] = None
    deleted_by: Optional[str] = None  # من قام بالحذف
    permanently_deleted: bool = False
    permanently_deleted_at: Optional[datetime] = None
    permanently_deleted_by: Optional[str] = None
    review_status: str = "بانتظار المراجعة"  # بانتظار المراجعة / تمت المراجعة
    reviewed_by: Optional[str] = None  # من قام بالمراجعة
    reviewed_at: Optional[datetime] = None  # تاريخ المراجعة


# ===== حالات البلاغ =====
class ReportStatus(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # اسم الحالة
    project: str  # المشروع المرتبط
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ReportStatusCreate(BaseModel):
    name: str
    project: str

# ===== أنواع البلاغات =====
class ReportType(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # اسم النوع (ترابي، بلاط، أسفلت)
    project: str  # المشروع المرتبط
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ReportTypeCreate(BaseModel):
    name: str
    project: str

# ===== مسميات البطاقات =====
class CardItem(BaseModel):
    """بطاقة واحدة"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    key: str  # مفتاح فريد للبطاقة (مثل: total, fixed, asphalt)
    label: str  # العنوان المعروض

class ProjectCards(BaseModel):
    """بطاقات المشروع"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project: str
    cards: List[CardItem] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CardItemCreate(BaseModel):
    label: str  # العنوان فقط، المفتاح يُنشأ تلقائياً

class CardItemUpdate(BaseModel):
    label: str


class Message(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sender_id: str
    receiver_id: str
    message: Optional[str] = None
    file_url: Optional[str] = None  # للصور/الصوت
    file_type: Optional[str] = None  # image/audio
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    edited_at: Optional[datetime] = None
    is_edited: bool = False
    is_delivered: bool = False
    is_read: bool = False
    delivered_at: Optional[datetime] = None
    read_at: Optional[datetime] = None


class TeamMember(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: str
    position: str
    project: str
    email: Optional[str] = None
    profile_picture: Optional[str] = None
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TeamMemberResponse(BaseModel):
    id: str
    name: str
    phone: str
    position: str
    project: str
    email: Optional[str] = None
    profile_picture: Optional[str] = None
    created_by: Optional[str] = None
    created_at: datetime



class Contractor(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    project: str
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ContractorCreate(BaseModel):
    name: str
    project: str


class ContractorResponse(BaseModel):
    id: str
    name: str
    project: str
    created_by: Optional[str] = None
    created_at: datetime


class ReportCreate(BaseModel):
    report_number: str
    report_date: Optional[datetime] = None
    license_number: str
    report_type: str
    status: str
    governorate: str
    project: str
    depth_meters: float
    diameter_mm: float
    contractor: str
    latitude: Optional[str] = None
    longitude: Optional[str] = None
    asphalt_license_issued: bool = False
    wfm_closed: bool = False
    notes: Optional[str] = None
    start_date: Optional[datetime] = None  # تاريخ مباشرة البلاغ


class ReportUpdate(BaseModel):
    report_number: Optional[str] = None
    report_date: Optional[datetime] = None
    license_number: Optional[str] = None
    report_type: Optional[str] = None
    status: Optional[str] = None
    governorate: Optional[str] = None
    project: Optional[str] = None
    depth_meters: Optional[float] = None
    diameter_mm: Optional[float] = None
    contractor: Optional[str] = None
    notes: Optional[str] = None
    start_date: Optional[datetime] = None  # تاريخ مباشرة البلاغ


class ReportResponse(BaseModel):
    id: str
    report_number: str
    report_date: Optional[datetime] = None
    license_number: str
    report_type: str
    status: str
    governorate: str
    project: str
    depth_meters: float
    diameter_mm: float
    contractor: str
    latitude: Optional[str]
    longitude: Optional[str]
    asphalt_license_issued: bool
    wfm_closed: bool = False
    wfm_closed_by: Optional[str] = None
    notes: Optional[str] = None
    consultant_note: Optional[str] = None
    consultant_note_reply: Optional[str] = None
    consultant_note_replied_by: Optional[str] = None
    consultant_note_processed: bool = False
    consultant_note_date: Optional[str] = None
    consultant_note_processed_date: Optional[str] = None
    images: List[str] = []
    created_by: str
    created_by_name: Optional[str] = None
    created_at: datetime
    start_date: Optional[datetime] = None  # تاريخ مباشرة البلاغ
    updated_at: Optional[datetime] = None
    closed_at: Optional[datetime] = None
    is_deleted: bool = False
    deleted_at: Optional[datetime] = None
    deleted_by: Optional[str] = None
    deleted_by_name: Optional[str] = None
    review_status: str = "بانتظار المراجعة"  # بانتظار المراجعة / تمت المراجعة
    reviewed_by: Optional[str] = None
    reviewed_at: Optional[datetime] = None
    reviewed_by_name: Optional[str] = None


class Extract(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    # الحقول الأساسية الجديدة
    extract_number: Optional[str] = None  # رقم المستخلص
    invoice_number: Optional[str] = None  # رقم الفاتورة (حروف إنجليزية + أرقام + underscore)
    extract_date: Optional[str] = None  # تاريخ المستخلص
    work_unit: Optional[str] = None  # وحدة الأعمال
    po_number: Optional[str] = None  # رقم PO
    project: str  # المشروع
    
    # الحقول المالية
    amount: Optional[float] = None  # المبلغ (من مدير المشروع)
    actual_value: Optional[float] = None  # قيمة المستخلص الفعلية
    advance_deduction: Optional[float] = None  # خصم الدفعة المقدمة
    net_after_deduction: Optional[float] = None  # الصافي بعد خصم الدفعة (محسوب)
    tax: Optional[float] = None  # الضريبة
    penalties: Optional[float] = None  # الغرامات
    total_submitted: Optional[float] = None  # إجمالي المستخلص المقدم (محسوب)
    total_collected: Optional[float] = None  # إجمالي المبلغ المحصل
    difference: Optional[float] = None  # الفرق (محسوب)
    
    # حالة الصرف
    is_paid: bool = False  # تم الصرف
    collection_date: Optional[str] = None  # تاريخ تحصيل المستخلص
    
    # الحقول القديمة للتوافق
    month: Optional[int] = None  # 1-12
    year: Optional[int] = None  # 2025, 2026, etc.
    images: List[str] = []
    status: str = "pending"  # pending, approved, rejected
    created_by: str
    created_by_name: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_deleted: bool = False
    
    # حقول الصرف القديمة - للتوافق
    total_amount: Optional[float] = None
    paid_amount: Optional[float] = None
    remaining_amount: Optional[float] = None
    paid_at: Optional[datetime] = None
    paid_by: Optional[str] = None
    paid_by_name: Optional[str] = None
    payment_notes: Optional[str] = None


class ExtractResponse(BaseModel):
    id: str
    extract_number: Optional[str] = None
    invoice_number: Optional[str] = None
    extract_date: Optional[str] = None
    work_unit: Optional[str] = None
    po_number: Optional[str] = None
    project: str
    
    # الحقول المالية
    amount: Optional[float] = None  # المبلغ (من مدير المشروع)
    actual_value: Optional[float] = None
    advance_deduction: Optional[float] = None
    net_after_deduction: Optional[float] = None
    tax: Optional[float] = None
    penalties: Optional[float] = None
    total_submitted: Optional[float] = None
    total_collected: Optional[float] = None
    difference: Optional[float] = None
    
    is_paid: bool = False
    collection_date: Optional[str] = None
    
    month: Optional[int] = None
    year: Optional[int] = None
    images: List[str] = []
    status: str
    created_by: str
    created_by_name: Optional[str]
    created_at: datetime
    updated_at: datetime
    is_deleted: bool
    
    # حقول الصرف القديمة
    total_amount: Optional[float] = None
    paid_amount: Optional[float] = None
    remaining_amount: Optional[float] = None
    paid_at: Optional[datetime] = None
    paid_by: Optional[str] = None
    paid_by_name: Optional[str] = None
    payment_notes: Optional[str] = None


# ============= HELPER FUNCTIONS =============

def verify_password(plain_password, hashed_password):
    # Truncate password to 72 bytes to avoid bcrypt limitation
    if len(plain_password.encode('utf-8')) > 72:
        plain_password = plain_password.encode('utf-8')[:72].decode('utf-8', errors='ignore')
    return pwd_context.verify(plain_password, hashed_password)


def get_password_hash(password):
    # Truncate password to 72 bytes to avoid bcrypt limitation
    if len(password.encode('utf-8')) > 72:
        password = password.encode('utf-8')[:72].decode('utf-8', errors='ignore')
    return pwd_context.hash(password)


def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt


def compress_image_data(image_data: str, max_size_mb: float = 3.0) -> str:
    """
    دالة للتوافق مع الكود القديم - ترجع الصورة كما هي
    """
    try:
        return image_data
    except Exception as e:
        # print error خطأ في ضغط الصورة
        return image_data


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> User:
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user_doc is None:
        raise HTTPException(status_code=401, detail="User not found")
        
    
    if isinstance(user_doc.get('created_at'), str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    
    # Ensure projects field exists (for backward compatibility)
    if 'projects' not in user_doc:
        user_doc['projects'] = []
    # Ensure governorates field exists (for backward compatibility)
    if 'governorates' not in user_doc:
        user_doc['governorates'] = []
    
    # Handle legacy 'password' field name (should be 'hashed_password')
    if 'password' in user_doc and 'hashed_password' not in user_doc:
        user_doc['hashed_password'] = user_doc.pop('password')
    
    user = User(**user_doc)
    if not user.is_active:
        raise HTTPException(status_code=400, detail="Inactive user")
    
    return user


async def get_current_admin_user(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not enough permissions")
    return current_user


async def get_all_subordinate_user_ids(user_id: str, include_self: bool = True) -> List[str]:
    """
    الحصول على جميع المستخدمين الفرعيين بشكل هرمي
    
    Args:
        user_id: معرف المستخدم
        include_self: هل نضيف المستخدم نفسه في القائمة
    
    Returns:
        قائمة بمعرفات جميع المستخدمين الفرعيين
    """
    user_ids = [user_id] if include_self else []
    
    # جلب المستخدمين الذين تم إنشاؤهم بواسطة هذا المستخدم
    direct_subordinates = await db.users.find(
        {"created_by": user_id},
        {"_id": 0, "id": 1}
    ).to_list(1000)
    
    # إضافة المستخدمين المباشرين
    for sub in direct_subordinates:
        sub_id = sub['id']
        if sub_id not in user_ids:
            user_ids.append(sub_id)
            # البحث بشكل تكراري عن المستخدمين الفرعيين للمستخدمين الفرعيين
            nested_ids = await get_all_subordinate_user_ids(sub_id, include_self=False)
            for nested_id in nested_ids:
                if nested_id not in user_ids:
                    user_ids.append(nested_id)
    
    return user_ids


async def get_hierarchy_filter(current_user: User) -> dict:
    """
    بناء فلتر MongoDB للهيكلية الإدارية بشكل تكراري (Recursive).
    يضمن أن كل مستخدم يرى بياناته وبيانات جميع التابعين له في جميع المستويات الأدنى.
    كما يسمح لجميع المستخدمين برؤية البلاغات المرفوعة من قبل الأدمن (بيت الخبرة) في مناطقهم.
    """
    if current_user.role == "admin":
        return {}
        
    user_projects = current_user.projects if hasattr(current_user, 'projects') and current_user.projects else []
    user_governorates = current_user.governorates if hasattr(current_user, 'governorates') and current_user.governorates else []
    has_all_govs = any(g in ["الكل", "جميع المحافظات", "كل المحافظات"] for g in user_governorates) or len(user_governorates) >= 8
    has_all_projects = len(user_projects) >= 3
    
    if has_all_govs and has_all_projects:
        return {}  # يعامل كمدير عام يرى جميع البلاغات
        
    # بناء فلتر المحافظات
    gov_filter = {}
    if not has_all_govs and len(user_governorates) > 0:
        gov_patterns = [normalize_arabic_regex(g) for g in user_governorates]
        gov_filter = {'governorate': {'$regex': f"({'|'.join(gov_patterns)})", '$options': 'i'}}
        
    # المستوى الثاني (مدير منطقة/محافظة) يرى جميع بلاغات المحافظات والمشاريع المسندة إليه
    # وكذلك المستخدمين الاستثنائيين الذين لديهم صلاحية رؤية إجمالي المحافظة
    can_view_all = has_project_permission(current_user, None, "view_governorate_data")
    if getattr(current_user, 'can_create_subusers', False) or can_view_all:
        return gov_filter
        
    # المستوى الثالث: يرى بلاغاته فقط، نجلب معرفاته
    all_subordinate_ids = await get_all_subordinate_user_ids(current_user.id, include_self=True)
    
    # تجميع كل المعرفات الممكنة (IDs و Usernames) للتوافق مع طرق التخزين المختلفة
    all_identifiers = set(all_subordinate_ids)
    if hasattr(current_user, 'username') and current_user.username:
        all_identifiers.add(current_user.username)
    if hasattr(current_user, 'id') and current_user.id:
        all_identifiers.add(current_user.id)
    
    # جلب أسماء المستخدمين (usernames) لجميع التابعين لضمان الشمولية في حقل created_by
    if all_subordinate_ids:
        sub_users_cursor = db.users.find({"id": {"$in": all_subordinate_ids}}, {"username": 1})
        async for u in sub_users_cursor:
            if u.get("username"):
                all_identifiers.add(u["username"])
                
    hierarchy_filter = {"created_by": {"$in": list(all_identifiers)}}
    
    if gov_filter:
        return {"$and": [gov_filter, hierarchy_filter]}
    return hierarchy_filter


# ============= نقاط نهاية رفع وعرض الصور (Object Storage) =============

@api_router.post("/uploads/image")
async def upload_image_endpoint(
    file: UploadFile = File(...),
    category: str = Form("general"),
    current_user: User = Depends(get_current_user)
):
    """رفع صورة إلى Object Storage وإرجاع URL للاستخدام"""
    content = await file.read()
    # ضغط تلقائي إلى 300KB قبل الرفع
    try:
        loop = asyncio.get_event_loop()
        content = await loop.run_in_executor(thread_pool, compress_image, content)
    except Exception:
        pass
    ext = _guess_ext(file.filename, file.content_type)
    path = _upload_image(content, category=category, ext=ext, content_type=f"image/{ext if ext != 'jpg' else 'jpeg'}")
    return {"url": path, "path": path}


@api_router.get("/images/{path:path}")
async def download_image(
    path: str,
    auth: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None)
):
    """تحميل صورة من التخزين - يدعم Bearer token أو ?auth= لـ <img src>"""
    token = None
    if authorization and authorization.startswith("Bearer "):
        token = authorization.split(" ", 1)[1]
    elif auth:
        token = auth
    if not token:
        raise HTTPException(status_code=401, detail="Unauthorized")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if not payload.get("sub"):
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    try:
        data, content_type = _get_object(path)
        return Response(content=data, media_type=content_type)
    except Exception:
        raise HTTPException(status_code=404, detail="Image not found")


def _store_image_bytes(content: bytes, category: str = "reports", filename: Optional[str] = None, content_type: Optional[str] = None) -> str:
    """مساعد: ارفع bytes الصورة وارجع URL داخل التطبيق"""
    ext = _guess_ext(filename, content_type)
    ct = f"image/{ext if ext != 'jpg' else 'jpeg'}"
    path = _upload_image(content, category=category, ext=ext, content_type=ct)
    return path


async def _resolve_logo_path(logo_url: Optional[str], default_filename: str = "bayt-alkhibra-logo.png") -> Optional[str]:
    """
    تحويل رابط شعار من إعدادات المنصة إلى مسار فعلي على القرص.
    يدعم:
      - مسار محلي قديم (/bayt-alkhibra-logo.png) → fonts folder
      - مسار /api/images/xxx → تحميل من Object Storage إلى /tmp
      - مسار قديم /api/storage/files/xxx → تحميل من Object Storage إلى /tmp
      - رابط HTTP كامل → تحميل
      - None / فارغ → الشعار الافتراضي
    """
    fonts_dir = os.path.join(os.path.dirname(__file__), 'fonts')
    default_path = os.path.join(fonts_dir, default_filename)
    
    if not logo_url:
        return default_path if os.path.exists(default_path) else None
    
    logo_url = str(logo_url).strip()
    
    # Strip full URL prefix إذا كان من نفس التطبيق
    backend_url = os.environ.get("REACT_APP_BACKEND_URL", "")
    if backend_url and logo_url.startswith(backend_url):
        logo_url = logo_url[len(backend_url):]
    elif logo_url.startswith("http") and "/api/storage/files/" in logo_url:
        # استخراج path بعد /api/storage/files/
        idx = logo_url.find("/api/storage/files/")
        logo_url = logo_url[idx:]
    elif logo_url.startswith("http") and "/api/images/" in logo_url:
        idx = logo_url.find("/api/images/")
        logo_url = logo_url[idx:]
    
    # مسار محلي قديم (مثل /bayt-alkhibra-logo.png أو /nwc-logo.png)
    if logo_url.startswith("/") and not logo_url.startswith("/api/"):
        local_name = logo_url.lstrip("/")
        local_path = os.path.join(fonts_dir, local_name)
        if os.path.exists(local_path):
            return local_path
        return default_path if os.path.exists(default_path) else None
    
    # شعار من Object Storage (تنسيقان: /api/images/ و /api/storage/files/)
    storage_prefix = None
    if logo_url.startswith("/api/images/"):
        storage_prefix = "/api/images/"
    elif logo_url.startswith("/api/storage/files/"):
        storage_prefix = "/api/storage/files/"
    
    if storage_prefix:
        try:
            storage_path = logo_url.replace(storage_prefix, "", 1)
            # إزالة query string إن وجد
            storage_path = storage_path.split("?")[0]
            data, _ct = _get_object(storage_path)
            cache_path = f"/tmp/_logo_{abs(hash(storage_path))}.png"
            with open(cache_path, "wb") as f:
                f.write(data)
            return cache_path
        except Exception as e:
            logger.warning(f"Failed to fetch logo from storage {logo_url}: {e}")
            return default_path if os.path.exists(default_path) else None
    
    # رابط HTTP/HTTPS خارجي كامل
    if logo_url.startswith("http"):
        try:
            import requests as _req
            resp = _req.get(logo_url, timeout=5)
            if resp.status_code == 200:
                cache_path = f"/tmp/_logo_{abs(hash(logo_url))}.png"
                with open(cache_path, "wb") as f:
                    f.write(resp.content)
                return cache_path
        except Exception as e:
            logger.warning(f"Failed to download logo {logo_url}: {e}")
        return default_path if os.path.exists(default_path) else None
    
    return default_path if os.path.exists(default_path) else None




# ربط المشاريع بالمحافظات (من الفرونت إند)
# ربط المشاريع بالمحافظات (من قاعدة البيانات)
async def get_total_governorates_count_for_projects(projects: List[str]) -> int:
    """
    حساب إجمالي عدد المحافظات في المشاريع المحددة من قاعدة البيانات
    
    Args:
        projects: قائمة المشاريع
    
    Returns:
        إجمالي عدد المحافظات (بدون تكرار)
    """
    query = {}
    if projects and len(projects) > 0:
        query["project"] = {"$in": projects}
    
    # جلب المحافظات المخصصة
    project_govs = await db.project_governorates.find(query, {"_id": 0, "name": 1}).to_list(2000)
    governorates = [g['name'] for g in project_govs if g.get('name')]
    
    # جلب المحافظات المحذوفة لاستثنائها
    deleted_govs = await db.deleted_governorates.find(query, {"_id": 0, "name": 1}).to_list(2000)
    deleted_names = set(d['name'] for d in deleted_govs if d.get('name'))
    
    # تصفية المحافظات
    active_govs = [g for g in governorates if g not in deleted_names]
    
    return len(set(active_govs))  # إزالة التكرار


# ============= AUTH ROUTES =============

@api_router.post("/auth/register", response_model=UserResponse)
async def register(user_data: UserCreate, current_user: User = Depends(get_current_user)):
    existing_user = await db.users.find_one({"username": user_data.username})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # التحقق من البريد الإلكتروني فقط إذا تم إدخاله
    if user_data.email:
        existing_email = await db.users.find_one({"email": user_data.email})
        if existing_email:
            raise HTTPException(status_code=400, detail="Email already exists")
    
    hashed_password = get_password_hash(user_data.password)
    user_dict = user_data.model_dump()
    user_dict.pop('password')
    
    # تحديد إذا كان المستخدم الجديد يمكنه إنشاء مستخدمين فرعيين
    # - إذا كان المُنشئ admin → يمكنه
    # - إذا كان المُنشئ user → لا يمكنه (منع التداخل)
    # تحديد إذا كان المستخدم الجديد يمكنه إنشاء مستخدمين فرعيين
    # Level 1 (admin) ينشئ Level 2 (can_create=True)
    # Level 2 (manager) ينشئ Level 3 (can_create=False)
    can_create = current_user.role == 'admin'
    
    user_obj = User(**user_dict, hashed_password=hashed_password, created_by=current_user.id, can_create_subusers=can_create)
    
    doc = user_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    
    # إضافة المشاريع الجديدة إلى جدول المشاريع تلقائياً
    if user_data.projects:
        for project_name in user_data.projects:
            existing_project = await db.projects.find_one({"name": project_name})
            if not existing_project:
                await db.projects.insert_one({
                    "id": str(uuid4()),
                    "name": project_name,
                    "description": project_name
                })
    
    return UserResponse(**user_obj.model_dump())


@api_router.post("/auth/login", response_model=Token)
async def login(login_data: UserLogin):
    user_doc = await db.users.find_one({"username": login_data.username}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Incorrect username or password")
    
    if isinstance(user_doc.get('created_at'), str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    
    # Ensure projects field exists (for backward compatibility)
    if 'projects' not in user_doc:
        user_doc['projects'] = []
    # Ensure governorates field exists (for backward compatibility)
    if 'governorates' not in user_doc:
        user_doc['governorates'] = []
    
    # Handle legacy 'password' field name (should be 'hashed_password')
    if 'password' in user_doc and 'hashed_password' not in user_doc:
        user_doc['hashed_password'] = user_doc.pop('password')
    
    user = User(**user_doc)
    
    if not verify_password(login_data.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Incorrect username or password")
    
    if not user.is_active:
        raise HTTPException(status_code=400, detail="Inactive user")
    
    access_token = create_access_token(data={"sub": user.id})
    
    await db.users.update_one({"id": user.id}, {"$set": {"current_session_token": access_token}})
    
    # حساب إذا كان المستخدم لديه فعلاً موظفين تحت إدارته (لتحديد Level 2 حقيقي)
    sub_user_count = await db.users.count_documents({"created_by": user.id})
    user_dict = user.model_dump()
    user_dict["has_sub_users"] = sub_user_count > 0
    
    return Token(
        access_token=access_token,
        token_type="bearer",
        user=UserResponse(**user_dict)
    )


@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: User = Depends(get_current_user)):
    sub_user_count = await db.users.count_documents({"created_by": current_user.id})
    user_dict = current_user.model_dump()
    user_dict["has_sub_users"] = sub_user_count > 0
    return UserResponse(**user_dict)


# ============= PASSWORD RESET =============

class PasswordResetRequest(BaseModel):
    username: str

class PasswordResetVerify(BaseModel):
    username: str
    code: str
    new_password: str

@api_router.post("/auth/forgot-password")
async def forgot_password(data: PasswordResetRequest):
    """إرسال كود التحقق إلى رسائل الدعم للمهندس محمود هارون"""
    # البحث عن المستخدم باسم المستخدم
    user = await db.users.find_one({"username": data.username}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="اسم المستخدم غير موجود")
    
    # إنشاء كود عشوائي من 6 أرقام
    code = str(random.randint(100000, 999999))
    
    # حفظ الكود مع وقت انتهاء (10 دقائق)
    password_reset_codes[data.username] = {
        "code": code,
        "expires": datetime.now(timezone.utc) + timedelta(minutes=10)
    }
    
    # إرسال الكود كرسالة دعم للمهندس محمود هارون
    support_message = {
        "id": str(uuid.uuid4()),
        "name": user.get('full_name', data.username),
        "email": f"reset-{data.username}@system.local",
        "subject": f"🔐 كود استعادة كلمة المرور - {data.username}",
        "message": f"""
طلب استعادة كلمة المرور

المستخدم: {data.username}
الاسم: {user.get('full_name', data.username)}
المشروع: {', '.join(user.get('projects', [])) or 'غير محدد'}

كود التحقق: {code}

⚠️ الكود صالح لمدة 10 دقائق فقط
        """.strip(),
        "status": "جديدة",
        "type": "password_reset",
        "request_username": data.username,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.support_messages.insert_one(support_message)
    logging.info(f"Password reset code created for user {data.username}")
    
    return {"message": "تم إرسال طلب استعادة كلمة المرور، يرجى التواصل مع الدعم الفني"}


@api_router.get("/auth/ticket-status/{username}")
async def get_ticket_status(username: str):
    """التحقق من حالة تذكرة استعادة كلمة المرور"""
    ticket = await db.support_messages.find_one(
        {"request_username": username, "type": "password_reset"},
        {"_id": 0}
    )
    if not ticket:
        return {"has_ticket": False}
    
    return {
        "has_ticket": True,
        "status": ticket.get("status", "جديدة"),
        "created_at": ticket.get("created_at"),
        "resolved": ticket.get("status") == "تم الحل"
    }


@api_router.post("/auth/reset-password")
async def reset_password(data: PasswordResetVerify):
    """التحقق من الكود وتغيير كلمة المرور"""
    # التحقق من وجود الكود
    if data.username not in password_reset_codes:
        raise HTTPException(status_code=400, detail="لم يتم طلب استعادة كلمة المرور لهذا المستخدم")
    
    reset_data = password_reset_codes[data.username]
    
    # التحقق من انتهاء الصلاحية
    if datetime.now(timezone.utc) > reset_data["expires"]:
        del password_reset_codes[data.username]
        raise HTTPException(status_code=400, detail="انتهت صلاحية كود التحقق")
    
    # التحقق من الكود
    if reset_data["code"] != data.code:
        raise HTTPException(status_code=400, detail="كود التحقق غير صحيح")
    
    # تحديث كلمة المرور
    hashed_password = get_password_hash(data.new_password)
    result = await db.users.update_one(
        {"username": data.username},
        {"$set": {"hashed_password": hashed_password}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="فشل في تحديث كلمة المرور")
    
    # حذف الكود بعد الاستخدام
    del password_reset_codes[data.username]
    
    return {"message": "تم تغيير كلمة المرور بنجاح"}


@api_router.put("/auth/update-profile")
async def update_profile(
    profile_data: ProfileUpdate,
    current_user: User = Depends(get_current_user)
):
    update_data = {}
    
    # تحديث الاسم
    if profile_data.full_name:
        update_data["full_name"] = profile_data.full_name
    
    # تحديث كلمة المرور
    if profile_data.current_password and profile_data.new_password:
        # التحقق من كلمة المرور الحالية
        if not verify_password(profile_data.current_password, current_user.hashed_password):
            raise HTTPException(status_code=400, detail="كلمة المرور الحالية غير صحيحة")
        
        # تحديث كلمة المرور
        update_data["hashed_password"] = get_password_hash(profile_data.new_password)
    
    if not update_data:
        raise HTTPException(status_code=400, detail="لا توجد بيانات للتحديث")
    
    # تحديث في قاعدة البيانات
    await db.users.update_one({"id": current_user.id}, {"$set": update_data})
    
    return {"message": "تم تحديث الملف الشخصي بنجاح"}


@api_router.post("/auth/upload-profile-picture")
async def upload_profile_picture(
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """رفع الصورة الشخصية للمستخدم (base64)"""
    try:
        picture = data.get('picture')
        
        if not picture:
            raise HTTPException(status_code=400, detail="لم يتم إرسال الصورة")
        
        # إذا كانت الصورة base64، قم برفعها لـ Cloudinary
        final_picture_url = picture
        if picture.startswith('data:image'):
            # استخراج الـ bytes من base64
            header, encoded = picture.split(",", 1)
            image_data = base64.b64decode(encoded)
            
            # رفع لـ Cloudinary
            final_picture_url = _upload_image(
                image_data, 
                category="profiles", 
                content_type="image/jpeg"
            )
        
        # تحديث الصورة في قاعدة البيانات
        await db.users.update_one(
            {"id": current_user.id},
            {"$set": {"profile_picture": final_picture_url}}
        )
        
        return {
            "message": "تم رفع الصورة الشخصية بنجاح",
            "profile_picture": final_picture_url
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"خطأ في رفع الصورة: {str(e)}")

@api_router.get("/auth/activity-logs")
async def get_activity_logs(current_user: User = Depends(get_current_user)):
    """جلب سجل النشاطات للمستخدم الحالي"""
    logs = await db.activity_logs.find({"user_id": current_user.id}).sort("timestamp", -1).to_list(length=50)
    
    if not logs:
        import datetime
        now = datetime.datetime.now(datetime.timezone.utc)
        mock_logs = [
            {
                "user_id": current_user.id,
                "username": current_user.username,
                "action": "تحديث إعدادات المظهر",
                "details": "تم تغيير الوضع الشخصي للمنصة بنجاح وتخصيص تفضيلات العرض.",
                "timestamp": (now - datetime.timedelta(minutes=15)).isoformat()
            },
            {
                "user_id": current_user.id,
                "username": current_user.username,
                "action": "مراجعة إعدادات الحساب والأمان",
                "details": "تم الدخول لصفحة الإعدادات وتأمين الجلسة الحالية بنجاح.",
                "timestamp": (now - datetime.timedelta(minutes=30)).isoformat()
            },
            {
                "user_id": current_user.id,
                "username": current_user.username,
                "action": "استعراض لوحة التحكم الرئيسية",
                "details": "تمت مراجعة مؤشرات الأداء ومتابعة إحصائيات المشاريع الفعّالة.",
                "timestamp": (now - datetime.timedelta(hours=1, minutes=20)).isoformat()
            },
            {
                "user_id": current_user.id,
                "username": current_user.username,
                "action": "تسجيل الدخول للنظام",
                "details": f"تم تسجيل الدخول بنجاح إلى المنصة للمستخدم {current_user.username} من جهاز Windows.",
                "timestamp": (now - datetime.timedelta(hours=2, minutes=5)).isoformat()
            }
        ]
        await db.activity_logs.insert_many(mock_logs)
        logs = await db.activity_logs.find({"user_id": current_user.id}).sort("timestamp", -1).to_list(length=50)
        
    for log in logs:
        if "_id" in log:
            log["_id"] = str(log["_id"])
        if isinstance(log.get("timestamp"), str):
            pass # Keep it as string
        elif hasattr(log.get("timestamp"), "isoformat"):
            log["timestamp"] = log["timestamp"].isoformat()
    return logs

@api_router.post("/auth/logout-others")
async def logout_others(current_user: User = Depends(get_current_user)):
    """تسجيل الخروج من جميع الأجهزة الأخرى"""
    access_token = create_access_token(data={"sub": current_user.id})
    await db.users.update_one({"id": current_user.id}, {"$set": {"current_session_token": access_token}})
    return {
        "message": "تم تسجيل الخروج من الأجهزة الأخرى بنجاح",
        "access_token": access_token
    }


# ===== إدارة المشاريع =====
@api_router.get("/projects")
async def get_all_projects(archived: bool = Query(False), current_user: User = Depends(get_current_user)):
    """جلب المشاريع حسب صلاحيات المستخدم"""
    filter_query = {"is_archived": True} if archived else {"is_archived": {"$ne": True}}
    
    # إذا كان المستخدم admin أو ليس لديه مشاريع محددة، أعد جميع المشاريع (بناءً على حالة الأرشيف)
    if current_user.role == "admin" or not current_user.projects:
        projects = await db.projects.find(filter_query, {"_id": 0}).to_list(100)
    else:
        # إذا كان المستخدم عادي ولديه مشاريع محددة، أعد فقط مشاريعه (مطابقة مرنة)
        query = get_loose_in_query(current_user.projects, "name")
        query.update(filter_query)
        projects = await db.projects.find(query, {"_id": 0}).to_list(100)
    return projects


@api_router.get("/projects/types")
async def get_project_types(current_user: User = Depends(get_current_user)):
    """تحديد نوع كل مشروع حسب صلاحيات المستخدمين (يأخذ في الاعتبار project_permissions لكل مشروع)"""
    users = await db.users.find({}, {"_id": 0, "projects": 1, "permissions": 1, "project_permissions": 1, "role": 1}).to_list(500)
    project_map = {}
    for u in users:
        global_perms = u.get("permissions", []) or []
        pp = u.get("project_permissions") or {}
        for proj in u.get("projects", []) or []:
            if proj not in project_map:
                project_map[proj] = {"connections": False, "reports": False}
            # نحدد الصلاحيات الفعّالة لكل مشروع بناءً على منطق override
            proj_specific = pp.get(proj) or []
            effective_perms = set(proj_specific) if proj_specific else set(global_perms)
            
            if "water_connections" in effective_perms or "sewage_connections" in effective_perms:
                project_map[proj]["connections"] = True
            if "reports_view" in effective_perms or "reports_add" in effective_perms:
                project_map[proj]["reports"] = True
    result = {}
    for proj, types in project_map.items():
        if types["connections"]:
            result[proj] = "connections"
        else:
            result[proj] = "reports"
    return result


@api_router.get("/dashboard/init-all")
async def dashboard_init_all(
    month: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """Batch endpoint: returns all projects + their stats in one call for fast dashboard loading"""
    # 1. Determine which projects to load
    if current_user.role == "admin" or not current_user.projects:
        projects_docs = await db.projects.find({"is_archived": {"$ne": True}}, {"_id": 0}).to_list(100)
        projects_list = [p.get("name") for p in projects_docs if p.get("name")]
    else:
        # بالنسبة للمستخدم العادي نجلب مشاريعه ونتأكد أنها غير مؤرشفة
        query = get_loose_in_query(current_user.projects, "name")
        query.update({"is_archived": {"$ne": True}})
        projects_docs = await db.projects.find(query, {"_id": 0}).to_list(100)
        projects_list = [p.get("name") for p in projects_docs if p.get("name")]

    allowed_projects = projects_list

    # 2. For each project fetch stats and project_cards in parallel
    async def fetch_project_data(project_name: str):
        try:
            # Build query filter
            q = {"is_deleted": {"$ne": True}, "project": get_flexible_project_query(project_name)}
            # Apply hierarchy filter for non-admin users
            if current_user.role != "admin":
                h_filter = await get_hierarchy_filter(current_user)
                q.update(h_filter)

            # Month filter
            if month:
                year, month_num = month.split('-')
                from datetime import datetime as dt_local
                date_from_obj = dt_local(int(year), int(month_num), 1, 0, 0, 0)
                if int(month_num) == 12:
                    date_to_obj = dt_local(int(year) + 1, 1, 1, 0, 0, 0)
                else:
                    date_to_obj = dt_local(int(year), int(month_num) + 1, 1, 0, 0, 0)
                start_str = f"{month}-01T00:00:00"
                end_str = date_to_obj.strftime("%Y-%m-%dT00:00:00")
                date_filter = {"$or": [
                    {"created_at": {"$gte": start_str, "$lt": end_str}},
                    {"created_at": {"$gte": date_from_obj, "$lt": date_to_obj}}
                ]}
                if "$or" in q:
                    q["$and"] = [{"$or": q.pop("$or")}, date_filter]
                else:
                    q.update(date_filter)

            # Aggregate stats - including license counts per type
            pipeline = [
                {"$match": q},
                {"$facet": {
                    "total": [{"$count": "count"}],
                    "fixed": [{"$match": {"status": "تم الإصلاح"}}, {"$count": "count"}],
                    "asphalt_remaining": [{"$match": {"$or": [{"status": "بانتظار الأسفلت"}, {"status": {"$regex": ".*متبقي.*[اأإ]سفلت.*", "$options": "i"}}]}}, {"$count": "count"}],
                    "by_type": [{"$group": {"_id": "$report_type", "count": {"$sum": 1}}}],
                    "asphalt_reports": [
                        {"$match": {"report_type": {"$in": ["أسفلت", "اسفلت", "إسفلت", "asphalt", "Asphalt"]}}},
                        {"$group": {
                            "_id": None,
                            "licensed": {"$sum": {"$cond": [{"$regexMatch": {"input": {"$ifNull": ["$license_number", ""]}, "regex": "[0-9]"}}, 1, 0]}},
                            "unlicensed": {"$sum": {"$cond": [{"$not": {"$regexMatch": {"input": {"$ifNull": ["$license_number", ""]}, "regex": "[0-9]"}}}, 1, 0]}}
                        }}
                    ],
                    "tile_reports": [
                        {"$match": {"report_type": {"$in": ["بلاط", "tile", "Tile"]}}},
                        {"$group": {
                            "_id": None,
                            "licensed": {"$sum": {"$cond": [{"$regexMatch": {"input": {"$ifNull": ["$license_number", ""]}, "regex": "[0-9]"}}, 1, 0]}},
                            "unlicensed": {"$sum": {"$cond": [{"$not": {"$regexMatch": {"input": {"$ifNull": ["$license_number", ""]}, "regex": "[0-9]"}}}, 1, 0]}}
                        }}
                    ],
                    "terrestrial_reports": [
                        {"$match": {"report_type": {"$in": ["ترابي", "terrestrial", "Terrestrial"]}}},
                        {"$group": {
                            "_id": None,
                            "licensed": {"$sum": {"$cond": [{"$regexMatch": {"input": {"$ifNull": ["$license_number", ""]}, "regex": "[0-9]"}}, 1, 0]}},
                            "unlicensed": {"$sum": {"$cond": [{"$not": {"$regexMatch": {"input": {"$ifNull": ["$license_number", ""]}, "regex": "[0-9]"}}}, 1, 0]}}
                        }}
                    ]
                }}
            ]
            agg_result = await db.reports.aggregate(pipeline, allowDiskUse=True).to_list(1)
            data = agg_result[0] if agg_result else {}

            total_reports = data.get("total", [{}])[0].get("count", 0) if data.get("total") else 0
            fixed_reports = data.get("fixed", [{}])[0].get("count", 0) if data.get("fixed") else 0
            asphalt_remaining = data.get("asphalt_remaining", [{}])[0].get("count", 0) if data.get("asphalt_remaining") else 0
            by_type_raw = {item["_id"]: item["count"] for item in data.get("by_type", []) if item.get("_id")}

            # استخراج إحصائيات الرخص لكل نوع
            asphalt_stats = data["asphalt_reports"][0] if data.get("asphalt_reports") else {}
            licensed = asphalt_stats.get("licensed", 0)
            unlicensed = asphalt_stats.get("unlicensed", 0)

            tile_stats = data["tile_reports"][0] if data.get("tile_reports") else {}
            tile_licensed = tile_stats.get("licensed", 0)
            tile_unlicensed = tile_stats.get("unlicensed", 0)

            terrestrial_stats = data["terrestrial_reports"][0] if data.get("terrestrial_reports") else {}
            terrestrial_licensed = terrestrial_stats.get("licensed", 0)
            terrestrial_unlicensed = terrestrial_stats.get("unlicensed", 0)

            # Water & sewage connections
            conn_q = q.copy()
            if "governorate" in conn_q:
                conn_q["area"] = conn_q.pop("governorate")
            water_total = await db.water_connections.count_documents(conn_q)
            sewage_total = await db.sewage_connections.count_documents(conn_q)

            # Project cards labels
            project_cards_doc = await db.project_cards.find_one(
                {"project": {"$regex": project_name.replace(" ", ".*"), "$options": "i"}},
                {"_id": 0, "cards": 1}
            )
            cards_docs = project_cards_doc.get("cards", []) if project_cards_doc else []

            return project_name, {
                "total": total_reports + water_total + sewage_total,
                "fixed": fixed_reports,
                "asphalt_remaining": asphalt_remaining,
                "licensed": licensed,
                "unlicensed": unlicensed,
                "tile_licensed": tile_licensed,
                "tile_unlicensed": tile_unlicensed,
                "terrestrial_licensed": terrestrial_licensed,
                "terrestrial_unlicensed": terrestrial_unlicensed,
                "terrestrial": by_type_raw.get("ترابي", 0),
                "tile": by_type_raw.get("بلاط", 0),
                "asphalt": by_type_raw.get("أسفلت", 0) + by_type_raw.get("إسفلت", 0) + by_type_raw.get("اسفلت", 0),
                "connections": water_total + sewage_total,
                "by_type": by_type_raw,
                "cards": cards_docs
            }

        except Exception as e:
            logging.warning(f"dashboard/init-all: failed for {project_name}: {e}")
            return project_name, {"total": 0, "fixed": 0, "asphalt_remaining": 0, "licensed": 0,
                                  "unlicensed": 0, "connections": 0, "by_type": {}, "cards": []}

    import asyncio as _asyncio
    tasks = [fetch_project_data(p) for p in projects_list]
    results = await _asyncio.gather(*tasks)

    projects_data = {name: stats for name, stats in results}

    return {
        "allowed_projects": allowed_projects,
        "projects": projects_data
    }


@api_router.post("/projects")
async def create_project(
    name: str = Form(...),
    description: str = Form(None),
    current_user: User = Depends(get_current_user)
):
    """إضافة مشروع جديد (للأدمن فقط)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من عدم تكرار الاسم
    existing = await db.projects.find_one({"name": name})
    if existing:
        raise HTTPException(status_code=400, detail="المشروع موجود بالفعل")
    
    project = {
        "id": str(uuid4()),
        "name": name,
        "description": description or name
    }
    await db.projects.insert_one(project)
    # إرجاع بدون _id
    return {"message": "تم إضافة المشروع بنجاح", "project": {"id": project["id"], "name": project["name"], "description": project["description"]}}


@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, current_user: User = Depends(get_current_user)):
    """حذف مشروع (للأدمن فقط)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # الحصول على اسم المشروع قبل حذفه
    project = await db.projects.find_one({"$or": [{"id": project_id}, {"name": project_id}]})
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    project_name = project.get("name")
    
    result = await db.projects.delete_one({"$or": [{"id": project_id}, {"name": project_id}]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    # حذف نهائي لجميع محافظات المشروع + تنظيف أي بيانات قديمة
    if project_name:
        await db.project_governorates.delete_many({"project": project_name})
        await db.deleted_governorates.delete_many({"project": project_name})  # تنظيف أي بقايا
        # إزالة المشروع من جميع المستخدمين
        await db.users.update_many(
            {},
            {"$pull": {"projects": project_name}}
        )
    
    return {"message": "تم حذف المشروع بنجاح"}

@api_router.post("/projects/{project_id}/archive")
async def archive_project(project_id: str, payload: dict = Body(...), current_user: User = Depends(get_current_user)):
    """أرشفة مشروع (للأدمن فقط)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    archive = payload.get("archive", True)
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    await db.projects.update_one({"id": project_id}, {"$set": {"is_archived": archive}})
    return {"message": "تم أرشفة المشروع بنجاح" if archive else "تم إلغاء أرشفة المشروع بنجاح"}

@api_router.get("/archive/projects")
async def get_archived_projects_with_counts(current_user: User = Depends(get_current_user)):
    """جلب المشاريع المؤرشفة مع عدد البلاغات (للأدمن فقط)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    projects = await db.projects.find({"is_archived": True}, {"_id": 0}).to_list(100)
    
    results = []
    for p in projects:
        name = p.get("name")
        q = {"project": get_flexible_project_query(name), "is_deleted": {"$ne": True}}
        # عد التقارير من جدول البلاغات
        count = await db.reports.count_documents(q)
        results.append({
            "name": name,
            "description": p.get("description"),
            "reports_count": count
        })
    
    return results

@api_router.put("/projects/{project_id}")
async def update_project(
    project_id: str,
    name: str = Form(...),
    current_user: User = Depends(get_current_user)
):
    """تعديل اسم مشروع (للأدمن فقط)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # البحث عن المشروع
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    old_name = project["name"]
    new_name = name.strip()
    
    # تحديث اسم المشروع
    await db.projects.update_one({"id": project_id}, {"$set": {"name": new_name}})
    
    # تحديث المشروع في جدول المحافظات
    await db.project_governorates.update_many(
        {"project": old_name},
        {"$set": {"project": new_name}}
    )
    
    # تحديث المشروع في جميع المستخدمين
    await db.users.update_many(
        {"projects": old_name},
        {"$set": {"projects.$": new_name}}
    )
    
    # تحديث project_permissions للمستخدمين
    users_with_perms = await db.users.find({f"project_permissions.{old_name}": {"$exists": True}}).to_list(1000)
    for u in users_with_perms:
        perms = u.get("project_permissions", {}).get(old_name)
        if perms is not None:
            await db.users.update_one(
                {"id": u["id"]},
                {
                    "$set": {f"project_permissions.{new_name}": perms},
                    "$unset": {f"project_permissions.{old_name}": ""}
                }
            )
            
    # تحديث المشروع في البلاغات
    await db.reports.update_many({"project": old_name}, {"$set": {"project": new_name}})
    
    # تحديث المشروع في أنواع وحالات البلاغات
    await db.report_types.update_many({"project": old_name}, {"$set": {"project": new_name}})
    await db.report_statuses.update_many({"project": old_name}, {"$set": {"project": new_name}})
    
    # تحديث المشروع في التوصيلات
    await db.water_connections.update_many({"project": old_name}, {"$set": {"project": new_name}})
    await db.sewage_connections.update_many({"project": old_name}, {"$set": {"project": new_name}})
    
    # تحديث المشروع في الفواتير
    await db.invoices.update_many({"project": old_name}, {"$set": {"project": new_name}})
    
    # تحديث المشروع في المستخلصات
    await db.extracts.update_many({"project": old_name}, {"$set": {"project": new_name}})
    
    # تحديث المشروع في طلبات الموظفين
    await db.employee_requests.update_many({"project": old_name}, {"$set": {"project": new_name}})
    
    # تحديث المشروع في المقاولين
    await db.contractors.update_many({"project": old_name}, {"$set": {"project": new_name}})
    
    return {"message": "تم تعديل المشروع بنجاح", "old_name": old_name, "new_name": new_name}


# ============= USER MANAGEMENT =============

@api_router.get("/users", response_model=List[UserResponse])
async def get_all_users(current_user: User = Depends(get_current_user)):
    """جلب قائمة المستخدمين - بدون صور البروفايل لتسريع التحميل"""
    if current_user.role == "admin":
        # الأدمن يرى جميع المستخدمين لتجنب مشكلة عدم رؤية المستخدمين المحذوفين أو المنشأين بواسطة أدمن آخر
        users = await db.users.find(
            {}, 
            {"_id": 0, "profile_picture": 0}
        ).to_list(1000)
    else:
        # Level 2 يرى فقط من قام بإنشائهم
        users = await db.users.find(
            {"created_by": current_user.id}, 
            {"_id": 0, "profile_picture": 0}
        ).to_list(1000)
    
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
        # Ensure projects field exists (for backward compatibility)
        if 'projects' not in user:
            user['projects'] = []
        # Ensure governorates field exists (for backward compatibility)
        if 'governorates' not in user:
            user['governorates'] = []
        # Ensure title field exists (for backward compatibility)
        if 'title' not in user:
            user['title'] = None
        # Ensure created_by field exists (for backward compatibility)
        if 'created_by' not in user:
            user['created_by'] = None
        # Ensure can_create_subusers field exists (for backward compatibility)
        if 'can_create_subusers' not in user:
            user['can_create_subusers'] = user.get('role') == 'admin'  # admin يمكنه دائماً
    
    return [UserResponse(**user) for user in users]



class ChatMessage(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sender_id: str
    receiver_id: str
    text: Optional[str] = None
    image_url: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    is_deleted: bool = False

class ChatMessageCreate(BaseModel):
    receiver_id: str
    text: Optional[str] = None
    image_url: Optional[str] = None

class UserUpdate(BaseModel):
    username: Optional[str] = None
    full_name: Optional[str] = None
    title: Optional[str] = None
    password: Optional[str] = None
    governorates: Optional[List[str]] = None  # تحديث المحافظات
    projects: Optional[List[str]] = None  # تحديث المشاريع
    permissions: Optional[List[str]] = None  # تحديث الصلاحيات
    allowed_chat_users: Optional[List[str]] = None  # تحديث صلاحيات التحدث الاستثنائية
# API لجلب قائمة الصلاحيات المتاحة
@api_router.get("/permissions")
async def get_all_permissions(current_user: User = Depends(get_current_user)):
    """جلب قائمة جميع الصلاحيات المتاحة"""
    return ALL_PERMISSIONS

# Model لتحديث الصلاحيات والمشاريع
class PermissionsUpdate(BaseModel):
    permissions: List[str]
    projects: Optional[List[str]] = None  # المشاريع المتاحة للمستخدم
    project_permissions: Optional[Dict[str, List[str]]] = None  # صلاحيات لكل مشروع

# دالة لتحديث صلاحيات ومشاريع المستخدمين التابعين (هرمياً)
async def update_subusers_permissions(user_id: str, new_permissions: List[str], new_projects: Optional[List[str]] = None):
    """إزالة الصلاحيات والمشاريع من المستخدمين التابعين التي لم تعد متاحة للمدير"""
    # جلب جميع المستخدمين الذين أنشأهم هذا المستخدم
    subusers = await db.users.find({"created_by": user_id}, {"_id": 0}).to_list(1000)
    
    for subuser in subusers:
        sub_id = subuser.get('id')
        updates = {}
        
        # 1. تحديث الصلاحيات العامة
        sub_perms = subuser.get('permissions', [])
        filtered_perms = [p for p in sub_perms if p in new_permissions]
        if filtered_perms != sub_perms:
            updates["permissions"] = filtered_perms
            
        # 2. تحديث المشاريع المتاحة
        if new_projects is not None:
            sub_projects = subuser.get('projects', [])
            norm_new_projs = {normalize_arabic(p) for p in new_projects}
            filtered_projs = [p for p in sub_projects if normalize_arabic(p) in norm_new_projs]
            if filtered_projs != sub_projects:
                updates["projects"] = filtered_projs
                
        # 3. تحديث الصلاحيات لكل مشروع
        sub_pp = subuser.get('project_permissions') or {}
        if sub_pp and new_projects is not None:
            norm_new_projs = {normalize_arabic(p) for p in new_projects}
            new_pp = {}
            changed_pp = False
            for proj, perms in sub_pp.items():
                if normalize_arabic(proj) in norm_new_projs:
                    new_pp[proj] = perms
                else:
                    changed_pp = True
            if changed_pp:
                updates["project_permissions"] = new_pp

        if updates:
            await db.users.update_one({"id": sub_id}, {"$set": updates})
            # تحديث هرمي للمستوى التالي
            await update_subusers_permissions(
                sub_id, 
                updates.get("permissions", sub_perms),
                updates.get("projects", subuser.get('projects', []))
            )

# API لتحديث صلاحيات ومشاريع مستخدم
@api_router.put("/users/{user_id}/permissions")
async def update_user_permissions(
    user_id: str,
    data: PermissionsUpdate,
    current_user: User = Depends(get_current_user)
):
    """تحديث صلاحيات ومشاريع المستخدم - Admin فقط أو المستوى 2 للمستخدمين التابعين"""
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    is_admin = current_user.role == "admin"
    is_creator = user_doc.get('created_by') == current_user.id
    is_level2 = current_user.can_create_subusers or is_creator
    
    # البيانات للتحديث
    update_data = {"permissions": data.permissions}
    if data.projects is not None:
        update_data["projects"] = data.projects
    if data.project_permissions is not None:
        # فلترة: فقط المشاريع المعروفة والصلاحيات المرتبطة بمشروع
        filtered_pp = {}
        allowed_projects = set(data.projects) if data.projects is not None else set(user_doc.get("projects") or [])
        for proj, perms in (data.project_permissions or {}).items():
            if proj in allowed_projects:
                filtered_pp[proj] = [p for p in (perms or []) if p in PROJECT_SCOPED_PERMISSIONS]
        update_data["project_permissions"] = filtered_pp
    
    # Admin يمكنه تعديل أي مستخدم
    if is_admin:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
        # تحديث صلاحيات المستخدمين التابعين هرمياً
        await update_subusers_permissions(user_id, data.permissions, data.projects)
        return {"message": "تم تحديث الصلاحيات والمشاريع بنجاح"}
    
    # المستوى 2 يمكنه تعديل المستخدمين التابعين له فقط
    if is_level2 and is_creator:
        my_permissions = set(current_user.permissions or [])
        my_project_perms = current_user.project_permissions or {}
        my_projects = current_user.projects or []
        is_all_projects = len(my_projects) == 0
        
        # تطبيع أسماء المشاريع للمدير لسهولة البحث
        norm_my_projects = {normalize_arabic(p): p for p in my_projects}
        
        # جمع كل صلاحيات المستوى 2 (عامة + لكل مشروع)
        my_all_perms = set(my_permissions)
        for pp_list in my_project_perms.values():
            my_all_perms.update(pp_list or [])
        
        # الحالة الحالية للمستخدم المستهدف (قبل التعديل)
        existing_perms = set(user_doc.get("permissions") or [])
        existing_pp = user_doc.get("project_permissions") or {}
        existing_projects = set(user_doc.get("projects") or [])
        
        # التحقق فقط من الصلاحيات الجديدة المُضافة (يُسمح بالحفاظ على/إلغاء القديمة ولو منحها الأدمن)
        new_global_perms = set(data.permissions) - existing_perms
        for perm in new_global_perms:
            if perm not in my_all_perms and perm != "view_governorate_data":
                raise HTTPException(status_code=403, detail=f"لا يمكنك إعطاء صلاحية {perm} لأنها غير متاحة لك")
        
        # التحقق من الصلاحيات الجديدة لكل مشروع
        if data.project_permissions:
            for proj, perms in data.project_permissions.items():
                norm_proj = normalize_arabic(proj)
                
                # 1. المشروع يجب أن يكون ضمن مشاريع المستوى 2 المسموح بها (باستخدام التطبيع)
                if not is_all_projects and norm_proj not in norm_my_projects:
                    raise HTTPException(status_code=403, detail=f"لا يمكنك تعديل صلاحيات مشروع {proj} لأنه غير مسند إليك")
                
                # جلب اسم المشروع الأصلي كما هو عند المدير
                original_proj_name = norm_my_projects.get(norm_proj, proj)
                
                # 2. التحقق من كل صلاحية داخل هذا المشروع
                existing_proj_perms = set(existing_pp.get(proj) or [])
                new_proj_perms = set(perms) - existing_proj_perms
                
                # صلاحيات المدير لهذا المشروع = (عامة) + (مخصصة لهذا المشروع)
                my_perms_for_this_proj = set(current_user.permissions or []) | set(my_project_perms.get(original_proj_name) or [])
                
                for p in new_proj_perms:
                    if p not in my_perms_for_this_proj and p not in my_all_perms:
                        raise HTTPException(status_code=403, detail=f"لا يمكنك منح صلاحية {p} في مشروع {proj} لأنك لا تملكها")
        
        # التحقق فقط من المشاريع الجديدة المضافة (يُسمح بإزالة أي مشروع موجود)
        if data.projects is not None and not is_all_projects:
            new_projects = set(data.projects) - existing_projects
            norm_my_projs_set = set(norm_my_projects.keys())
            
            invalid_new_projects = []
            for p in new_projects:
                if normalize_arabic(p) not in norm_my_projs_set:
                    invalid_new_projects.append(p)
                    
            if invalid_new_projects:
                raise HTTPException(status_code=403, detail=f"لا يمكنك منح مشاريع غير متاحة لك: {', '.join(invalid_new_projects)}")
        
        await db.users.update_one({"id": user_id}, {"$set": update_data})
        # تحديث صلاحيات المستخدمين التابعين هرمياً (الصلاحيات والمشاريع)
        await update_subusers_permissions(user_id, data.permissions, data.projects)
        return {"message": "تم تحديث الصلاحيات والمشاريع بنجاح"}
    
    raise HTTPException(status_code=403, detail="ليس لديك صلاحية")

@api_router.put("/users/{user_id}")
async def update_user(
    user_id: str,
    user_update: UserUpdate,
    current_user: User = Depends(get_current_user)
):
    """تحديث معلومات المستخدم - يمكن للـ Admin والمستوى 2 تعديل المستخدمين التابعين لهم"""
    # البحث عن المستخدم المراد تحديثه
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    # التحقق من الصلاحيات
    is_admin = current_user.role == "admin"
    is_own_user = current_user.id == user_id
    is_creator = user_doc.get('created_by') == current_user.id
    is_level2 = current_user.can_create_subusers or is_creator
    
    # السماح بالتعديل في الحالات التالية:
    # 1. Admin يمكنه تعديل أي مستخدم
    # 2. المستوى 2 يمكنه تعديل المستخدمين الذين أنشأهم
    # 3. المستخدم يمكنه تعديل معلوماته الخاصة
    if not (is_admin or (is_level2 and is_creator) or is_own_user):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية لتعديل هذا المستخدم")
    
    update_data = {}
    
    # تحديث اسم المستخدم
    if user_update.username is not None and user_update.username.strip():
        # التحقق من عدم وجود اسم مستخدم مكرر
        existing = await db.users.find_one({"username": user_update.username, "id": {"$ne": user_id}})
        if existing:
            raise HTTPException(status_code=400, detail="اسم المستخدم موجود بالفعل")
        update_data["username"] = user_update.username.strip()
    
    # تحديث الاسم الكامل
    if user_update.full_name is not None and user_update.full_name.strip():
        update_data["full_name"] = user_update.full_name.strip()
    
    # تحديث اللقب
    if user_update.title is not None:
        update_data["title"] = user_update.title.strip()
    
    # تحديث كلمة المرور
    if user_update.password is not None and user_update.password.strip():
        from passlib.context import CryptContext
        pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
        update_data["hashed_password"] = pwd_context.hash(user_update.password)
    
    # تحديث المحافظات (للـ Admin والمستوى 2 فقط)
    if user_update.governorates is not None and (is_admin or is_level2):
        update_data["governorates"] = user_update.governorates
    
    # تحديث المشاريع (للـ Admin والمستوى 2 فقط)
    if user_update.projects is not None and (is_admin or is_level2):
        update_data["projects"] = user_update.projects
    
    # تحديث الصلاحيات (للـ Admin أو المستوى 2 للمستخدمين التابعين)
    if user_update.permissions is not None:
        if is_admin:
            update_data["permissions"] = user_update.permissions
        elif is_level2 and is_creator:
            # المستوى 2 لا يمكنه إعطاء صلاحيات ليست لديه
            my_permissions = current_user.permissions or []
            for perm in user_update.permissions:
                if perm not in my_permissions:
                    raise HTTPException(status_code=403, detail=f"لا يمكنك إعطاء صلاحية غير متاحة لك")
            update_data["permissions"] = user_update.permissions
    
    # تحديث تصاريح المحادثة الاستثنائية (للأدمن فقط)
    if user_update.allowed_chat_users is not None and is_admin:
        update_data["allowed_chat_users"] = user_update.allowed_chat_users

    # تنفيذ التحديث
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
        return {"message": "تم تحديث المستخدم بنجاح"}
    else:
        return {"message": "لا توجد تعديلات للحفظ"}


class ConnectionPermissionsUpdate(BaseModel):
    connection_permissions: dict  # {"project_id": {"water_connections": true, "sewage_connections": true}}

@api_router.put("/users/{user_id}/connection-permissions")
async def update_user_connection_permissions(
    user_id: str,
    data: ConnectionPermissionsUpdate,
    current_user: User = Depends(get_current_user)
):
    """تحديث صلاحيات مشاريع الإيصال للمستخدم (legacy) - Admin أو المستوى 2 للتابعين له"""
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    is_admin = current_user.role == "admin"
    is_level2_creator = (current_user.can_create_subusers or user_doc.get('created_by') == current_user.id) and user_doc.get('created_by') == current_user.id
    
    if not (is_admin or is_level2_creator):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية لتعديل هذا المستخدم")
    
    await db.users.update_one(
        {"id": user_id}, 
        {"$set": {"connection_permissions": data.connection_permissions}}
    )
    return {"message": "تم تحديث صلاحيات مشاريع الإيصال بنجاح"}


@api_router.put("/users/{user_id}/toggle-active")
async def toggle_user_active(user_id: str, current_user: User = Depends(get_current_user)):
    # السماح للمسؤول والمستوى 2 أو منشئ الحساب بالتعطيل/التفعيل
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
        
    is_creator = user_doc.get('created_by') == current_user.id
    if not (current_user.role == "admin" or current_user.can_create_subusers or is_creator):
        raise HTTPException(status_code=403, detail="Not authorized")
    
    new_status = not user_doc['is_active']
    await db.users.update_one({"id": user_id}, {"$set": {"is_active": new_status}})
    
    return {"message": "User status updated", "is_active": new_status}


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    # السماح للمسؤول والمستوى 2 أو منشئ الحساب بالحذف
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
        
    is_creator = user_doc.get('created_by') == current_user.id
    if not (current_user.role == "admin" or current_user.can_create_subusers or is_creator):
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}



@api_router.delete("/users/bulk/delete-all")
async def delete_all_users(current_user: User = Depends(get_current_admin_user)):
    """حذف جميع المستخدمين ما عدا المسؤول الحالي"""
    result = await db.users.delete_many({"id": {"$ne": current_user.id}})
    
    return {
        "message": f"تم حذف {result.deleted_count} مستخدم بنجاح",
        "deleted_count": result.deleted_count
    }




async def get_next_sequence_value(sequence_name: str):

    """جلب القيمة التالية للتسلسل وتحديثها في قاعدة البيانات"""
    result = await db.counters.find_one_and_update(
        {"id": sequence_name},
        {"$inc": {"sequence_value": 1}},
        upsert=True,
        return_document=True
    )
    return result["sequence_value"]

# ============= REPORTS ROUTES =============


@api_router.post("/reports", response_model=ReportResponse)
async def create_report(
    report_number: str = Form(...),
    license_number: str = Form(...),
    report_type: str = Form(...),
    status: str = Form(...),
    governorate: str = Form(...),
    project: str = Form(...),
    depth_meters: float = Form(...),
    diameter_mm: float = Form(...),
    contractor: str = Form(...),
    latitude: Optional[str] = Form(None),
    longitude: Optional[str] = Form(None),
    asphalt_license_issued: Optional[str] = Form("false"),
    wfm_closed: Optional[str] = Form("false"),
    notes: Optional[str] = Form(None),
    created_at: Optional[str] = Form(None),
    start_date: Optional[str] = Form(None),
    closed_at: Optional[str] = Form(None),
    remove_closed_at: Optional[str] = Form(None),
    images: List[UploadFile] = File(default=[]),
    current_user: User = Depends(get_current_user)
):
    # ⚡ معالجة الترقيم التلقائي
    if report_number == "AUTO" or not report_number.strip():
        seq = await get_next_sequence_value("reports_global")
        report_number = f"CCB-R-{seq:05d}" # تنسيق CCB-R-00001
    
    # ⚡ فحص سريع وبسيط
    existing_report = await db.reports.find_one({
        "report_number": report_number
    }, {"_id": 1})
    
    if existing_report:
        raise HTTPException(
            status_code=400,
            detail="هذا الرقم موجود مسبقاً"
        )
    
    # التحقق من تكرار رقم الرخصة - فقط للأرقام الفعلية (ليست نصوص تعريفية مثل "لم يتم إصدار رخصة")
    def _is_actual_license(val: str) -> bool:
        """يُعتبر رقم رخصة فعلياً إذا احتوى رقم (حتى لو معه نص)"""
        if not val:
            return False
        cleaned = val.strip()
        if not cleaned:
            return False
        # نصوص شائعة تعني "لا توجد رخصة" - لا نتعامل معها كأرقام
        placeholders = {"لم يتم إصدار رخصة", "لم يتم", "-", "0", "nan", "none", "غير محدد", "بدون رخصة"}
        if cleaned in placeholders:
            return False
        # يجب أن يحتوي على رقم واحد على الأقل (وإلا نعتبره نصاً عاماً)
        return any(ch.isdigit() for ch in cleaned)
    
    if _is_actual_license(license_number):
        existing_license = await db.reports.find_one({
            "license_number": license_number
        }, {"_id": 1, "report_number": 1})
        if existing_license:
            raise HTTPException(
                status_code=400,
                detail="هذا الرقم موجود مسبقاً"
            )
    
    # ⚡ معالجة الصور بشكل متوازي فائق السرعة
    image_data = []
    if images:
        async def process_image_fast(image):
            """معالجة صورة واحدة - رفع إلى Object Storage بدلاً من base64"""
            try:
                content = await image.read()
                # ضغط الصورة إلى 300KB
                loop = asyncio.get_event_loop()
                content = await loop.run_in_executor(thread_pool, compress_image, content)
                # رفع إلى التخزين
                ext = _guess_ext(image.filename, image.content_type)
                url = await loop.run_in_executor(
                    thread_pool,
                    lambda: _store_image_bytes(content, category="reports", filename=image.filename, content_type=image.content_type)
                )
                return url
            except Exception as e:
                print(f"Image upload error: {e}")
                return None
        
        # معالجة جميع الصور بشكل متوازي
        results = await asyncio.gather(*[process_image_fast(img) for img in images], return_exceptions=True)
        image_data = [r for r in results if r and not isinstance(r, Exception)]
    
    # تحديد تاريخ الاستلام
    if created_at:
        created_at_dt = datetime.fromisoformat(created_at)
    else:
        created_at_dt = datetime.now(timezone.utc)
    
    # تحديد تاريخ مباشرة البلاغ
    start_date_dt = None
    if start_date:
        start_date_dt = datetime.fromisoformat(start_date)
    
    # تحديد تاريخ الإغلاق
    closed_at_dt = None
    if closed_at:
        closed_at_dt = datetime.fromisoformat(closed_at)
    
    # ⚡ إنشاء الوثيقة مباشرة - أسرع!
    doc = {
        "id": str(uuid4()),
        "report_number": report_number,
        "report_date": None,
        "license_number": license_number,
        "report_type": report_type,
        "status": normalize_asphalt_status(status),
        "governorate": governorate,
        "project": project,
        "depth_meters": depth_meters,
        "diameter_mm": diameter_mm,
        "contractor": contractor,
        "latitude": latitude,
        "longitude": longitude,
        "asphalt_license_issued": (asphalt_license_issued.lower() == "true" if asphalt_license_issued else False),
        "wfm_closed": (wfm_closed.lower() == "true" if wfm_closed else False),
        "notes": notes or '',
        "images": image_data,
        "created_by": current_user.id,
        "created_by_name": current_user.full_name,
        "created_at": created_at_dt,
        "start_date": start_date_dt,
        "added_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc),
        "closed_at": closed_at_dt,
        "deleted_at": None,
        "is_deleted": False,
        "review_status": "قيد المراجعة",
        "reviewed_by": None,
        "reviewed_at": None
    }
    
    # ⚡ حفظ مباشر - بدون Pydantic overhead
    await db.reports.insert_one(doc)
    
    # ⚡ مسح الـ cache لتحديث الإحصائيات فوراً
    global stats_cache
    stats_cache.clear()
    
    # إرجاع مع ReportResponse للتوافق
    return ReportResponse(**doc)


@api_router.get("/reports")
async def get_reports(
    search: Optional[str] = Query(None),
    exact: Optional[bool] = Query(False),  # للبحث الدقيق (من الإشعارات)
    license_number: Optional[str] = Query(None),
    governorate: Optional[str] = Query(None),
    project: Optional[str] = Query(None),
    contractor: Optional[str] = Query(None),
    report_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    license_status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    start_date_from: Optional[str] = Query(None),
    start_date_to: Optional[str] = Query(None),
    my_reports: Optional[bool] = Query(False),
    created_by: Optional[str] = Query(None),
    page: Optional[int] = Query(1, ge=1),
    limit: Optional[int] = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user)
):
    # print debug
    query = {"is_deleted": {"$ne": True}}
    
    # قائمة المستخدمين الذين يرون بلاغاتهم فقط تلقائياً (بدون تفعيل "بلاغاتي")
    restricted_users = ["Mohamed Esmat", "ElShazly"]
    
    # ========== منطق صلاحيات المشروعات والمحافظات ==========
    if current_user.role == "admin":
        # Admin يرى كل شيء بدون قيود (إلا إذا فعّل my_reports)
        pass
    else:
        # المستخدم ليس Admin
        # فلترة حسب المشاريع التي يملك فيها صلاحية reports_view أو reports_add
        allowed_view_projects = []
        if len(current_user.projects) > 0:
            allowed_view_projects = list(set(get_projects_with_permission(current_user, "reports_view")) | \
                                         set(get_projects_with_permission(current_user, "reports_add")))
            if not allowed_view_projects:
                allowed_view_projects = current_user.projects
        
        # فلترة حسب المحافظات والتسلسل الهرمي
        user_governorates = current_user.governorates if hasattr(current_user, 'governorates') and current_user.governorates else []
        
        # تطبيق الفلترة الهرمية الشاملة (Recursive)
        hierarchy_filter = await get_hierarchy_filter(current_user)
        
        # فلترة حسب المحافظات والتسلسل الهرمي
        user_governorates = current_user.governorates if hasattr(current_user, 'governorates') and current_user.governorates else []
        has_all_govs = any(g in ["الكل", "جميع المحافظات", "كل المحافظات"] for g in user_governorates)
        
        if allowed_view_projects:
            query.update(get_flexible_in_query(allowed_view_projects, "project"))

        if has_all_govs:
            # إذا كان لديه صلاحية "الكل"، يرى جميع بلاغات تابعيه في مشاريعه
            query.update(hierarchy_filter)
        else:
            # فلترة بالمحافظات المخصصة + التابعين هرمياً
            if len(user_governorates) > 0:
                gov_patterns = []
                for g in user_governorates:
                    p = normalize_arabic_regex(g)
                    gov_patterns.append(p)
                gov_regex = f"({'|'.join(gov_patterns)})"
                
                # تطبيق الفلترة الصارمة: يجب أن يكون ضمن المحافظة المسندة AND من إنتاج المستخدم أو تابعيه
                query.update(hierarchy_filter)
                query['governorate'] = {'$regex': gov_regex, '$options': 'i'}
            else:
                # لا توجد محافظات محددة - يرى بيانات تابعيه فقط
                query.update(hierarchy_filter)
    
    # ========== منطق فلترة my_reports ==========
    if my_reports:
        # المستخدم طلب رؤية بلاغاته فقط
        query["$or"] = [
            {"created_by": current_user.id},
            {"created_by": current_user.username}
        ]
    
    # ========== فلاتر البحث ==========
    if search:
        # إذا كان البحث دقيق (من الإشعارات) - نبحث بالمطابقة التامة
        if exact:
            search_condition = {
                "$or": [
                    {"report_number": search},
                    {"id": search}
                ]
            }
        else:
            # البحث العادي بالجزء من الرقم
            search_condition = {
                "$or": [
                    {"report_number": {"$regex": search, "$options": "i"}},
                    {"license_number": {"$regex": search, "$options": "i"}}
                ]
            }
        if "$or" in query:
            existing_or = query.pop("$or")
            query["$and"] = [{"$or": existing_or}, search_condition]
        else:
            query.update(search_condition)
    
    if license_number:
        query["license_number"] = {"$regex": license_number, "$options": "i"}
    
    if governorate:
        gov_clean = governorate.strip()
        if gov_clean and gov_clean not in ["الكل", "جميع المحافظات", "كل المحافظات", "الكل ", "جميع المحافظات ", "جميع محافظات المشروع"]:
            # استخدام البحث المرن للمحافظة المحددة في الفلتر أيضاً
            gov_p = normalize_arabic_regex(gov_clean)
            query["governorate"] = {"$regex": f"({gov_p})", "$options": "i"}
    
    # فلترة حسب المشروع المحدد (مع التأكد من الصلاحيات والبحث المرن)
    if project:
        regex_query = get_flexible_project_query(project)
        if current_user.role != "admin" and len(current_user.projects) > 0:
            # التحقق من الصلاحية بمرونة عالية (تبادلية)
            has_permission = False
            for up in current_user.projects:
                up_keywords = [k for k in up.replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'أعمال', 'إصلاح']]
                proj_keywords = [k for k in project.replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'أعمال', 'إصلاح']]
                if any(k in project for k in up_keywords) or any(k in up for k in proj_keywords):
                    has_permission = True
                    break
            
            if not has_permission:
                # إذا لم يكن لديه صلاحية على المشروع المختار، نرجع قائمة فارغة
                query["project"] = "___NONE___"
            else:
                query["project"] = regex_query
        else:
            query["project"] = regex_query
    
    if contractor:
        contractor_clean = contractor.strip()
        if contractor_clean and contractor_clean not in ["الكل", "جميع المقاولين"]:
            query["contractor"] = contractor_clean
    
    if report_type:
        type_clean = report_type.strip()
        if type_clean and type_clean not in ["الكل", "جميع الأنواع"]:
            query["report_type"] = type_clean
    
    if status:
        status_clean = status.strip()
        if status_clean and status_clean not in ["الكل", "جميع الحالات"]:
            query["status"] = status_clean
    
    # فلتر حسب المستخدم (للـ Admin ومستوى 2 ومحمود هارون ومدحت)
    if created_by:
        created_by_clean = created_by.strip()
        if created_by_clean and created_by_clean not in ["الكل", "جميع المستخدمين"]:
            # البحث بـ username أو user_id
            if "$and" in query:
                # إذا كان هناك $and موجود، نضيف شرط جديد
                query["$and"].append({
                    "$or": [
                        {"created_by": created_by_clean},
                        {"created_by": {"$regex": created_by_clean, "$options": "i"}}
                    ]
                })
            else:
                # إذا لم يكن هناك $and، نستخدم created_by مباشرة
                query["created_by"] = created_by_clean
    
    # فلترة حسب حالة الرخصة أو الحالة
    if license_status == 'status_fixed':
        # تم الإصلاح فقط
        query["status"] = "تم الإصلاح"
    elif license_status == 'status_asphalt':
        # تم الإصلاح - ومتبقي الأسفلت
        query["status"] = {"$in": [_ASPHALT_CANONICAL, "بانتظار الأسفلت"]}
    elif license_status == 'status_in_progress':
        # قيد المعالجة - بلاغات لم تغلق بعد
        query["wfm_closed"] = {"$ne": True}
    elif license_status == 'status_wfm_closed':
        # مغلقة بواسطة الاستشاري
        query["wfm_closed"] = True
    elif license_status == 'review_pending':
        query["review_status"] = {"$in": ["بانتظار المراجعة", "قيد المراجعة", None]}
    elif license_status == 'license_issued':
        # تم إصدار رخص (يحتوي رقم - أي يحوي digit واحد على الأقل)
        query["license_number"] = {"$regex": "[0-9]"}
    elif license_status == 'license_not_issued':
        # لم يتم إصدار رخصة (لا يحتوي أي رقم)
        query["$or"] = [
            {"license_number": {"$exists": False}},
            {"license_number": None},
            {"license_number": {"$not": {"$regex": "[0-9]"}}}
        ]
    elif license_status and license_status.startswith('custom_'):
        # حالة مخصصة ديناميكية (custom_اسم الحالة)
        custom_status_name = license_status[len('custom_'):]
        query["status"] = custom_status_name
    
    # فلترة بتاريخ استلام البلاغ (created_at) - يدعم string و datetime
    if date_from or date_to:
        if date_from and not date_to:
            date_to = date_from
        elif date_to and not date_from:
            date_from = date_to
            
        from datetime import datetime as dt, timedelta
        
        try:
            if date_from and date_to:
                # كلا التاريخين موجودان
                date_from_obj = dt.fromisoformat(date_from)
                date_to_obj = dt.fromisoformat(date_to)
                next_day = date_to_obj + timedelta(days=1)
                
                # فلتر يدعم string و datetime
                date_filter = {
                    "$or": [
                        {"created_at": {"$gte": f"{date_from}T00:00:00", "$lt": next_day.strftime("%Y-%m-%dT00:00:00")}},
                        {"created_at": {"$gte": date_from_obj, "$lt": next_day}}
                    ]
                }

            
            # إضافة فلتر التاريخ للـ query
            if "$and" in query:
                query["$and"].append(date_filter)
            elif "$or" in query:
                query["$and"] = [{"$or": query.pop("$or")}, date_filter]
            else:
                if "$or" in date_filter:
                    query["$and"] = [date_filter]
                else:
                    query.update(date_filter)
        except Exception as e:
            print(f"Date filter error: {e}")
    
    # فلترة بتاريخ مباشرة البلاغ (start_date)
    if start_date_from or start_date_to:
        if start_date_from and not start_date_to:
            start_date_to = start_date_from
        elif start_date_to and not start_date_from:
            start_date_from = start_date_to
            
        from datetime import datetime as dt, timedelta
        try:
            if start_date_from and start_date_to:
                s_from = dt.fromisoformat(start_date_from)
                s_to = dt.fromisoformat(start_date_to)
                s_next = s_to + timedelta(days=1)
                sdate_filter = {
                    "$or": [
                        {"start_date": {"$gte": f"{start_date_from}T00:00:00", "$lt": s_next.strftime("%Y-%m-%dT00:00:00")}},
                        {"start_date": {"$gte": s_from, "$lt": s_next}}
                    ]
                }

            
            if "$and" in query:
                query["$and"].append(sdate_filter)
            elif "$or" in query:
                query["$and"] = [{"$or": query.pop("$or")}, sdate_filter]
            else:
                query["$and"] = [sdate_filter]
        except Exception as e:
            print(f"Start date filter error: {e}")
            
    with open("debug_log.txt", "a", encoding="utf-8") as f:
        f.write(f"\n[{datetime.now()}] User: {current_user.username}, Role: {current_user.role}, Requested Project: '{project}'\n")
        f.write(f"Final Query: {query}\n")
    # حساب العدد الكلي للتقارير
    total_count = await db.reports.count_documents(query)
    with open("debug_log.txt", "a", encoding="utf-8") as f:
        f.write(f"Total Count: {total_count}\n")
    
    # حساب عدد الصفحات وموقع البداية
    skip = (page - 1) * limit
    
    # جلب البلاغات مع الترقيم (بدون الصور لتسريع الاستجابة)
    projection = {"_id": 0, "images": 0}  # استبعاد الصور
    reports = await db.reports.find(query, projection).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # ⚡ تحسين: جلب أسماء المستخدمين مرة واحدة
    creator_ids = list(set(r.get('created_by') for r in reports if r.get('created_by')))
    users_map = {}
    if creator_ids:
        users_cursor = db.users.find(
            {"id": {"$in": creator_ids}}, 
            {"_id": 0, "id": 1, "full_name": 1, "title": 1}
        )
        async for user in users_cursor:
            title = user.get('title', '')
            name = user.get('full_name', 'غير معروف')
            users_map[user['id']] = f"{title} {name}".strip() if title else name
    
    for report in reports:
        if isinstance(report.get('report_date'), str):
            report['report_date'] = datetime.fromisoformat(report['report_date'])
        if isinstance(report.get('created_at'), str):
            report['created_at'] = datetime.fromisoformat(report['created_at'])
        if isinstance(report.get('added_at'), str):
            report['added_at'] = datetime.fromisoformat(report['added_at'])
        elif not report.get('added_at'):
            report['added_at'] = report.get('created_at')
        if isinstance(report.get('updated_at'), str):
            report['updated_at'] = datetime.fromisoformat(report['updated_at'])
        if report.get('deleted_at') and isinstance(report['deleted_at'], str):
            report['deleted_at'] = datetime.fromisoformat(report['deleted_at'])
        if 'closed_at' not in report:
            report['closed_at'] = None
        elif report.get('closed_at') and isinstance(report['closed_at'], str):
            report['closed_at'] = datetime.fromisoformat(report['closed_at'])
        if 'project' not in report:
            report['project'] = 'مشروع إصلاح أعمال المحافظات الغربية - القطاع الأوسط'
        if 'latitude' not in report:
            report['latitude'] = None
        if 'longitude' not in report:
            report['longitude'] = None
        if 'asphalt_license_issued' not in report:
            report['asphalt_license_issued'] = False
        if 'is_deleted' not in report:
            report['is_deleted'] = False
        
        # ⚡ استخدام الـ map بدلاً من استعلام لكل report
        report['created_by_name'] = users_map.get(report.get('created_by'), 'غير معروف')
    
    # إرجاع البيانات مع معلومات الترقيم
    return {
        "reports": [ReportResponse(**report).model_dump() for report in reports],
        "total_count": total_count,
        "page": page,
        "limit": limit,
        "total_pages": (total_count + limit - 1) // limit  # حساب عدد الصفحات
    }


@api_router.get("/governorates")
async def get_governorates(
    project: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """Get governorates based on project and user permissions"""
    def get_fuzzy_regex_pattern(text: str) -> str:
        if not text:
            return ""
        chars = []
        for c in text:
            if c in ('أ', 'إ', 'آ', 'ا'):
                chars.append('[أإآا]')
            elif c == ' ':
                chars.append('.*')
            else:
                chars.append(c)
        return "".join(chars)

    try:
        # جلب المحافظات المخصصة من قاعدة البيانات
        query = {}
        if project:
            # استخدام مطابقة مرنة (Fuzzy Match) لتجنب الفروقات في الهمزات والمسافات لاسم المشروع
            proj_pattern = get_fuzzy_regex_pattern(project)
            query["project"] = {"$regex": f"^{proj_pattern}$", "$options": "i"}
            
        custom_govs = await db.project_governorates.find(query, {"_id": 0}).to_list(2000)
        
        # جلب المحافظات المحذوفة
        deleted_govs = await db.deleted_governorates.find(query, {"_id": 0}).to_list(2000)
        deleted_set = {(d.get('name'), d.get('project')) for d in deleted_govs}
        
        # بناء قائمة المحافظات الكاملة من قاعدة البيانات
        PROJECT_GOVERNORATES = {}
        for custom in custom_govs:
            proj_name = custom.get('project')
            gov_name = custom.get('name')
            if proj_name and gov_name and (gov_name, proj_name) not in deleted_set:
                if proj_name not in PROJECT_GOVERNORATES:
                    PROJECT_GOVERNORATES[proj_name] = []
                if gov_name not in PROJECT_GOVERNORATES[proj_name]:
                    PROJECT_GOVERNORATES[proj_name].append(gov_name)
        
        if not project:
            # إذا لم يتم تحديد مشروع، نجمع كل المحافظات المتاحة للمستخدم عبر جميع مشاريعه
            all_govs_list = []
            
            # تحديد المشاريع التي يحق للمستخدم رؤيتها
            target_projects = []
            if current_user.role == "admin":
                target_projects = list(PROJECT_GOVERNORATES.keys())
            else:
                target_projects = current_user.projects or []
            
            for proj in target_projects:
                # محاولة المطابقة المباشرة أو المرنة لكل مشروع
                if proj in PROJECT_GOVERNORATES:
                    all_govs_list.extend(PROJECT_GOVERNORATES[proj])
                else:
                    from re import search
                    proj_pattern = get_fuzzy_regex_pattern(proj)
                    for p_key, p_govs in PROJECT_GOVERNORATES.items():
                        if search(proj_pattern, p_key, re.IGNORECASE):
                            all_govs_list.extend(p_govs)
                            break
            
            # إذا كان لدى المستخدم قائمة محافظات محددة (Level 3)، نفلتر بها أيضاً
            if current_user.role != "admin" and current_user.governorates:
                all_govs_list = [g for g in all_govs_list if g in current_user.governorates]
                
            return sorted(list(set(all_govs_list)))
        
        # Get all governorates for the selected project
        # البحث عن المحافظات للمشروع المختار (استخدام مطابقة مرنة)
        all_governorates = []
        if project:
            # محاولة المطابقة المباشرة أولاً
            if project in PROJECT_GOVERNORATES:
                all_governorates = PROJECT_GOVERNORATES[project]
            else:
                # محاولة المطابقة المرنة إذا فشلت المباشرة
                from re import search
                proj_pattern = get_fuzzy_regex_pattern(project)
                for p_key, p_govs in PROJECT_GOVERNORATES.items():
                    if search(proj_pattern, p_key, re.IGNORECASE):
                        all_governorates = p_govs
                        break
        
        # Apply user permissions
        # 1. Admin - sees all governorates of selected project
        if current_user.role == "admin":
            return sorted(all_governorates)
        
        # 2. Check if user has project access (using fuzzy match for hamzas/spaces)
        if not current_user.projects:
            return []
        
        has_project_access = False
        from re import search
        proj_pattern = get_fuzzy_regex_pattern(project)
        for user_proj in current_user.projects:
            if user_proj == project or search(proj_pattern, user_proj, re.IGNORECASE):
                has_project_access = True
                break
                
        if not has_project_access:
            return []
        
        # 3. Users with ALL governorates of their project (Level 2 or special Level 3)
        # If user has ALL governorates of the project assigned, show all
        if current_user.governorates and set(current_user.governorates) == set(all_governorates):
            return sorted(all_governorates)
        
        # 4. Users with specific governorates (Level 3)
        if current_user.governorates:
            filtered = [g for g in all_governorates if g in current_user.governorates]
            return sorted(filtered)
        
        # 5. Users with project but no specific governorates (Level 2)
        # If user has project but no governorates list, show all governorates
        return sorted(all_governorates)
        
    except Exception as e:
        import logging
        logging.error(f"Error fetching governorates: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ============= إدارة المحافظات الديناميكية =============

class GovernorateCreate(BaseModel):
    name: str
    project: str

class GovernorateUpdate(BaseModel):
    old_name: str
    new_name: str
    project: str

@api_router.get("/project-governorates")
async def get_all_project_governorates(current_user: User = Depends(get_current_user)):
    """جلب جميع المحافظات من قاعدة البيانات مع ربطها بالمشاريع"""
    try:
        # جلب المحافظات مباشرة - لا يوجد deleted_governorates (حذف نهائي)
        custom_govs = await db.project_governorates.find({}, {"_id": 0}).to_list(2000)
        
        result = {}
        for custom in custom_govs:
            project = custom.get('project')
            gov_name = custom.get('name')
            if project and gov_name:
                if project not in result:
                    result[project] = []
                if gov_name not in result[project]:
                    result[project].append(gov_name)
        
        return result
    except Exception as e:
        logging.error(f"Error fetching project governorates: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/project-governorates")
async def add_governorate(
    data: GovernorateCreate,
    current_user: User = Depends(get_current_user)
):
    """إضافة محافظة جديدة لمشروع - للأدمن فقط"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="فقط المسؤول يمكنه إضافة محافظات")
    
    try:
        # التحقق من عدم وجود المحافظة مسبقاً
        existing = await db.project_governorates.find_one({
            "name": data.name,
            "project": data.project
        })
        
        if existing:
            raise HTTPException(status_code=400, detail="المحافظة موجودة مسبقاً في هذا المشروع")
        
        # إضافة المحافظة الجديدة
        await db.project_governorates.insert_one({
            "id": str(uuid4()),
            "name": data.name,
            "project": data.project,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user.id
        })
        
        return {"message": f"تم إضافة محافظة {data.name} بنجاح"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error adding governorate: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.put("/project-governorates")
async def update_governorate(
    data: GovernorateUpdate,
    current_user: User = Depends(get_current_user)
):
    """تعديل اسم محافظة - للأدمن فقط"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="فقط المسؤول يمكنه تعديل المحافظات")
    
    try:
        # 1. تحديث المحافظة في قاعدة البيانات الأساسية للمشروع
        result = await db.project_governorates.update_one(
            {"name": data.old_name, "project": data.project},
            {"$set": {"name": data.new_name}}
        )
        
        # 2. تحديث جميع البلاغات المرتبطة بهذا المشروع وهذه المحافظة حصراً
        await db.reports.update_many(
            {"governorate": data.old_name, "project": data.project},
            {"$set": {"governorate": data.new_name}}
        )
        
        # 3. تحديث توصيلات المياه المرتبطة بهذا المشروع
        await db.water_connections.update_many(
            {"governorate": data.old_name, "project": data.project},
            {"$set": {"governorate": data.new_name}}
        )
        
        # 4. تحديث توصيلات الصرف الصحي المرتبطة بهذا المشروع
        await db.sewage_connections.update_many(
            {"governorate": data.old_name, "project": data.project},
            {"$set": {"governorate": data.new_name}}
        )
        
        # 5. تحديث الفواتير المرتبطة بهذا المشروع
        await db.invoices.update_many(
            {"governorate": data.old_name, "project": data.project},
            {"$set": {"governorate": data.new_name}}
        )
        
        # 6. تحديث المستخدمين (منطق ذكي للتعامل مع الأسماء المكررة في مشاريع مختلفة)
        # جلب المستخدمين الذين لديهم هذه المحافظة
        cursor = db.users.find({"governorates": data.old_name})
        async for user_doc in cursor:
            # فقط إذا كان المستخدم لديه وصول لهذا المشروع
            if data.project in (user_doc.get("projects") or []):
                new_govs = list(user_doc.get("governorates") or [])
                
                # إضافة الاسم الجديد إذا لم يكن موجوداً
                if data.new_name not in new_govs:
                    new_govs.append(data.new_name)
                
                # هل لا يزال المستخدم بحاجة للاسم القديم لمشاريع أخرى؟
                needs_old = False
                for p in (user_doc.get("projects") or []):
                    if p == data.project: continue
                    # التحقق إذا كان المشروع الآخر لا يزال يحتوي على المحافظة القديمة
                    other_gov = await db.project_governorates.find_one({"name": data.old_name, "project": p})
                    if other_gov:
                        needs_old = True
                        break
                
                if not needs_old:
                    if data.old_name in new_govs:
                        new_govs.remove(data.old_name)
                
                await db.users.update_one({"id": user_doc["id"]}, {"$set": {"governorates": new_govs}})
        
        return {"message": f"تم تعديل المحافظة من {data.old_name} إلى {data.new_name} بنجاح في جميع السجلات المرتبطة بمشروع {data.project}"}
    except Exception as e:
        logging.error(f"Error updating governorate: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/project-governorates/{project}/{governorate}")
async def delete_governorate(
    project: str,
    governorate: str,
    current_user: User = Depends(get_current_user)
):
    """حذف محافظة نهائياً من المشروع - للأدمن فقط"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="فقط المسؤول يمكنه حذف المحافظات")
    
    try:
        # حذف نهائي من project_governorates - بدون تخزين في deleted_governorates
        await db.project_governorates.delete_one({
            "name": governorate,
            "project": project
        })
        
        # إزالة المحافظة من المستخدمين المرتبطين بهذا المشروع فقط
        # (إذا كانت المحافظة موجودة في مشروع آخر للمستخدم، تبقى معه)
        cursor = db.users.find({"governorates": governorate})
        async for user_doc in cursor:
            if project in (user_doc.get("projects") or []):
                still_needed = False
                for p in (user_doc.get("projects") or []):
                    if p == project:
                        continue
                    other_gov = await db.project_governorates.find_one({"name": governorate, "project": p})
                    if other_gov:
                        still_needed = True
                        break
                if not still_needed:
                    await db.users.update_one(
                        {"id": user_doc["id"]},
                        {"$pull": {"governorates": governorate}}
                    )
        
        return {"message": f"تم حذف محافظة {governorate} من مشروع {project} نهائياً"}
    except Exception as e:
        logging.error(f"Error deleting governorate: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/reports/stats")
async def get_reports_stats(
    project: Optional[str] = Query(None),
    month: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """الحصول على إحصائيات البلاغات بسرعة مع caching - تدعم التسلسل الهرمي"""
    
    # Check cache first
    cache_key = f"stats_{current_user.id}_{project}_{month}"
    current_time = time.time()
    
    if cache_key in stats_cache:
        cached_data, cached_time = stats_cache[cache_key]
        if current_time - cached_time < CACHE_TTL:
            return cached_data
    
    query_filter = {"is_deleted": {"$ne": True}}
    # التصفية الهرمية الشاملة
    hierarchy_filter = await get_hierarchy_filter(current_user)
    query_filter.update(hierarchy_filter)
    
    if current_user.role != "admin" and current_user.projects:
        # التحقق من الصلاحية بمرونة
        if project:
            has_permission = False
            for up in current_user.projects:
                up_keywords = [k for k in up.replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'أعمال', 'إصلاح']]
                proj_keywords = [k for k in project.replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'أعمال', 'إصلاح']]
                is_match = any(k in project for k in up_keywords) or any(k in up for k in proj_keywords)
                if is_match:
                    has_permission = True
                    break
            
            if not has_permission:
                return {"total": 0, "fixed": 0, "asphalt_remaining": 0, "licensed": 0, "unlicensed": 0, "tile_licensed": 0, "tile_unlicensed": 0, "terrestrial_licensed": 0, "terrestrial_unlicensed": 0, "terrestrial": 0, "tile": 0, "asphalt": 0, "by_type": {}}
            
            query_filter["project"] = get_flexible_project_query(project)
        else:
            query_filter.update(get_flexible_in_query(current_user.projects, "project"))
    elif project:
        query_filter["project"] = get_flexible_project_query(project)
    
    # تصفية حسب الشهر
    if month:
        from datetime import datetime as dt, timedelta
        
        # month format: "2024-01" or "2025-12"
        year, month_num = month.split('-')
        
        # Calculate start and end dates (Naive for DB matching)
        from datetime import datetime as dt
        date_from_obj = dt(int(year), int(month_num), 1, 0, 0, 0)
        
        if int(month_num) == 12:
            date_to_obj = dt(int(year) + 1, 1, 1, 0, 0, 0)
        else:
            date_to_obj = dt(int(year), int(month_num) + 1, 1, 0, 0, 0)
        
        # فلتر يدعم string و datetime
        start_str = f"{month}-01T00:00:00"
        end_str = date_to_obj.strftime("%Y-%m-%dT00:00:00")
        
        date_filter = {
            "$or": [
                {"created_at": {"$gte": start_str, "$lt": end_str}},
                {"created_at": {"$gte": date_from_obj, "$lt": date_to_obj}}
            ]
        }
        
        # إضافة فلتر التاريخ للـ query
        if "$or" in query_filter:
            query_filter["$and"] = [{"$or": query_filter.pop("$or")}, date_filter]
        else:
            query_filter.update(date_filter)
    
    # استعلامات aggregate للسرعة
    pipeline = [
        {'$match': query_filter},
        {
            '$facet': {
                'total': [{'$count': 'count'}],
                'fixed': [
                    {'$match': {'status': 'تم الإصلاح'}},
                    {'$count': 'count'}
                ],
                'asphalt_remaining': [
                    {'$match': {'$or': [
                        {'status': 'بانتظار الأسفلت'},
                        {'status': {'$regex': '.*متبقي.*[اأإ]سفلت.*', '$options': 'i'}}
                    ]}},
                    {'$count': 'count'}
                ],
                'asphalt_reports': [
                    {'$match': {'report_type': {'$in': ['أسفلت', 'اسفلت', 'asphalt', 'Asphalt']}}},
                    {
                        '$group': {
                            '_id': None,
                            'total': {'$sum': 1},
                            'licensed': {
                                '$sum': {
                                    '$cond': [
                                        {
                                            '$regexMatch': {
                                                'input': {'$ifNull': ['$license_number', '']},
                                                'regex': '[0-9]'
                                            }
                                        },
                                        1,
                                        0
                                    ]
                                }
                            },
                            'unlicensed': {
                                '$sum': {
                                    '$cond': [
                                        {
                                            '$not': {
                                                '$regexMatch': {
                                                    'input': {'$ifNull': ['$license_number', '']},
                                                    'regex': '[0-9]'
                                                }
                                            }
                                        },
                                        1,
                                        0
                                    ]
                                }
                            }
                        }
                    }
                ],
                'tile_reports': [
                    {'$match': {'report_type': {'$in': ['بلاط', 'tile', 'Tile']}}},
                    {
                        '$group': {
                            '_id': None,
                            'total': {'$sum': 1},
                            'licensed': {
                                '$sum': {
                                    '$cond': [
                                        {
                                            '$regexMatch': {
                                                'input': {'$ifNull': ['$license_number', '']},
                                                'regex': '[0-9]'
                                            }
                                        },
                                        1,
                                        0
                                    ]
                                }
                            },
                            'unlicensed': {
                                '$sum': {
                                    '$cond': [
                                        {
                                            '$not': {
                                                '$regexMatch': {
                                                    'input': {'$ifNull': ['$license_number', '']},
                                                    'regex': '[0-9]'
                                                }
                                            }
                                        },
                                        1,
                                        0
                                    ]
                                }
                            }
                        }
                    }
                ],
                'terrestrial_reports': [
                    {'$match': {'report_type': {'$in': ['ترابي', 'terrestrial', 'Terrestrial']}}},
                    {
                        '$group': {
                            '_id': None,
                            'total': {'$sum': 1},
                            'licensed': {
                                '$sum': {
                                    '$cond': [
                                        {
                                            '$regexMatch': {
                                                'input': {'$ifNull': ['$license_number', '']},
                                                'regex': '[0-9]'
                                            }
                                        },
                                        1,
                                        0
                                    ]
                                }
                            },
                            'unlicensed': {
                                '$sum': {
                                    '$cond': [
                                        {
                                            '$not': {
                                                '$regexMatch': {
                                                    'input': {'$ifNull': ['$license_number', '']},
                                                    'regex': '[0-9]'
                                                }
                                            }
                                        },
                                        1,
                                        0
                                    ]
                                }
                            }
                        }
                    }
                ],
                'by_type': [
                    {
                        '$group': {
                            '_id': '$report_type',
                            'count': {'$sum': 1}
                        }
                    }
                ]
            }
        }
    ]
    
    # استخدام allowDiskUse للعمليات الكبيرة
    result = await db.reports.aggregate(pipeline, allowDiskUse=True).to_list(1)
    
    # جلب إحصائيات التوصيلات (مياه وصرف) مع تطبيق نفس الفلاتر الهرمية
    conn_filter = query_filter.copy()
    # تحويل governorate إلى area في استعلام التوصيلات
    if "governorate" in conn_filter:
        conn_filter["area"] = conn_filter.pop("governorate")
    if "$or" in conn_filter:
        for branch in conn_filter["$or"]:
            if "governorate" in branch:
                branch["area"] = branch.pop("governorate")

    # إضافة فلترة التاريخ للتوصيلات إذا وجد الشهر
    if month:
        date_filter_conn = {
            "$or": [
                {"created_at": {"$gte": start_str, "$lt": end_str}},
                {"created_at": {"$gte": date_from_obj, "$lt": date_to_obj}}
            ]
        }
        conn_filter.update(date_filter_conn)

    water_total = await db.water_connections.count_documents(conn_filter)
    water_fixed = await db.water_connections.count_documents({**conn_filter, "request_status": "مكتمل"})
    
    sewage_total = await db.sewage_connections.count_documents(conn_filter)
    sewage_fixed = await db.sewage_connections.count_documents({**conn_filter, "request_status": "مكتمل"})

    if not result:
        total = water_total + sewage_total
        fixed = water_fixed + sewage_fixed
        return {
            'total': total,
            'fixed': fixed,
            'asphalt_remaining': 0,
            'licensed': 0,
            'unlicensed': 0,
            'terrestrial': 0,
            'tile': 0,
            'asphalt': 0,
            'tile_licensed': 0,
            'tile_unlicensed': 0,
            'terrestrial_licensed': 0,
            'terrestrial_unlicensed': 0,
            'water_connections': water_total,
            'sewage_connections': sewage_total,
            'by_type': {'توصيلة مياه': water_total, 'توصيلة صرف صحي': sewage_total}
        }
    
    data = result[0]
    
    # استخراج النتائج
    total = (data['total'][0]['count'] if data['total'] else 0) + water_total + sewage_total
    fixed = (data['fixed'][0]['count'] if data['fixed'] else 0) + water_fixed + sewage_fixed
    asphalt_remaining = data['asphalt_remaining'][0]['count'] if data['asphalt_remaining'] else 0
    
    asphalt_stats = data['asphalt_reports'][0] if data['asphalt_reports'] else {}
    licensed = asphalt_stats.get('licensed', 0)
    unlicensed = asphalt_stats.get('unlicensed', 0)
    
    tile_stats = data['tile_reports'][0] if data['tile_reports'] else {}
    tile_licensed = tile_stats.get('licensed', 0)
    tile_unlicensed = tile_stats.get('unlicensed', 0)
    
    # إحصائيات التراب
    terrestrial_stats = data['terrestrial_reports'][0] if data['terrestrial_reports'] else {}
    terrestrial_licensed = terrestrial_stats.get('licensed', 0)
    terrestrial_unlicensed = terrestrial_stats.get('unlicensed', 0)
    
    # حساب الأنواع
    by_type = {item['_id']: item['count'] for item in data['by_type']}
    terrestrial = by_type.get('ترابي', 0)
    tile = by_type.get('بلاط', 0)
    # دعم كل من "أسفلت" و "إسفلت"
    asphalt = by_type.get('أسفلت', 0) + by_type.get('إسفلت', 0) + by_type.get('اسفلت', 0)
    
    result = {
        'total': total,
        'fixed': fixed,
        'asphalt_remaining': asphalt_remaining,
        'licensed': licensed,
        'unlicensed': unlicensed,
        'tile_licensed': tile_licensed,
        'tile_unlicensed': tile_unlicensed,
        'terrestrial_licensed': terrestrial_licensed,
        'terrestrial_unlicensed': terrestrial_unlicensed,
        'terrestrial': terrestrial,
        'tile': tile,
        'asphalt': asphalt,
        'water_connections': water_total,
        'sewage_connections': sewage_total,
        'by_type': {**by_type, 'توصيلة مياه': water_total, 'توصيلة صرف صحي': sewage_total}
    }
    
    # Store in cache
    stats_cache[cache_key] = (result, current_time)
    
    return result


@api_router.get("/reports/governorate-72h-counts")
async def get_governorate_48h_counts(
    project: Optional[str] = Query(None),
    category: Optional[str] = Query("reports"),
    base_date: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """جلب عدد البلاغات لكل محافظة في آخر 72 ساعة بناءً على timestamp"""
    # جلب جميع المحافظات المتاحة للمستخدم لضمان ظهورها جميعاً حتى لو كان العدد 0
    available_govs = []
    try:
        # جلب المحافظات من قاعدة البيانات
        custom_govs = await db.project_governorates.find({}, {"_id": 0}).to_list(2000)
        
        # دمج المحافظات المخصصة
        all_possible_govs = {}
        for c in custom_govs:
            p = c.get('project')
            n = c.get('name')
            if p and n:
                if p not in all_possible_govs: all_possible_govs[p] = []
                if n not in all_possible_govs[p]: all_possible_govs[p].append(n)
        
        # تحديد المشاريع المستهدفة بناءً على فلتر المستخدم وصلاحياته
        target_projects = []
        if current_user.role == "admin":
            if project: target_projects = [project]
            else: target_projects = list(all_possible_govs.keys())
        else:
            if project: target_projects = [project]
            else: target_projects = current_user.projects or []
            
        for tp in target_projects:
            # مطابقة مرنة للمشروع
            for p_key, p_govs in all_possible_govs.items():
                if tp == p_key or (project and tp in p_key):
                    for g in p_govs:
                        # التحقق من صلاحيات المحافظة للمستوى 3
                        if current_user.role == "admin" or not current_user.governorates or g in current_user.governorates:
                            if (g, p_key) not in available_govs:
                                available_govs.append((g, p_key))
    except Exception as e:
        import logging
        logging.error(f"Error building available govs list: {str(e)}")

    from datetime import timedelta, datetime
    
    # حساب الوقت المرجعي (الآن أو التاريخ المختار)
    if base_date:
        try:
            base_dt = datetime.fromisoformat(base_date.replace('Z', '+00:00'))
            if base_dt.tzinfo:
                base_dt = base_dt.replace(tzinfo=None)
            reference_time = base_dt
        except:
            reference_time = datetime.utcnow()
    else:
        reference_time = datetime.utcnow()

    # حساب الوقت قبل 24 ساعة بالضبط
    seventy_two_hours_ago = reference_time - timedelta(hours=24)

    start_time = None
    end_time = None
    if base_date:
        try:
            base_dt = datetime.fromisoformat(base_date.replace('Z', '+00:00'))
            if base_dt.tzinfo:
                base_dt = base_dt.replace(tzinfo=None)
            start_time = base_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            end_time = base_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
        except:
            pass
    
    # بناء الاستعلام الأساسي 
    query = {"is_deleted": {"$ne": True}}
    
    # إضافة فلتر المشروع - بحث مرن جداً بالكلمات المفتاحية
    if project:
        query["project"] = get_flexible_project_query(project)
    
    # التصفية الهرمية حسب الصلاحيات
    if current_user.role != 'admin':
        # للمستخدمين (مستوى 2 و 3): رؤية كل ما هو متاح في نطاق مشروعاتهم ومحافظاتهم
        if current_user.projects:
            # فلترة بالمشاريع المتاحة
            if project:
                # التحقق من أن المشروع المختار ضمن مشاريع المستخدم بمرونة
                has_permission = False
                for up in current_user.projects:
                    up_keywords = [k for k in up.replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'أعمال', 'إصلاح']]
                    proj_keywords = [k for k in project.replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'أعمال', 'إصلاح']]
                    is_match = any(k in project for k in up_keywords) or any(k in up for k in proj_keywords)
                    
                    if is_match:
                        has_permission = True
                        break
                if not has_permission:
                    return []
            else:
                # إذا لم يختر مشروع، نفلتر بجميع مشاريعه بمرونة
                query.update(get_flexible_in_query(current_user.projects, "project"))
        
        # التصفية الهرمية
        user_governorates = current_user.governorates if hasattr(current_user, 'governorates') and current_user.governorates else []
        
        # تطبيق الفلترة الهرمية الشاملة (Recursive)
        hierarchy_filter = await get_hierarchy_filter(current_user)
        
        if len(user_governorates) > 0:
            gov_patterns = []
            for g in user_governorates:
                p = normalize_arabic_regex(g)
                gov_patterns.append(p)
            gov_regex = f"({'|'.join(gov_patterns)})"
            
            query.update(hierarchy_filter)
            query['governorate'] = {'$regex': gov_regex, '$options': 'i'}
        else:
            query.update(hierarchy_filter)
    
    # تحديد ما إذا كان المشروع مخصص للتوصيلات
    is_connection_only = False
    if project and any(kw in project for kw in ['ايصال', 'إيصال', 'توصيل']):
        is_connection_only = True

    # جلب البلاغات إذا كانت الفئة 'reports' أو 'all' أو إذا كان المشروع ليس مخصصاً للتوصيلات فقط
    all_reports = []
    
    # استعلام التوصيلات (تستخدم area بدلاً من governorate)
    conn_query = query.copy()
    if "governorate" in conn_query:
        conn_query["area"] = conn_query.pop("governorate")
    if "$or" in conn_query:
        for branch in conn_query["$or"]:
            if "governorate" in branch:
                branch["area"] = branch.pop("governorate")

    # جلب البلاغات العادية
    if category in ['reports', 'all', None]:
        reports_list = await db.reports.find(query, {"_id": 0, "governorate": 1, "created_at": 1, "added_at": 1, "start_date": 1, "project": 1}).to_list(5000)
        for r in reports_list:
            r['item_type'] = 'report'
            all_reports.append(r)
    
    # جلب التوصيلات مع الترتيب
    if category in ['water_connections', 'all'] or (is_connection_only and category in ['reports', None]):
        water_conns = await db.water_connections.find(conn_query, {"_id": 0, "area": 1, "governorate": 1, "created_at": 1, "added_at": 1, "project": 1}).sort("created_at", -1).to_list(2000)
        for c in water_conns:
            c['governorate'] = c.get('governorate') or c.get('area')
            c['item_type'] = 'water_connection'
            all_reports.append(c)
        
    if category in ['sewage_connections', 'all'] or (is_connection_only and category in ['reports', None]):
        sewage_conns = await db.sewage_connections.find(conn_query, {"_id": 0, "area": 1, "governorate": 1, "created_at": 1, "added_at": 1, "project": 1}).sort("created_at", -1).to_list(2000)
        for c in sewage_conns:
            c['governorate'] = c.get('governorate') or c.get('area')
            c['item_type'] = 'sewage_connection'
            all_reports.append(c)
    
    # فلترة وتجميع حسب 72 ساعة بشكل يدوي
    group_counts = {} # (gov, proj) -> count
    
    # تهيئة العدادات لجميع المحافظات المتاحة بـ 0 لضمان ظهورها
    for gov, proj in available_govs:
        group_counts[(gov, proj)] = 0
    for report in all_reports:
        start_date_val = report.get('start_date')
        created_at = report.get('created_at')
        added_at = report.get('added_at')
        governorate = report.get('governorate')
        project_val = report.get('project', 'غير محدد')
        
        if not governorate:
            continue
        
        # تحويل التاريخ إلى datetime للمقارنة
        report_date = None
        
        # أولوية 1: start_date
        if isinstance(start_date_val, datetime):
            report_date = start_date_val.replace(tzinfo=None) if start_date_val.tzinfo else start_date_val
        elif isinstance(start_date_val, str) and start_date_val:
            try:
                clean_date = start_date_val.replace('+00:00', '').replace('Z', '').split('.')[0]
                report_date = datetime.fromisoformat(clean_date)
            except Exception:
                pass
        
        # أولوية 2: created_at
        if not report_date and isinstance(created_at, datetime):
            report_date = created_at.replace(tzinfo=None) if created_at.tzinfo else created_at
        elif not report_date and isinstance(created_at, str):
            try:
                clean_date = created_at.replace('+00:00', '').replace('Z', '').split('.')[0]
                report_date = datetime.fromisoformat(clean_date)
            except Exception:
                pass
        
        if not report_date and isinstance(added_at, datetime):
            report_date = added_at.replace(tzinfo=None) if added_at.tzinfo else added_at
            
        if base_date:
            if report_date and start_time and end_time and start_time <= report_date <= end_time:
                key = (governorate, project_val)
                group_counts[key] = group_counts.get(key, 0) + 1
            continue
        
        # التحقق من أن التاريخ ضمن 72 ساعة
        if report_date and report_date >= seventy_two_hours_ago:
            key = (governorate, project_val)
            group_counts[key] = group_counts.get(key, 0) + 1
    
    # تحويل النتيجة إلى قائمة
    result = [
        {"governorate": gov, "project": proj, "count": count} 
        for (gov, proj), count in group_counts.items()
    ]
    
    # إخفاء المحافظات التي عددها 0 لتقليل الزحام دائماً
    result = [r for r in result if r['count'] > 0]
        
    result.sort(key=lambda x: x['count'], reverse=True)
    return result


@api_router.get("/reports/last-72-hours-list")
async def get_reports_last_72_hours_list(
    project: Optional[str] = Query(None),
    governorate: Optional[str] = Query(None),
    category: Optional[str] = Query("reports"),
    base_date: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """جلب قائمة البلاغات خلال آخر 72 ساعة (بناءً على created_at timestamp)"""
    from datetime import timedelta, datetime
    
    # حساب الوقت المرجعي
    if base_date:
        try:
            base_dt = datetime.fromisoformat(base_date.replace('Z', '+00:00'))
            if base_dt.tzinfo:
                base_dt = base_dt.replace(tzinfo=None)
            reference_time = base_dt
        except:
            reference_time = datetime.utcnow()
    else:
        reference_time = datetime.utcnow()

    # حساب الوقت قبل 24 ساعة بالضبط
    seventy_two_hours_ago = reference_time - timedelta(hours=24)

    start_time = None
    end_time = None
    if base_date:
        try:
            base_dt = datetime.fromisoformat(base_date.replace('Z', '+00:00'))
            if base_dt.tzinfo:
                base_dt = base_dt.replace(tzinfo=None)
            start_time = base_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            end_time = base_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
        except:
            pass
    
    # بناء الاستعلام الأساسي 
    query = {"is_deleted": {"$ne": True}}
    
    # إضافة فلاتر المشروع والمحافظة (مع البحث المرن)
    if project:
        query["project"] = get_flexible_project_query(project)
            
    if governorate and governorate not in ["الكل", "جميع المحافظات", "كل المحافظات"]:
        query["governorate"] = {'$regex': f"({normalize_arabic_regex(governorate)})", '$options': 'i'}
    
    # التصفية الهرمية حسب الصلاحيات
    if current_user.role != "admin":
        user_governorates = current_user.governorates if hasattr(current_user, 'governorates') and current_user.governorates else []
        has_all_govs = any(g in ["الكل", "جميع المحافظات", "كل المحافظات"] for g in user_governorates)
        
        # 1. التحقق من صلاحية المشروع بمرونة
        if current_user.projects:
            has_permission = False
            for up in current_user.projects:
                up_keywords = [k for k in up.replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'أعمال', 'إصلاح']]
                proj_keywords = [k for k in (project or "").replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'أعمال', 'إصلاح']]
                if project and (any(k in project for k in up_keywords) or any(k in up for k in proj_keywords)):
                    has_permission = True
                    break
            
            if project and not has_permission:
                return {"reports": []}
            elif not project:
                # استخدام البحث المرن بدلاً من المطابقة التامة لضمان ظهور جميع البلاغات
                query.update(get_flexible_in_query(current_user.projects, "project"))
        
        # 2. تجميع معرفات المستخدمين (الهرمية)
        sub_users_docs = await db.users.find(
            {"$or": [{"created_by": current_user.id}, {"created_by": current_user.username}]}, 
            {"_id": 0, "id": 1, "username": 1}
        ).to_list(1000)
        sub_user_ids = []
        for u in sub_users_docs:
            if u.get("id"): sub_user_ids.append(u["id"])
            if u.get("username"): sub_user_ids.append(u["username"])
        all_authorized_creators = [current_user.id, current_user.username] + sub_user_ids

        # 3. تطبيق فلتر الهرمية (يشمل البلاغات المضافة من الأدمن والمديرين في محافظات المستخدم)
        permissions = getattr(current_user, 'permissions', [])
        has_reports_review = "reports_review" in permissions or len(get_projects_with_permission(current_user, "reports_review")) > 0
        # التحقق من reports_view سواء في الصلاحيات العامة أو في صلاحيات المشروع
        has_reports_view = "reports_view" in permissions or len(get_projects_with_permission(current_user, "reports_view")) > 0
        if not getattr(current_user, 'can_create_subusers', False) and not has_reports_view and not has_reports_review:
            # المستوى الثالث: يرى بلاغاته + بلاغات الأدمن والمديرين في محافظاته
            # جلب معرفات الأدمن والمديرين
            admin_ids = [current_user.id, current_user.username]
            admins_cursor = db.users.find(
                {"$or": [{"role": "admin"}, {"can_create_subusers": True}]},
                {"id": 1, "username": 1}
            )
            async for adm in admins_cursor:
                if adm.get("id"): admin_ids.append(adm["id"])
                if adm.get("username"): admin_ids.append(adm["username"])
            admin_ids.append("مكتب بيت الخبرة للاستشارات الهندسية")
            query['created_by'] = {'$in': list(set(admin_ids))}
            if not has_all_govs and user_governorates:
                if governorate and governorate not in ["الكل", "جميع المحافظات", "كل المحافظات"]:
                    norm_req = normalize_arabic(governorate)
                    if not any(normalize_arabic(g) == norm_req for g in user_governorates):
                        return {"reports": []}
                else:
                    gov_patterns = [normalize_arabic_regex(g) for g in user_governorates]
                    query['governorate'] = {'$regex': f"({'|'.join(gov_patterns)})", '$options': 'i'}
        else:
            # المستوى الثاني يرى كافة البلاغات في المحافظات المسندة إليه
            if has_all_govs:
                pass
            elif user_governorates:
                if governorate and governorate not in ["الكل", "جميع المحافظات", "كل المحافظات"]:
                    norm_req = normalize_arabic(governorate)
                    if not any(normalize_arabic(g) == norm_req for g in user_governorates):
                        return {"reports": []}
                else:
                    gov_patterns = [normalize_arabic_regex(g) for g in user_governorates]
                    query['governorate'] = {'$regex': f"({'|'.join(gov_patterns)})", '$options': 'i'}
    
    # تحديد ما إذا كان المشروع مخصص للتوصيلات
    is_connection_only = False
    if project and any(kw in project for kw in ['ايصال', 'إيصال', 'توصيل']):
        is_connection_only = True

    # جلب البيانات بناءً على الفئة
    all_reports = []
    projection = {"_id": 0, "images": 0}

    if category in ['reports', 'all', None]:
        reports_list = await db.reports.find(query, projection).sort("created_at", -1).to_list(10000)
        all_reports.extend(reports_list)
    
    # جلب التوصيلات خلال آخر 72 ساعة (مع التحقق من مشاريع الإيصال)
    if category in ['water_connections', 'sewage_connections', 'all'] or (is_connection_only and category in ['reports', None]):
        conn_query = query.copy()
        if "governorate" in conn_query:
            conn_query["area"] = conn_query.pop("governorate")
        
        # الاعتماد على الفلترة اليدوية اللاحقة لضمان الدقة وتوحيد المنطق مع إحصائيات المحافظات
        pass
        
        if category in ['water_connections', 'all']:
            water_conns = await db.water_connections.find(conn_query, projection).sort("created_at", -1).to_list(1000)
            for c in water_conns:
                c['report_type'] = 'توصيلة مياه'
                c['report_number'] = c.get('request_number') or c.get('ccb_report_number')
                c['status'] = c.get('request_status')
                all_reports.append(c)
        
        if category in ['sewage_connections', 'all']:
            sewage_conns = await db.sewage_connections.find(conn_query, projection).sort("created_at", -1).to_list(1000)
            for c in sewage_conns:
                c['report_type'] = 'توصيلة صرف صحي'
                c['report_number'] = c.get('request_number') or c.get('ccb_report_number')
                c['status'] = c.get('request_status')
                all_reports.append(c)
    
    # فلترة البلاغات حسب 72 ساعة بشكل يدوي (يستخدم start_date = تاريخ مباشرة البلاغ)
    reports = []
    for report in all_reports:
        start_date_val = report.get('start_date')
        created_at = report.get('created_at')
        added_at = report.get('added_at')
        
        # تحويل التاريخ إلى datetime للمقارنة
        report_date = None
        
        # أولوية 1: start_date (تاريخ مباشرة البلاغ)
        if isinstance(start_date_val, datetime):
            report_date = start_date_val.replace(tzinfo=None) if start_date_val.tzinfo else start_date_val
        elif isinstance(start_date_val, str) and start_date_val:
            try:
                clean_date = start_date_val.replace('+00:00', '').replace('Z', '').split('.')[0]
                report_date = datetime.fromisoformat(clean_date)
            except Exception:
                pass
        
        # أولوية 2: created_at fallback
        if not report_date and isinstance(created_at, datetime):
            report_date = created_at.replace(tzinfo=None) if created_at.tzinfo else created_at
        elif not report_date and isinstance(created_at, str):
            try:
                clean_date = created_at.replace('+00:00', '').replace('Z', '').split('.')[0]
                report_date = datetime.fromisoformat(clean_date)
            except Exception:
                pass
        
        if not report_date and isinstance(added_at, datetime):
            report_date = added_at.replace(tzinfo=None) if added_at.tzinfo else added_at
            
        if base_date:
            if report_date and start_time and end_time and start_time <= report_date <= end_time:
                reports.append(report)
            continue
        
        if report_date and report_date >= seventy_two_hours_ago:
            reports.append(report)
    
    # معالجة التواريخ للإرجاع
    for report in reports:
        if isinstance(report.get('created_at'), str):
            try:
                report['created_at'] = datetime.fromisoformat(report['created_at'].replace('Z', '+00:00'))
            except:
                pass
        if isinstance(report.get('report_date'), str):
            try:
                report['report_date'] = datetime.fromisoformat(report['report_date'].replace('Z', '+00:00'))
            except:
                pass
    
    return {"reports": reports}


@api_router.get("/reports/last-72-hours")
async def get_reports_last_72_hours(
    project: Optional[str] = Query(None),
    governorate: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """الحصول على عدد البلاغات خلال 24 ساعة لكل محافظة أو محافظة محددة"""
    from datetime import timedelta
    
    # حساب الوقت قبل 24 ساعة (timestamp)
    seventy_two_hours_ago = datetime.now(timezone.utc) - timedelta(hours=24)
    
    # Use $or to search in both created_at and added_at fields
    query = {
        "is_deleted": {"$ne": True},
        "$or": [
            {"created_at": {"$gte": seventy_two_hours_ago.isoformat()}},
            {"added_at": {"$gte": seventy_two_hours_ago}}
        ]
    }
    
    # إضافة فلاتر المشروع والمحافظة (مع البحث المرن)
    if project:
        query["project"] = get_flexible_project_query(project)
            
    if governorate:
        query["governorate"] = {'$regex': f"({normalize_arabic_regex(governorate)})", '$options': 'i'}
        
    # استعلام التوصيلات
    conn_query = {
        "$or": [
            {"created_at": {"$gte": seventy_two_hours_ago.isoformat()}},
            {"added_at": {"$gte": seventy_two_hours_ago}}
        ]
    }
    if project:
        conn_query["project"] = get_flexible_project_query(project)
            
    if governorate and governorate not in ["الكل", "جميع المحافظات", "كل المحافظات"]:
        conn_query["area"] = {'$regex': f"({normalize_arabic_regex(governorate)})", '$options': 'i'} 
    
    # التصفية الهرمية حسب الصلاحيات
    if current_user.role != "admin":
        permissions = getattr(current_user, 'permissions', [])
        has_reports_review = "reports_review" in permissions or len(get_projects_with_permission(current_user, "reports_review")) > 0
        if not getattr(current_user, 'can_create_subusers', False) and "reports_view" not in permissions and not has_reports_review:
            query["created_by"] = {"$in": [current_user.id, current_user.username]}
            conn_query["created_by"] = {"$in": [current_user.id, current_user.username]}
            if current_user.governorates and not any(g in ["الكل", "جميع المحافظات", "كل المحافظات"] for g in current_user.governorates):
                if governorate and governorate not in ["الكل", "جميع المحافظات", "كل المحافظات"]:
                    norm_req = normalize_arabic(governorate)
                    if not any(normalize_arabic(g) == norm_req for g in current_user.governorates):
                        return {"governorate": governorate, "count": 0} if governorate else []
                else:
                    gov_patterns = [normalize_arabic_regex(g) for g in current_user.governorates]
                    query['governorate'] = {'$regex': f"({'|'.join(gov_patterns)})", '$options': 'i'}
            if current_user.projects:
                if project:
                    has_proj_perm = False
                    for up in current_user.projects:
                        up_kws = [k for k in up.replace('-', ' ').split() if len(k)>2 and k not in ['مشروع','أعمال','إصلاح']]
                        p_kws = [k for k in project.replace('-', ' ').split() if len(k)>2 and k not in ['مشروع','أعمال','إصلاح']]
                        if any(k in project for k in up_kws) or any(k in up for k in p_kws):
                            has_proj_perm = True; break
                    if not has_proj_perm:
                        return {"governorate": governorate, "count": 0} if governorate else []
                else:
                    query["project"] = {"$in": current_user.projects}
        else:
            # Level 2 and above
            if current_user.projects:
                if project:
                    has_proj_perm = False
                    for up in current_user.projects:
                        up_kws = [k for k in up.replace('-', ' ').split() if len(k)>2 and k not in ['مشروع','أعمال','إصلاح']]
                        p_kws = [k for k in project.replace('-', ' ').split() if len(k)>2 and k not in ['مشروع','أعمال','إصلاح']]
                        if any(k in project for k in up_kws) or any(k in up for k in p_kws):
                            has_proj_perm = True; break
                    if not has_proj_perm:
                        return {"governorate": governorate, "count": 0} if governorate else []
                else:
                    query["project"] = {"$in": current_user.projects}
            
            if current_user.governorates and not any(g in ["الكل", "جميع المحافظات", "كل المحافظات"] for g in current_user.governorates):
                if governorate and governorate not in ["الكل", "جميع المحافظات", "كل المحافظات"]:
                    norm_req = normalize_arabic(governorate)
                    if not any(normalize_arabic(g) == norm_req for g in current_user.governorates):
                        return {"governorate": governorate, "count": 0}
                else:
                    gov_patterns = [normalize_arabic_regex(g) for g in current_user.governorates]
                    query['governorate'] = {'$regex': f"({'|'.join(gov_patterns)})", '$options': 'i'}
    
    # إذا تم تحديد محافظة، نرجع العدد مباشرة
    if governorate and governorate not in ["الكل", "جميع المحافظات", "كل المحافظات"]:
        count = await db.reports.count_documents(query)
        return {"governorate": governorate, "count": count, "project": project}
    
    # جمع البلاغات وتجميعها حسب المحافظة
    pipeline = [
        {"$match": query},
        {
            "$group": {
                "_id": "$governorate",
                "count": {"$sum": 1}
            }
        }
    ]
    
    result = await db.reports.aggregate(pipeline).to_list(100)
    
    # إضافة التوصيلات للتجميع
    water_count = await db.water_connections.count_documents(conn_query)
    sewage_count = await db.sewage_connections.count_documents(conn_query)
    
    # تحويل النتيجة إلى قائمة
    governorate_counts = [{"governorate": item['_id'], "count": item['count']} for item in result]
    
    # إذا كان هناك توصيلات، يمكننا إضافتها تحت محافظة "توصيلات" أو توزيعها إذا كان لدينا معلومات المحافظة
    # للتبسيط الآن سنضيفها للعدد الإجمالي إذا لم يكن هناك فلتر محافظة، أو نضيفها لنتائج البحث
    if not governorate:
        # إضافة عنصر للتوصيلات
        if water_count > 0:
            governorate_counts.append({"governorate": "توصيلات مياه", "count": water_count})
        if sewage_count > 0:
            governorate_counts.append({"governorate": "توصيلات صرف", "count": sewage_count})
    
    return governorate_counts


@api_router.get("/reports/{report_id}/images")
async def get_report_images(report_id: str, current_user: User = Depends(get_current_user)):
    """جلب صور بلاغ معين فقط (Lazy Loading)"""
    query = {"id": report_id, "is_deleted": {"$ne": True}}
    
    # فلترة حسب صلاحيات المحافظات
    if current_user.role != "admin" or len(current_user.governorates) > 0:
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
    
    report_doc = await db.reports.find_one(query, {"_id": 0, "images": 1})
    if not report_doc:
        raise HTTPException(status_code=404, detail="Report not found")
    
    return {"images": report_doc.get('images', [])}


@api_router.get("/users/level3")
async def get_level3_users(
    project: Optional[str] = Query(None),
    governorate: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """
    إرجاع قائمة المستخدمين حسب المشروع والمحافظة - ديناميكي.
    - Admin: بدون قيود
    - غير Admin: يُقيَّد بمشاريع ومحافظات المستخدم الحالي
    """
    try:
        is_admin = current_user.role == "admin"
        is_level_2 = bool(current_user.can_create_subusers)
        user_permissions = current_user.permissions or []
        has_view_all = "view_all_invoices" in user_permissions or user_has_any_project_permission(current_user, "view_all_invoices")
        has_reviewer_role = user_has_any_project_permission(current_user, "reports_review") or "reports_review" in user_permissions
        has_view_gov = "view_governorate_data" in user_permissions or user_has_any_project_permission(current_user, "view_governorate_data")
        
        if not (is_admin or is_level_2 or has_view_all or has_reviewer_role or has_view_gov):
            return {"users": []}
        
        my_projects = list(current_user.projects or [])
        my_governorates = list(current_user.governorates or [])
        
        # التحقق من أن المشروع المطلوب ضمن مشاريع المستخدم (إلا Admin)
        if project and not is_admin and project not in my_projects:
            return {"users": []}
        
        # التحقق من أن المحافظة ضمن محافظات المستخدم (إلا إذا كان لديه كل المحافظات = قائمة فارغة)
        if governorate and not is_admin and my_governorates and governorate not in my_governorates:
            return {"users": []}
        
        # القاعدة: جميع المستخدمين النشطين (ما عدا الأدمن) ولديهم محافظات محددة (المراقبين الميدانيين)
        query = {
            "is_active": True,
            "role": {"$ne": "admin"},
            "can_create_subusers": {"$ne": True},
            "governorates": {"$exists": True, "$ne": [], "$not": {"$size": 0}}
        }
        
        # فلترة حسب المشروع
        if project:
            query["projects"] = project
        elif not is_admin:
            # إذا لم يُحدَّد مشروع: قيِّد بمشاريع المستخدم الحالي
            if not my_projects:
                return {"users": []}
            query["projects"] = {"$in": my_projects}
        
        all_users = await db.users.find(
            query,
            {"_id": 0, "id": 1, "username": 1, "full_name": 1, "governorates": 1, "projects": 1, "created_by": 1}
        ).to_list(500)
        
        # فلترة حسب محافظة محددة (على مستوى Python)
        if governorate:
            filtered = []
            for u in all_users:
                user_govs = u.get("governorates") or []
                if governorate in user_govs:
                    filtered.append(u)
            all_users = filtered
        elif not is_admin and my_governorates:
            # لم تُحدَّد محافظة: قيِّد بمحافظات المستخدم الحالي
            filtered = []
            for u in all_users:
                user_govs = u.get("governorates") or []
                if any(g in my_governorates for g in user_govs):
                    filtered.append(u)
            all_users = filtered
        
        # المستوى 2 بدون صلاحية مراجعة: يرى تابعيه + نفسه فقط
        if not is_admin and is_level_2 and not has_reviewer_role:
            all_users = [u for u in all_users if u.get("created_by") == current_user.id or u.get("id") == current_user.id]
        
        # تنظيف الحقول المرجعة
        result = [{"id": u["id"], "username": u["username"], "full_name": u.get("full_name", u["username"])} for u in all_users]
        
        return {"users": result}
        
    except Exception as e:
        logger.error(f"Error getting level3 users: {str(e)}")
        return {"users": []}


@api_router.get("/reports/my-count")
async def get_my_reports_count(current_user: User = Depends(get_current_user)):
    """
    إرجاع عدد البلاغات التي أدخلها المستخدم الحالي
    """
    try:
        # البحث بالـ user_id أو username
        query = {
            "is_deleted": {"$ne": True},
            "$or": [
                {"created_by": current_user.id},
                {"created_by": current_user.username}
            ]
        }
        
        count = await db.reports.count_documents(query)
        
        return {"count": count}
        
    except Exception as e:
        logger.error(f"Error getting my reports count: {str(e)}")
        return {"count": 0}


@api_router.get("/reports/pending-review-count")
async def get_pending_review_count(current_user: User = Depends(get_current_user)):
    """
    إرجاع عدد البلاغات التي بانتظار المراجعة - ديناميكي حسب الصلاحيات والمحافظات
    - Admin: يرى جميع البلاغات
    - المراجع: يرى البلاغات في المشاريع التي يملك فيها reports_review (+ فلتر المحافظات)
    - أي مستخدم آخر: يرى بلاغاته الشخصية التي لا تزال بانتظار المراجعة
    """
    try:
        base = {
            "is_deleted": {"$ne": True},
            "review_status": {"$in": ["بانتظار المراجعة", "قيد المراجعة", None]}
        }
        
        if current_user.role == "admin":
            count = await db.reports.count_documents(base)
            return {"count": count}
            
        # للمستوى الثالث العادي، يرون فقط بلاغاتهم بانتظار المراجعة
        permissions = getattr(current_user, 'permissions', [])
        has_review = "reports_review" in permissions or len(get_projects_with_permission(current_user, "reports_review")) > 0
        if not getattr(current_user, 'can_create_subusers', False) and "reports_view" not in permissions and not has_review:
            query = {**base, "created_by": {"$in": [current_user.id, current_user.username]}}
            count = await db.reports.count_documents(query)
            return {"count": count}
        
        # المشاريع التي يملك فيها المستخدم صلاحية reports_review
        allowed_projects = get_projects_with_permission(current_user, "reports_review")
        governorates = current_user.governorates or []
        
        or_clauses = []
        # ✅ مراجع: كل بلاغات مشاريعه المصرح له بها (+ فلتر المحافظات)
        if allowed_projects:
            reviewer_clause = {"project": {"$in": allowed_projects}}
            if governorates:
                reviewer_clause["governorate"] = {"$in": governorates}
            or_clauses.append(reviewer_clause)
        
        # ✅ المنشئ: يرى بلاغاته بانتظار المراجعة دائماً
        or_clauses.append({"created_by": current_user.id})
        
        query = {**base, "$or": or_clauses}
        count = await db.reports.count_documents(query)
        return {"count": count}
        
    except Exception as e:
        logger.error(f"Error getting pending review count: {str(e)}")
        return {"count": 0}


@api_router.get("/reports/pending-review-by-governorate")
async def get_pending_review_by_governorate(current_user: User = Depends(get_current_user)):
    """
    إرجاع البلاغات الجديدة (بانتظار المراجعة) مجموعة حسب المحافظة - ديناميكي حسب الصلاحيات
    يشمل: بلاغات المشاريع التي لديه فيها صلاحية مراجعة + بلاغاته الشخصية
    """
    try:
        base = {
            "is_deleted": {"$ne": True},
            "review_status": {"$in": ["بانتظار المراجعة", "قيد المراجعة", None]}
        }
        
        if current_user.role == "admin":
            query = base
        elif not current_user.can_create_subusers and not ("reports_review" in getattr(current_user, 'permissions', []) or len(get_projects_with_permission(current_user, "reports_review")) > 0):
            # للمستوى الثالث العادي (لا يملكون صلاحية مراجعة)، يرون فقط بلاغاتهم
            query = {**base, "created_by": {"$in": [current_user.id, current_user.username]}}
        else:
            # جلب المشاريع التي يملك فيها المستخدم أي صلاحية (رؤية أو مراجعة أو إضافة)
            allowed_projects = set(get_projects_with_permission(current_user, "reports_view")) | \
                               set(get_projects_with_permission(current_user, "reports_review")) | \
                               set(get_projects_with_permission(current_user, "reports_add"))
            
            governorates = current_user.governorates or []
            
            # بناء الاستعلام بناءً على النطاق الجغرافي والمشاريع
            query_parts = []
            
            if allowed_projects:
                p_query = get_flexible_in_query(list(allowed_projects), "project")
                if governorates:
                    p_query["governorate"] = {"$in": governorates}
                query_parts.append(p_query)
            
            # دائماً يرى بلاغاته الشخصية
            query_parts.append({"created_by": current_user.id})
            
            query = {**base, "$or": query_parts}
        
        # Aggregation pipeline للحصول على العدد لكل محافظة ومشروع ومستخدم
        pipeline = [
            {"$match": query},
            {
                "$group": {
                    "_id": {
                        "governorate": "$governorate",
                        "project": "$project",
                        "created_by": "$created_by"
                    },
                    "count": {"$sum": 1}
                }
            },
            {"$sort": {"count": -1}}
        ]
        
        result = await db.reports.aggregate(pipeline).to_list(100)
        
        # تحويل النتيجة
        governorate_counts = []
        for item in result:
            if not item.get('_id'):
                continue
                
            # إيجاد اسم المستخدم إذا أمكن
            created_by = item['_id'].get('created_by')
            created_by_name = None
            if created_by:
                # إذا كان معرف، نجلب الاسم
                user_obj = await db.users.find_one({"$or": [{"id": created_by}, {"username": created_by}]})
                if user_obj:
                    created_by_name = user_obj.get("full_name") or user_obj.get("username")
                else:
                    created_by_name = str(created_by)
                    
            governorate_counts.append({
                "governorate": item['_id'].get('governorate', 'غير محدد'), 
                "project": item['_id'].get('project', 'غير محدد'),
                "created_by": created_by,
                "created_by_name": created_by_name,
                "count": item['count']
            })
        
        return {"data": governorate_counts, "total": sum(item['count'] for item in governorate_counts)}
        
    except Exception as e:
        logger.error(f"Error getting pending review by governorate: {str(e)}")
        return {"data": [], "total": 0}


# ============= نظام إشعارات البلاغات الجديدة =============

@api_router.get("/reports/notifications/unseen")
async def get_unseen_reports(current_user: User = Depends(get_current_user)):
    """
    جلب العناصر الجديدة (غير المرئية) للمستخدم الحالي - يشمل البلاغات والتوصيلات
    - Admin (بيت الخبرة): يرى جميع العناصر من كل المشاريع
    - المستخدم: يرى حسب صلاحياته لكل مشروع
      * البلاغات: reports_notifications
      * توصيلات المياه: water_connections
      * توصيلات الصرف: sewage_connections
    - يُستبعد دائماً العنصر الذي أنشأه المستخدم نفسه
    """
    try:
        is_admin = current_user.role == "admin"
        # المشاريع المسموحة لكل نوع
        # تحسين: إذا كان لديه صلاحية إشعارات البلاغات، يرى التوصيلات أيضاً لهذا المشروع
        report_projects = get_projects_with_permission(current_user, "reports_notifications")
        water_projects = list(set(get_projects_with_permission(current_user, "water_connections") + report_projects))
        sewage_projects = list(set(get_projects_with_permission(current_user, "sewage_connections") + report_projects))
        
        governorates = current_user.governorates or []
        
        # ===== البلاغات =====
        reports = []
        if is_admin or report_projects:
            base_query = {
                "is_deleted": {"$ne": True},
                "$and": [
                    {"$or": [
                        {"seen_by": {"$exists": False}},
                        {"seen_by": {"$ne": current_user.id}}
                    ]},
                    {"$or": [
                        {"deleted_notifications": {"$exists": False}},
                        {"deleted_notifications": {"$ne": current_user.id}}
                    ]}
                ]
            }
            if not is_admin:
                expanded_reports = []
                for p in report_projects:
                    expanded_reports.extend([p, f"مشروع {p}", p.replace("مشروع ", "").strip()])
                base_query["project"] = {"$in": expanded_reports}
                
                if governorates:
                    expanded_govs = []
                    for g in governorates:
                        clean_g = g.replace("محافظة ", "").replace("محافظه ", "").strip()
                        expanded_govs.extend([g, clean_g, f"محافظة {clean_g}", f"محافظه {clean_g}"])
                    base_query["governorate"] = {"$in": expanded_govs}
            reports = await db.reports.find(
                base_query,
                {"_id": 0, "id": 1, "report_number": 1, "governorate": 1, "project": 1,
                 "report_type": 1, "created_at": 1, "added_at": 1, "contractor": 1, "created_by": 1, "created_by_name": 1}
            ).sort("added_at", -1).limit(300).to_list(300)
            for r in reports:
                r["item_type"] = "report"
        
        # ===== توصيلات المياه =====
        water_conns = []
        if is_admin or water_projects:
            wq = {
                "is_deleted": {"$ne": True},
                "$and": [
                    {"$or": [
                        {"seen_by": {"$exists": False}},
                        {"seen_by": {"$ne": current_user.id}}
                    ]},
                    {"$or": [
                        {"deleted_notifications": {"$exists": False}},
                        {"deleted_notifications": {"$ne": current_user.id}}
                    ]}
                ]
            }
            if not is_admin:
                expanded_water = []
                for p in water_projects:
                    expanded_water.extend([p, f"مشروع {p}", p.replace("مشروع ", "").strip()])
                wq["project"] = {"$in": expanded_water}
                if governorates:
                    expanded_govs = []
                    for g in governorates:
                        clean_g = g.replace("محافظة ", "").replace("محافظه ", "").strip()
                        expanded_govs.extend([g, clean_g, f"محافظة {clean_g}", f"محافظه {clean_g}"])
                    wq["area"] = {"$in": expanded_govs}
            water_conns = await db.water_connections.find(
                wq,
                {"_id": 0, "id": 1, "request_number": 1, "account_number": 1, "project": 1,
                 "customer_name": 1, "created_at": 1, "created_by": 1, "created_by_name": 1, "area": 1}
            ).sort("created_at", -1).limit(300).to_list(300)
            for c in water_conns:
                c["item_type"] = "water_connection"
                c["governorate"] = c.get("area") or "غير محدد"
        
        # ===== توصيلات الصرف =====
        sewage_conns = []
        if is_admin or sewage_projects:
            sq = {
                "is_deleted": {"$ne": True},
                "$and": [
                    {"$or": [
                        {"seen_by": {"$exists": False}},
                        {"seen_by": {"$ne": current_user.id}}
                    ]},
                    {"$or": [
                        {"deleted_notifications": {"$exists": False}},
                        {"deleted_notifications": {"$ne": current_user.id}}
                    ]}
                ]
            }
            if not is_admin:
                expanded_sewage = []
                for p in sewage_projects:
                    expanded_sewage.extend([p, f"مشروع {p}", p.replace("مشروع ", "").strip()])
                sq["project"] = {"$in": expanded_sewage}
                if governorates:
                    expanded_govs = []
                    for g in governorates:
                        clean_g = g.replace("محافظة ", "").replace("محافظه ", "").strip()
                        expanded_govs.extend([g, clean_g, f"محافظة {clean_g}", f"محافظه {clean_g}"])
                    sq["area"] = {"$in": expanded_govs}
            sewage_conns = await db.sewage_connections.find(
                sq,
                {"_id": 0, "id": 1, "request_number": 1, "account_number": 1, "project": 1,
                 "customer_name": 1, "created_at": 1, "created_by": 1, "created_by_name": 1, "area": 1}
            ).sort("created_at", -1).limit(300).to_list(300)
            for c in sewage_conns:
                c["item_type"] = "sewage_connection"
                c["governorate"] = c.get("area") or "غير محدد"
        
        # تجميع العناصر حسب المحافظة والمشروع (لإظهار اسم المشروع تحت المحافظة في الهيدر)
        by_gov_project = {} # (gov, project) -> items
        for report in reports:
            gov = report.get('governorate') or 'غير محدد'
            proj = report.get('project') or 'غير محدد'
            by_gov_project.setdefault((gov, proj), []).append(report)
        
        for conn in water_conns:
            gov = conn.get('governorate') or 'غير محدد'
            proj = conn.get('project') or 'غير محدد'
            by_gov_project.setdefault((gov, proj), []).append(conn)
            
        for conn in sewage_conns:
            gov = conn.get('governorate') or 'غير محدد'
            proj = conn.get('project') or 'غير محدد'
            by_gov_project.setdefault((gov, proj), []).append(conn)
        
        total = len(reports) + len(water_conns) + len(sewage_conns)
        
        return {
            "total": total,
            "reports": reports,
            "water_connections": water_conns,
            "sewage_connections": sewage_conns,
            "by_governorate": [
                {"governorate": gov, "project": proj, "count": len(items), "reports": items}
                for (gov, proj), items in by_gov_project.items()
            ],
            "counts": {
                "reports": len(reports),
                "water_connections": len(water_conns),
                "sewage_connections": len(sewage_conns),
            }
        }
        
    except Exception as e:
        logger.error(f"Error getting unseen items: {str(e)}")
        return {"total": 0, "reports": [], "water_connections": [], "sewage_connections": [], "by_governorate": [], "counts": {"reports": 0, "water_connections": 0, "sewage_connections": 0}}


@api_router.get("/reports/notifications/seen")
async def get_seen_reports(current_user: User = Depends(get_current_user)):
    """
    جلب البلاغات المقروءة (التي شاهدها المستخدم الحالي) - آخر 50 بلاغ
    المشروع هو الفيصل وليس المحافظة
    يستبعد البلاغات التي حذفها المستخدم من الإشعارات
    """
    try:
        allowed_projects = list(set(get_projects_with_permission(current_user, "reports_notifications") + get_projects_with_permission(current_user, "reports_review")))
        
        if current_user.role != "admin" and not allowed_projects:
            return {"total": 0, "reports": []}
        
        # الشرط الأساسي: البلاغات المقروءة من قبل المستخدم واستبعاد المحذوفة
        base_query = {
            "is_deleted": {"$ne": True},
            "seen_by": current_user.id,
            "$or": [
                {"deleted_notifications": {"$exists": False}},
                {"deleted_notifications": {"$ne": current_user.id}}
            ]
        }
        
        if current_user.role == "admin":
            # Admin يرى جميع البلاغات المقروءة
            query = base_query
        else:
            # المستخدم يرى بلاغات مشاريعه المصرح له بإشعاراتها فقط
            expanded_reports = []
            for p in allowed_projects:
                expanded_reports.extend([p, f"مشروع {p}", p.replace("مشروع ", "").strip()])
            query = {**base_query, "project": {"$in": expanded_reports}}
            governorates = current_user.governorates or []
            if governorates:
                expanded_govs = []
                for g in governorates:
                    clean_g = g.replace("محافظة ", "").replace("محافظه ", "").strip()
                    expanded_govs.extend([g, clean_g, f"محافظة {clean_g}", f"محافظه {clean_g}"])
                query["governorate"] = {"$in": expanded_govs}
        
        # جلب آخر 50 بلاغ مقروء
        reports = await db.reports.find(
            query,
            {"_id": 0, "id": 1, "report_number": 1, "governorate": 1, "project": 1, 
             "report_type": 1, "created_at": 1, "added_at": 1, "contractor": 1}
        ).sort("added_at", -1).limit(50).to_list(50)
        
        return {
            "total": len(reports),
            "reports": reports
        }
        
    except Exception as e:
        logger.error(f"Error getting seen reports: {str(e)}")
        return {"total": 0, "reports": []}


@api_router.post("/reports/{report_id}/mark-seen")
async def mark_report_as_seen(report_id: str, current_user: User = Depends(get_current_user)):
    """
    تحديد البلاغ كـ 'تمت رؤيته' للمستخدم الحالي
    """
    try:
        # إضافة المستخدم لقائمة من شاهدوا البلاغ
        result = await db.reports.update_one(
            {"id": report_id},
            {"$addToSet": {"seen_by": current_user.id}}
        )
        
        if result.modified_count == 0:
            # قد يكون البلاغ غير موجود أو المستخدم شاهده مسبقاً
            report = await db.reports.find_one({"id": report_id})
            if not report:
                raise HTTPException(status_code=404, detail="البلاغ غير موجود")
        
        return {"success": True, "message": "تم تحديد البلاغ كمرئي"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error marking report as seen: {str(e)}")
        raise HTTPException(status_code=500, detail="فشل في تحديث حالة البلاغ")


@api_router.post("/reports/{report_id}/mark-unseen")
async def mark_report_as_unseen(report_id: str, current_user: User = Depends(get_current_user)):
    """
    إرجاع البلاغ إلى 'غير مرئي' للمستخدم الحالي
    """
    try:
        # إزالة المستخدم من قائمة من شاهدوا البلاغ
        result = await db.reports.update_one(
            {"id": report_id},
            {"$pull": {"seen_by": current_user.id}}
        )
        
        if result.modified_count == 0:
            report = await db.reports.find_one({"id": report_id})
            if not report:
                raise HTTPException(status_code=404, detail="البلاغ غير موجود")
        
        return {"success": True, "message": "تم إرجاع البلاغ إلى غير مرئي"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error marking report as unseen: {str(e)}")
        raise HTTPException(status_code=500, detail="فشل في تحديث حالة البلاغ")


@api_router.post("/reports/mark-all-seen")
async def mark_all_reports_as_seen(current_user: User = Depends(get_current_user)):
    """
    تحديد جميع البلاغات والتوصيلات الجديدة كـ 'تمت رؤيتها' للمستخدم الحالي
    """
    try:
        report_projects = get_projects_with_permission(current_user, "reports_notifications")
        water_projects = list(set(get_projects_with_permission(current_user, "water_connections") + report_projects))
        sewage_projects = list(set(get_projects_with_permission(current_user, "sewage_connections") + report_projects))
        
        governorates = current_user.governorates or []
        is_admin = current_user.role == "admin"
        
        # Reports
        rq = {
            "is_deleted": {"$ne": True},
            "$and": [
                {"$or": [
                    {"seen_by": {"$exists": False}},
                    {"seen_by": {"$ne": current_user.id}}
                ]}
            ]
        }
        if not is_admin:
            if report_projects:
                expanded_reports = []
                for p in report_projects:
                    expanded_reports.extend([p, f"مشروع {p}", p.replace("مشروع ", "").strip()])
                rq["project"] = {"$in": expanded_reports}
            if governorates:
                expanded_govs = []
                for g in governorates:
                    clean_g = g.replace("محافظة ", "").replace("محافظه ", "").strip()
                    expanded_govs.extend([g, clean_g, f"محافظة {clean_g}", f"محافظه {clean_g}"])
                rq["governorate"] = {"$in": expanded_govs}
        
        result_reports = await db.reports.update_many(rq, {"$addToSet": {"seen_by": current_user.id}})
        
        # Water
        wq = {
            "is_deleted": {"$ne": True},
            "$and": [
                {"$or": [
                    {"seen_by": {"$exists": False}},
                    {"seen_by": {"$ne": current_user.id}}
                ]}
            ]
        }
        if not is_admin:
            if water_projects:
                expanded_water = []
                for p in water_projects:
                    expanded_water.extend([p, f"مشروع {p}", p.replace("مشروع ", "").strip()])
                wq["project"] = {"$in": expanded_water}
            if governorates:
                expanded_govs = []
                for g in governorates:
                    clean_g = g.replace("محافظة ", "").replace("محافظه ", "").strip()
                    expanded_govs.extend([g, clean_g, f"محافظة {clean_g}", f"محافظه {clean_g}"])
                wq["area"] = {"$in": expanded_govs}
                
        result_water = await db.water_connections.update_many(wq, {"$addToSet": {"seen_by": current_user.id}})
        
        # Sewage
        sq = {
            "is_deleted": {"$ne": True},
            "$and": [
                {"$or": [
                    {"seen_by": {"$exists": False}},
                    {"seen_by": {"$ne": current_user.id}}
                ]}
            ]
        }
        if not is_admin:
            if sewage_projects:
                expanded_sewage = []
                for p in sewage_projects:
                    expanded_sewage.extend([p, f"مشروع {p}", p.replace("مشروع ", "").strip()])
                sq["project"] = {"$in": expanded_sewage}
            if governorates:
                expanded_govs = []
                for g in governorates:
                    clean_g = g.replace("محافظة ", "").replace("محافظه ", "").strip()
                    expanded_govs.extend([g, clean_g, f"محافظة {clean_g}", f"محافظه {clean_g}"])
                sq["area"] = {"$in": expanded_govs}
                
        result_sewage = await db.sewage_connections.update_many(sq, {"$addToSet": {"seen_by": current_user.id}})
        
        total_modified = result_reports.modified_count + result_water.modified_count + result_sewage.modified_count
        return {"success": True, "count": total_modified, "message": f"تم تحديد {total_modified} عنصر كمرئي"}
        
    except Exception as e:
        logger.error(f"Error marking all reports as seen: {str(e)}")
        raise HTTPException(status_code=500, detail="فشل في تحديث حالة الإشعارات")


@api_router.delete("/reports/notifications/clear-all")
async def clear_all_read_notifications(current_user: User = Depends(get_current_user)):
    """
    حذف جميع الإشعارات (المقروءة فقط) للمستخدم
    """
    try:
        report_projects = get_projects_with_permission(current_user, "reports_notifications")
        water_projects = list(set(get_projects_with_permission(current_user, "water_connections") + report_projects))
        sewage_projects = list(set(get_projects_with_permission(current_user, "sewage_connections") + report_projects))
        
        governorates = current_user.governorates or []
        is_admin = current_user.role == "admin"
        
        # Reports
        rq = {
            "is_deleted": {"$ne": True},
            "seen_by": current_user.id
        }
        if not is_admin:
            if report_projects:
                expanded_reports = []
                for p in report_projects:
                    expanded_reports.extend([p, f"مشروع {p}", p.replace("مشروع ", "").strip()])
                rq["project"] = {"$in": expanded_reports}
            if governorates:
                expanded_govs = []
                for g in governorates:
                    clean_g = g.replace("محافظة ", "").replace("محافظه ", "").strip()
                    expanded_govs.extend([g, clean_g, f"محافظة {clean_g}", f"محافظه {clean_g}"])
                rq["governorate"] = {"$in": expanded_govs}
        
        result_reports = await db.reports.update_many(rq, {"$addToSet": {"deleted_notifications": current_user.id}})
        
        # Water
        wq = {
            "is_deleted": {"$ne": True},
            "seen_by": current_user.id
        }
        if not is_admin:
            if water_projects:
                expanded_water = []
                for p in water_projects:
                    expanded_water.extend([p, f"مشروع {p}", p.replace("مشروع ", "").strip()])
                wq["project"] = {"$in": expanded_water}
            if governorates:
                expanded_govs = []
                for g in governorates:
                    clean_g = g.replace("محافظة ", "").replace("محافظه ", "").strip()
                    expanded_govs.extend([g, clean_g, f"محافظة {clean_g}", f"محافظه {clean_g}"])
                wq["area"] = {"$in": expanded_govs}
                
        result_water = await db.water_connections.update_many(wq, {"$addToSet": {"deleted_notifications": current_user.id}})
        
        # Sewage
        sq = {
            "is_deleted": {"$ne": True},
            "seen_by": current_user.id
        }
        if not is_admin:
            if sewage_projects:
                expanded_sewage = []
                for p in sewage_projects:
                    expanded_sewage.extend([p, f"مشروع {p}", p.replace("مشروع ", "").strip()])
                sq["project"] = {"$in": expanded_sewage}
            if governorates:
                expanded_govs = []
                for g in governorates:
                    clean_g = g.replace("محافظة ", "").replace("محافظه ", "").strip()
                    expanded_govs.extend([g, clean_g, f"محافظة {clean_g}", f"محافظه {clean_g}"])
                sq["area"] = {"$in": expanded_govs}
                
        result_sewage = await db.sewage_connections.update_many(sq, {"$addToSet": {"deleted_notifications": current_user.id}})
        
        total_modified = result_reports.modified_count + result_water.modified_count + result_sewage.modified_count
        return {"success": True, "count": total_modified, "message": f"تم حذف {total_modified} إشعار"}
        
    except Exception as e:
        logger.error(f"Error clearing notifications: {str(e)}")
        raise HTTPException(status_code=500, detail="فشل في حذف الإشعارات")


@api_router.delete("/reports/notifications/{report_id}")
async def delete_report_notification(report_id: str, current_user: User = Depends(get_current_user)):
    """
    حذف إشعار بلاغ واحد (إضافة المستخدم لقائمة deleted_notifications)
    هذا لا يحذف البلاغ نفسه، فقط يخفي الإشعار من القائمة
    """
    try:
        # إضافة المستخدم لقائمة deleted_notifications للبلاغ المحدد
        result = await db.reports.update_one(
            {"id": report_id},
            {"$addToSet": {"deleted_notifications": current_user.id}}
        )
        
        if result.modified_count == 0:
            result = await db.water_connections.update_one(
                {"id": report_id},
                {"$addToSet": {"deleted_notifications": current_user.id}}
            )
            
        if result.modified_count == 0:
            result = await db.sewage_connections.update_one(
                {"id": report_id},
                {"$addToSet": {"deleted_notifications": current_user.id}}
            )
        
        return {"success": True, "message": "تم حذف الإشعار"}
        
    except Exception as e:
        logger.error(f"Error deleting notification: {str(e)}")
        raise HTTPException(status_code=500, detail="فشل في حذف الإشعار")


# ============= إشعارات التوصيلات =============

@api_router.post("/water-connections/{conn_id}/mark-seen")
async def mark_water_connection_seen(conn_id: str, current_user: User = Depends(get_current_user)):
    """تحديد توصيلة مياه كـ مرئية للمستخدم الحالي"""
    try:
        await db.water_connections.update_one(
            {"id": conn_id},
            {"$addToSet": {"seen_by": current_user.id}}
        )
        return {"success": True}
    except Exception as e:
        logger.error(f"Error marking water connection seen: {str(e)}")
        raise HTTPException(status_code=500, detail="فشل")


@api_router.post("/sewage-connections/{conn_id}/mark-seen")
async def mark_sewage_connection_seen(conn_id: str, current_user: User = Depends(get_current_user)):
    """تحديد توصيلة صرف كـ مرئية للمستخدم الحالي"""
    try:
        await db.sewage_connections.update_one(
            {"id": conn_id},
            {"$addToSet": {"seen_by": current_user.id}}
        )
        return {"success": True}
    except Exception as e:
        logger.error(f"Error marking sewage connection seen: {str(e)}")
        raise HTTPException(status_code=500, detail="فشل")


@api_router.delete("/water-connections/{conn_id}/notification")
async def delete_water_connection_notification(conn_id: str, current_user: User = Depends(get_current_user)):
    """حذف إشعار توصيلة مياه من قائمة المستخدم"""
    try:
        await db.water_connections.update_one(
            {"id": conn_id},
            {"$addToSet": {"deleted_notifications": current_user.id}}
        )
        return {"success": True}
    except Exception as e:
        logger.error(f"Error deleting water connection notification: {str(e)}")
        raise HTTPException(status_code=500, detail="فشل")


@api_router.delete("/sewage-connections/{conn_id}/notification")
async def delete_sewage_connection_notification(conn_id: str, current_user: User = Depends(get_current_user)):
    """حذف إشعار توصيلة صرف من قائمة المستخدم"""
    try:
        await db.sewage_connections.update_one(
            {"id": conn_id},
            {"$addToSet": {"deleted_notifications": current_user.id}}
        )
        return {"success": True}
    except Exception as e:
        logger.error(f"Error deleting sewage connection notification: {str(e)}")
        raise HTTPException(status_code=500, detail="فشل")




@api_router.get("/reports/consultant-notes")
async def get_consultant_notes(
    project: str = Query(None),
    governorate: str = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
    search: str = Query(None),
    status_filter: str = Query(None),
    count_only: Optional[bool] = False,
    current_user: User = Depends(get_current_user)
):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    role = user_doc.get("role")
    
    # Check if user has consultant_notes permission
    if role != "admin":
        has_perm = False
        if "consultant_notes" in user_perms:
            has_perm = True
        else:
            for p in pp.values():
                if "consultant_notes" in p:
                    has_perm = True
                    break
        if not has_perm:
            raise HTTPException(status_code=403, detail="Forbidden")
            
    # Base query
    query = {
        "consultant_note": {"$exists": True, "$ne": "", "$type": "string"},
        "is_deleted": {"$ne": True}
    }
    
    if status_filter == "processed":
        query["consultant_note_processed"] = True
    elif status_filter == "unprocessed":
        query["consultant_note_processed"] = {"$ne": True}
    
    user_projs = user_doc.get("projects", [])
    user_govs = user_doc.get("governorates", [])
    
    and_conditions = []
    
    if role != "admin":
        if not project and user_projs:
            proj_filter = get_loose_in_query(user_projs, "project")
            if proj_filter: and_conditions.append(proj_filter)
        elif project:
            proj_filter = get_flexible_in_query([project], "project")
            if proj_filter: and_conditions.append(proj_filter)
            
        if not governorate and user_govs:
            gov_filter = get_flexible_in_query(user_govs, "governorate")
            if gov_filter: and_conditions.append(gov_filter)
        elif governorate:
            gov_filter = get_flexible_in_query([governorate], "governorate")
            if gov_filter: and_conditions.append(gov_filter)
            
    else:
        if project:
            proj_filter = get_flexible_in_query([project], "project")
            if proj_filter: and_conditions.append(proj_filter)
        if governorate:
            gov_filter = get_flexible_in_query([governorate], "governorate")
            if gov_filter: and_conditions.append(gov_filter)
            
    if search:
        search_clean = search.strip()
        and_conditions.append({
            "$or": [
                {"report_number": {"$regex": search_clean, "$options": "i"}},
                {"id": {"$regex": search_clean, "$options": "i"}}
            ]
        })
        
    if and_conditions:
        query["$and"] = and_conditions

    print("=== DEBUG get_consultant_notes ===")
    print("User:", current_user.username)
    print("Query:", query)
    
    skip = (page - 1) * limit
    
    total = await db.reports.count_documents(query)
    
    reports = await db.reports.find(
        query, 
        {"_id": 0, "id": 1, "report_number": 1, "project": 1, "governorate": 1, "contractor": 1, "consultant_note": 1, "consultant_note_by": 1, "consultant_note_reply": 1, "consultant_note_replied_by": 1, "consultant_note_processed": 1, "consultant_note_date": 1, "consultant_note_processed_date": 1, "created_at": 1, "status": 1}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    print("Found Reports:", len(reports), "Total Count:", total)
    
    return {"reports": reports, "total": total, "page": page, "limit": limit}


@api_router.put("/reports/{report_id}/consultant_note_processed")
async def toggle_consultant_note_processed(report_id: str, current_user: User = Depends(get_current_user)):
    report = await db.reports.find_one({"id": report_id})
    if not report: raise HTTPException(status_code=404, detail="Report not found")
    
    user_role = getattr(current_user, "role", current_user.get("role") if isinstance(current_user, dict) else None)
    can_create = getattr(current_user, "can_create_subusers", current_user.get("can_create_subusers") if isinstance(current_user, dict) else False)
    
    if user_role != "admin" and not can_create:
        raise HTTPException(status_code=403, detail="Only Level 1 and Level 2 can process consultant notes")
    current_status = report.get("consultant_note_processed", False)
    new_status = not current_status
    
    from datetime import datetime, timezone
    update_data = {
        "consultant_note_processed": new_status,
        "report_note_processed": new_status
    }
    if new_status:
        now_iso = datetime.now(timezone.utc).isoformat()
        update_data["consultant_note_processed_date"] = now_iso
        update_data["report_note_processed_date"] = now_iso
    else:
        update_data["consultant_note_processed_date"] = ""
        update_data["report_note_processed_date"] = ""
        
    await db.reports.update_one({"id": report_id}, {"$set": update_data})
    return {"success": True, "consultant_note_processed": new_status, "consultant_note_processed_date": update_data.get("consultant_note_processed_date", "")}


@api_router.get("/reports/report-notes")
async def get_report_notes(
    project: str = Query(None),
    governorate: str = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
    search: str = Query(None),
    count_only: Optional[bool] = False,
    current_user: User = Depends(get_current_user)
):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    role = user_doc.get("role")
    
    if role != "admin":
        has_perm = False
        if "report_notes" in user_perms:
            has_perm = True
        else:
            for p in pp.values():
                if "report_notes" in p:
                    has_perm = True
                    break
        if not has_perm:
            raise HTTPException(status_code=403, detail="Forbidden")
            
    query = {
        "notes": {"$exists": True, "$ne": "", "$type": "string"},
        "is_deleted": {"$ne": True}
    }
    

    user_projs = user_doc.get("projects", [])
    user_govs = user_doc.get("governorates", [])
    
    and_conditions = []
    
    if role != "admin":
        if not project and user_projs:
            proj_filter = get_loose_in_query(user_projs, "project")
            if proj_filter: and_conditions.append(proj_filter)
        elif project:
            proj_filter = get_flexible_in_query([project], "project")
            if proj_filter: and_conditions.append(proj_filter)
            
        if not governorate and user_govs:
            gov_filter = get_flexible_in_query(user_govs, "governorate")
            if gov_filter: and_conditions.append(gov_filter)
        elif governorate:
            gov_filter = get_flexible_in_query([governorate], "governorate")
            if gov_filter: and_conditions.append(gov_filter)
            
    else:
        if project:
            proj_filter = get_flexible_in_query([project], "project")
            if proj_filter: and_conditions.append(proj_filter)
        if governorate:
            gov_filter = get_flexible_in_query([governorate], "governorate")
            if gov_filter: and_conditions.append(gov_filter)
            
    if search:
        search_clean = search.strip()
        and_conditions.append({
            "$or": [
                {"report_number": {"$regex": search_clean, "$options": "i"}},
                {"id": {"$regex": search_clean, "$options": "i"}}
            ]
        })
        
    if and_conditions:
        query["$and"] = and_conditions

    skip = (page - 1) * limit
    total = await db.reports.count_documents(query)
    
    reports = await db.reports.find(
        query, 
        {"_id": 0, "id": 1, "report_number": 1, "project": 1, "governorate": 1, "contractor": 1, "notes": 1, "report_note_reply": 1, "report_note_replied_by": 1, "report_note_replies": 1, "report_note_processed": 1, "report_note_processed_date": 1, "created_at": 1, "status": 1}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {"reports": reports, "total": total, "page": page, "limit": limit}

class ReportNoteUpdate(BaseModel):
    notes: str

@api_router.put("/reports/{report_id}/report-note")
async def edit_report_note(report_id: str, payload: ReportNoteUpdate, current_user: User = Depends(get_current_user)):
    report = await db.reports.find_one({"id": report_id})
    if not report: raise HTTPException(status_code=404, detail="Report not found")
    
    user_role = getattr(current_user, "role", current_user.get("role") if isinstance(current_user, dict) else None)
    can_create = getattr(current_user, "can_create_subusers", current_user.get("can_create_subusers") if isinstance(current_user, dict) else False)
    
    user_doc = current_user.model_dump() if hasattr(current_user, 'model_dump') else current_user
    has_owner_perm = has_project_permission(user_doc, report.get("project"), "owner_notes")
    has_note_perm = has_project_permission(user_doc, report.get("project"), "report_notes")
    
    if user_role != "admin" and not can_create and not has_owner_perm and not has_note_perm:
        raise HTTPException(status_code=403, detail="Not authorized to edit report notes")
        
    update_data = {"notes": payload.notes}
    if payload.notes:
        update_data["report_note_processed"] = False
    await db.reports.update_one({"id": report_id}, {"$set": update_data})
    return {"success": True, "notes": payload.notes}

@api_router.delete("/reports/{report_id}/report-note")
async def delete_report_note(report_id: str, current_user: User = Depends(get_current_user)):
    report = await db.reports.find_one({"id": report_id})
    if not report: raise HTTPException(status_code=404, detail="Report not found")
    
    user_role = getattr(current_user, "role", current_user.get("role") if isinstance(current_user, dict) else None)
    can_create = getattr(current_user, "can_create_subusers", current_user.get("can_create_subusers") if isinstance(current_user, dict) else False)
    
    if user_role != "admin" and not can_create:
        raise HTTPException(status_code=403, detail="Only Level 1 and Level 2 can delete report notes")
        
    # User requested to clear the note from the text box, not delete the report
    await db.reports.update_one({"id": report_id}, {"$set": {"notes": ""}})
    return {"success": True}

@api_router.put("/reports/{report_id}/report_note_processed")
async def toggle_report_note_processed(report_id: str, current_user: User = Depends(get_current_user)):
    report = await db.reports.find_one({"id": report_id})
    if not report: raise HTTPException(status_code=404, detail="Report not found")
    
    user_role = getattr(current_user, "role", current_user.get("role") if isinstance(current_user, dict) else None)
    can_create = getattr(current_user, "can_create_subusers", current_user.get("can_create_subusers") if isinstance(current_user, dict) else False)
    
    if user_role != "admin" and not can_create:
        raise HTTPException(status_code=403, detail="Only Level 1 and Level 2 can process report notes")
    
    current_status = report.get("report_note_processed", False)
    new_status = not current_status
    
    from datetime import datetime, timezone
    update_data = {
        "report_note_processed": new_status,
        "consultant_note_processed": new_status
    }
    if new_status:
        now_iso = datetime.now(timezone.utc).isoformat()
        update_data["report_note_processed_date"] = now_iso
        update_data["consultant_note_processed_date"] = now_iso
    else:
        update_data["report_note_processed_date"] = ""
        update_data["consultant_note_processed_date"] = ""
        
    await db.reports.update_one({"id": report_id}, {"$set": update_data})
    return {"success": True, "report_note_processed": new_status, "report_note_processed_date": update_data.get("report_note_processed_date", "")}

class ReportNoteReply(BaseModel):
    reply: str

@api_router.put("/reports/{report_id}/report_note_reply")
async def update_report_note_reply(report_id: str, payload: ReportNoteReply, current_user: User = Depends(get_current_user)):
    report = await db.reports.find_one({"id": report_id})
    if not report: raise HTTPException(status_code=404, detail="Report not found")
    
    user_doc = current_user.model_dump() if hasattr(current_user, 'model_dump') else current_user
    if not has_project_permission(user_doc, report.get("project"), "consultant_notes") and user_doc.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only consultant can reply to this note")
        
    replier_name = current_user.full_name if getattr(current_user, 'full_name', None) else current_user.username
    from datetime import datetime, timezone
    import uuid
    
    if not payload.reply.strip():
        # Clear replies
        await db.reports.update_one({"id": report_id}, {"$set": {"report_note_replies": [], "report_note_processed": False}})
        return {"success": True, "report_note_replies": []}
    
    new_reply = {
        "id": str(uuid.uuid4()),
        "reply": payload.reply,
        "replied_by": replier_name,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.reports.update_one(
        {"id": report_id}, 
        {"$push": {"report_note_replies": new_reply}, "$set": {"report_note_processed": False}}
    )
    
    # fetch updated document to return full array
    updated_report = await db.reports.find_one({"id": report_id})
    return {"success": True, "report_note_replies": updated_report.get("report_note_replies", [])}


class ConsultantNoteReply(BaseModel):
    reply: str

@api_router.put("/reports/{report_id}/consultant_note_reply")
async def update_consultant_note_reply(report_id: str, payload: ConsultantNoteReply, current_user: User = Depends(get_current_user)):
    report = await db.reports.find_one({"id": report_id})
    if not report: raise HTTPException(status_code=404, detail="Report not found")
    
    if not has_project_permission(current_user.model_dump() if hasattr(current_user, 'model_dump') else current_user, report.get("project"), "consultant_notes") and \
       not has_project_permission(current_user.model_dump() if hasattr(current_user, 'model_dump') else current_user, report.get("project"), "owner_notes") and \
       not getattr(current_user, 'role', '') == 'admin':
        raise HTTPException(status_code=403, detail="Not authorized to reply")
        
    replier_name = current_user.full_name if getattr(current_user, 'full_name', None) else current_user.username
    update_data = {
        "consultant_note_reply": payload.reply,
        "consultant_note_replied_by": replier_name if payload.reply else ""
    }
    
    if payload.reply:
        # Reset processed status so admin/consultant sees the new reply
        update_data["consultant_note_processed"] = False
        
    await db.reports.update_one({"id": report_id}, {"$set": update_data})
    
    return {"success": True, "reply": payload.reply, "replied_by": update_data["consultant_note_replied_by"]}


@api_router.delete("/reports/{report_id}/consultant_note")
async def delete_consultant_note(report_id: str, current_user: User = Depends(get_current_user)):
    report = await db.reports.find_one({"id": report_id})
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
        
    user_doc = current_user.model_dump() if hasattr(current_user, 'model_dump') else current_user
    if not has_project_permission(user_doc, report.get("project"), "consultant_notes"):
        raise HTTPException(status_code=403, detail="Not authorized to delete consultant notes")
        
    await db.reports.update_one(
        {"id": report_id},
        {"$unset": {
            "consultant_note": "",
            "consultant_note_by": "",
            "consultant_note_reply": "",
            "consultant_note_replied_by": "",
            "consultant_note_processed": ""
        }}
    )
    return {"success": True, "message": "Consultant note deleted successfully"}


@api_router.get("/reports/{report_id}", response_model=ReportResponse)
async def get_report(
    report_id: str, 
    exclude_images: bool = Query(False, description="استبعاد الصور لتسريع التحميل"),
    current_user: User = Depends(get_current_user)
):
    query = {"id": report_id, "is_deleted": {"$ne": True}}
    
    # فلترة حسب صلاحيات المحافظات — الأدمن يرى كل البلاغات بدون قيود
    if current_user.role != "admin":
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
    
    # ⚡ استبعاد الصور إذا طُلب ذلك لتسريع التحميل
    projection = {"_id": 0}
    if exclude_images:
        projection["images"] = 0
    
    report_doc = await db.reports.find_one(query, projection)
    if not report_doc:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # إضافة مصفوفة فارغة للصور إذا تم استبعادها
    if exclude_images:
        report_doc["images"] = []
    
    if isinstance(report_doc.get('report_date'), str):
        report_doc['report_date'] = datetime.fromisoformat(report_doc['report_date'])
    if isinstance(report_doc.get('created_at'), str):
        report_doc['created_at'] = datetime.fromisoformat(report_doc['created_at'])
    if isinstance(report_doc.get('updated_at'), str):
        report_doc['updated_at'] = datetime.fromisoformat(report_doc['updated_at'])
    # Ensure closed_at field exists (for backward compatibility)
    if 'closed_at' not in report_doc:
        report_doc['closed_at'] = None
    elif report_doc.get('closed_at') and isinstance(report_doc['closed_at'], str):
        report_doc['closed_at'] = datetime.fromisoformat(report_doc['closed_at'])
    # Ensure project field exists (for backward compatibility)
    if 'project' not in report_doc:
        report_doc['project'] = 'مشروع إصلاح أعمال المحافظات الغربية - القطاع الأوسط'
    # Ensure new fields exist (for backward compatibility)
    if 'latitude' not in report_doc:
        report_doc['latitude'] = None
    if 'longitude' not in report_doc:
        report_doc['longitude'] = None
    if 'asphalt_license_issued' not in report_doc:
        report_doc['asphalt_license_issued'] = False
    
    return ReportResponse(**report_doc)


    return ReportResponse(**report_doc)


@api_router.put("/reports/{report_id}", response_model=ReportResponse)
async def update_report(
    report_id: str,
    report_number: Optional[str] = Form(None),
    license_number: Optional[str] = Form(None),
    report_type: Optional[str] = Form(None),
    status: Optional[str] = Form(None),
    governorate: Optional[str] = Form(None),
    project: Optional[str] = Form(None),
    depth_meters: Optional[float] = Form(None),
    diameter_mm: Optional[float] = Form(None),
    contractor: Optional[str] = Form(None),
    latitude: Optional[str] = Form(None),
    longitude: Optional[str] = Form(None),
    wfm_closed: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    created_at: Optional[str] = Form(None),
    start_date: Optional[str] = Form(None),
    closed_at: Optional[str] = Form(None),
    remove_closed_at: Optional[str] = Form(None),  # فلاج لحذف التاريخ
    images: List[UploadFile] = File(default=[]),
    current_user: User = Depends(get_current_user)
):
    query = {"id": report_id, "is_deleted": {"$ne": True}}
    
    # فلترة حسب صلاحيات المشاريع والمحافظات — الأدمن يرى كل البلاغات بدون قيود
    if current_user.role != "admin":
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
    
    report_doc = await db.reports.find_one(query, {"_id": 0})
    if not report_doc:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # التحقق من تكرار رقم البلاغ (إذا تم تغييره)
    if report_number is not None and report_number != report_doc.get('report_number'):
        existing_report = await db.reports.find_one({
            "report_number": report_number,
            "id": {"$ne": report_id}  # استثناء البلاغ الحالي
        })
        
        if existing_report:
            raise HTTPException(
                status_code=400,
                detail="هذا الرقم موجود مسبقاً"
            )
    
    # التحقق من تكرار رقم الرخصة - فقط للأرقام الفعلية (ليست نصوص مثل "لم يتم إصدار رخصة")
    def _is_actual_license_edit(val):
        if not val:
            return False
        cleaned = str(val).strip()
        placeholders = {"لم يتم إصدار رخصة", "لم يتم", "-", "0", "nan", "none", "غير محدد", "بدون رخصة"}
        if not cleaned or cleaned in placeholders:
            return False
        return any(ch.isdigit() for ch in cleaned)
    
    if license_number is not None and license_number != report_doc.get('license_number'):
        if _is_actual_license_edit(license_number):
            existing_license = await db.reports.find_one({
                "license_number": license_number,
                "id": {"$ne": report_id}  # استثناء البلاغ الحالي
            })
            
            if existing_license:
                raise HTTPException(
                    status_code=400,
                    detail="هذا الرقم موجود مسبقاً"
                )
    
    update_data = {}
    if report_number is not None:
        update_data["report_number"] = report_number
    if license_number is not None:
        update_data["license_number"] = license_number
    if report_type is not None:
        update_data["report_type"] = report_type
    if status is not None:
        update_data["status"] = normalize_asphalt_status(status)
    if governorate is not None:
        update_data["governorate"] = governorate
    if project is not None:
        update_data["project"] = project
    if depth_meters is not None:
        update_data["depth_meters"] = depth_meters
    if diameter_mm is not None:
        update_data["diameter_mm"] = diameter_mm
    if contractor is not None:
        update_data["contractor"] = contractor
    if latitude is not None:
        update_data["latitude"] = latitude
    if longitude is not None:
        update_data["longitude"] = longitude
    if wfm_closed is not None:
        update_data["wfm_closed"] = (wfm_closed.lower() == "true" if wfm_closed else False)
    # تحديث الملاحظات دائماً (السماح بالحذف عند إرسال string فارغ)
    update_data["notes"] = notes if notes else ""
    if created_at is not None:
        update_data["created_at"] = datetime.fromisoformat(created_at).isoformat()
    
    # تحديث تاريخ مباشرة البلاغ
    if start_date is not None and start_date.strip():
        try:
            from datetime import datetime as dt
            update_data["start_date"] = dt.fromisoformat(start_date).isoformat()
        except:
            pass
    
    # تحديث تاريخ الإغلاق - السماح بالتعديل أو الحذف
    if remove_closed_at and remove_closed_at.lower() == "true":
        # حذف التاريخ صراحة
        update_data["closed_at"] = None
    elif closed_at is not None and closed_at.strip():
        # إذا كان التاريخ موجود وغير فارغ
        try:
            from datetime import datetime as dt
            update_data["closed_at"] = dt.fromisoformat(closed_at).isoformat()
        except:
            pass  # تجاهل التواريخ غير الصحيحة
    
    if images and images[0].filename:
        # إضافة الصور الجديدة - معالجة متوازية
        import asyncio
        existing_images = report_doc.get('images', [])
        
        async def process_single_image(image):
            content = await image.read()
            loop = asyncio.get_event_loop()
            content = await loop.run_in_executor(thread_pool, compress_image, content)
            url = await loop.run_in_executor(
                thread_pool,
                lambda: _store_image_bytes(content, category="reports", filename=image.filename, content_type=image.content_type)
            )
            return url
        
        # معالجة جميع الصور بشكل متوازي
        image_data = await asyncio.gather(*[process_single_image(img) for img in images])
        
        # دمج الصور القديمة مع الجديدة
        update_data["images"] = existing_images + image_data
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.reports.update_one({"id": report_id}, {"$set": update_data})
    
    updated_report = await db.reports.find_one({"id": report_id}, {"_id": 0})
    
    # ⚡ مسح الـ cache لتحديث الإحصائيات
    global stats_cache
    stats_cache.clear()
    
    if isinstance(updated_report.get('report_date'), str):
        updated_report['report_date'] = datetime.fromisoformat(updated_report['report_date'])
    if isinstance(updated_report.get('created_at'), str):
        updated_report['created_at'] = datetime.fromisoformat(updated_report['created_at'])
    if isinstance(updated_report.get('updated_at'), str):
        updated_report['updated_at'] = datetime.fromisoformat(updated_report['updated_at'])
    
    return ReportResponse(**updated_report)


@api_router.put("/reports/{report_id}/review")
async def update_report_review_status(
    report_id: str,
    current_user: User = Depends(get_current_user)
):
    """تحديث حالة مراجعة البلاغ - ديناميكي حسب صلاحية reports_review لكل مشروع"""
    
    report = await db.reports.find_one({"id": report_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    report_project = report.get("project", "")
    
    # Admin يمكنه مراجعة أي بلاغ
    if current_user.role != "admin":
        # التحقق من أن المستخدم يملك صلاحية reports_review في هذا المشروع
        if not has_project_permission(current_user, report_project, "reports_review"):
            # ديناميكياً: إيجاد المستخدمين الذين لديهم صلاحية reports_review في هذا المشروع
            all_users = await db.users.find(
                {"is_active": True, "role": {"$ne": "admin"}},
                {"_id": 0, "id": 1, "username": 1, "full_name": 1, "permissions": 1, "project_permissions": 1, "projects": 1}
            ).to_list(500)
            
            authorized_users = []
            for u in all_users:
                if report_project not in (u.get("projects") or []):
                    continue
                if has_project_permission(u, report_project, "reports_review"):
                    authorized_users.append(u.get("full_name") or u.get("username"))
            
            if authorized_users:
                raise HTTPException(
                    status_code=403, 
                    detail="لا يمكنك المراجعه فقط م/محمود هارون مهندس نظم المعلومات وتحليل البيانات لديه صلاحيه مراجعه بلاغات هذا المشروع"
                )
            else:
                raise HTTPException(
                    status_code=403, 
                    detail="لا يمكنك المراجعه فقط م/محمود هارون مهندس نظم المعلومات وتحليل البيانات لديه صلاحيه مراجعه بلاغات هذا المشروع"
                )
    
    # تحديث حالة المراجعة
    is_pending = report.get("review_status") in ["بانتظار المراجعة", "قيد المراجعة", None]
    new_status = "تمت المراجعة" if is_pending else "قيد المراجعة"
    reviewed_by_name = (current_user.full_name or current_user.username) if new_status == "تمت المراجعة" else None
    
    await db.reports.update_one(
        {"id": report_id},
        {
            "$set": {
                "review_status": new_status,
                "reviewed_by": current_user.id if new_status == "تمت المراجعة" else None,
                "reviewed_by_name": reviewed_by_name,
                "reviewed_at": datetime.now(timezone.utc).isoformat() if new_status == "تمت المراجعة" else None,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "wfm_closed": True if new_status == "تمت المراجعة" else False,
                "wfm_closed_by": "م/ مدحت حسين محمد" if new_status == "تمت المراجعة" else None,
                "wfm_closed_at": datetime.now(timezone.utc).isoformat() if new_status == "تمت المراجعة" else None
            }
        }
    )
    
    msg = "تم مراجعة البلاغ بنجاح" if new_status == "تمت المراجعة" else "تم اعادة فتح حالة المراجعة"
    return {
        "message": msg, 
        "review_status": new_status,
        "reviewed_by_name": reviewed_by_name,
        "wfm_closed": True if new_status == "تمت المراجعة" else False,
        "wfm_closed_by": "م/ مدحت حسين محمد" if new_status == "تمت المراجعة" else None
    }


@api_router.delete("/reports/{report_id}/images/{image_index}")
async def delete_report_image(
    report_id: str, 
    image_index: int,
    current_user: User = Depends(get_current_user)
):
    """حذف صورة معينة من البلاغ"""
    query = {"id": report_id, "is_deleted": {"$ne": True}}
    
    # فلترة حسب صلاحيات المشاريع والمحافظات
    if current_user.role != "admin" or len(current_user.projects) > 0:
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
    
    if current_user.role != "admin" or len(current_user.governorates) > 0:
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
    
    report_doc = await db.reports.find_one(query, {"_id": 0})
    if not report_doc:
        raise HTTPException(status_code=404, detail="Report not found")
    
    images = report_doc.get('images', [])
    if image_index < 0 or image_index >= len(images):
        raise HTTPException(status_code=400, detail="Invalid image index")
    
    # حذف الصورة
    images.pop(image_index)
    
    await db.reports.update_one(
        {"id": report_id},
        {"$set": {"images": images, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "تم حذف الصورة بنجاح"}


@api_router.post("/reports/{report_id}/images")
async def add_report_image(
    report_id: str,
    image: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """إضافة صورة واحدة للبلاغ (رفع في الخلفية)"""
    query = {"id": report_id, "is_deleted": {"$ne": True}}
    
    if current_user.role != "admin" or len(current_user.projects) > 0:
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
    
    if current_user.role != "admin" or len(current_user.governorates) > 0:
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
    
    report_doc = await db.reports.find_one(query, {"_id": 0})
    if not report_doc:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # قراءة الصورة وضغطها ورفعها إلى Object Storage
    content = await image.read()
    loop = asyncio.get_event_loop()
    content = await loop.run_in_executor(thread_pool, compress_image, content)
    image_url = await loop.run_in_executor(
        thread_pool,
        lambda: _store_image_bytes(content, category="reports", filename=image.filename, content_type=image.content_type)
    )
    image_base64 = image_url
    
    # إضافة الصورة للمصفوفة الحالية
    await db.reports.update_one(
        {"id": report_id},
        {
            "$push": {"images": image_base64},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        }
    )
    
    return {"message": "تم إضافة الصورة بنجاح"}


@api_router.delete("/reports/{report_id}")
async def delete_report(report_id: str, current_user: User = Depends(get_current_user)):
    """حذف بلاغ (نقله إلى سلة المحذوفات)"""
    # أولاً: احصل على البلاغ
    report_doc = await db.reports.find_one({"id": report_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    
    if not report_doc:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # ثانياً: تحقق من الصلاحيات
    has_permission = False
    
    if current_user.role == "admin":
        # Admin يمكنه حذف أي بلاغ
        has_permission = True
    elif current_user.can_create_subusers:
        # Level 2 يمكنه حذف أي بلاغ
        has_permission = True
    else:
        # Level 3 وغيره: تحقق من المحافظات والمشاريع
        report_governorate = report_doc.get('governorate')
        report_project = report_doc.get('project')
        
        # تحقق من المحافظة
        if current_user.governorates and report_governorate in current_user.governorates:
            has_permission = True
        
        # تحقق من المشروع
        if current_user.projects and report_project in current_user.projects:
            has_permission = True
    
    if not has_permission:
        raise HTTPException(status_code=403, detail="You don't have permission to delete this report")
    
    # ثالثاً: حذف البلاغ (نقله إلى سلة المحذوفات)
    result = await db.reports.update_one(
        {"id": report_id},
        {"$set": {
            "is_deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.id
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=500, detail="Failed to delete report")
    
    # ⚡ مسح الـ cache لتحديث الإحصائيات
    global stats_cache
    stats_cache.clear()
    
    return {"message": "تم نقل البلاغ إلى سلة المحذوفات بنجاح"}


# ============= TRASH =============

@api_router.get("/reports-trash")
async def get_deleted_reports(
    page: int = 1,
    limit: int = 15,
    current_user: User = Depends(get_current_user)
):
    """
    صلاحيات عرض البلاغات المحذوفة:
    - Admin: يرى الكل
    - Level 2: يرى الكل
    - Level 3 مع جميع المحافظات: يرى الكل
    - Level 3 عادي: يرى محافظاته + ما حذفه
    """
    if current_user.role != "admin" and "trash" not in current_user.permissions:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية لعرض سلة المحذوفات")
        
    query = {"is_deleted": True}
    
    # Admin يرى كل شيء
    if current_user.role == "admin":
        pass
    
    # Level 2 (can_create_subusers وبدون محافظات محددة)
    elif current_user.can_create_subusers and (not current_user.governorates or len(current_user.governorates) == 0):
        pass
    
    # Level 3 وغيره
    else:
        or_conditions = []
        
        # دائماً: البلاغات التي حذفها المستخدم نفسه
        or_conditions.append({"deleted_by": current_user.id})
        
        # البلاغات في محافظات المستخدم
        if current_user.governorates and len(current_user.governorates) > 0:
            or_conditions.append({"governorate": {"$in": current_user.governorates}})
        
        # تطبيق الشروط
        if or_conditions:
            query["$or"] = or_conditions
        else:
            query["deleted_by"] = current_user.id
    
    total_count = await db.reports.count_documents(query)
    skip = (page - 1) * limit
    reports = await db.reports.find(query, {"_id": 0}).sort("deleted_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # جلب أسماء المستخدمين الذين حذفوا البلاغات
    deleted_by_ids = [r.get('deleted_by') for r in reports if r.get('deleted_by')]
    deleted_by_users = {}
    if deleted_by_ids:
        users = await db.users.find(
            {'id': {'$in': list(set(deleted_by_ids))}},
            {'_id': 0, 'id': 1, 'full_name': 1}
        ).to_list(1000)
        deleted_by_users = {u['id']: u.get('full_name', 'غير معروف') for u in users}
    
    for report in reports:
        if isinstance(report.get('report_date'), str):
            report['report_date'] = datetime.fromisoformat(report['report_date'])
        if isinstance(report.get('created_at'), str):
            report['created_at'] = datetime.fromisoformat(report['created_at'])
        if isinstance(report.get('added_at'), str):
            report['added_at'] = datetime.fromisoformat(report['added_at'])
        elif not report.get('added_at'):
            # للبلاغات القديمة التي لا تحتوي على added_at، استخدم created_at
            report['added_at'] = report.get('created_at')
        if isinstance(report.get('updated_at'), str):
            report['updated_at'] = datetime.fromisoformat(report['updated_at'])
        if report.get('deleted_at') and isinstance(report['deleted_at'], str):
            report['deleted_at'] = datetime.fromisoformat(report['deleted_at'])
        # إضافة اسم من حذف البلاغ
        if report.get('deleted_by'):
            report['deleted_by_name'] = deleted_by_users.get(report['deleted_by'], 'غير معروف')
        else:
            report['deleted_by_name'] = 'غير معروف'
        # Ensure closed_at field exists (for backward compatibility)
        if 'closed_at' not in report:
            report['closed_at'] = None
        elif report.get('closed_at') and isinstance(report['closed_at'], str):
            report['closed_at'] = datetime.fromisoformat(report['closed_at'])
        # Ensure project field exists (for backward compatibility)
        if 'project' not in report:
            report['project'] = 'مشروع إصلاح أعمال المحافظات الغربية - القطاع الأوسط'
        # Ensure new fields exist (for backward compatibility)
        if 'latitude' not in report:
            report['latitude'] = None
        if 'longitude' not in report:
            report['longitude'] = None
        if 'asphalt_license_issued' not in report:
            report['asphalt_license_issued'] = False
        # Ensure report_date exists
        if 'report_date' not in report or not report['report_date']:
            report['report_date'] = datetime.now(timezone.utc)
        # Ensure license_number exists
        if 'license_number' not in report:
            report['license_number'] = ''
        # Ensure other required fields
        if 'status' not in report:
            report['status'] = 'معلقة'
        if 'governorate' not in report:
            report['governorate'] = ''
        if 'report_type' not in report:
            report['report_type'] = 'بلاغ عام'
    
    return {"items": [ReportResponse(**report) for report in reports], "total": total_count}


@api_router.post("/reports-trash/{report_id}/restore")
async def restore_report(report_id: str, current_user: User = Depends(get_current_user)):
    query = {"id": report_id, "is_deleted": True}
    
    # فلترة حسب صلاحيات المحافظات (إلا إذا كان admin بدون محافظات محددة)
    if current_user.role != "admin" or len(current_user.governorates) > 0:
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
    
    report_doc = await db.reports.find_one(query, {"_id": 0})
    if not report_doc:
        raise HTTPException(status_code=404, detail="Deleted report not found")
    
    await db.reports.update_one(
        {"id": report_id},
        {"$set": {
            "is_deleted": {"$ne": True},
            "deleted_at": None
        }}
    )
    
    return {"message": "Report restored successfully"}


@api_router.delete("/reports/{report_id}/permanent")
async def permanently_delete_report_v2(report_id: str, current_user: User = Depends(get_current_user)):
    """حذف البلاغ نهائياً من قاعدة البيانات"""
    query = {"id": report_id, "is_deleted": True}
    
    # فلترة حسب صلاحيات المشاريع والمحافظات
    if current_user.role != "admin":
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
    
    report_doc = await db.reports.find_one(query, {"_id": 0})
    if not report_doc:
        raise HTTPException(status_code=404, detail="Deleted report not found or no permission")
    
    # حذف البلاغ نهائياً من قاعدة البيانات
    result = await db.reports.delete_one({"id": report_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Report not found")
    
    return {"message": "تم الحذف النهائي بنجاح"}


@api_router.delete("/reports-trash/all")
async def permanently_delete_all_trashed(current_user: User = Depends(get_current_user)):
    """حذف نهائي لجميع البلاغات في سلة المحذوفات (حسب صلاحيات المستخدم)"""
    has_trash_perm = current_user.role == "admin" or "trash" in (current_user.permissions or []) or user_has_any_project_permission(current_user, "trash")
    if not has_trash_perm:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية الحذف النهائي")
    
    query = {"is_deleted": True}
    if current_user.role != "admin":
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
    
    # نقل إلى الحذف النهائي
    docs = await db.reports.find(query, {"_id": 0}).to_list(10000)
    if not docs:
        return {"message": "لا توجد بلاغات للحذف", "deleted_count": 0}
    
    now = datetime.now(timezone.utc).isoformat()
    for doc in docs:
        doc["permanently_deleted"] = True
        doc["permanently_deleted_at"] = now
        doc["permanently_deleted_by"] = current_user.id
    
    await db.permanently_deleted_reports.insert_many(docs)
    delete_result = await db.reports.delete_many(query)
    
    return {"message": f"تم حذف {delete_result.deleted_count} بلاغ نهائياً", "deleted_count": delete_result.deleted_count}



@api_router.delete("/reports-trash/{report_id}/permanent")
async def permanently_delete_report(report_id: str, current_user: User = Depends(get_current_user)):
    # التحقق من الصلاحيات
    has_trash_perm = current_user.role == "admin" or "trash" in (current_user.permissions or []) or user_has_any_project_permission(current_user, "trash")
    if not has_trash_perm and current_user.username != "Eng Mahmoud Haroun":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية الحذف النهائي")
    
    query = {"id": report_id, "is_deleted": True}
    
    # فلترة حسب صلاحيات المشاريع والمحافظات
    if current_user.role != "admin":
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
    
    report_doc = await db.reports.find_one(query, {"_id": 0})
    if not report_doc:
        raise HTTPException(status_code=404, detail="Deleted report not found")
    
    # نقل البلاغ إلى مجموعة الحذف النهائي
    permanently_deleted_report = {
        **report_doc,
        "permanently_deleted": True,
        "permanently_deleted_at": datetime.now(timezone.utc).isoformat(),
        "permanently_deleted_by": current_user.id
    }
    
    await db.permanently_deleted_reports.insert_one(permanently_deleted_report)
    
    # حذف البلاغ من المجموعة الأساسية
    await db.reports.delete_one({"id": report_id})
    
    return {"message": "تم الحذف النهائي بنجاح"}


@api_router.get("/reports-permanently-deleted")
async def get_permanently_deleted_reports(current_user: User = Depends(get_current_admin_user)):
    """الحصول على سجل البلاغات المحذوفة نهائياً"""
    query = {}
    
    # فلترة حسب صلاحيات المشاريع والمحافظات
    if current_user.role != "admin" or len(current_user.projects) > 0:
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
    
    if current_user.role != "admin" or len(current_user.governorates) > 0:
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
    
    reports = await db.permanently_deleted_reports.find(query, {"_id": 0}).sort("permanently_deleted_at", -1).to_list(1000)
    
    # الحصول على معلومات المستخدمين الذين قاموا بالحذف
    for report in reports:
        if report.get('permanently_deleted_by'):
            user = await db.users.find_one({"id": report['permanently_deleted_by']}, {"_id": 0, "full_name": 1})
            report['permanently_deleted_by_name'] = user.get('full_name') if user else 'غير معروف'
        else:
            report['permanently_deleted_by_name'] = 'غير معروف'
    
    return reports


# ============= WATER & SEWAGE CONNECTIONS TRASH =============

@api_router.get("/water-connections-trash")
async def get_deleted_water_connections(
    page: int = 1,
    limit: int = 15,
    current_user: User = Depends(get_current_user)
):
    """جلب توصيلات المياه المحذوفة مؤقتاً"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "water_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        
    query = {"is_deleted": True}
    
    # فلترة حسب المشاريع وصاحب الحذف ("كل مشروع وكل مستخدم خاصة بسلة محذوفاتة")
    if not is_admin:
        allowed_projects = get_projects_with_permission(current_user, "water_connections")
        query["project"] = {"$in": allowed_projects}
        query["deleted_by"] = current_user.id
        
    total_count = await db.water_connections.count_documents(query)
    skip = (page - 1) * limit
    connections = await db.water_connections.find(query, {"_id": 0}).sort("deleted_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # جلب أسماء المستخدمين الذين حذفوا التوصيلات
    for conn in connections:
        if conn.get('deleted_by'):
            user_doc = await db.users.find_one({"id": conn['deleted_by']}, {"_id": 0, "full_name": 1})
            conn['deleted_by_name'] = user_doc.get('full_name') if user_doc else 'غير معروف'
        else:
            conn['deleted_by_name'] = 'غير معروف'
            
    return {"items": connections, "total": total_count}


@api_router.post("/water-connections-trash/{conn_id}/restore")
async def restore_water_connection(conn_id: str, current_user: User = Depends(get_current_user)):
    """استعادة توصيلة مياه محذوفة"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "water_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        
    query = {"id": conn_id, "is_deleted": True}
    if not is_admin:
        allowed_projects = get_projects_with_permission(current_user, "water_connections")
        query["project"] = {"$in": allowed_projects}
        query["deleted_by"] = current_user.id
        
    result = await db.water_connections.update_one(
        query,
        {"$unset": {"is_deleted": "", "deleted_at": "", "deleted_by": ""}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة في سلة المحذوفات أو ليست تابعة لك")
        
    return {"message": "تم استعادة التوصيلة بنجاح"}


@api_router.delete("/water-connections-trash/{conn_id}/permanent")
async def permanently_delete_water_connection(conn_id: str, current_user: User = Depends(get_current_user)):
    """حذف توصيلة مياه نهائياً"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "water_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        
    query = {"id": conn_id, "is_deleted": True}
    if not is_admin:
        allowed_projects = get_projects_with_permission(current_user, "water_connections")
        query["project"] = {"$in": allowed_projects}
        query["deleted_by"] = current_user.id
        
    result = await db.water_connections.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة في سلة المحذوفات أو ليست تابعة لك")
        
    return {"message": "تم حذف التوصيلة نهائياً بنجاح"}


@api_router.get("/sewage-connections-trash")
async def get_deleted_sewage_connections(
    page: int = 1,
    limit: int = 15,
    current_user: User = Depends(get_current_user)
):
    """جلب توصيلات الصرف الصحي المحذوفة مؤقتاً"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "sewage_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        
    query = {"is_deleted": True}
    
    # فلترة حسب المشاريع وصاحب الحذف ("كل مشروع وكل مستخدم خاصة بسلة محذوفاتة")
    if not is_admin:
        allowed_projects = get_projects_with_permission(current_user, "sewage_connections")
        query["project"] = {"$in": allowed_projects}
        query["deleted_by"] = current_user.id
        
    total_count = await db.sewage_connections.count_documents(query)
    skip = (page - 1) * limit
    connections = await db.sewage_connections.find(query, {"_id": 0}).sort("deleted_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # جلب أسماء المستخدمين الذين حذفوا التوصيلات
    for conn in connections:
        if conn.get('deleted_by'):
            user_doc = await db.users.find_one({"id": conn['deleted_by']}, {"_id": 0, "full_name": 1})
            conn['deleted_by_name'] = user_doc.get('full_name') if user_doc else 'غير معروف'
        else:
            conn['deleted_by_name'] = 'غير معروف'
            
    return {"items": connections, "total": total_count}


@api_router.post("/sewage-connections-trash/{conn_id}/restore")
async def restore_sewage_connection(conn_id: str, current_user: User = Depends(get_current_user)):
    """استعادة توصيلة صرف صحي محذوفة"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "sewage_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        
    query = {"id": conn_id, "is_deleted": True}
    if not is_admin:
        allowed_projects = get_projects_with_permission(current_user, "sewage_connections")
        query["project"] = {"$in": allowed_projects}
        query["deleted_by"] = current_user.id
        
    result = await db.sewage_connections.update_one(
        query,
        {"$unset": {"is_deleted": "", "deleted_at": "", "deleted_by": ""}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة في سلة المحذوفات أو ليست تابعة لك")
        
    return {"message": "تم استعادة التوصيلة بنجاح"}


@api_router.delete("/sewage-connections-trash/{conn_id}/permanent")
async def permanently_delete_sewage_connection(conn_id: str, current_user: User = Depends(get_current_user)):
    """حذف توصيلة صرف صحي نهائياً"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "sewage_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        
    query = {"id": conn_id, "is_deleted": True}
    if not is_admin:
        allowed_projects = get_projects_with_permission(current_user, "sewage_connections")
        query["project"] = {"$in": allowed_projects}
        query["deleted_by"] = current_user.id
        
    result = await db.sewage_connections.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة في سلة المحذوفات أو ليست تابعة لك")
        
    return {"message": "تم حذف التوصيلة نهائياً بنجاح"}


@api_router.get("/reports/check-duplicate")
async def check_duplicate_report(
    report_number: str = Query(...),
    project: str = Query(None),
    license_number: str = Query(None),
    report_type: str = Query(None),
    current_user: User = Depends(get_current_user)
):
    """فحص وجود بلاغ مكرر - يفحص رقم البلاغ + المشروع + رقم الرخصة + نوع البلاغ"""
    query = {
        "report_number": report_number,
        "is_deleted": {"$ne": True}
    }
    
    # إضافة فلاتر إضافية إذا كانت موجودة
    if project:
        query["project"] = project
    if license_number:
        query["license_number"] = license_number
    if report_type:
        query["report_type"] = report_type
    
    existing = await db.reports.find_one(query)
    
    return {
        "exists": existing is not None,
        "duplicate_info": {
            "report_number": existing.get("report_number") if existing else None,
            "governorate": existing.get("governorate") if existing else None,
            "status": existing.get("status") if existing else None
        } if existing else None
    }


@api_router.post("/reports/bulk-delete")
async def bulk_delete_reports(
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """حذف عدة بلاغات دفعة واحدة"""
    ids = data.get('ids', [])
    
    if not ids:
        raise HTTPException(status_code=400, detail="No report IDs provided")
    
    # التحقق من الصلاحيات للمستخدم العادي (يمكن تركها أو تطبيق نفس شروط الحذف المفرد)
    # لكن الأهم هو النقل لسلة المحذوفات بدلاً من الحذف النهائي
    now = datetime.now(timezone.utc).isoformat()
    result = await db.reports.update_many(
        {"id": {"$in": ids}},
        {"$set": {
            "is_deleted": True,
            "deleted_at": now,
            "deleted_by": current_user.id
        }}
    )
    
    # ⚡ مسح الـ cache لتحديث الإحصائيات
    global stats_cache
    stats_cache.clear()
    
    return {
        "message": f"تم نقل {result.modified_count} بلاغ إلى سلة المحذوفات بنجاح",
        "deleted_count": result.modified_count
    }


@api_router.post("/reports/fix-license-status")
async def fix_reports_license_status(
    current_user: User = Depends(get_current_user)
):
    """تحديث حالة الرخص والإغلاق لجميع البلاغات"""
    if current_user.role != 'admin':
        raise HTTPException(status_code=403, detail="هذه العملية متاحة للمسؤول فقط")
    
    # تحديث البلاغات التي لها رقم رخصة صادر (يحوي رقم)
    license_result = await db.reports.update_many(
        {
            "license_number": {"$regex": "[0-9]"},
            "is_deleted": {"$ne": True}
        },
        {
            "$set": {"asphalt_license_issued": True}
        }
    )
    
    # تحديث البلاغات التي ليس لها رخصة
    no_license_result = await db.reports.update_many(
        {
            "$or": [
                {"license_number": {"$exists": False}},
                {"license_number": None},
                {"license_number": {"$not": {"$regex": "[0-9]"}}}
            ],
            "is_deleted": {"$ne": True}
        },
        {
            "$set": {"asphalt_license_issued": False}
        }
    )
    
    # تحديث البلاغات المغلقة
    closed_result = await db.reports.update_many(
        {
            "status": {"$in": ["تم الإصلاح", "مغلق", "منتهي"]},
            "is_deleted": {"$ne": True}
        },
        {
            "$set": {"wfm_closed": True, "review_status": "تمت المراجعة"}
        }
    )
    
    return {
        "success": True,
        "updated_with_license": license_result.modified_count,
        "updated_without_license": no_license_result.modified_count,
        "updated_closed": closed_result.modified_count,
        "message": f"تم تحديث {license_result.modified_count} بلاغ برخصة، {no_license_result.modified_count} بلاغ بدون رخصة، {closed_result.modified_count} بلاغ مغلق"
    }


@api_router.post("/reports/import-excel")
async def import_reports_from_excel(
    file: UploadFile = File(...),
    selected_project: Optional[str] = Form(None),
    selected_governorate: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user)
):
    """استيراد بلاغات من ملف Excel مع إمكانية تحديد المشروع والمحافظة"""
    # التحقق من صلاحية الاستيراد
    if not user_has_any_project_permission(current_user, 'reports_import'):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية استيراد البلاغات")
    
    # التحقق من نوع الملف
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="يجب رفع ملف Excel (.xlsx أو .xls)")
    
    try:
        import pandas as pd
        from io import BytesIO
        
        # قراءة الملف
        content = await file.read()
        content_buffer = BytesIO(content)
        
        # قراءة الملف بدون header أولاً لتحليل البنية
        df_raw = pd.read_excel(content_buffer, header=None)
        
        # البحث عن صف الأسماء (يحتوي على كلمات مثل "بلاغ" أو "محافظ")
        header_row = None
        for i in range(min(10, len(df_raw))):
            row_values = [str(v) for v in df_raw.iloc[i].tolist()]
            row_str = ' '.join(row_values)
            if 'بلاغ' in row_str and ('محافظ' in row_str or 'رخص' in row_str):
                header_row = i
                print(f"✅ تم العثور على أسماء الأعمدة في الصف {i}: {row_values}")
                break
        
        if header_row is not None:
            # إعادة القراءة مع تحديد صف الأسماء
            content_buffer.seek(0)
            df = pd.read_excel(content_buffer, header=header_row)
        else:
            # استخدام القراءة العادية
            content_buffer.seek(0)
            df = pd.read_excel(content_buffer)
        
        # إزالة الأعمدة الفارغة (Unnamed)
        cols_to_drop = [col for col in df.columns if 'Unnamed' in str(col) or pd.isna(col)]
        df = df.drop(columns=cols_to_drop, errors='ignore')
        
        # إزالة الصفوف الفارغة
        df = df.dropna(how='all')
        
        if df is None or len(df) == 0:
            raise HTTPException(status_code=400, detail="لم يتم العثور على بيانات في الملف")
        
        # طباعة معلومات الملف للتصحيح
        original_columns = [str(col).replace(' *', '').replace('*', '').strip() for col in df.columns]
        print(f"📋 الأعمدة النهائية: {original_columns}")
        print(f"📊 عدد الصفوف: {len(df)}")
        print(f"📄 أول 3 صفوف:\n{df.head(3).to_string()}")
        
        # تنظيف أسماء الأعمدة (إزالة * والمسافات الزائدة)
        df.columns = [str(col).replace(' *', '').replace('*', '').strip() for col in df.columns]
        
        # تحويل أسماء الأعمدة - دعم عدة صيغ لكل عمود
        column_mapping = {}
        for col in df.columns:
            col_clean = str(col).strip()
            
            # الترتيب مهم! الأكثر تحديداً أولاً
            
            # رقم الرخصة (قبل رقم البلاغ)
            if 'رخص' in col_clean or 'license' in col_clean.lower():
                column_mapping[col] = 'license_number'
            # نوع البلاغ
            elif 'نوع' in col_clean or 'type' in col_clean.lower():
                column_mapping[col] = 'report_type'
            # الحالة
            elif 'حال' in col_clean or 'status' in col_clean.lower():
                column_mapping[col] = 'status'
            # المحافظة
            elif 'محافظ' in col_clean or 'governorate' in col_clean.lower():
                column_mapping[col] = 'governorate'
            # المشروع
            elif 'مشروع' in col_clean or 'project' in col_clean.lower():
                column_mapping[col] = 'project'
            # العمق
            elif 'عمق' in col_clean or 'depth' in col_clean.lower():
                column_mapping[col] = 'depth_meters'
            # القطر
            elif 'قطر' in col_clean or 'diameter' in col_clean.lower():
                column_mapping[col] = 'diameter_mm'
            # المقاول
            elif 'مقاول' in col_clean or 'contractor' in col_clean.lower():
                column_mapping[col] = 'contractor'
            # الملاحظات
            elif 'ملاحظ' in col_clean or 'notes' in col_clean.lower():
                column_mapping[col] = 'notes'
            # رقم البلاغ (آخر شيء - لأنه يحتوي على كلمات عامة)
            elif 'بلاغ' in col_clean or col_clean == 'رقم' or 'report_number' in col_clean.lower():
                column_mapping[col] = 'report_number'
        
        print(f"🔄 تحويل الأعمدة: {column_mapping}")
        
        # تطبيق التحويل
        df = df.rename(columns=column_mapping)
        print(f"📋 الأعمدة بعد التحويل: {list(df.columns)}")
        
        # إذا لم يتم العثور على عمود report_number، استخدم أول عمود
        if 'report_number' not in df.columns and len(df.columns) > 0:
            first_col = df.columns[0]
            df = df.rename(columns={first_col: 'report_number'})
            print(f"⚠️ استخدام العمود الأول '{first_col}' كرقم البلاغ")
        
        imported = 0
        skipped = 0
        errors = []
        
        # الحصول على المشروع الافتراضي من صلاحيات المستخدم
        default_project = current_user.projects[0] if current_user.projects else 'مشروع إصلاح أعمال المحافظات الغربية - القطاع الأوسط'
        
        # طباعة المشروع والمحافظة المختارة للتتبع
        print(f"📌 المشروع المختار من الواجهة: '{selected_project}'")
        print(f"📌 المحافظة المختارة من الواجهة: '{selected_governorate}'")
        print(f"📌 المشروع الافتراضي: '{default_project}'")
        
        print(f"🚀 بدء معالجة {len(df)} صف...")
        
        for idx, row in df.iterrows():
            try:
                # قراءة رقم البلاغ
                report_number = None
                if 'report_number' in df.columns:
                    val = row['report_number']
                    if hasattr(val, 'iloc'):
                        val = val.iloc[0] if len(val) > 0 else None
                    report_number = val
                elif len(df.columns) > 0:
                    val = row.iloc[0]
                    report_number = val
                
                # تنظيف رقم البلاغ
                if report_number is not None:
                    report_number = str(report_number).strip()
                
                # تخطي الصفوف الفارغة
                if not report_number or report_number.lower() in ['nan', 'none', '', 'null']:
                    continue
                
                print(f"📝 معالجة صف {idx + 2}: رقم البلاغ = {report_number}")
                
                # دالة مساعدة لقراءة القيم بشكل آمن
                def safe_str(col_name, default=''):
                    try:
                        if col_name in row.index:
                            val = row[col_name]
                            if pd.isna(val) or val is None or str(val).strip().lower() in ['nan', 'none', '']:
                                return default
                            return str(val).strip()
                        return default
                    except:
                        return default
                
                # الحصول على المشروع من الصف أو من الـ parameters أو الافتراضي
                # الأولوية: 1. الـ Excel 2. المشروع المختار 3. الافتراضي
                row_project = safe_str('project', '') or safe_str('المشروع', '') or selected_project or default_project
                
                # الحصول على المحافظة من الصف أو من الـ parameters أو الافتراضي
                # الأولوية: 1. الـ Excel 2. المحافظة المختارة 3. الافتراضية
                default_governorate = current_user.governorates[0] if current_user.governorates else ''
                row_governorate = safe_str('governorate', '') or safe_str('المحافظة', '') or selected_governorate or default_governorate
                
                # التحقق من عدم وجود البلاغ في نفس المشروع فقط
                existing = await db.reports.find_one({
                    "report_number": report_number, 
                    "project": row_project,
                    "is_deleted": {"$ne": True}
                }, {"_id": 1})
                
                if existing:
                    print(f"⏭️ تخطي صف {idx + 2}: البلاغ {report_number} موجود مسبقاً في المشروع {row_project}")
                    skipped += 1
                    continue
                
                def safe_float(col_name, default=0):
                    try:
                        if col_name in row.index:
                            val = row[col_name]
                            if pd.isna(val) or val is None or str(val).strip().lower() in ['nan', 'none', '']:
                                return default
                            return float(val)
                    except:
                        return default
                
                # قراءة رقم الرخصة والتحقق منه
                license_num = safe_str('license_number', '')
                # التحقق إذا كانت الرخصة صادرة (ليست فارغة وليست "لم يتم إصدار رخصة")
                has_license = bool(license_num and license_num not in ['لم يتم إصدار رخصة', 'لم يتم', '-', '0', 'nan', 'none', ''])
                
                # قراءة نوع البلاغ والحالة
                report_type_val = safe_str('report_type', 'تسرب')
                status_val = safe_str('status', 'جديد')
                
                # تحديد إذا كانت رخصة الأسفلت صادرة بناءً على نوع البلاغ والرخصة
                # البلاغات من نوع أسفلت أو بلاط تحتاج رخصة
                is_asphalt_type = 'اسفلت' in report_type_val.lower() or 'أسفلت' in report_type_val.lower() or 'asphalt' in report_type_val.lower()
                is_tile_type = 'بلاط' in report_type_val.lower() or 'tile' in report_type_val.lower()
                
                # تحضير البيانات
                now = datetime.now(timezone.utc)
                report_data = {
                    "id": str(uuid.uuid4()),
                    "report_number": report_number,
                    "license_number": license_num if has_license else 'لم يتم إصدار رخصة',
                    "report_type": report_type_val,
                    "status": status_val,
                    "governorate": row_governorate,  # استخدام المحافظة المحددة
                    "project": row_project,  # استخدام المشروع المحدد من الصف
                    "depth_meters": safe_float('depth_meters', 0),
                    "diameter_mm": safe_float('diameter_mm', 0),
                    "contractor": safe_str('contractor', ''),
                    "latitude": safe_str('latitude', ''),
                    "longitude": safe_str('longitude', ''),
                    "notes": safe_str('notes', ''),
                    "images": [],
                    "is_deleted": {"$ne": True},
                    "created_by": current_user.id,
                    "created_by_name": current_user.full_name or current_user.username,
                    "created_at": now,
                    "updated_at": now,
                    "added_at": now,  # وقت الإضافة للنظام
                    "report_date": now,  # تاريخ البلاغ
                    "review_status": "تمت المراجعة" if status_val in ['تم الإصلاح', 'مغلق', 'منتهي'] else "قيد المراجعة",
                    "asphalt_license_issued": has_license,  # صدرت الرخصة إذا كان هناك رقم رخصة
                    "wfm_closed": status_val in ['تم الإصلاح', 'مغلق', 'منتهي'],  # مغلق في WFM إذا كانت الحالة منتهية
                    "closed_at": now if status_val in ['تم الإصلاح', 'مغلق', 'منتهي'] else None
                }
                
                # معالجة التاريخ إذا وجد في Excel
                date_columns = ['created_at', 'تاريخ', 'التاريخ', 'date', 'report_date', 'تاريخ البلاغ']
                for date_col in date_columns:
                    if date_col in row.index and pd.notna(row.get(date_col)):
                        try:
                            date_val = row[date_col]
                            if isinstance(date_val, str):
                                # محاولة تحويل التاريخ النصي
                                parsed_date = pd.to_datetime(date_val, dayfirst=True, errors='coerce')
                                if pd.notna(parsed_date):
                                    report_data['created_at'] = parsed_date.to_pydatetime()
                                    report_data['report_date'] = parsed_date.to_pydatetime()
                            else:
                                parsed_date = pd.to_datetime(date_val, errors='coerce')
                                if pd.notna(parsed_date):
                                    report_data['created_at'] = parsed_date.to_pydatetime()
                                    report_data['report_date'] = parsed_date.to_pydatetime()
                            break
                        except:
                            pass
                
                await db.reports.insert_one(report_data)
                imported += 1
                print(f"✅ تم إضافة البلاغ: {report_number}")
                
            except Exception as e:
                # print error ❌ خطأ في صف {idx + 2}
                errors.append(f"سطر {idx + 2}: {str(e)}")
        
        print(f"📊 النتيجة: تم استيراد {imported}، تم تخطي {skipped}، أخطاء {len(errors)}")
        
        # إعداد رسالة مفصلة
        message = f"تم استيراد {imported} بلاغ بنجاح"
        if skipped > 0:
            message += f"، تم تخطي {skipped} بلاغ موجود مسبقاً"
        if imported == 0 and skipped == 0:
            message = f"لم يتم استيراد أي بلاغ. الأعمدة الموجودة: {original_columns}"
        
        return {
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "total": len(df),
            "errors": errors[:10] if errors else [],
            "columns_found": original_columns,
            "columns_mapped": list(column_mapping.values()),
            "message": message
        }
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"❌ خطأ: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"خطأ في معالجة الملف: {str(e)}")


# ============= فواتير العهدة =============

@api_router.post("/invoices")
async def create_invoice(
    data: InvoiceCreate,
    current_user: User = Depends(get_current_user)
):
    """رفع فاتورة جديدة - للمستوى 3"""
    # الحصول على المدير المباشر (created_by)
    manager = None
    if current_user.created_by:
        manager = await db.users.find_one({"id": current_user.created_by}, {"_id": 0})
    
    # معالجة الصور ورفعها لـ Cloudinary إذا كانت base64
    processed_images = await process_images_for_storage(data.images or ([data.image] if data.image else []), category="invoices")

    invoice = {
        "id": str(uuid4()),
        "invoice_number": data.invoice_number,
        "amount": data.amount,
        "description": data.description,
        "images": processed_images,
        "uploaded_by": current_user.id,
        "uploaded_by_name": current_user.full_name,
        "project": data.project,
        "governorate": data.governorate,
        "invoice_date": data.invoice_date or datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        "status": "pending",
        "manager_id": current_user.created_by,
        "manager_name": manager.get("full_name") if manager else None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.invoices.insert_one(invoice)
    return {"message": "تم رفع الفاتورة بنجاح", "id": invoice["id"]}


@api_router.get("/invoices")
async def get_invoices(
    status: Optional[str] = None,
    project: Optional[str] = None,
    month: Optional[str] = None,
    date: Optional[str] = None,  # فلتر التاريخ المحدد (اليوم)
    page: int = 1,
    limit: int = 10,
    for_review: bool = False,  # جلب الفواتير للمراجعة فقط
    current_user: User = Depends(get_current_user)
):
    """
    جلب الفواتير حسب صلاحية المستخدم والمشروع مع الترقيم
    
    نظام المراجعة المتدرج:
    - review_invoices_1: يرى الفواتير بحالة approved_by_manager
    - review_invoices_2: يرى الفواتير بحالة approved_by_reviewer_1
    - review_invoices_3: يرى الفواتير بحالة approved_by_reviewer_2
    """
    query = {"is_deleted": {"$ne": True}}
    user_permissions = current_user.permissions or []
    
    # صلاحية عرض جميع الفواتير (عامة أو في أي مشروع)
    can_view_all = "view_all_invoices" in user_permissions or user_has_any_project_permission(current_user, "view_all_invoices")
    
    # المشاريع التي يملك فيها المستخدم صلاحية view_all_invoices
    projects_view_all = get_projects_with_permission(current_user, "view_all_invoices")
    
    # التحقق إذا كان المستخدم لديه موظفين (مدير مشروع)
    sub_users = await db.users.find({"created_by": current_user.id}, {"_id": 0, "id": 1}).to_list(1000)
    sub_user_ids = [u["id"] for u in sub_users]
    has_sub_users = len(sub_user_ids) > 0
    
    # التحقق من صلاحيات الاعتماد النهائي
    has_invoices_perm = "invoices" in user_permissions or can_view_all or "review_invoices_3" in user_permissions
    has_multiple_projects = len(current_user.projects or []) >= 2
    can_final_approve = has_invoices_perm and has_multiple_projects
    
    # المراجعة المتدرجة
    if for_review:
        if current_user.role == "admin":
            # Admin يرى الفواتير المعتمدة من مدير المشروع (جاهزة للاعتماد النهائي)
            query["status"] = "approved_by_manager"
        elif has_sub_users and ("review_invoices" in user_permissions or current_user.can_create_subusers):
            # مدير المشروع الذي لديه موظفين: يرى فواتير موظفيه pending للمراجعة الأولى
            query["status"] = "pending"
            query["uploaded_by"] = {"$in": sub_user_ids}
        elif can_final_approve:
            # من لديه صلاحية جميع المشاريع + صلاحية الفواتير: يرى الفواتير المعتمدة من المديرين
            query["status"] = "approved_by_manager"
            # فلترة حسب مشاريعه
            if current_user.projects and len(current_user.projects) > 0:
                query["project"] = {"$in": current_user.projects}
        else:
            return {"invoices": [], "total_count": 0, "total_pages": 0, "current_page": 1}
        
        # إذا تم تحديد مشروع معين في الفلتر، استخدمه
        if project:
            query["project"] = project
    
    # العرض العادي - كل مستخدم يرى فواتيره + فواتير موظفيه + الفواتير المعتمدة حسب صلاحياته
    # جميع الفواتير تبقى ظاهرة بغض النظر عن حالتها (مرفوضة، ملغاة، معتمدة)
    elif current_user.role == "admin":
        # الأدمن يرى جميع الفواتير (ما عدا pending التي تحتاج مراجعة المدير أولاً)
        if status and status != 'all':
            query["status"] = status
        else:
            query["status"] = {"$in": ["approved_by_manager", "approved_final", "approved_by_admin", "rejected", "cancelled"]}
        if project:
            query["project"] = project
    elif can_view_all:
        # المستخدم مع صلاحية view_all_invoices - يرى جميع الفواتير بكل الحالات
        if status and status != 'all':
            query["status"] = status
        if project:
            query["project"] = project
        else:
            # إذا كانت الصلاحية عامة: يرى جميع مشاريعه
            # إذا كانت لبعض المشاريع فقط: يرى فقط تلك المشاريع
            if "view_all_invoices" in user_permissions:
                if current_user.projects and len(current_user.projects) > 0:
                    query["project"] = {"$in": current_user.projects}
            elif projects_view_all:
                query["project"] = {"$in": projects_view_all}
    elif has_sub_users:
        # مدير المشروع (لديه موظفين): يرى فواتير موظفيه + فواتيره الشخصية (كل الحالات)
        query["uploaded_by"] = {"$in": sub_user_ids + [current_user.id]}
        # لا نفلتر على الحالة افتراضياً - يرى كل شيء بما فيه المرفوضة والملغاة
        if status and status != 'all':
            query["status"] = status
        if project:
            query["project"] = project
        elif current_user.projects and len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
    else:
        # الموظف (المستوى 3) يرى فواتيره فقط (كل الحالات بما فيها المرفوضة والملغاة)
        query["uploaded_by"] = current_user.id
        # لا نفلتر على الحالة افتراضياً - يرى كل فواتيره
        if status and status != 'all':
            query["status"] = status
        if project:
            query["project"] = project
        elif current_user.projects and len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
    
    # فلترة حسب تاريخ محدد (يوم)
    if date:
        # date format: "2025-12-15"
        date_filter = {"$or": [
            {"invoice_date": date},
            {"invoice_date": {"$regex": f"^{date}"}}
        ]}
        if "$or" in query:
            existing_or = query.pop("$or")
            query["$and"] = [{"$or": existing_or}, date_filter]
        else:
            query.update(date_filter)
    # فلترة حسب الشهر (إذا لم يتم تحديد تاريخ محدد)
    elif month:
        # month format: "2025-12" or just "12"
        if "-" in month:
            month_regex = f"^{month}"
        else:
            month_regex = f"-{month.zfill(2)}-"
        
        if "$or" in query:
            # إذا كان هناك $or موجود، نحتاج لإعادة هيكلة الاستعلام
            existing_or = query.pop("$or")
            query["$and"] = [
                {"$or": existing_or},
                {"$or": [
                    {"invoice_date": {"$regex": month_regex}},
                    {"created_at": {"$regex": month_regex}}
                ]}
            ]
        else:
            query["$or"] = [
                {"invoice_date": {"$regex": month_regex}},
                {"created_at": {"$regex": month_regex}}
            ]
    
    # Get total count
    total_count = await db.invoices.count_documents(query)
    total_pages = (total_count + limit - 1) // limit if limit > 0 else 1
    
    # Apply pagination
    skip = (page - 1) * limit
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "invoices": invoices,
        "total_count": total_count,
        "total_pages": total_pages,
        "current_page": page,
        "limit": limit
    }


@api_router.get("/invoices/months")
async def get_invoice_months(
    current_user: User = Depends(get_current_user)
):
    """جلب الشهور المتاحة للفواتير"""
    invoices = await db.invoices.find({}, {"_id": 0, "invoice_date": 1, "created_at": 1}).to_list(1000)
    months = set()
    for inv in invoices:
        date_str = inv.get("invoice_date") or inv.get("created_at", "")
        if date_str and len(date_str) >= 7:
            month_year = date_str[:7]  # "2025-12"
            months.add(month_year)
    return sorted(list(months), reverse=True)



@api_router.put("/invoices/{invoice_id}")
async def update_invoice(
    invoice_id: str,
    data: InvoiceCreate,
    current_user: User = Depends(get_current_user)
):
    """تعديل فاتورة - فقط لصاحب الفاتورة قبل الاعتماد"""
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="الفاتورة غير موجودة")
    
    # فقط صاحب الفاتورة يمكنه التعديل
    if invoice.get("uploaded_by") != current_user.id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل هذه الفاتورة")
    
    # لا يمكن التعديل بعد الاعتماد
    if invoice.get("status") in ["approved_by_manager", "approved_by_admin"]:
        raise HTTPException(status_code=400, detail="لا يمكن تعديل الفاتورة بعد الاعتماد")
    
    update_data = {
        "invoice_number": data.invoice_number,
        "amount": data.amount,
        "description": data.description,
        "project": data.project,
        "governorate": data.governorate,
        "invoice_date": data.invoice_date,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # تحديث الصور - دعم صور متعددة
    if data.images:
        update_data["images"] = data.images
    elif data.image:
        # دمج الصورة الجديدة مع الموجودة
        existing_images = invoice.get("images", [])
        if data.image not in existing_images:
            existing_images.append(data.image)
        update_data["images"] = existing_images
    
    await db.invoices.update_one({"id": invoice_id}, {"$set": update_data})
    return {"message": "تم تعديل الفاتورة بنجاح"}


@api_router.put("/invoices/{invoice_id}/approve")
async def approve_invoice(
    invoice_id: str,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    نظام اعتماد الفواتير المتدرج:
    1. الموظف (المستوى 3) يرفع الفاتورة (status: pending)
    2. مدير المشروع (المستوى 2) يعتمدها اعتماد أولي (status: approved_by_manager)
    3. بيت الخبرة (الأدمن) فقط يعتمدها نهائياً (status: approved_final) أو يلغي الاعتماد
    
    - المستوى 2 لا يمكنه الاعتماد النهائي أبداً
    - الأدمن يمكنه الاعتماد النهائي من أي مرحلة
    - الأدمن يمكنه إلغاء الاعتماد وإعادته
    """
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="الفاتورة غير موجودة")
    
    current_status = invoice.get("status", "pending")
    user_permissions = current_user.permissions or []
    invoice_project = invoice.get("project")
    
    # التحقق من تفويض الاعتماد النهائي (نيابة عن بيت الخبرة)
    # يُمنح عبر صلاحية review_invoices_3 أو view_all_invoices من قِبَل الأدمن
    has_final_review_delegation = (
        "review_invoices_3" in user_permissions or
        "view_all_invoices" in user_permissions or
        has_project_permission(current_user, invoice_project, "review_invoices_3") or
        has_project_permission(current_user, invoice_project, "view_all_invoices")
    )
    
    # فصل المهام: من قام بالاعتماد الأولي لا يمكنه الاعتماد النهائي لنفس الفاتورة
    if current_status == "approved_by_manager" and invoice.get("reviewed_by_manager") == current_user.id:
        has_final_review_delegation = False
    
    update_data = {}
    
    # Admin (بيت الخبرة) يعتمد الفاتورة نهائياً من أي حالة
    if current_user.role == "admin":
        update_data = {
            "status": "approved_final",
            "final_approved_by": current_user.id,
            "final_approved_by_name": current_user.full_name or current_user.username,
            "final_approved_at": datetime.now(timezone.utc).isoformat(),
            "final_notes": notes
        }
    
    # المستخدم المفوض بالاعتماد النهائي (نيابة عن بيت الخبرة)
    elif current_status == "approved_by_manager" and has_final_review_delegation:
        update_data = {
            "status": "approved_final",
            "final_approved_by": current_user.id,
            "final_approved_by_name": current_user.full_name or current_user.username,
            "final_approved_at": datetime.now(timezone.utc).isoformat(),
            "final_notes": notes,
            "final_approved_on_behalf": True,  # تم الاعتماد نيابة عن بيت الخبرة
            "delegated_by_admin": True
        }
    
    # المستوى 2 (مدير المشروع) - اعتماد أولي فقط
    elif current_status == "pending":
        # الفاتورة قيد الانتظار - يراجعها مدير المشروع (الذي أنشأ الموظف) أو يعتمد فاتورته بنفسه
        uploader_id = invoice.get("uploaded_by")
        
        # السماح للمدير بالاعتماد إذا كانت الفاتورة:
        # 1. من أحد موظفيه (المستوى 3) الذين أنشأهم
        # 2. أو مرفوعة من قبله هو نفسه (اعتماد ذاتي للذهاب إلى الاعتماد النهائي)
        is_own_invoice = uploader_id == current_user.id
        
        if not is_own_invoice:
            uploader = await db.users.find_one({"id": uploader_id}, {"_id": 0, "created_by": 1})
            if not uploader or uploader.get("created_by") != current_user.id:
                raise HTTPException(status_code=403, detail="هذه الفاتورة ليست من موظفيك - يجب أن يراجعها مديرهم المباشر")
        
        # التحقق من صلاحية المراجعة
        if not ("review_invoices" in user_permissions or current_user.can_create_subusers):
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية مراجعة الفواتير")
        
        update_data = {
            "status": "approved_by_manager",
            "reviewed_by_manager": current_user.id,
            "reviewed_by_manager_name": current_user.full_name or current_user.username,
            "reviewed_by_manager_at": datetime.now(timezone.utc).isoformat(),
            "manager_notes": notes,
            "self_approved": is_own_invoice  # علامة للاعتماد الذاتي
        }
    
    elif current_status == "approved_by_manager":
        # الفاتورة معتمدة من المدير - فقط الأدمن أو المفوض يعتمدها نهائياً
        raise HTTPException(status_code=403, detail="الاعتماد النهائي متاح فقط لبيت الخبرة أو من فوضه")
    
    elif current_status == "approved_final":
        raise HTTPException(status_code=400, detail="هذه الفاتورة معتمدة نهائياً بالفعل")
    
    elif current_status == "rejected":
        # إذا كانت مرفوضة، فقط الأدمن يمكنه إعادة اعتمادها
        raise HTTPException(status_code=403, detail="هذه الفاتورة ملغاة - فقط بيت الخبرة يمكنه إعادة اعتمادها")
    
    else:
        raise HTTPException(status_code=400, detail=f"حالة الفاتورة غير معروفة: {current_status}")
    
    await db.invoices.update_one({"id": invoice_id}, {"$set": update_data})
    
    # تحديث سجل المراجعات
    approval_entry = {
        "approved_by": current_user.id,
        "approved_by_name": current_user.full_name or current_user.username,
        "approved_at": datetime.now(timezone.utc).isoformat(),
        "status": update_data.get("status"),
        "notes": notes
    }
    await db.invoices.update_one(
        {"id": invoice_id}, 
        {"$push": {"approval_history": approval_entry}}
    )
    
    return {"message": "تم اعتماد الفاتورة بنجاح", "new_status": update_data.get("status")}


@api_router.put("/invoices/{invoice_id}/reject")
async def reject_invoice(
    invoice_id: str,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    إلغاء اعتماد الفاتورة - متاح فقط لبيت الخبرة (الأدمن)
    الفاتورة تبقى موجودة في النظام ويمكن إعادة اعتمادها
    """
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="الفاتورة غير موجودة")
    
    # فقط الأدمن يمكنه إلغاء الاعتماد
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="فقط بيت الخبرة (الأدمن) يمكنه إلغاء الاعتماد")
    
    update_data = {
        "status": "rejected",
        "rejected_by": current_user.id,
        "rejected_by_name": current_user.full_name or current_user.username,
        "rejected_at": datetime.now(timezone.utc).isoformat(),
        "rejection_notes": notes
    }
    
    await db.invoices.update_one({"id": invoice_id}, {"$set": update_data})
    
    # تسجيل في السجل
    rejection_entry = {
        "action": "rejected",
        "by": current_user.id,
        "by_name": current_user.full_name or current_user.username,
        "at": datetime.now(timezone.utc).isoformat(),
        "notes": notes
    }
    await db.invoices.update_one(
        {"id": invoice_id}, 
        {"$push": {"approval_history": rejection_entry}}
    )
    
    return {"message": "تم إلغاء اعتماد الفاتورة - يمكن للأدمن إعادة اعتمادها"}


@api_router.put("/invoices/{invoice_id}/cancel")
async def cancel_invoice(
    invoice_id: str,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """إلغاء اعتماد فاتورة - إعادتها لحالة pending"""
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="الفاتورة غير موجودة")
    
    user_permissions = current_user.permissions or []
    has_cancel_permission = (
        current_user.role == "admin" or 
        "view_all_invoices" in user_permissions or
        "review_invoices_3" in user_permissions
    )
    
    if not has_cancel_permission:
        raise HTTPException(status_code=403, detail="غير مصرح لك بإلغاء هذه الفاتورة")
    
    current_status = invoice.get("status", "pending")
    if current_status == "pending":
        raise HTTPException(status_code=400, detail="هذه الفاتورة قيد المراجعة بالفعل")
    
    update_data = {
        "status": "pending",
        "cancelled_by": current_user.id,
        "cancelled_by_name": current_user.full_name or current_user.username,
        "cancelled_at": datetime.now(timezone.utc).isoformat(),
        "cancellation_notes": notes,
        # مسح بيانات المراجعة السابقة
        "reviewed_by_manager": None,
        "reviewed_by_manager_name": None,
        "reviewed_by_manager_at": None,
        "manager_notes": None,
        "final_approved_by": None,
        "final_approved_by_name": None,
        "final_approved_at": None,
        "final_notes": None,
        # مسح بيانات الرفض السابقة
        "rejected_by": None,
        "rejected_by_name": None,
        "rejected_at": None,
        "rejection_notes": None
    }
    
    await db.invoices.update_one({"id": invoice_id}, {"$set": update_data})
    return {"message": "تم إلغاء الفاتورة وإعادتها لحالة قيد المراجعة"}


@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(
    invoice_id: str,
    current_user: User = Depends(get_current_user)
):
    """حذف فاتورة - للأدمن ومدراء المشاريع"""
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="الفاتورة غير موجودة")
    
    # الأدمن ومدراء المشاريع (can_create_subusers) يمكنهم الحذف
    if current_user.role != "admin" and not current_user.can_create_subusers:
        # المستوى 3 يمكنه حذف فواتيره فقط إذا كانت pending
        if invoice.get("uploaded_by") != current_user.id or invoice.get("status") != "pending":
            raise HTTPException(status_code=403, detail="غير مصرح لك بحذف هذه الفاتورة")
    
    await db.invoices.update_one({"id": invoice_id}, {"$set": {
        "is_deleted": True,
        "deleted_by": current_user.id,
        "deleted_at": datetime.now(timezone.utc).isoformat()
    }})
    return {"message": "تم حذف الفاتورة"}



# ============= EMPLOYEE REQUESTS (طلبات الموظفين) =============

class EmployeeRequestCreate(BaseModel):
    request_type: str  # نوع الطلب
    reason: str  # السبب
    images: Optional[List[str]] = []  # الصور المرفقة
    project: str  # المشروع
    notes: Optional[str] = None  # ملاحظات إضافية
    amount: Optional[float] = None # مبلغ السلفة
    monthly_deduction: Optional[float] = None # الخصم الشهري

class EmployeeRequestUpdate(BaseModel):
    reason: Optional[str] = None
    images: Optional[List[str]] = None
    notes: Optional[str] = None
    amount: Optional[float] = None
    monthly_deduction: Optional[float] = None


@api_router.get("/request-templates")
async def get_request_templates():
    """جلب قوالب الطلبات المتاحة للتحميل"""
    import os
    
    # النماذج الافتراضية
    default_templates = [
        {"id": "custody", "name": "إقرار استلام عهدة", "file": "اقرار_استلام_عهدة.pdf", "is_default": True},
        {"id": "vacation", "name": "طلب إجازة", "file": "طلب_اجازة.pdf", "is_default": True},
        {"id": "clearance", "name": "نموذج إخلاء طرف", "file": "نموذج_اخلاء_طرف.pdf", "is_default": True},
        {"id": "family_visit", "name": "نموذج إقرار زيارة عائلة", "file": "نموذج_اقرار_زيارة_عائلة.pdf", "is_default": True},
        {"id": "delegation", "name": "نموذج الانتداب", "file": "نموذج_الانتداب.pdf", "is_default": True},
        {"id": "employee_request", "name": "نموذج طلب موظف", "file": "نموذج_طلب_موظف.pdf", "is_default": True},
        {"id": "guarantee", "name": "نموذج كفالة غرامية سلفة", "file": "نموذج_كفالة_غرامية_سلفة.pdf", "is_default": True},
        {"id": "work_start", "name": "نموذج مباشرة عمل", "file": "نموذج_مباشرة_عمل.pdf", "is_default": True},
    ]
    
    # جلب قائمة النماذج المحذوفة
    deleted_templates = await db.deleted_templates.find({}, {"_id": 0, "template_id": 1}).to_list(100)
    deleted_ids = [d["template_id"] for d in deleted_templates]
    
    # فلترة النماذج الافتراضية - استبعاد المحذوفة أو التي لا يوجد لها ملف
    active_default_templates = []
    for template in default_templates:
        if template["id"] not in deleted_ids:
            file_path = f"/app/backend/static/templates/{template['file']}"
            if os.path.exists(file_path):
                active_default_templates.append(template)
    
    # جلب النماذج المضافة من قاعدة البيانات
    db_templates = await db.request_templates.find({}, {"_id": 0}).to_list(100)
    
    # دمج النماذج: المضافة أولاً ثم الافتراضية
    all_templates = db_templates + active_default_templates
    
    return all_templates


@api_router.post("/request-templates")
async def add_request_template(
    name: str = Form(...),
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """إضافة نموذج جديد (للأدمن فقط)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك بإضافة نماذج")
    
    # إنشاء ID فريد
    template_id = str(uuid4())
    
    # حفظ الملف
    import os
    os.makedirs("/app/backend/static/templates", exist_ok=True)
    file_path = f"/app/backend/static/templates/{file.filename}"
    
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    # حفظ في قاعدة البيانات
    template_doc = {
        "id": template_id,
        "name": name,
        "file": file.filename,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.request_templates.insert_one(template_doc)
    
    # إرجاع النموذج بدون _id
    return {"message": "تم إضافة النموذج بنجاح", "template": {"id": template_id, "name": name, "file": file.filename}}


@api_router.delete("/request-templates/{template_id}")
async def delete_request_template(
    template_id: str,
    current_user: User = Depends(get_current_user)
):
    """حذف نموذج (للأدمن فقط)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك بحذف النماذج")
    
    import os
    
    # محاولة حذف من قاعدة البيانات أولاً (النماذج المضافة)
    result = await db.request_templates.delete_one({"id": template_id})
    
    if result.deleted_count > 0:
        # حذف الملف المرتبط إن وجد
        template_doc = await db.request_templates.find_one({"id": template_id})
        if template_doc and template_doc.get("file"):
            file_path = f"/app/backend/static/templates/{template_doc['file']}"
            if os.path.exists(file_path):
                os.remove(file_path)
        return {"message": "تم حذف النموذج بنجاح"}
    
    # إذا لم يتم العثور عليه، قد يكون نموذج افتراضي
    default_ids = ["custody", "vacation", "clearance", "family_visit", "delegation", "employee_request", "guarantee", "work_start"]
    
    if template_id in default_ids:
        # حذف ملف النموذج الافتراضي
        default_files = {
            "custody": "اقرار_استلام_عهدة.pdf",
            "vacation": "طلب_اجازة.pdf",
            "clearance": "نموذج_اخلاء_طرف.pdf",
            "family_visit": "نموذج_اقرار_زيارة_عائلة.pdf",
            "delegation": "نموذج_الانتداب.pdf",
            "employee_request": "نموذج_طلب_موظف.pdf",
            "guarantee": "نموذج_كفالة_غرامية_سلفة.pdf",
            "work_start": "نموذج_مباشرة_عمل.pdf",
        }
        
        # حذف الملف
        if template_id in default_files:
            file_path = f"/app/backend/static/templates/{default_files[template_id]}"
            if os.path.exists(file_path):
                os.remove(file_path)
        
        # تسجيل النموذج كمحذوف في قاعدة البيانات
        await db.deleted_templates.update_one(
            {"template_id": template_id},
            {"$set": {"template_id": template_id, "deleted_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        
        return {"message": "تم حذف النموذج بنجاح"}
    
    raise HTTPException(status_code=404, detail="النموذج غير موجود")


@api_router.get("/request-templates/{template_id}/download")
async def download_template(template_id: str):
    """تحميل قالب طلب"""
    from fastapi.responses import FileResponse
    import os
    
    # النماذج الافتراضية
    default_templates = {
        "custody": "اقرار_استلام_عهدة.pdf",
        "vacation": "طلب_اجازة.pdf",
        "clearance": "نموذج_اخلاء_طرف.pdf",
        "family_visit": "نموذج_اقرار_زيارة_عائلة.pdf",
        "delegation": "نموذج_الانتداب.pdf",
        "employee_request": "نموذج_طلب_موظف.pdf",
        "guarantee": "نموذج_كفالة_غرامية_سلفة.pdf",
        "work_start": "نموذج_مباشرة_عمل.pdf",
    }
    
    # التحقق من النماذج الافتراضية أولاً
    if template_id in default_templates:
        file_path = f"/app/backend/static/templates/{default_templates[template_id]}"
        if os.path.exists(file_path):
            return FileResponse(file_path, filename=default_templates[template_id], media_type="application/pdf")
    
    # البحث في النماذج المضافة من قاعدة البيانات
    db_template = await db.request_templates.find_one({"id": template_id}, {"_id": 0})
    if db_template:
        file_path = f"/app/backend/static/templates/{db_template['file']}"
        if os.path.exists(file_path):
            return FileResponse(file_path, filename=db_template['file'], media_type="application/pdf")
        else:
            raise HTTPException(status_code=404, detail="ملف النموذج غير موجود على الخادم")
    
    raise HTTPException(status_code=404, detail="القالب غير موجود")


@api_router.post("/employee-requests")
async def create_employee_request(
    data: EmployeeRequestCreate,
    current_user: User = Depends(get_current_user)
):
    """إنشاء طلب جديد"""
    # الحصول على المدير المباشر
    manager = None
    if current_user.created_by:
        manager = await db.users.find_one({"id": current_user.created_by}, {"_id": 0})
    
    # معالجة الصور ورفعها لـ Cloudinary إذا كانت base64
    processed_images = await process_images_for_storage(data.images or [], category="requests")

    request = {
        "id": str(uuid4()),
        "request_type": data.request_type,
        "reason": data.reason,
        "images": processed_images,
        "project": data.project,
        "notes": data.notes,
        "amount": data.amount,
        "monthly_deduction": data.monthly_deduction,
        "uploaded_by": current_user.id,
        "uploaded_by_name": current_user.full_name,
        "manager_id": current_user.created_by,
        "manager_name": manager.get("full_name") if manager else None,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.employee_requests.insert_one(request)
    return {"message": "تم رفع الطلب بنجاح", "id": request["id"]}


@api_router.get("/employee-requests")
async def get_employee_requests(
    status: Optional[str] = None,
    project: Optional[str] = None,
    request_type: Optional[str] = None,
    month: Optional[str] = None,
    date: Optional[str] = None,  # فلتر التاريخ المحدد
    for_review: bool = False,  # للمراجعة فقط
    page: int = 1,
    limit: int = 10,
    current_user: User = Depends(get_current_user)
):
    """
    جلب طلبات الموظفين حسب الصلاحيات مع الترقيم
    صلاحية view_all_employee_requests: تتيح للمستخدم رؤية جميع طلبات الموظفين
    صلاحية employee_requests + مشاريع متعددة: يرى الطلبات للاعتماد النهائي
    for_review=True: يعرض الطلبات الجاهزة للمراجعة حسب المستوى
    """
    query = {"is_deleted": {"$ne": True}}
    user_permissions = current_user.permissions or []
    
    # صلاحية عرض جميع طلبات الموظفين
    can_view_all = "view_all_employee_requests" in user_permissions
    can_review = "review_employee_requests" in user_permissions
    has_employee_requests_perm = "employee_requests" in user_permissions
    has_multiple_projects = len(current_user.projects or []) >= 3  # لديه 3 مشاريع أو أكثر
    
    # المسؤول عن الاعتماد النهائي: لديه صلاحية طلبات الموظفين + مشاريع متعددة
    can_final_approve = can_view_all or (has_employee_requests_perm and has_multiple_projects)
    
    # التحقق من وجود موظفين تحت هذا المستخدم
    sub_users = await db.users.find({"created_by": current_user.id}, {"_id": 0, "id": 1}).to_list(1000)
    sub_user_ids = [u["id"] for u in sub_users]
    has_sub_users = len(sub_user_ids) > 0
    
    if for_review:
        # وضع المراجعة - حسب المستوى
        if current_user.role == "admin":
            # الأدمن: يرى الطلبات المعتمدة من المديرين + السلف التي لم تعتمد بعد
            query["$or"] = [
                {"status": {"$in": ["approved_by_manager", "reviewing"]}},
                {"request_type": {"$in": ["advance_request", "advance"]}, "status": "pending"}
            ]
        elif has_sub_users:
            # مدير المشروع الذي لديه موظفين: يرى طلبات موظفيه قيد الانتظار (باستثناء السلف التي تذهب للأدمن مباشرة)
            query["status"] = "pending"
            query["uploaded_by"] = {"$in": sub_user_ids}
            query["request_type"] = {"$ne": "advance_request"}
        elif can_final_approve:
            # المسؤول عن الاعتماد النهائي (صلاحية طلبات الموظفين + مشاريع متعددة)
            query["status"] = "approved_by_manager"
            # فلترة حسب مشاريعه إذا لم يكن لديه view_all
            if not can_view_all and current_user.projects:
                query["project"] = {"$in": current_user.projects}
        else:
            return {"requests": [], "total_count": 0, "total_pages": 0, "current_page": 1, "limit": limit}
    else:
        # العرض العادي - جميع الطلبات تبقى ظاهرة بغض النظر عن حالتها (مرفوضة، ملغاة، معتمدة)
        if current_user.role == "admin":
            # الأدمن يرى جميع الطلبات (ما عدا pending التي تحتاج مراجعة المدير أولاً)
            if status and status != 'all':
                query["status"] = status
            else:
                # الأدمن يرى كل شيء ما عدا pending (إلا إذا كان نوع الطلب سلفة)
                query["$or"] = [
                    {"status": {"$in": ["approved_by_manager", "approved_by_admin", "approved_final", "rejected", "cancelled", "reviewing"]}},
                    {"request_type": {"$in": ["advance_request", "advance"]}, "status": "pending"}
                ]
            if project:
                query["project"] = project
        elif can_view_all:
            # المستخدم مع صلاحية view_all_employee_requests
            if status and status != 'all':
                query["status"] = status
            else:
                query["status"] = {"$in": ["approved_by_manager", "approved_by_admin", "approved_final", "rejected", "cancelled"]}
            if project:
                query["project"] = project
        elif has_sub_users:
            # المدير الذي لديه موظفين يرى طلباته الشخصية + طلبات موظفيه (باستثناء السلف التي تذهب للأدمن)
            query["$or"] = [
                {"uploaded_by": current_user.id}, # طلباته هو
                {"uploaded_by": {"$in": sub_user_ids}, "request_type": {"$nin": ["advance_request", "advance"]}} # طلبات موظفيه (بدون السلف)
            ]
            # لا نفلتر على الحالة افتراضياً - يرى كل شيء
            if status and status != 'all':
                query["status"] = status
            if project:
                query["project"] = project
        elif can_final_approve:
            # المسؤول عن الاعتماد النهائي (بدون موظفين) يرى طلبات مشاريعه
            if status and status != 'all':
                query["status"] = status
            else:
                query["status"] = {"$in": ["approved_by_manager", "approved_by_admin", "approved_final", "rejected", "cancelled"]}
            if current_user.projects:
                query["project"] = {"$in": current_user.projects}
            if project:
                query["project"] = project
        else:
            # الموظف (المستوى 3) يرى طلباته فقط (كل الحالات بما فيها المرفوضة والملغاة)
            query["uploaded_by"] = current_user.id
            # لا نفلتر على الحالة افتراضياً - يرى كل طلباته
            if status and status != 'all':
                query["status"] = status
    
    if request_type:
        if request_type in ["advance_request", "advance"]:
            query["request_type"] = {"$in": ["advance_request", "advance"]}
        else:
            query["request_type"] = request_type
    
    # فلترة حسب تاريخ محدد (يوم)
    if date:
        # date format: "2025-12-15"
        date_filter = {"created_at": {"$regex": f"^{date}"}}
        if "$or" in query:
            existing_or = query.pop("$or")
            if "$and" not in query:
                query["$and"] = []
            query["$and"].append({"$or": existing_or})
            query["$and"].append(date_filter)
        else:
            query.update(date_filter)
    # فلترة حسب الشهر
    elif month:
        # month format: "2025-12"
        month_filter = {"created_at": {"$regex": f"^{month}"}}
        if "$or" in query:
            existing_or = query.pop("$or")
            if "$and" not in query:
                query["$and"] = []
            query["$and"].append({"$or": existing_or})
            query["$and"].append(month_filter)
        else:
            query.update(month_filter)
    
    # Get total count
    total_count = await db.employee_requests.count_documents(query)
    total_pages = (total_count + limit - 1) // limit if limit > 0 else 1
    
    # Apply pagination
    skip = (page - 1) * limit
    requests = await db.employee_requests.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "requests": requests,
        "total_count": total_count,
        "total_pages": total_pages,
        "current_page": page,
        "limit": limit
    }


@api_router.put("/employee-requests/{request_id}")
async def update_employee_request(
    request_id: str,
    data: EmployeeRequestUpdate,
    current_user: User = Depends(get_current_user)
):
    """تعديل طلب"""
    request = await db.employee_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # فقط صاحب الطلب أو الأدمن
    if request.get("uploaded_by") != current_user.id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل هذا الطلب")
    
    # لا يمكن التعديل بعد الاعتماد
    if request.get("status") in ["approved_by_manager", "approved_by_admin"]:
        raise HTTPException(status_code=400, detail="لا يمكن تعديل الطلب بعد الاعتماد")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if data.reason is not None:
        update_data["reason"] = data.reason
    if data.images is not None:
        update_data["images"] = await process_images_for_storage(data.images, category="requests")
    if data.notes is not None:
        update_data["notes"] = data.notes
    if data.amount is not None:
        update_data["amount"] = data.amount
    if data.monthly_deduction is not None:
        update_data["monthly_deduction"] = data.monthly_deduction
    
    await db.employee_requests.update_one({"id": request_id}, {"$set": update_data})
    return {"message": "تم تعديل الطلب بنجاح"}


@api_router.put("/employee-requests/{request_id}/approve")
async def approve_employee_request(
    request_id: str,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    نظام اعتماد طلبات الموظفين المتدرج:
    1. الموظف (المستوى 3) يرفع الطلب (status: pending)
    2. مدير المشروع (المستوى 2) يعتمده اعتماد أولي (status: approved_by_manager)
    3. بيت الخبرة (الأدمن) فقط يعتمده نهائياً (status: approved_by_admin) أو يلغي الاعتماد
    
    - المستوى 2 لا يمكنه الاعتماد النهائي أبداً
    - الأدمن يمكنه الاعتماد النهائي من أي مرحلة
    - الأدمن يمكنه إلغاء الاعتماد وإعادته
    """
    request = await db.employee_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    current_status = request.get("status", "pending")
    update_data = {}
    user_permissions = current_user.permissions or []
    request_project = request.get("project")
    
    # التحقق من وجود موظفين تحت هذا المستخدم
    sub_users = await db.users.find({"created_by": current_user.id}, {"_id": 0, "id": 1}).to_list(1000)
    sub_user_ids = [u["id"] for u in sub_users]
    has_sub_users = len(sub_user_ids) > 0
    
    # التحقق من تفويض الاعتماد النهائي (نيابة عن بيت الخبرة) لطلبات الموظفين
    has_final_review_delegation = (
        "review_employee_requests" in user_permissions or
        "view_all_employee_requests" in user_permissions or
        has_project_permission(current_user, request_project, "review_employee_requests") or
        has_project_permission(current_user, request_project, "view_all_employee_requests")
    )
    
    # فصل المهام: من اعتمد أولياً لا يعتمد نهائياً نفس الطلب
    if current_status == "approved_by_manager" and request.get("reviewed_by_manager") == current_user.id:
        has_final_review_delegation = False
    
    if current_user.role == "admin":
        # الأدمن (بيت الخبرة): يعتمد الطلبات نهائياً من أي حالة (بما فيها جاري المراجعة)
        update_data = {
            "status": "approved_by_admin",
            "approved_by": current_user.id,
            "approved_by_name": current_user.full_name or current_user.username,
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "admin_notes": notes
        }
    elif current_status == "approved_by_manager" and has_final_review_delegation:
        # المفوض بالاعتماد النهائي (نيابة عن بيت الخبرة)
        update_data = {
            "status": "approved_by_admin",
            "approved_by": current_user.id,
            "approved_by_name": current_user.full_name or current_user.username,
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "admin_notes": notes,
            "final_approved_on_behalf": True,
            "delegated_by_admin": True
        }
    elif current_status == "pending" and has_sub_users:
        # مدير المشروع (المستوى 2): يراجع طلبات موظفيه قيد الانتظار أو طلبه هو نفسه
        uploader_id = request.get("uploaded_by")
        is_own_request = uploader_id == current_user.id
        if not is_own_request and uploader_id not in sub_user_ids:
            raise HTTPException(status_code=403, detail="هذا الطلب ليس من موظفيك")
        
        # منع مدير المشروع من اعتماد السلف - تذهب للأدمن مباشرة
        if request.get("request_type") in ["advance_request", "advance"] and not is_own_request:
            raise HTTPException(status_code=403, detail="طلبات السلف تعتمد مباشرة من بيت الخبرة")

        update_data = {
            "status": "approved_by_manager",
            "reviewed_by_manager": current_user.id,
            "reviewed_by_manager_name": current_user.full_name or current_user.username,
            "reviewed_by_manager_at": datetime.now(timezone.utc).isoformat(),
            "manager_notes": notes,
            "self_approved": is_own_request
        }
    elif current_status == "approved_by_manager":
        # الطلب معتمد من المدير - فقط الأدمن أو المفوض يعتمده نهائياً
        raise HTTPException(status_code=403, detail="الاعتماد النهائي متاح فقط لبيت الخبرة أو من فوضه")
    elif current_status == "pending":
        raise HTTPException(status_code=403, detail="هذا الطلب ليس من موظفيك")
    elif current_status == "approved_by_admin":
        raise HTTPException(status_code=400, detail="هذا الطلب معتمد نهائياً بالفعل")
    elif current_status == "rejected":
        raise HTTPException(status_code=403, detail="هذا الطلب ملغي - فقط بيت الخبرة يمكنه إعادة اعتماده")
    else:
        raise HTTPException(status_code=403, detail="غير مصرح لك باعتماد الطلبات")
    
    await db.employee_requests.update_one({"id": request_id}, {"$set": update_data})
    
    # تسجيل في السجل
    approval_entry = {
        "action": "approved",
        "status": update_data.get("status"),
        "by": current_user.id,
        "by_name": current_user.full_name or current_user.username,
        "at": datetime.now(timezone.utc).isoformat(),
        "notes": notes
    }
    await db.employee_requests.update_one(
        {"id": request_id}, 
        {"$push": {"approval_history": approval_entry}}
    )
    
    return {"message": "تم اعتماد الطلب بنجاح", "new_status": update_data.get("status")}


@api_router.put("/employee-requests/{request_id}/reject")
async def reject_employee_request(
    request_id: str,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    إلغاء اعتماد الطلب - متاح فقط لبيت الخبرة (الأدمن)
    الطلب يبقى موجوداً في النظام ويمكن إعادة اعتماده
    """
    request = await db.employee_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # فقط الأدمن يمكنه إلغاء الاعتماد
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="فقط بيت الخبرة (الأدمن) يمكنه إلغاء الاعتماد")
    
    update_data = {
        "status": "rejected",
        "rejected_by": current_user.id,
        "rejected_by_name": current_user.full_name or current_user.username,
        "rejected_at": datetime.now(timezone.utc).isoformat(),
        "rejection_notes": notes
    }
    
    await db.employee_requests.update_one({"id": request_id}, {"$set": update_data})
    
    # تسجيل في السجل
    rejection_entry = {
        "action": "rejected",
        "by": current_user.id,
        "by_name": current_user.full_name or current_user.username,
        "at": datetime.now(timezone.utc).isoformat(),
        "notes": notes
    }
    await db.employee_requests.update_one(
        {"id": request_id}, 
        {"$push": {"approval_history": rejection_entry}}
    )
    
    return {"message": "تم إلغاء اعتماد الطلب - يمكن للأدمن إعادة اعتماده"}
    

@api_router.put("/employee-requests/{request_id}/reviewing")
async def mark_employee_request_reviewing(
    request_id: str,
    current_user: User = Depends(get_current_user)
):
    """تحويل الطلب إلى حالة جاري المراجعة - متاح فقط للأدمن"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="فقط بيت الخبرة يمكنه وضع الطلب قيد المراجعة")
        
    request = await db.employee_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
        
    await db.employee_requests.update_one(
        {"id": request_id}, 
        {"$set": {
            "status": "reviewing",
            "reviewing_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "الطلب الآن قيد المراجعة"}


@api_router.put("/employee-requests/{request_id}/upload-signed")
async def upload_signed_document(
    request_id: str,
    data: dict, # Expecting {"signed_document": "..."}
    current_user: User = Depends(get_current_user)
):
    """رفع المستند الموقع - متاح لصاحب الطلب أو الأدمن"""
    request = await db.employee_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # التحقق من الصلاحية: صاحب الطلب أو الأدمن
    if request.get("uploaded_by") != current_user.id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك برفع المستند لهذا الطلب")

    signed_doc = data.get("signed_document")
    if not signed_doc:
        raise HTTPException(status_code=400, detail="المستند مطلوب")

    # إذا كان المستند base64، قم برفعه لـ Cloudinary
    if signed_doc.startswith("data:"):
        processed = await process_images_for_storage([signed_doc], category="signed")
        signed_doc = processed[0]

    await db.employee_requests.update_one(
        {"id": request_id},
        {"$set": {
            "signed_document": signed_doc,
            "signed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "تم رفع المستند الموقع بنجاح"}


@api_router.put("/employee-requests/{request_id}/verify")
async def verify_signed_request(
    request_id: str,
    current_user: User = Depends(get_current_user)
):
    """تأكيد استلام المستند الموقع وإغلاق الطلب نهائياً - للأدمن فقط"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك بتأكيد الطلب")
        
    await db.employee_requests.update_one(
        {"id": request_id},
        {"$set": {
            "admin_verified_at": datetime.now(timezone.utc).isoformat(),
            "admin_verified_by": current_user.full_name
        }}
    )
    return {"message": "تم تأكيد الاستلام بنجاح"}


@api_router.put("/employee-requests/{request_id}/cancel")
async def cancel_employee_request(
    request_id: str,
    notes: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """إلغاء اعتماد طلب - إعادته لحالة pending"""
    request = await db.employee_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    user_permissions = current_user.permissions or []
    has_cancel_permission = (
        current_user.role == "admin" or 
        "view_all_employee_requests" in user_permissions
    )
    
    if not has_cancel_permission:
        raise HTTPException(status_code=403, detail="غير مصرح لك بإلغاء هذا الطلب")
    
    current_status = request.get("status", "pending")
    if current_status == "pending":
        raise HTTPException(status_code=400, detail="هذا الطلب قيد المراجعة بالفعل")
    
    update_data = {
        "status": "pending",
        "cancelled_by": current_user.id,
        "cancelled_by_name": current_user.full_name or current_user.username,
        "cancelled_at": datetime.now(timezone.utc).isoformat(),
        "cancellation_notes": notes,
        # مسح بيانات المراجعة السابقة
        "reviewed_by_manager": None,
        "reviewed_by_manager_name": None,
        "reviewed_by_manager_at": None,
        "manager_notes": None,
        "approved_by": None,
        "approved_by_name": None,
        "approved_at": None,
        "admin_notes": None
    }
    
    await db.employee_requests.update_one({"id": request_id}, {"$set": update_data})
    return {"message": "تم إلغاء الطلب وإعادته لحالة قيد المراجعة"}


@api_router.delete("/employee-requests/{request_id}")
async def delete_employee_request(
    request_id: str,
    current_user: User = Depends(get_current_user)
):
    """حذف طلب"""
    request = await db.employee_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # الأدمن والمدراء يمكنهم الحذف
    if current_user.role != "admin" and not current_user.can_create_subusers:
        if request.get("uploaded_by") != current_user.id or request.get("status") != "pending":
            raise HTTPException(status_code=403, detail="غير مصرح لك بحذف هذا الطلب")
    
    await db.employee_requests.update_one({"id": request_id}, {"$set": {
        "is_deleted": True,
        "deleted_by": current_user.id,
        "deleted_at": datetime.now(timezone.utc).isoformat()
    }})
    return {"message": "تم حذف الطلب"}


@api_router.get("/notifications/pending-count")
async def get_pending_notifications_count(current_user: User = Depends(get_current_user)):
    """
    جلب عدد الفواتير والطلبات والمستخلصات المعلقة التي تحتاج مراجعة
    للمستوى 2 (المدراء) والأدمن
    المستخلصات الواردة = التي في حالة pending وليس لها قيمة فعلية (لم يتم تسجيلها بعد من بيت الخبرة)
    
    المنطق:
    - المدير يرى فقط الطلبات pending التي تحتاج مراجعته (لم يعتمدها بعد)
    - بعد اعتماد المدير، الإشعار يختفي من عنده ويظهر عند الأدمن
    - الأدمن يرى فقط الطلبات approved_by_manager التي تحتاج موافقته النهائية
    """
    try:
        pending_invoices = 0
        pending_requests = 0
        pending_extracts = 0
        signed_requests = 0
        
        user_permissions = current_user.permissions or []
        has_all_projects = "all_projects" in user_permissions
        
        # ما هي المشاريع التي يملك فيها المستخدم صلاحية view_all_invoices أو review_invoices_3 (للاعتماد النهائي)
        invoice_delegation_projects = set(get_projects_with_permission(current_user, "view_all_invoices")) | \
                                       set(get_projects_with_permission(current_user, "review_invoices_3"))
        request_delegation_projects = set(get_projects_with_permission(current_user, "view_all_employee_requests")) | \
                                       set(get_projects_with_permission(current_user, "review_employee_requests"))
        
        # حساب الموظفين التابعين
        sub_users = await db.users.find({"created_by": current_user.id}, {"_id": 0, "id": 1}).to_list(1000)
        sub_user_ids = [u["id"] for u in sub_users]
        has_sub_users = len(sub_user_ids) > 0
        
        if current_user.role == "admin" or has_all_projects:
            # الأدمن: يرى الفواتير والطلبات المعتمدة من المدير (تحتاج اعتماد نهائي)
            pending_invoices = await db.invoices.count_documents({
                "status": "approved_by_manager",
                "is_deleted": {"$ne": True}
            })
            pending_requests = await db.employee_requests.count_documents({
                "$or": [
                    {"status": "approved_by_manager"},
                    {"request_type": {"$in": ["advance_request", "advance"]}, "status": "pending"}
                ],
                "is_deleted": {"$ne": True}
            })
            pending_extracts = await db.extracts.count_documents({
                "status": "pending",
                "$or": [
                    {"actual_value": {"$exists": False}},
                    {"actual_value": None},
                    {"actual_value": 0},
                    {"actual_value": ""}
                ],
                "is_deleted": {"$ne": True}
            })
            # طلبات الموظفين التي تم توقيعها ولم يتم تأكيد استلامها نهائياً
            signed_requests = await db.employee_requests.count_documents({
                "signed_document": {"$exists": True, "$ne": None},
                "admin_verified_at": {"$exists": False},
                "is_deleted": {"$ne": True}
            })
        else:
            # غير الأدمن: اجمع إشعارات حسب دور المستخدم
            inv_count = 0
            req_count = 0
            
            # 1. المدير (لديه موظفين): الفواتير/الطلبات المعلقة من موظفيه (اعتماد أولي)
            if has_sub_users:
                inv_count += await db.invoices.count_documents({
                    "status": "pending",
                    "uploaded_by": {"$in": sub_user_ids + [current_user.id]},
                    "is_deleted": {"$ne": True}
                })
                # طلبات موظفيه (باستثناء السلف) + طلباته الشخصية (بدون السلف)
                req_count += await db.employee_requests.count_documents({
                    "$or": [
                        {
                            "uploaded_by": current_user.id, 
                            "status": "pending",
                            "request_type": {"$nin": ["advance_request", "advance"]}
                        },
                        {
                            "uploaded_by": {"$in": sub_user_ids}, 
                            "status": "pending", 
                            "request_type": {"$nin": ["advance_request", "advance"]}
                        }
                    ],
                    "is_deleted": {"$ne": True}
                })
            else:
                # ليس لديه موظفين - يرى فقط الطلبات التي هو مديرها (حالة نادرة)
                inv_count += await db.invoices.count_documents({
                    "status": "pending",
                    "manager_id": current_user.id,
                    "is_deleted": {"$ne": True}
                })
                req_count += await db.employee_requests.count_documents({
                    "status": "pending",
                    "manager_id": current_user.id,
                    "request_type": {"$nin": ["advance_request", "advance"]},
                    "is_deleted": {"$ne": True}
                })
            
            # 2. المفوض بالاعتماد النهائي (نيابة عن بيت الخبرة): يرى approved_by_manager
            has_invoice_final_delegation = (
                "review_invoices_3" in user_permissions or
                "view_all_invoices" in user_permissions or
                len(invoice_delegation_projects) > 0
            )
            has_request_final_delegation = (
                "review_employee_requests" in user_permissions or
                "view_all_employee_requests" in user_permissions or
                len(request_delegation_projects) > 0
            )
            
            if has_invoice_final_delegation:
                # فواتير معتمدة من المدير، ليست معتمدة من نفس المستخدم (فصل المهام)
                inv_query = {
                    "status": "approved_by_manager",
                    "is_deleted": {"$ne": True},
                    "reviewed_by_manager": {"$ne": current_user.id}
                }
                # إذا كان التفويض لمشاريع محددة فقط، فلتر بها
                if ("review_invoices_3" not in user_permissions and 
                    "view_all_invoices" not in user_permissions and
                    invoice_delegation_projects):
                    inv_query["project"] = {"$in": list(invoice_delegation_projects)}
                inv_count += await db.invoices.count_documents(inv_query)
            
            if has_request_final_delegation:
                req_query = {
                    "status": "approved_by_manager",
                    "is_deleted": {"$ne": True},
                    "reviewed_by_manager": {"$ne": current_user.id}
                }
                if ("review_employee_requests" not in user_permissions and
                    "view_all_employee_requests" not in user_permissions and
                    request_delegation_projects):
                    req_query["project"] = {"$in": list(request_delegation_projects)}
                req_count += await db.employee_requests.count_documents(req_query)
            
            pending_invoices = inv_count
            pending_requests = req_count
            pending_extracts = 0
        
        return {
            "pending_invoices": pending_invoices,
            "pending_requests": pending_requests,
            "pending_extracts": pending_extracts,
            "signed_requests": signed_requests,
            "total": pending_invoices + pending_requests + pending_extracts + signed_requests
        }
    except Exception as e:
        logger.error(f"Error getting notifications count: {str(e)}")
        return {"pending_invoices": 0, "pending_requests": 0, "pending_extracts": 0, "total": 0}



# ============= EXPORT =============


@api_router.get("/dashboard-stats")
async def get_dashboard_stats(
    current_user: User = Depends(get_current_user)
):
    """الحصول على إحصائيات لوحة التحكم"""
    query = {"is_deleted": {"$ne": True}}
    
    # قائمة المستخدمين الذين يرون بلاغاتهم فقط تلقائياً
    restricted_users = ["Mohamed Esmat", "ElShazly"]
    
    # التصفية الهرمية الشاملة
    hierarchy_filter = await get_hierarchy_filter(current_user)
    
    # دمج الفلاتر بمرونة
    if current_user.role != "admin":
        user_projects_filter = get_flexible_in_query(current_user.projects, "project")
        if user_projects_filter:
            query = {"$and": [
                {"is_deleted": {"$ne": True}},
                hierarchy_filter,
                user_projects_filter
            ]}
        else:
            query.update(hierarchy_filter)
    else:
        # للأدمن، لا يوجد قيود إلا الـ is_deleted
        query = {"is_deleted": {"$ne": True}}
    
    # إحصائيات عامة
    total_reports = await db.reports.count_documents(query)
    
    # إحصائيات حسب النوع والحالة
    fixed_asphalt = await db.reports.count_documents({
        **query,
        "report_type": "أسفلت",
        "status": "تم الإصلاح"
    })
    
    fixed_dirt = await db.reports.count_documents({
        **query,
        "report_type": "ترابي",
        "status": "تم الإصلاح"
    })
    
    fixed_tiles = await db.reports.count_documents({
        **query,
        "report_type": "بلاط",
        "status": "تم الإصلاح"
    })
    
    return {
        "total_reports": total_reports,
        "fixed_asphalt": fixed_asphalt,
        "fixed_dirt": fixed_dirt,
        "fixed_tiles": fixed_tiles
    }


@api_router.post("/reports/export-selected/excel")
async def export_selected_reports_excel(
    request: dict,
    current_user: User = Depends(get_current_user)
):
    """تصدير بلاغات محددة إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    # الحصول على report_ids من request body
    report_ids = request.get('report_ids', [])
    project_filter = request.get('project')  # إضافة فلتر المشروع من Frontend
    
    if not report_ids:
        raise HTTPException(status_code=400, detail="No report IDs provided")
    
    # جلب البلاغات المحددة
    query = {
        "id": {"$in": report_ids},
        "is_deleted": {"$ne": True}
    }
    
    # تطبيق الصلاحيات
    if current_user.role != "admin":
        if current_user.projects:
            query["project"] = {"$in": current_user.projects}
        if current_user.governorates:
            query["governorate"] = {"$in": current_user.governorates}
    
    # تطبيق فلتر المشروع إذا تم إرساله (حتى للـ admin)
    if project_filter:
        query["project"] = project_filter
    
    # ⚡ استثناء حقل الصور الكامل لتسريع الاستعلام - نجلب فقط عدد الصور
    projection = {
        "_id": 0,
        "images": 0  # استثناء الصور الكاملة
    }
    reports = await db.reports.find(query, projection).sort("created_at", -1).to_list(5000)
    
    # جلب عدد الصور لكل بلاغ بشكل منفصل وسريع
    selected_report_ids = [r.get('id') for r in reports if r.get('id')]
    images_counts = {}
    if selected_report_ids:
        pipeline = [
            {"$match": {"id": {"$in": selected_report_ids}}},
            {"$project": {"id": 1, "images_count": {"$size": {"$ifNull": ["$images", []]}}}}
        ]
        counts_cursor = db.reports.aggregate(pipeline)
        async for doc in counts_cursor:
            images_counts[doc.get('id')] = doc.get('images_count', 0)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "البلاغات المحددة"
    
    headers = [
        "رقم", "المحافظة", "المشروع", "رقم البلاغ", "رقم الرخصة",
        "حالة المعالجة", "الحالة", "نوع البلاغ", "العمق (سم)", "القطر (ملم)",
        "اسم المقاول", "خط العرض", "خط الطول", "رخصة أسفلت",
        "الملاحظات", "تاريخ الاستلام", "تاريخ المباشرة", "تاريخ الإغلاق", "عدد الصور", "مراقب الاستشاري"
    ]
    
    # تنسيق الرأس
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    
    column_widths = [6, 15, 35, 15, 15, 18, 18, 12, 12, 12, 25, 12, 12, 12, 30, 18, 18, 10, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    
    # البيانات
    for row_idx, report in enumerate(reports, 2):
        created_at = report.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        elif not isinstance(created_at, datetime):
            created_at = None
        
        start_date = report.get('start_date')
        if isinstance(start_date, str):
            try:
                start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            except:
                start_date = None
        elif not isinstance(start_date, datetime):
            start_date = None
            
        closed_at = report.get('closed_at')
        if isinstance(closed_at, str):
            try:
                closed_at = datetime.fromisoformat(closed_at.replace('Z', '+00:00'))
            except:
                closed_at = None
        elif not isinstance(closed_at, datetime):
            closed_at = None
        
        images_count = images_counts.get(report.get('id'), 0)
        
        _rn = report.get('report_number', '')
        if _rn:
            _rn = str(_rn).replace('CCP-', '').replace('CCB-', '')
            proj_name = report.get('project', '')
            if 'الغربي' in proj_name:
                _rn = f"CCB-{_rn}"
            else:
                _rn = f"CCP-{_rn}"
        row_data = [
            row_idx - 1,
            report.get('governorate', ''),
            report.get('project', ''),
            _rn,
            report.get('license_number', ''),
            "مغلقة بواسطة الاستشاري" if report.get('wfm_closed') else "قيد المعالجة",
            report.get('status', ''),
            report.get('report_type', ''),
            report.get('depth_meters', ''),
            report.get('diameter_mm', ''),
            report.get('contractor', ''),
            report.get('latitude', ''),
            report.get('longitude', ''),
            'نعم' if report.get('asphalt_license_issued') else 'لا',
            report.get('notes', ''),
            created_at.strftime('%Y-%m-%d %H:%M') if created_at else '',
            start_date.strftime('%Y-%m-%d') if start_date else '',
            closed_at.strftime('%Y-%m-%d %H:%M') if closed_at else '',
            images_count,
            report.get('created_by_name', '')
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"selected_reports_{len(reports)}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.post("/reports/export-selected/pdf")
async def export_selected_reports_pdf(
    request: dict,
    current_user: User = Depends(get_current_user)
):
    lang = request.get("lang", "ar")
    """تصدير بلاغات محددة إلى PDF بدعم كامل للعربية"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.units import mm
    from io import BytesIO
    from arabic_reshaper import reshape
    from bidi.algorithm import get_display
    
    # الحصول على report_ids من request body
    report_ids = request.get('report_ids', [])
    project_filter = request.get('project')  # إضافة فلتر المشروع من Frontend
    
    if not report_ids:
        raise HTTPException(status_code=400, detail="No report IDs provided")
    
    # جلب البلاغات المحددة
    query = {
        "id": {"$in": report_ids},
        "is_deleted": {"$ne": True}
    }
    
    # تطبيق الصلاحيات
    if current_user.role != "admin":
        if current_user.projects:
            query["project"] = {"$in": current_user.projects}
        if current_user.governorates:
            query["governorate"] = {"$in": current_user.governorates}
    
    # تطبيق فلتر المشروع إذا تم إرساله (حتى للـ admin)
    if project_filter:
        query["project"] = project_filter
    
    # جلب البلاغات بدون حد (None) لدعم أي عدد من البلاغات
    # ⚡ استثناء حقل الصور الكامل لتسريع الاستعلام - نجلب فقط عدد الصور
    projection = {
        "_id": 0,
        "images": 0  # استثناء الصور الكاملة
    }
    reports = await db.reports.find(query, projection).sort("created_at", -1).to_list(None)
    
    # جلب عدد الصور لكل بلاغ بشكل منفصل وسريع
    report_ids = [r.get('id') for r in reports if r.get('id')]
    images_counts = {}
    if report_ids:
        pipeline = [
            {"$match": {"id": {"$in": report_ids}}},
            {"$project": {"id": 1, "images_count": {"$size": {"$ifNull": ["$images", []]}}}}
        ]
        counts_cursor = db.reports.aggregate(pipeline)
        async for doc in counts_cursor:
            images_counts[doc.get('id')] = doc.get('images_count', 0)
    
    # دالة لاختصار اسم المشروع
    def shorten_project_name(project_name):
        if not project_name:
            return ''
        if 'الغربية' in project_name:
            return 'مشروع المحافظات الغربية'
        elif 'الشمالية' in project_name:
            return 'مشروع المحافظات الشمالية'
        elif 'الجنوبية' in project_name:
            return 'مشروع المحافظات الجنوبية'
        return project_name

    def t(text):
        if lang != 'en': return text
        trans = {
            "رقم": "No.", "المحافظة": "Governorate", "المشروع": "Project", 
            "رقم البلاغ": "Report No.", "رقم الرخصة": "License No.", "حالة الرخصة": "License Status",
            "الحالة": "Status", "خط العرض": "Latitude", "خط الطول": "Longitude", "نوع البلاغ": "Type", "العمق": "Depth", "القطر": "Diameter",
            "المقاول": "Contractor", "تاريخ الإنشاء": "Created At",
            "بلاغات محافظة": "Reports of Governorate", "لشهر": "for Month", "تقرير البلاغات": "Reports Report",
            "تنفيذ م-محمود محمد هارون مدير النظام وتحليل البيانات": "Implemented by Eng. Mahmoud Haroon - System & Data Manager",
            "شركة المياه الوطنية": "National Water Company", "مكتب بيت الخبرة للاستشارات الهندسية": "Bayt Al Khibra Eng. Consultancy",
            "قيد المعالجة": "In Progress", "مغلقة بواسطة الاستشاري": "Closed by Consultant",
            "لم يتم إصدار رخصة": "License Not Issued",
            "ايصال": "Esal", "ايصال الرياض": "Esal Riyadh",
            "مشروع كشف التسربات وإصلاحها": "Leak Detection & Repair Project",
            "مشروع المحافظات الغربية -القطاع الأوسط": "Western Governorates Project - Middle Sector",
            "مشروع المحافظات الشمالية": "Northern Governorates Project",
            "مشروع المحافظات الجنوبية": "Southern Governorates Project"
        }
        for k, v in trans.items():
            if text == k: return v
        
        if text.startswith("الاسم:"): return text.replace("الاسم:", "Name:")
        if text.startswith("بلاغات محافظة"): return text.replace("بلاغات محافظة", "Reports of Governorate").replace("لشهر", "for Month")
        
        return text

    def arabic_text(text):
        if not text: return ""
        text = t(text)
        try:
            if any("؀" <= c <= "ۿ" for c in text):
                from arabic_reshaper import reshape
                from bidi.algorithm import get_display
                return get_display(reshape(text))
            return text
        except:
            return text

    
    # إنشاء PDF - تقليل الهوامش لإتاحة مساحة أكبر
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=15, leftMargin=15, topMargin=15, bottomMargin=15)
    elements = []
    
    # تسجيل خط عربي - استخدام NotoSansArabic الذي يدعم العربية بالكامل
    try:
        arabic_font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'NotoSansArabic-Regular.ttf')
        latin_font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'NotoSans-Regular.ttf')
        pdfmetrics.registerFont(TTFont('Arabic', arabic_font_path))
        if os.path.exists(latin_font_path):
            pdfmetrics.registerFont(TTFont('NotoSans', latin_font_path))
        font_name = 'Arabic'
    except Exception as _e:
        # fallback - DejaVu لا يدعم العربية لكنه آخر خيار
        try:
            pdfmetrics.registerFont(TTFont('Arabic', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            font_name = 'Arabic'
        except:
            font_name = 'Helvetica'
    
    # ========== إضافة الشعارات (ديناميكية من إعدادات المنصة) ==========
    # استخراج روابط الشعارات من الإعدادات
    branding = await db.platform_settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    bayt_logo_path = await _resolve_logo_path(branding.get("company_logo_url"), default_filename="bayt-alkhibra-logo.png")
    nwc_logo_path = await _resolve_logo_path(branding.get("partner_logo_url"), default_filename="nwc-logo.png")
    
    logo_data = []
    logo_row = []
    
    # شعار بيت الخبرة (يسار)
    if bayt_logo_path and os.path.exists(bayt_logo_path):
        bayt_logo = Image(bayt_logo_path, width=40*mm, height=20*mm)
        logo_row.append(bayt_logo)
    else:
        logo_row.append('')
    
    # مسافة فارغة في المنتصف
    logo_row.append('')
    
    # شعار شركة المياه الوطنية (يمين) - تصغير الشعارات
    if nwc_logo_path and os.path.exists(nwc_logo_path):
        nwc_logo = Image(nwc_logo_path, width=40*mm, height=20*mm)
        logo_row.append(nwc_logo)
    else:
        logo_row.append('')
    
    logo_data.append(logo_row)
    
    # إنشاء جدول الشعارات
    page_width = landscape(A4)[0] - 40
    logo_table = Table(logo_data, colWidths=[page_width*0.3, page_width*0.4, page_width*0.3])
    logo_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(logo_table)
    elements.append(Spacer(1, 5*mm))  # تقليل المسافة بعد الشعارات
    
    # ========== إنشاء العنوان الديناميكي ==========
    # استخراج المحافظة والشهر من البلاغات
    governorate_name = ""
    month_name = ""
    
    if reports:
        # الحصول على المحافظة (أخذ المحافظة الأكثر تكراراً)
        gov_counts = {}
        for r in reports:
            gov = r.get('governorate', '')
            if gov:
                gov_counts[gov] = gov_counts.get(gov, 0) + 1
        if gov_counts:
            governorate_name = max(gov_counts, key=gov_counts.get)
        
        # الحصول على الشهر من أول بلاغ
        first_report_date = reports[0].get('created_at')
        if first_report_date:
            try:
                if isinstance(first_report_date, str):
                    first_report_date = datetime.fromisoformat(first_report_date.replace('Z', '+00:00'))
                
                # أسماء الأشهر بالعربية
                months_ar = {
                    1: "يناير", 2: "فبراير", 3: "مارس", 4: "أبريل",
                    5: "مايو", 6: "يونيو", 7: "يوليو", 8: "أغسطس",
                    9: "سبتمبر", 10: "أكتوبر", 11: "نوفمبر", 12: "ديسمبر"
                }
                month_name = months_ar.get(first_report_date.month, "")
            except:
                pass
    
    # استخراج اسم المشروع من البلاغات
    project_name_full = ""
    if reports:
        project_counts = {}
        for r in reports:
            proj = r.get('project', '')
            if proj:
                project_counts[proj] = project_counts.get(proj, 0) + 1
        if project_counts:
            project_name_full = max(project_counts, key=project_counts.get)
    
    # تحديد اسم المشروع الكامل للعرض
    def get_full_project_name(project):
        return project if project else ""
    
    # إنشاء العنوان الديناميكي
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    
    title_style = ParagraphStyle(
        'TitleStyle',
        fontName=font_name,
        fontSize=16,  # تصغير حجم العنوان
        alignment=TA_CENTER,
        textColor=colors.HexColor('#366092'),
        spaceBefore=0,
        spaceAfter=8,  # تقليل المسافة بعد العنوان
        leading=20
    )
    
    # بناء العنوان الديناميكي
    if governorate_name and month_name:
        dynamic_title = f"بلاغات محافظة {governorate_name} لشهر {month_name}"
    elif governorate_name:
        dynamic_title = f"بلاغات محافظة {governorate_name}"
    else:
        dynamic_title = "تقرير البلاغات"
    
    title_text = arabic_text(dynamic_title)
    title = Paragraph(title_text, title_style)
    elements.append(title)
    
    # إضافة اسم المشروع الكامل (في سطر منفصل) - تقليل المسافات
    project_style = ParagraphStyle(
        'ProjectStyle',
        fontName=font_name,
        fontSize=10,  # تصغير حجم الخط
        alignment=TA_CENTER,
        textColor=colors.HexColor('#333333'),
        spaceBefore=0,
        spaceAfter=5,  # تقليل المسافة
        leading=14
    )
    
    full_project_display = get_full_project_name(project_name_full)
    if full_project_display:
        project_text = arabic_text(full_project_display)
        project_para = Paragraph(project_text, project_style)
        elements.append(project_para)
    
    # إضافة م / محمود هارون فوق الجدول
    author_style = ParagraphStyle(
        'AuthorStyle',
        fontName=font_name,
        fontSize=10,
        alignment=TA_RIGHT,
        textColor=colors.HexColor('#555555'),
        spaceBefore=2,
        spaceAfter=5,
        leading=14
    )
    author_text = arabic_text("تنفيذ م-محمود محمد هارون مدير النظام وتحليل البيانات")
    author_para = Paragraph(author_text, author_style)
    elements.append(author_para)
    
    # إضافة مسافة صغيرة قبل الجدول
    elements.append(Spacer(1, 3*mm))
    
    # الرأس - 11 عمود (مع العمق والقطر) - معكوسة من اليمين لليسار
    headers = [
        "رقم", "المحافظة", "المشروع", "رقم البلاغ", "رقم الرخصة", 
        "الحالة", "خط العرض", "خط الطول", "نوع البلاغ", "المقاول", "تاريخ الإنشاء"
    ]
    
    # عكس ترتيب الأعمدة لتبدأ من اليمين (RTL)
    headers_reversed = headers[::-1]
    
    # ستايلات خلايا الجدول لضمان التنسيق والتفاف النص والترميز الصحيح
    header_style = ParagraphStyle(
        'HeaderStyle',
        fontName=font_name,
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.whitesmoke,
        leading=10
    )
    
    cell_style = ParagraphStyle(
        'CellStyle',
        fontName=font_name,
        fontSize=8,
        alignment=TA_CENTER,
        leading=10
    )
    
    latin_cell_style = ParagraphStyle(
        'LatinCellStyle',
        fontName='NotoSans' if os.path.exists(latin_font_path) else 'Helvetica',
        fontSize=8,
        alignment=TA_CENTER,
        leading=10
    )
    
    data = [[Paragraph(arabic_text(h), header_style) for h in headers_reversed]]
    
    # البيانات
    for idx, report in enumerate(reports, 1):
        created_at = report.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        date_str = created_at.strftime('%Y-%m-%d') if created_at else ''
        
        # تكوين الإحداثيات
        lat = report.get('latitude', '')
        lng = report.get('longitude', '')
        coordinates = f"{lat}, {lng}" if lat and lng else ''
        
        # إضافة CCB- أو CCP- حسب المشروع
        report_num = report.get('report_number', '')
        if report_num:
            report_num = str(report_num).replace('CCP-', '').replace('CCB-', '')
            proj_name = report.get('project', '')
            if 'الغربي' in proj_name:
                report_num = f"CCB-{report_num}"
            else:
                report_num = f"CCP-{report_num}"
        
        # جلب قيم العمق والقطر
        depth_val = report.get('depth_meters', '')
        diameter_val = report.get('diameter_mm', '')
        depth_str = str(depth_val) if depth_val else '-'
        diameter_str = str(diameter_val) if diameter_val else '-'
        
        # ترتيب الصف من اليمين لليسار مباشرة
        # معالجة رقم الرخصة - إذا كان فارغ نعرض "لم يتم إصدار رخصة"
        license_num = report.get('license_number') or ''
        if not license_num or str(license_num).strip() == '':
            license_num = 'لم يتم إصدار رخصة'
        
        row = [
            Paragraph(date_str, latin_cell_style),  # تاريخ الإنشاء - يمين
            Paragraph(arabic_text(report.get('contractor') or ''), cell_style),  # المقاول
            Paragraph(arabic_text(report.get('report_type') or ''), cell_style),  # نوع البلاغ
            Paragraph(str(lng) if lng else '-', latin_cell_style),  # خط الطول
            Paragraph(str(lat) if lat else '-', latin_cell_style),  # خط العرض
            Paragraph(arabic_text(report.get('status') or ''), cell_style),  # الحالة
            Paragraph(arabic_text(license_num), cell_style),  # رقم الرخصة
            Paragraph(report_num, latin_cell_style),  # رقم البلاغ مع CCP-
            Paragraph(arabic_text(shorten_project_name(report.get('project') or '')), cell_style),  # المشروع
            Paragraph(arabic_text(report.get('governorate') or ''), cell_style),  # المحافظة
            Paragraph(str(idx), latin_cell_style)  # رقم - يسار
        ]
        data.append(row)
    
    # إنشاء الجدول - 12 عمود متوازن مع توسيع خانة رقم البلاغ والحالة
    # العروض: تاريخ(50), مقاول(70), قطر(35), عمق(35), نوع(50), حالة(125), حالة رخصة(85), رقم رخصة(50), رقم بلاغ(110), مشروع(120), محافظة(35), رقم(20)
    col_widths = [50, 70, 50, 55, 55, 125, 95, 110, 140, 35, 20]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    table.setStyle(TableStyle([
        # تنسيق الرأس
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        
        # تنسيق الجسم
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # ========== استخراج اسم المراقب من البلاغات ==========
    supervisor_name = ""
    if reports:
        # جلب أسماء المستخدمين الذين أنشأوا البلاغات
        created_by_ids = {}
        for r in reports:
            # أولاً نحاول الحصول على الاسم مباشرة
            creator_name = r.get('created_by_name', '')
            creator_id = r.get('created_by', '')
            
            if creator_name and creator_name != creator_id:
                # الاسم موجود ومختلف عن الـ ID
                created_by_ids[creator_name] = created_by_ids.get(creator_name, 0) + 1
            elif creator_id:
                # نحتاج لجلب الاسم من قاعدة البيانات
                created_by_ids[creator_id] = created_by_ids.get(creator_id, 0) + 1
        
        # أخذ الـ ID/الاسم الأكثر تكراراً
        if created_by_ids:
            most_common = max(created_by_ids, key=created_by_ids.get)
            
            # إذا كان UUID، نجلب الاسم من قاعدة البيانات
            if len(most_common) == 36 and '-' in most_common:  # UUID format
                user_doc = await db.users.find_one({"id": most_common}, {"_id": 0, "full_name": 1, "username": 1})
                if user_doc:
                    supervisor_name = user_doc.get('full_name') or user_doc.get('username', '')
            else:
                supervisor_name = most_common
    
    # ========== إضافة قسم التوقيعات مباشرة تحت الجدول ==========
    # بيانات التوقيع مع اسم المراقب
    platform_settings = await db.platform_settings.find_one({"key": "platform_name"}, {"_id": 0}) or {}
    platform_name = platform_settings.get("value", "مكتب بيت الخبرة للاستشارات الهندسية")
    
    company_name = branding.get("company_name")
    if not company_name:
        company_name = platform_name
        
    partner_company = branding.get("partner_company_name")
    if not partner_company:
        partner_company = "شركة المياه الوطنية"
        
    sig_data = [
        [
            arabic_text(partner_company),
            "",
            "",
            arabic_text(company_name)
        ],
        [
            arabic_text("الاسم: ........................"),
            "",
            "",
            arabic_text(f"الاسم: {supervisor_name}") if supervisor_name else arabic_text("الاسم: ........................")
        ],
        [
            arabic_text("التوقيع: ........................"),
            "",
            "",
            arabic_text("التوقيع: ........................")
        ]
    ]
    
    sig_table = Table(sig_data, colWidths=[page_width*0.30, page_width*0.20, page_width*0.20, page_width*0.30])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    
    # استخدام KeepTogether لضمان بقاء التوقيع في صفحة واحدة
    from reportlab.platypus import KeepTogether
    elements.append(KeepTogether([Spacer(1, 5*mm), sig_table]))
    
    doc.build(elements)
    
    buffer.seek(0)
    filename = f"selected_reports_{len(reports)}.pdf"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/reports/export-72h/excel")
async def export_72h_reports_excel(
    project: Optional[str] = Query(None),
    governorate: Optional[str] = Query(None),
    category: Optional[str] = Query("reports"),
    base_date: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """تصدير بلاغات آخر 24 ساعة إلى Excel"""
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io
    
    # جلب البيانات باستخدام نفس منطق العرض لضمان تطابق الفلاتر
    reports_data = await get_reports_last_72_hours_list(
        project=project,
        governorate=governorate,
        category=category,
        base_date=base_date,
        current_user=current_user
    )
    reports = reports_data.get("reports", [])
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "بلاغات 24 ساعة"
    
    # إضافة عنوان
    gov_text = governorate if governorate else "جميع المحافظات"
    ws.merge_cells('A1:K1')
    ws['A1'] = f"بلاغات 24 ساعة - {gov_text}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # رؤوس الأعمدة
    headers = ["#", "المحافظة", "المشروع", "رقم البلاغ", "رقم الرخصة", "حالة المعالجة", "الحالة", "نوع البلاغ", "العمق", "القطر", "المقاول", "تاريخ الاستلام", "تاريخ المباشرة"]
    header_row = 3
    
    # تنسيق رؤوس الأعمدة
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # إضافة البيانات
    for idx, report in enumerate(reports, 1):
        row = header_row + idx
        
        created_at = report.get('created_at', '')
        if isinstance(created_at, datetime):
            created_at = created_at.strftime('%d-%m-%Y %H:%M')
        elif isinstance(created_at, str):
            try:
                dt_obj = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                created_at = dt_obj.strftime('%d-%m-%Y %H:%M')
            except:
                pass
                
        start_date = report.get('start_date', '')
        if isinstance(start_date, datetime):
            start_date = start_date.strftime('%d-%m-%Y')
        elif isinstance(start_date, str):
            try:
                dt_obj = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                start_date = dt_obj.strftime('%d-%m-%Y')
            except:
                pass
        
        report_num = report.get('report_number', '')
        if report_num:
            report_num = str(report_num)
            if report_num.startswith('CCP-'):
                report_num = report_num.replace('CCP-', 'CCB-')
            elif not report_num.startswith('CCB-'):
                report_num = f"CCB-{report_num}"
        
        license_num = report.get('license_number', '')
        if not license_num or license_num.strip() == '':
            license_num = 'لم يتم إصدار رخصة'
        
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=report.get('governorate', ''))
        ws.cell(row=row, column=3, value=report.get('project', ''))
        ws.cell(row=row, column=4, value=report_num)
        ws.cell(row=row, column=5, value=license_num)
        ws.cell(row=row, column=6, value="مغلقة بواسطة الاستشاري" if report.get('wfm_closed') else "قيد المعالجة")
        ws.cell(row=row, column=7, value=report.get('status', ''))
        ws.cell(row=row, column=8, value=report.get('report_type', ''))
        ws.cell(row=row, column=9, value=report.get('depth_meters', ''))
        ws.cell(row=row, column=10, value=report.get('diameter_mm', ''))
        ws.cell(row=row, column=11, value=report.get('contractor', ''))
        ws.cell(row=row, column=12, value=created_at)
        ws.cell(row=row, column=13, value=start_date)
    
    # ضبط عرض الأعمدة
    from openpyxl.utils import get_column_letter
    column_widths = [5, 15, 30, 20, 20, 20, 15, 15, 10, 10, 20, 20, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # حفظ الملف
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # استخدام اسم ملف آمن بدون أحرف عربية في الـ header
    from urllib.parse import quote
    safe_gov = quote(governorate) if governorate else "all"
    filename = f"reports_72h_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.get("/reports/export/excel")
async def export_reports_excel(
    search: Optional[str] = Query(None),
    governorate: Optional[str] = Query(None),
    project: Optional[str] = Query(None),
    contractor: Optional[str] = Query(None),
    report_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    license_status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    start_date_from: Optional[str] = Query(None),
    start_date_to: Optional[str] = Query(None),
    created_by: Optional[str] = Query(None),
    lang: Optional[str] = Query("ar"),
    current_user: User = Depends(get_current_user)
):
    query = {"is_deleted": {"$ne": True}}
    
    # التصفية الهرمية: عرض تقارير المستخدم + المستخدمين الفرعيين
    if current_user.role != "admin":
        # تصفية حسب صلاحيات المشاريع إذا كانت محددة
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
        
        # تصفية حسب صلاحيات المحافظات إذا كانت محددة
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
        
        # كل يوزر لا يرى إلا بلاغاته التي أضافها بنفسه
        is_manager = getattr(current_user, "can_create_subusers", False)
        can_view_all = has_project_permission(current_user, project, "view_governorate_data")
        if not is_manager and not can_view_all:
            query["created_by"] = current_user.username
            
        # تطبيق الفلترة الهرمية للمستخدمين غير المديرين (مطابق لمنطق get_reports)
        hierarchy_filter = await get_hierarchy_filter(current_user)
        
        user_governorates = current_user.governorates if hasattr(current_user, 'governorates') and current_user.governorates else []
        has_all_govs = any(g in ["الكل", "جميع المحافظات", "كل المحافظات"] for g in user_governorates)
        
        if has_all_govs:
            query.update(hierarchy_filter)
            if "created_by" in query:
                del query["created_by"] # السماح برؤية البلاغات الهرمية
        else:
            if len(user_governorates) > 0:
                query.update(hierarchy_filter)
                if "created_by" in query:
                    del query["created_by"]
            else:
                query.update(hierarchy_filter)
                if "created_by" in query:
                    del query["created_by"]
    
    # --- DEBUG LOGGING ---
    try:
        with open("d:\\sery17-main\\sery17-main\\query_debug.log", "a", encoding="utf-8") as f:
            f.write(f"USER: {current_user.full_name} | ROLE: {current_user.role} | GOVS: {current_user.governorates} | PROJS: {current_user.projects} | PERMS: {current_user.permissions}\n")
            f.write(f"can_view_all: {can_view_all} | is_manager: {is_manager}\n")
            f.write(f"FINAL QUERY: {query}\n")
            f.write("-" * 50 + "\n")
    except Exception:
        pass
    # ---------------------
    
    if search:
        query["$or"] = [
            {"report_number": {"$regex": search, "$options": "i"}},
            {"license_number": {"$regex": search, "$options": "i"}}
        ]
    
    if governorate:
        query["governorate"] = governorate
    
    if project:
        query["project"] = project
    if contractor:
        query["contractor"] = contractor
    
    if report_type:
        query["report_type"] = report_type
    
    if status:
        query["status"] = status
    
    # فلتر حسب المستخدم (للـ Admin ومستوى 2 ومحمود هارون ومدحت)
    if created_by:
        # البحث بـ username أو user_id
        if "$and" in query:
            # إذا كان هناك $and موجود، نضيف شرط جديد
            query["$and"].append({
                "$or": [
                    {"created_by": created_by},
                    {"created_by": {"$regex": created_by, "$options": "i"}}
                ]
            })
        else:
            # إذا لم يكن هناك $and، نستخدم created_by مباشرة
            query["created_by"] = created_by
    
    # فلترة حسب حالة الرخصة أو الحالة
    if license_status == 'status_fixed':
        query["status"] = "تم الإصلاح"
    elif license_status == 'status_asphalt':
        query["status"] = {"$in": [_ASPHALT_CANONICAL, "بانتظار الأسفلت"]}
    elif license_status == 'status_in_progress':
        query["wfm_closed"] = {"$ne": True}
    elif license_status == 'status_wfm_closed':
        query["wfm_closed"] = True
    elif license_status == 'review_pending':
        query["review_status"] = {"$in": ["بانتظار المراجعة", "قيد المراجعة", None]}
    elif license_status == 'license_issued':
        query["license_number"] = {"$regex": "[0-9]"}
    elif license_status == 'license_not_issued':
        query["$or"] = [
            {"license_number": {"$exists": False}},
            {"license_number": None},
            {"license_number": {"$not": {"$regex": "[0-9]"}}}
        ]
    elif license_status and license_status.startswith('custom_'):
        custom_status_name = license_status[len('custom_'):]
        query["status"] = custom_status_name
    
    # فلترة بتاريخ استلام البلاغ (created_at) - يدعم string و datetime
    if date_from or date_to:
        from datetime import datetime as dt, timedelta
        
        try:
            # offset السعودية +3 ساعات عن UTC
            SAUDI_OFFSET = timedelta(hours=3)
            
            if date_from and date_to:
                date_from_obj = dt.fromisoformat(date_from)
                date_to_obj = dt.fromisoformat(date_to)
                # نهاية اليوم = بداية اليوم التالي (23:59:59 + 1 ثانية)
                next_day = date_to_obj + timedelta(days=1)
                
                # تحويل للـ UTC لمطابقة datetime objects المحفوظة
                from_utc = date_from_obj - SAUDI_OFFSET
                next_utc = next_day - SAUDI_OFFSET
                
                date_filter = {
                    "$or": [
                        # مطابقة string-based created_at
                        {"created_at": {"$gte": f"{date_from}T00:00:00", "$lt": next_day.strftime("%Y-%m-%dT00:00:00")}},
                        # مطابقة datetime-based created_at (UTC)
                        {"created_at": {"$gte": from_utc, "$lt": next_utc}}
                    ]
                }
            elif date_from:
                date_from_obj = dt.fromisoformat(date_from)
                from_utc = date_from_obj - SAUDI_OFFSET
                date_filter = {
                    "$or": [
                        {"created_at": {"$gte": f"{date_from}T00:00:00"}},
                        {"created_at": {"$gte": from_utc}}
                    ]
                }
            elif date_to:
                date_to_obj = dt.fromisoformat(date_to)
                next_day = date_to_obj + timedelta(days=1)
                next_utc = next_day - SAUDI_OFFSET
                date_filter = {
                    "$or": [
                        {"created_at": {"$lt": next_day.strftime("%Y-%m-%dT00:00:00")}},
                        {"created_at": {"$lt": next_utc}}
                    ]
                }
            
            # إضافة فلتر التاريخ للـ query
            if "$and" in query:
                query["$and"].append(date_filter)
            elif "$or" in query:
                query["$and"] = [{"$or": query.pop("$or")}, date_filter]
            else:
                if "$or" in date_filter:
                    query["$and"] = [date_filter]
                else:
                    query.update(date_filter)
        except Exception as e:
            print(f"Date filter error: {e}")

    # فلترة بتاريخ مباشرة البلاغ (start_date)
    if start_date_from or start_date_to:
        from datetime import datetime as dt, timedelta
        try:
            if start_date_from and start_date_to:
                s_from = dt.fromisoformat(start_date_from)
                s_to = dt.fromisoformat(start_date_to)
                s_next = s_to + timedelta(days=1)
                sdate_filter = {
                    "$or": [
                        {"start_date": {"$gte": f"{start_date_from}T00:00:00", "$lt": s_next.strftime("%Y-%m-%dT00:00:00")}},
                        {"start_date": {"$gte": s_from, "$lt": s_next}}
                    ]
                }
            elif start_date_from:
                s_from = dt.fromisoformat(start_date_from)
                sdate_filter = {
                    "$or": [
                        {"start_date": {"$gte": f"{start_date_from}T00:00:00"}},
                        {"start_date": {"$gte": s_from}}
                    ]
                }
            else:
                s_to = dt.fromisoformat(start_date_to)
                s_next = s_to + timedelta(days=1)
                sdate_filter = {
                    "$or": [
                        {"start_date": {"$lt": s_next.strftime("%Y-%m-%dT00:00:00")}},
                        {"start_date": {"$lt": s_next}}
                    ]
                }
            
            if "$and" in query:
                query["$and"].append(sdate_filter)
            elif "$or" in query:
                query["$and"] = [{"$or": query.pop("$or")}, sdate_filter]
            else:
                query["$and"] = [sdate_filter]
        except Exception as e:
            print(f"Start date filter error: {e}")
    
    # جلب البلاغات بدون حد (None) لدعم أي عدد من البلاغات
    # ⚡ استثناء حقل الصور الكامل لتسريع الاستعلام - نجلب فقط عدد الصور
    projection = {
        "_id": 0,
        "images": 0  # استثناء الصور الكاملة
    }
    reports = await db.reports.find(query, projection).sort("created_at", -1).to_list(None)
    
    # جلب عدد الصور لكل بلاغ بشكل منفصل وسريع
    report_ids = [r.get('id') for r in reports if r.get('id')]
    images_counts = {}
    if report_ids:
        pipeline = [
            {"$match": {"id": {"$in": report_ids}}},
            {"$project": {"id": 1, "images_count": {"$size": {"$ifNull": ["$images", []]}}}}
        ]
        counts_cursor = db.reports.aggregate(pipeline)
        async for doc in counts_cursor:
            images_counts[doc.get('id')] = doc.get('images_count', 0)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "تقارير البلاغات"
    
    # ⚡ إضافة عنوان رئيسي ديناميكي حسب الفلتر
    # استخراج الشهر من date_from إذا كان موجوداً
    month_text = ""
    if date_from:
        try:
            date_obj = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            month_num = date_obj.month
            # أسماء الأشهر بالعربية
            months_ar = ["", "يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو", 
                        "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]
            month_text = f" شهر {months_ar[month_num]}"
        except:
            pass
    
    # بناء العنوان
    if governorate and month_text:
        title = f"بلاغات محافظة {governorate} -{month_text}"
    elif governorate:
        title = f"بلاغات محافظة {governorate}"
    elif project and month_text:
        # اختصار اسم المشروع
        short_proj = project
        if 'الغربية' in project:
            short_proj = 'مشروع المحافظات الغربية'
        elif 'الشمالية' in project:
            short_proj = 'مشروع المحافظات الشمالية'
        title = f"بلاغات {short_proj} -{month_text}"
    elif month_text:
        title = f"البلاغات -{month_text}"
    else:
        title = "تقرير البلاغات"
    
    # دمج الخلايا للعنوان
    ws.merge_cells('A1:S1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )
    ws.row_dimensions[1].height = 30
    
    headers = [
        "رقم", "المحافظة", "المشروع", "رقم البلاغ", "رقم الرخصة",
        "حالة المعالجة", "الحالة", "نوع البلاغ", "العمق (سم)", "القطر (ملم)",
        "اسم المقاول", "خط العرض", "خط الطول", "رخصة أسفلت",
        "الملاحظات", "تاريخ الاستلام", "تاريخ المباشرة", "تاريخ الإغلاق", "عدد الصور", "مراقب الاستشاري"
    ]
    
    # تنسيق الرأس (السطر الثاني)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    
    # ضبط عرض الأعمدة
    column_widths = [6, 15, 35, 15, 15, 18, 18, 12, 12, 12, 25, 12, 12, 12, 30, 18, 18, 10, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=2, column=idx).column_letter].width = width
    
    # البيانات (تبدأ من السطر 3)
    for row_idx, report in enumerate(reports, 3):
        created_at = report.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        elif not isinstance(created_at, datetime):
            created_at = None
        
        start_date = report.get('start_date')
        if isinstance(start_date, str):
            try:
                start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            except:
                start_date = None
        elif not isinstance(start_date, datetime):
            start_date = None
            
        closed_at = report.get('closed_at')
        if isinstance(closed_at, str):
            try:
                closed_at = datetime.fromisoformat(closed_at.replace('Z', '+00:00'))
            except:
                closed_at = None
        elif not isinstance(closed_at, datetime):
            closed_at = None
        
        images_count = images_counts.get(report.get('id'), 0)
        _rn = report.get('report_number', '')
        if _rn:
            _rn = str(_rn).replace('CCP-', '').replace('CCB-', '')
            proj_name = report.get('project', '')
            if 'الغربي' in proj_name:
                _rn = f"CCB-{_rn}"
            else:
                _rn = f"CCP-{_rn}"
        
        row_data = [
            row_idx - 2,  # رقم تسلسلي يبدأ من 1 (row_idx يبدأ من 3، فنطرح 2)
            report.get('governorate', ''),
            report.get('project', ''),
            _rn,
            report.get('license_number', ''),
            "مغلقة بواسطة الاستشاري" if report.get('wfm_closed') else "قيد المعالجة",
            report.get('status', ''),
            report.get('report_type', ''),
            report.get('depth_meters', ''),
            report.get('diameter_mm', ''),
            report.get('contractor', ''),
            report.get('latitude', ''),
            report.get('longitude', ''),
            'نعم' if report.get('asphalt_license_issued') else 'لا',
            report.get('notes', ''),  # الملاحظات
            created_at.strftime('%d-%m-%Y %H:%M') if created_at else '',
            start_date.strftime('%d-%m-%Y') if start_date else '',
            closed_at.strftime('%d-%m-%Y %H:%M') if closed_at else '',
            images_count,
            report.get('created_by_name', '')
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # تلوين الصفوف بالتناوب
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reports.xlsx"}
    )


# ============= TEAM MANAGEMENT =============
@api_router.get("/team-members", response_model=List[TeamMemberResponse])
async def get_team_members(
    current_user: User = Depends(get_current_user),
    project: Optional[str] = Query(None, description="تصفية حسب المشروع")
):
    # بناء query حسب دور المستخدم
    query = {}
    
    # Admin (بيت الخبرة) - يرى جميع فرق العمل
    if current_user.role == 'admin':
        # إذا تم اختيار مشروع محدد، فلتر حسبه
        if project:
            query['project'] = project
        # وإلا، يرى الجميع (query فارغ)
    else:
        # باقي المستخدمين - يرون فقط فريق عملهم حسب مشاريعهم
        if current_user.projects and len(current_user.projects) > 0:
            if len(current_user.projects) == 1:
                query['project'] = current_user.projects[0]
            else:
                query['project'] = {'$in': current_user.projects}
        else:
            # إذا لم يكن لديه مشاريع، لا يرى شيء
            return []
    
    members = await db.team_members.find(query, {"_id": 0}).to_list(1000)
    
    for member in members:
        if isinstance(member.get('created_at'), str):
            member['created_at'] = datetime.fromisoformat(member['created_at'])
        if 'created_by' not in member:
            member['created_by'] = None
    
    return [TeamMemberResponse(**member) for member in members]


@api_router.post("/team-members", response_model=TeamMemberResponse)
async def add_team_member(member_data: dict, current_user: User = Depends(get_current_user)):
    if 'profile_picture' in member_data and member_data['profile_picture']:
        picture = member_data['profile_picture']
        if picture.startswith('data:image'):
            header, encoded = picture.split(",", 1)
            image_data = base64.b64decode(encoded)
            member_data['profile_picture'] = _upload_image(image_data, category="profiles", content_type="image/jpeg")
            
    member = TeamMember(**member_data, created_by=current_user.id)
    
    doc = member.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.team_members.insert_one(doc)
    
    return TeamMemberResponse(**member.model_dump())


@api_router.put("/team-members/{member_id}")
async def update_team_member(
    member_id: str,
    member_data: dict,
    current_user: User = Depends(get_current_user)
):
    """تحديث بيانات عضو الفريق"""
    update_data = {}
    
    if 'name' in member_data:
        update_data['name'] = member_data['name']
    if 'phone' in member_data:
        update_data['phone'] = member_data['phone']
    if 'position' in member_data:
        update_data['position'] = member_data['position']
    if 'project' in member_data:
        update_data['project'] = member_data['project']
    if 'email' in member_data:
        update_data['email'] = member_data['email']
    if 'profile_picture' in member_data:
        picture = member_data['profile_picture']
        if picture and picture.startswith('data:image'):
            header, encoded = picture.split(",", 1)
            image_data = base64.b64decode(encoded)
            update_data['profile_picture'] = _upload_image(image_data, category="profiles", content_type="image/jpeg")
        else:
            update_data['profile_picture'] = picture
    
    if not update_data:
        raise HTTPException(status_code=400, detail="لا توجد بيانات للتحديث")
    
    result = await db.team_members.update_one(
        {"id": member_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    
    return {"message": "تم تحديث بيانات العضو بنجاح"}


@api_router.delete("/team-members/{member_id}")
async def delete_team_member(member_id: str, current_user: User = Depends(get_current_user)):
    result = await db.team_members.delete_one({"id": member_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Member not found")
    
    return {"message": "تم حذف العضو بنجاح"}


# ============= CONTRACTORS ROUTES =============

@api_router.get("/contractors", response_model=List[ContractorResponse])
async def get_contractors(
    project: Optional[str] = Query(None),
    all_contractors: Optional[bool] = Query(False),
    current_user: User = Depends(get_current_user)
):
    """جلب قائمة المقاولين، مع إمكانية الفلترة حسب المشروع والصلاحيات"""
    query = {}
    
    # التحقق من الصلاحيات والمشاريع المسموحة لغير المسؤولين
    if current_user.role != "admin":
        # المقاولين يمكن جلبهم إذا كان لدى المستخدم صلاحية إدارة المقاولين، أو صلاحية إدخال/عرض التوصيلات (المياه أو الصرف) أو البلاغات أو المستخلصات/الفواتير
        allowed_contractor_projects = get_projects_with_permission(current_user, "contractors")
        allowed_water_projects = get_projects_with_permission(current_user, "water_connections")
        allowed_sewage_projects = get_projects_with_permission(current_user, "sewage_connections")
        allowed_reports_view_projects = get_projects_with_permission(current_user, "reports_view")
        allowed_reports_add_projects = get_projects_with_permission(current_user, "reports_add")
        allowed_reports_edit_projects = get_projects_with_permission(current_user, "reports_edit")
        allowed_reports_review_projects = get_projects_with_permission(current_user, "reports_review")
        allowed_extracts_projects = get_projects_with_permission(current_user, "extracts")
        allowed_invoices_projects = get_projects_with_permission(current_user, "invoices")
        
        allowed_projects = list(set(
            allowed_contractor_projects + 
            allowed_water_projects + 
            allowed_sewage_projects + 
            allowed_reports_view_projects + 
            allowed_reports_add_projects +
            allowed_reports_edit_projects +
            allowed_reports_review_projects +
            allowed_extracts_projects +
            allowed_invoices_projects
        ))
        
        if not allowed_projects:
            return []
            
        if project:
            # إذا حدد مشروعاً معيناً، نتحقق من صلاحيته عليه
            if project in allowed_projects:
                query["project"] = project
            else:
                return []
        else:
            # إذا لم يحدد مشروعاً
            if len(allowed_projects) == 1:
                # إذا كان لديه مشروع واحد فقط، نظهر مقاولي هذا المشروع فقط
                query["project"] = allowed_projects[0]
            else:
                # إذا كان لديه عدة مشاريع، نظهر مقاولي كل مشاريع المستخدم المسموحة
                query["project"] = {"$in": allowed_projects}
    else:
        # للـ Admin: إذا طُلب جميع المقاولين أو لم يُحدد مشروع
        if current_user.role == "admin" and all_contractors:
            # إرجاع جميع المقاولين بدون فلتر
            pass
        # فلترة حسب المشروع إذا تم تحديده
        elif project:
            query["project"] = project
            
    contractors = await db.contractors.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # تحويل التواريخ
    for contractor in contractors:
        if isinstance(contractor.get('created_at'), str):
            contractor['created_at'] = datetime.fromisoformat(contractor['created_at'])
            
    return [ContractorResponse(**contractor) for contractor in contractors]


@api_router.post("/contractors", response_model=ContractorResponse)
async def create_contractor(
    contractor_data: ContractorCreate,
    current_user: User = Depends(get_current_user)
):
    """إضافة مقاول جديد"""
    # التحقق من صلاحية المستخدم على هذا المشروع
    if current_user.role != "admin":
        allowed_projects = get_projects_with_permission(current_user, "contractors")
        if contractor_data.project not in allowed_projects:
            raise HTTPException(
                status_code=403, 
                detail="ليس لديك صلاحية لإضافة مقاول لهذا المشروع"
            )
            
    # التحقق من عدم وجود مقاول بنفس الاسم في نفس المشروع
    existing = await db.contractors.find_one({
        "name": contractor_data.name,
        "project": contractor_data.project
    })
    
    if existing:
        raise HTTPException(
            status_code=400, 
            detail="يوجد مقاول بنفس الاسم في هذا المشروع"
        )
        
    contractor = Contractor(
        **contractor_data.model_dump(),
        created_by=current_user.id
    )
    
    doc = contractor.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.contractors.insert_one(doc)
    
    return ContractorResponse(**contractor.model_dump())


@api_router.put("/contractors/{contractor_id}", response_model=ContractorResponse)
async def update_contractor(
    contractor_id: str,
    contractor_data: ContractorCreate,
    current_user: User = Depends(get_current_user)
):
    """تعديل بيانات مقاول"""
    contractor = await db.contractors.find_one({"id": contractor_id}, {"_id": 0})
    
    if not contractor:
        raise HTTPException(status_code=404, detail="المقاول غير موجود")
        
    # التحقق من صلاحية المستخدم على هذا المشروع
    if current_user.role != "admin":
        allowed_projects = get_projects_with_permission(current_user, "contractors")
        if contractor_data.project not in allowed_projects:
            raise HTTPException(
                status_code=403,
                detail="ليس لديك صلاحية لتعديل مقاول في هذا المشروع"
            )
        if contractor.get("project") not in allowed_projects:
            raise HTTPException(
                status_code=403,
                detail="ليس لديك صلاحية لتعديل مقاول في هذا المشروع"
            )
            
    # التحقق من عدم تكرار الاسم في نفس المشروع (باستثناء المقاول الحالي)
    existing = await db.contractors.find_one({
        "name": contractor_data.name,
        "project": contractor_data.project,
        "id": {"$ne": contractor_id}
    })
    
    if existing:
        raise HTTPException(
            status_code=400,
            detail="يوجد مقاول آخر بنفس الاسم في هذا المشروع"
        )
        
    update_data = contractor_data.model_dump()
    
    await db.contractors.update_one(
        {"id": contractor_id},
        {"$set": update_data}
    )
    
    # جلب البيانات المحدثة
    updated_contractor = await db.contractors.find_one({"id": contractor_id}, {"_id": 0})
    
    if isinstance(updated_contractor.get('created_at'), str):
        updated_contractor['created_at'] = datetime.fromisoformat(updated_contractor['created_at'])
        
    return ContractorResponse(**updated_contractor)


@api_router.delete("/contractors/{contractor_id}")
async def delete_contractor(
    contractor_id: str,
    current_user: User = Depends(get_current_user)
):
    """حذف مقاول"""
    contractor = await db.contractors.find_one({"id": contractor_id})
    if not contractor:
        raise HTTPException(status_code=404, detail="المقاول غير موجود")
        
    # التحقق من صلاحية المستخدم على هذا المشروع
    if current_user.role != "admin":
        allowed_projects = get_projects_with_permission(current_user, "contractors")
        if contractor.get("project") not in allowed_projects:
            raise HTTPException(
                status_code=403,
                detail="ليس لديك صلاحية لحذف مقاول في هذا المشروع"
            )
            
    await db.contractors.delete_one({"id": contractor_id})
    
    return {"message": "تم حذف المقاول بنجاح"}


# ===== حالات البلاغ =====
@api_router.get("/report-statuses")
async def get_report_statuses(
    project: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """جلب حالات البلاغ حسب المشروع"""
    query = {}
    if project:
        query["project"] = project
    
    statuses = await db.report_statuses.find(query, {"_id": 0}).to_list(100)
    
    # إضافة الحالات الافتراضية إذا لم توجد
    default_statuses = ["تم الإصلاح", "جاري العمل", "معلق", "ملغي"]
    
    if not statuses and project:
        # إنشاء الحالات الافتراضية للمشروع
        for status_name in default_statuses:
            status = ReportStatus(name=status_name, project=project)
            await db.report_statuses.insert_one(status.model_dump())
        statuses = await db.report_statuses.find(query, {"_id": 0}).to_list(100)
    
    return statuses

@api_router.post("/report-statuses")
async def create_report_status(
    status_data: ReportStatusCreate,
    current_user: User = Depends(get_current_user)
):
    """إضافة حالة بلاغ جديدة"""
    # التحقق من الصلاحيات
    if current_user.role != "admin" and not current_user.can_create_subusers:
        raise HTTPException(status_code=403, detail="غير مصرح لك بإضافة حالات")
    
    # التحقق من عدم التكرار
    existing = await db.report_statuses.find_one({
        "name": status_data.name,
        "project": status_data.project
    })
    if existing:
        raise HTTPException(status_code=400, detail="هذه الحالة موجودة بالفعل")
    
    status = ReportStatus(**status_data.model_dump())
    await db.report_statuses.insert_one(status.model_dump())
    
    return {"message": "تم إضافة الحالة بنجاح", "id": status.id}

@api_router.delete("/report-statuses/{status_id}")
async def delete_report_status(
    status_id: str,
    current_user: User = Depends(get_current_user)
):
    """حذف حالة بلاغ"""
    if current_user.role != "admin" and not current_user.can_create_subusers:
        raise HTTPException(status_code=403, detail="غير مصرح لك بالحذف")
    
    result = await db.report_statuses.delete_one({"id": status_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الحالة غير موجودة")
    
    return {"message": "تم حذف الحالة بنجاح"}


# ===== أنواع البلاغات =====
DEFAULT_REPORT_TYPES = ["ترابي", "بلاط", "أسفلت"]

@api_router.get("/report-types")
async def get_report_types(
    project: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """جلب أنواع البلاغات"""
    query = {}
    if project:
        query["project"] = project
    
    types = await db.report_types.find(query, {"_id": 0}).to_list(100)
    
    # إذا لم توجد أنواع للمشروع، أنشئ الافتراضية
    if project and not types:
        for type_name in DEFAULT_REPORT_TYPES:
            new_type = ReportType(name=type_name, project=project)
            await db.report_types.insert_one(new_type.model_dump())
        types = await db.report_types.find({"project": project}, {"_id": 0}).to_list(100)
    
    return types

@api_router.post("/report-types")
async def create_report_type(
    type_data: ReportTypeCreate,
    current_user: User = Depends(get_current_user)
):
    """إضافة نوع بلاغ جديد"""
    if current_user.role != "admin" and not current_user.can_create_subusers:
        raise HTTPException(status_code=403, detail="غير مصرح لك بالإضافة")
    
    # التحقق من عدم التكرار
    existing = await db.report_types.find_one({"name": type_data.name, "project": type_data.project})
    if existing:
        raise HTTPException(status_code=400, detail="هذا النوع موجود مسبقاً")
    
    new_type = ReportType(**type_data.model_dump())
    await db.report_types.insert_one(new_type.model_dump())
    
    return {"message": "تم إضافة النوع بنجاح", "id": new_type.id}

@api_router.delete("/report-types/{type_id}")
async def delete_report_type(
    type_id: str,
    current_user: User = Depends(get_current_user)
):
    """حذف نوع بلاغ"""
    if current_user.role != "admin" and not current_user.can_create_subusers:
        raise HTTPException(status_code=403, detail="غير مصرح لك بالحذف")
    
    result = await db.report_types.delete_one({"id": type_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="النوع غير موجود")
    
    return {"message": "تم حذف النوع بنجاح"}


# ===== بطاقات المشروع (CRUD كامل) =====

def get_default_cards():
    """البطاقات الافتراضية - تطابق البطاقات في لوحة التحكم"""
    return [
        {"id": str(uuid.uuid4()), "key": "asphalt_waiting", "label": "عدد البلاغات بانتظار الأسفلت"},
        {"id": str(uuid.uuid4()), "key": "asphalt_licensed", "label": "عدد الرخص الصادرة لبلاغات الأسفلت"},
        {"id": str(uuid.uuid4()), "key": "asphalt_unlicensed", "label": "عدد الرخص الغير صادرة لبلاغات الأسفلت"},
        {"id": str(uuid.uuid4()), "key": "tile_licensed", "label": "عدد الرخص الصادرة لبلاغات البلاط"},
        {"id": str(uuid.uuid4()), "key": "tile_unlicensed", "label": "عدد الرخص الغير صادرة لبلاغات البلاط"},
        {"id": str(uuid.uuid4()), "key": "fixed_title", "label": "البلاغات التي تم إصلاحها حسب النوع"},
        {"id": str(uuid.uuid4()), "key": "type_terrestrial", "label": "ترابي"},
        {"id": str(uuid.uuid4()), "key": "type_tile", "label": "بلاط"},
        {"id": str(uuid.uuid4()), "key": "type_asphalt", "label": "أسفلت"}
    ]

@api_router.get("/project-cards/{project}")
async def get_project_cards(
    project: str,
    current_user: User = Depends(get_current_user)
):
    """جلب بطاقات مشروع معين"""
    data = await db.project_cards.find_one({"project": project}, {"_id": 0})
    
    if not data:
        # إنشاء بطاقات افتراضية
        new_data = {
            "id": str(uuid.uuid4()),
            "project": project,
            "cards": get_default_cards(),
            "created_at": datetime.now(timezone.utc)
        }
        await db.project_cards.insert_one(new_data)
        data = new_data
    
    return data

@api_router.post("/project-cards/{project}")
async def add_project_card(
    project: str,
    card_data: CardItemCreate,
    current_user: User = Depends(get_current_user)
):
    """إضافة بطاقة جديدة لمشروع"""
    is_admin = current_user.role == "admin"
    is_level2 = current_user.can_create_subusers
    
    if not (is_admin or is_level2):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    if is_level2 and not is_admin and project not in (current_user.projects or []):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية على هذا المشروع")
    
    # إنشاء مفتاح تلقائي من UUID
    auto_key = f"card_{str(uuid.uuid4())[:8]}"
    
    new_card = {
        "id": str(uuid.uuid4()),
        "key": auto_key,
        "label": card_data.label
    }
    
    # التأكد من وجود المشروع
    existing = await db.project_cards.find_one({"project": project})
    
    if existing:
        await db.project_cards.update_one(
            {"project": project},
            {"$push": {"cards": new_card}}
        )
    else:
        new_data = {
            "id": str(uuid.uuid4()),
            "project": project,
            "cards": [new_card],
            "created_at": datetime.now(timezone.utc)
        }
        await db.project_cards.insert_one(new_data)
    
    return {"message": "تم إضافة البطاقة بنجاح", "card": new_card}

@api_router.put("/project-cards/{project}/{card_id}")
async def update_project_card(
    project: str,
    card_id: str,
    card_data: CardItemUpdate,
    current_user: User = Depends(get_current_user)
):
    """تعديل بطاقة"""
    is_admin = current_user.role == "admin"
    is_level2 = current_user.can_create_subusers
    
    if not (is_admin or is_level2):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    if is_level2 and not is_admin and project not in (current_user.projects or []):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية على هذا المشروع")
    
    result = await db.project_cards.update_one(
        {"project": project, "cards.id": card_id},
        {"$set": {"cards.$.label": card_data.label}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="البطاقة غير موجودة")
    
    return {"message": "تم تعديل البطاقة بنجاح"}

@api_router.delete("/project-cards/{project}/{card_id}")
async def delete_project_card(
    project: str,
    card_id: str,
    current_user: User = Depends(get_current_user)
):
    """حذف بطاقة"""
    is_admin = current_user.role == "admin"
    is_level2 = current_user.can_create_subusers
    
    if not (is_admin or is_level2):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    if is_level2 and not is_admin and project not in (current_user.projects or []):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية على هذا المشروع")
    
    result = await db.project_cards.update_one(
        {"project": project},
        {"$pull": {"cards": {"id": card_id}}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="البطاقة غير موجودة")
    
    return {"message": "تم حذف البطاقة بنجاح"}

@api_router.get("/project-cards")
async def get_all_project_cards(
    current_user: User = Depends(get_current_user)
):
    """جلب جميع بطاقات المشاريع"""
    data = await db.project_cards.find({}, {"_id": 0}).to_list(100)
    return data


# ============= WORK UNITS (وحدات الأعمال) ENDPOINTS =============

@api_router.get("/work-units")
async def get_work_units(current_user: User = Depends(get_current_user)):
    """الحصول على قائمة وحدات الأعمال"""
    units = await db.work_units.find({"is_deleted": {"$ne": True}}, {"_id": 0}).to_list(100)
    return units

@api_router.post("/work-units")
async def create_work_unit(
    name: str = Form(...),
    current_user: User = Depends(get_current_user)
):
    """إضافة وحدة أعمال جديدة"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="هذه الميزة متاحة للأدمن فقط")
    
    existing = await db.work_units.find_one({"name": name, "is_deleted": {"$ne": True}})
    if existing:
        raise HTTPException(status_code=400, detail="وحدة الأعمال موجودة مسبقاً")
    
    unit = {
        "id": str(uuid.uuid4()),
        "name": name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "is_deleted": {"$ne": True}
    }
    await db.work_units.insert_one(unit)
    return {"id": unit["id"], "name": name, "message": "تم إضافة وحدة الأعمال بنجاح"}

@api_router.delete("/work-units/{unit_id}")
async def delete_work_unit(unit_id: str, current_user: User = Depends(get_current_user)):
    """حذف وحدة أعمال"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="هذه الميزة متاحة للأدمن فقط")
    
    result = await db.work_units.update_one({"id": unit_id}, {"$set": {"is_deleted": True}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="وحدة الأعمال غير موجودة")
    return {"message": "تم حذف وحدة الأعمال بنجاح"}


# ============= EXTRACTS (المستخلصات) ENDPOINTS =============

@api_router.get("/extracts")
async def get_extracts(
    project: Optional[str] = None,
    month: Optional[int] = None,
    payment_status: Optional[str] = None,  # paid, unpaid, all
    my_extracts: Optional[bool] = None,  # عرض مستخلصاتي فقط
    extract_number: Optional[str] = None,  # بحث برقم المستخلص
    date: Optional[str] = None,  # بحث بالتاريخ
    page: int = 1,
    limit: int = 10,
    current_user: User = Depends(get_current_user)
):
    """الحصول على قائمة المستخلصات مع فلترة حسب المشروع والشهر وحالة الصرف"""
    query = {"is_deleted": {"$ne": True}}
    
    # فلترة حسب المشروع إذا تم تحديده
    if project:
        query["project"] = project
    
    # فلترة حسب الشهر إذا تم تحديده
    if month:
        query["month"] = month
    
    # فلترة حسب رقم المستخلص
    if extract_number:
        query["extract_number"] = {"$regex": extract_number, "$options": "i"}
    
    # فلترة حسب التاريخ
    if date:
        query["extract_date"] = date
    
    # فلترة حسب حالة الصرف (لبيت الخبرة فقط)
    if payment_status and current_user.role == "admin":
        if payment_status == "paid":
            query["is_paid"] = True
        elif payment_status == "unpaid":
            query["$or"] = [{"is_paid": False}, {"is_paid": {"$exists": False}}]
    
    # فلترة حسب صلاحيات المستخدم
    if current_user.role != "admin":
        if current_user.projects and len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
        # إذا طلب المستخدم مستخلصاته فقط
        if my_extracts:
            query["created_by"] = current_user.id
    
    # الحصول على إجمالي العدد
    total_count = await db.extracts.count_documents(query)
    total_pages = (total_count + limit - 1) // limit if limit > 0 else 1
    
    # تطبيق الترقيم
    skip = (page - 1) * limit
    extracts = await db.extracts.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # تحويل التواريخ
    for extract in extracts:
        if isinstance(extract.get('created_at'), str):
            extract['created_at'] = datetime.fromisoformat(extract['created_at'])
        if isinstance(extract.get('updated_at'), str):
            extract['updated_at'] = datetime.fromisoformat(extract['updated_at'])
        if isinstance(extract.get('paid_at'), str):
            extract['paid_at'] = datetime.fromisoformat(extract['paid_at'])
    
    return {
        "extracts": [ExtractResponse(**extract) for extract in extracts],
        "total_count": total_count,
        "total_pages": total_pages,
        "current_page": page,
        "limit": limit
    }


@api_router.post("/extracts", response_model=ExtractResponse)
async def create_extract(
    project: str = Form(...),
    extract_number: Optional[str] = Form(None),
    invoice_number: Optional[str] = Form(None),
    extract_date: Optional[str] = Form(None),
    work_unit: Optional[str] = Form(None),
    po_number: Optional[str] = Form(None),
    amount: Optional[float] = Form(None),  # المبلغ من مدير المشروع
    actual_value: Optional[float] = Form(None),
    advance_deduction: Optional[float] = Form(None),
    tax: Optional[float] = Form(None),
    penalties: Optional[float] = Form(None),
    total_collected: Optional[float] = Form(None),
    is_paid: Optional[bool] = Form(False),
    collection_date: Optional[str] = Form(None),
    month: Optional[int] = Form(None),
    year: Optional[int] = Form(None),
    images: List[UploadFile] = File(default=[]),
    current_user: User = Depends(get_current_user)
):
    """إنشاء مستخلص جديد"""
    
    # حساب المعادلات
    net_after_deduction = (actual_value or 0) - (advance_deduction or 0)
    total_submitted = net_after_deduction + (tax or 0) - (penalties or 0)
    # الفرق = الغرامات (إذا كانت موجودة) أو الفرق بين المقدم والمحصل
    if penalties and penalties > 0:
        difference = penalties
    else:
        difference = total_submitted - (total_collected or 0)
    
    # معالجة الصور - رفع إلى Object Storage
    import asyncio as _asyncio
    image_data = []
    for image in images:
        if image.filename:
            content = await image.read()
            loop = _asyncio.get_event_loop()
            content = await loop.run_in_executor(thread_pool, compress_image, content)
            url = await loop.run_in_executor(
                thread_pool,
                lambda c=content, fn=image.filename, ct=image.content_type:
                    _store_image_bytes(c, category="extracts", filename=fn, content_type=ct)
            )
            image_data.append(url)
    
    # إنشاء المستخلص
    extract = Extract(
        project=project,
        extract_number=extract_number,
        invoice_number=invoice_number,
        extract_date=extract_date,
        work_unit=work_unit,
        po_number=po_number,
        amount=amount,  # إضافة المبلغ
        actual_value=actual_value,
        advance_deduction=advance_deduction,
        net_after_deduction=net_after_deduction,
        tax=tax,
        penalties=penalties,
        total_submitted=total_submitted,
        total_collected=total_collected,
        difference=difference,
        is_paid=is_paid or False,
        collection_date=collection_date,
        month=month,
        year=year,
        images=image_data,
        status="pending",
        created_by=current_user.id,
        created_by_name=current_user.full_name
    )
    
    extract_dict = extract.model_dump()
    extract_dict['created_at'] = extract.created_at.isoformat()
    extract_dict['updated_at'] = extract.updated_at.isoformat()
    
    await db.extracts.insert_one(extract_dict)
    
    return ExtractResponse(**extract.model_dump())


@api_router.get("/extracts/payment-stats")
async def get_extract_payment_stats(
    project: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """إحصائيات صرف المستخلصات - لبيت الخبرة فقط"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="هذه الميزة متاحة لبيت الخبرة فقط")
    
    query = {"is_deleted": {"$ne": True}}
    if project:
        query["project"] = project
    
    # إجمالي المستخلصات
    total = await db.extracts.count_documents(query)
    
    # المستخلصات المصروفة
    paid_query = {**query, "is_paid": True}
    paid_count = await db.extracts.count_documents(paid_query)
    
    # المستخلصات غير المصروفة
    unpaid_query = {**query, "$or": [{"is_paid": False}, {"is_paid": {"$exists": False}}]}
    unpaid_count = await db.extracts.count_documents(unpaid_query)
    
    # إجمالي المبالغ المصروفة والمتبقية
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": None, 
            "total_paid": {"$sum": {"$cond": [{"$eq": ["$is_paid", True]}, "$paid_amount", 0]}},
            "total_amount": {"$sum": {"$ifNull": ["$total_amount", 0]}},
            "total_remaining": {"$sum": {"$ifNull": ["$remaining_amount", 0]}}
        }}
    ]
    result = await db.extracts.aggregate(pipeline).to_list(1)
    stats = result[0] if result else {"total_paid": 0, "total_amount": 0, "total_remaining": 0}
    
    return {
        "total": total,
        "paid_count": paid_count,
        "unpaid_count": unpaid_count,
        "total_paid_amount": stats.get("total_paid") or 0,
        "total_amount": stats.get("total_amount") or 0,
        "total_remaining_amount": stats.get("total_remaining") or 0
    }


@api_router.get("/extracts/{extract_id}", response_model=ExtractResponse)
async def get_extract(
    extract_id: str,
    current_user: User = Depends(get_current_user)
):
    """الحصول على تفاصيل مستخلص محدد"""
    extract = await db.extracts.find_one(
        {"id": extract_id, "is_deleted": {"$ne": True}},
        {"_id": 0}
    )
    
    if not extract:
        raise HTTPException(status_code=404, detail="المستخلص غير موجود")
    
    # تحويل التواريخ
    if isinstance(extract.get('created_at'), str):
        extract['created_at'] = datetime.fromisoformat(extract['created_at'])
    if isinstance(extract.get('updated_at'), str):
        extract['updated_at'] = datetime.fromisoformat(extract['updated_at'])
    
    return ExtractResponse(**extract)


@api_router.put("/extracts/{extract_id}", response_model=ExtractResponse)
async def update_extract(
    extract_id: str,
    project: Optional[str] = Form(None),
    extract_number: Optional[str] = Form(None),
    invoice_number: Optional[str] = Form(None),
    extract_date: Optional[str] = Form(None),
    work_unit: Optional[str] = Form(None),
    po_number: Optional[str] = Form(None),
    actual_value: Optional[float] = Form(None),
    advance_deduction: Optional[float] = Form(None),
    tax: Optional[float] = Form(None),
    penalties: Optional[float] = Form(None),
    total_collected: Optional[float] = Form(None),
    is_paid: Optional[bool] = Form(None),
    collection_date: Optional[str] = Form(None),
    month: Optional[int] = Form(None),
    year: Optional[int] = Form(None),
    status: Optional[str] = Form(None),
    images: List[UploadFile] = File(default=[]),
    current_user: User = Depends(get_current_user)
):
    """تحديث مستخلص"""
    extract = await db.extracts.find_one(
        {"id": extract_id, "is_deleted": {"$ne": True}},
        {"_id": 0}
    )
    
    if not extract:
        raise HTTPException(status_code=404, detail="المستخلص غير موجود")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    # تحديث الحقول الأساسية
    if project is not None:
        update_data["project"] = project
    if extract_number is not None:
        update_data["extract_number"] = extract_number
    if invoice_number is not None:
        update_data["invoice_number"] = invoice_number
    if extract_date is not None:
        update_data["extract_date"] = extract_date
    if work_unit is not None:
        update_data["work_unit"] = work_unit
    if po_number is not None:
        update_data["po_number"] = po_number
    if month is not None:
        update_data["month"] = month
    if year is not None:
        update_data["year"] = year
    if status is not None:
        update_data["status"] = status
    if collection_date is not None:
        update_data["collection_date"] = collection_date
    if is_paid is not None:
        update_data["is_paid"] = is_paid
    
    # تحديث الحقول المالية وحساب المعادلات
    current_actual = actual_value if actual_value is not None else extract.get('actual_value', 0) or 0
    current_advance = advance_deduction if advance_deduction is not None else extract.get('advance_deduction', 0) or 0
    current_tax = tax if tax is not None else extract.get('tax', 0) or 0
    current_penalties = penalties if penalties is not None else extract.get('penalties', 0) or 0
    current_collected = total_collected if total_collected is not None else extract.get('total_collected', 0) or 0
    
    if actual_value is not None:
        update_data["actual_value"] = actual_value
    if advance_deduction is not None:
        update_data["advance_deduction"] = advance_deduction
    if tax is not None:
        update_data["tax"] = tax
    if penalties is not None:
        update_data["penalties"] = penalties
    if total_collected is not None:
        update_data["total_collected"] = total_collected
    
    # حساب المعادلات
    net_after_deduction = current_actual - current_advance
    total_submitted = net_after_deduction + current_tax - current_penalties
    # الفرق = الغرامات (إذا كانت موجودة) أو الفرق بين المقدم والمحصل
    if current_penalties and current_penalties > 0:
        difference = current_penalties
    else:
        difference = total_submitted - current_collected
    
    update_data["net_after_deduction"] = net_after_deduction
    update_data["total_submitted"] = total_submitted
    update_data["difference"] = difference
    
    # إضافة صور جديدة - رفع إلى Object Storage
    if images and images[0].filename:
        import asyncio as _asyncio
        existing_images = extract.get('images', [])
        new_image_data = []
        for image in images:
            content = await image.read()
            loop = _asyncio.get_event_loop()
            content = await loop.run_in_executor(thread_pool, compress_image, content)
            url = await loop.run_in_executor(
                thread_pool,
                lambda c=content, fn=image.filename, ct=image.content_type:
                    _store_image_bytes(c, category="extracts", filename=fn, content_type=ct)
            )
            new_image_data.append(url)
        update_data["images"] = existing_images + new_image_data
    
    await db.extracts.update_one(
        {"id": extract_id},
        {"$set": update_data}
    )
    
    # الحصول على البيانات المحدثة
    updated_extract = await db.extracts.find_one(
        {"id": extract_id},
        {"_id": 0}
    )
    
    if isinstance(updated_extract.get('created_at'), str):
        updated_extract['created_at'] = datetime.fromisoformat(updated_extract['created_at'])
    if isinstance(updated_extract.get('updated_at'), str):
        updated_extract['updated_at'] = datetime.fromisoformat(updated_extract['updated_at'])
    
    return ExtractResponse(**updated_extract)


@api_router.delete("/extracts/{extract_id}")
async def delete_extract(
    extract_id: str,
    current_user: User = Depends(get_current_user)
):
    """حذف مستخلص (soft delete)"""
    result = await db.extracts.update_one(
        {"id": extract_id},
        {"$set": {"is_deleted": True}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="المستخلص غير موجود")
    
    return {"message": "تم حذف المستخلص بنجاح"}


@api_router.get("/extracts-trash")
async def get_deleted_extracts(
    page: int = 1,
    limit: int = 15,
    current_user: User = Depends(get_current_user)
):
    """جلب المستخلصات المحذوفة"""
    # التحقق من الصلاحيات للمستخلصات
    query = {"is_deleted": True}
    
    if current_user.role != "admin" and current_user.level not in ["1", "2"]:
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
        else:
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية عرض المستخلصات المحذوفة")
    
    total_count = await db.extracts.count_documents(query)
    skip = (page - 1) * limit
    extracts = await db.extracts.find(query, {"_id": 0}).sort("deleted_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # تحويل التواريخ
    for extract in extracts:
        if isinstance(extract.get('created_at'), str):
            extract['created_at'] = datetime.fromisoformat(extract['created_at'])
        if isinstance(extract.get('updated_at'), str):
            extract['updated_at'] = datetime.fromisoformat(extract['updated_at'])
            
    return {"items": [ExtractResponse(**ext) for ext in extracts], "total": total_count}


@api_router.post("/extracts-trash/{extract_id}/restore")
async def restore_deleted_extract(extract_id: str, current_user: User = Depends(get_current_user)):
    """استعادة مستخلص محذوف"""
    query = {"id": extract_id, "is_deleted": True}
    
    if current_user.role != "admin" and current_user.level not in ["1", "2"]:
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
    
    extract = await db.extracts.find_one(query, {"_id": 0})
    if not extract:
        raise HTTPException(status_code=404, detail="المستخلص غير موجود أو ليس لديك صلاحية")
        
    await db.extracts.update_one({"id": extract_id}, {"$set": {"is_deleted": {"$ne": True}}})
    return {"message": "تم استعادة المستخلص بنجاح"}


@api_router.delete("/extracts-trash/{extract_id}/permanent")
async def permanently_delete_extract(extract_id: str, current_user: User = Depends(get_current_user)):
    """حذف المستخلص نهائياً"""
    has_trash_perm = current_user.role == "admin" or "trash" in (current_user.permissions or []) or user_has_any_project_permission(current_user, "trash")
    if not has_trash_perm and current_user.username != "Eng Mahmoud Haroun":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية الحذف النهائي")
        
    query = {"id": extract_id, "is_deleted": True}
    
    if current_user.role != "admin" and current_user.level not in ["1", "2"]:
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
            
    extract = await db.extracts.find_one(query, {"_id": 0})
    if not extract:
        raise HTTPException(status_code=404, detail="المستخلص غير موجود أو ليس لديك صلاحية")
        
    await db.extracts.delete_one({"id": extract_id})
    return {"message": "تم الحذف النهائي بنجاح"}


@api_router.post("/extracts/recalculate-differences")
async def recalculate_extract_differences(
    current_user: User = Depends(get_current_user)
):
    """إعادة حساب الفرق لجميع المستخلصات التي لديها غرامات"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    # تحديث الفرق = الغرامات للمستخلصات التي لديها غرامات > 0
    result = await db.extracts.update_many(
        {"penalties": {"$gt": 0}},
        [{"$set": {"difference": "$penalties"}}]
    )
    
    return {"message": f"تم تحديث {result.modified_count} مستخلص"}



# ============= API صرف المستخلصات (لبيت الخبرة فقط) =============

class ExtractPaymentRequest(BaseModel):
    is_paid: bool
    total_amount: Optional[float] = None  # المبلغ الكلي للمستخلص
    paid_amount: Optional[float] = None  # المبلغ المصروف
    paid_date: Optional[str] = None  # تاريخ الصرف (يحدده المستخدم)
    payment_notes: Optional[str] = None


@api_router.put("/extracts/{extract_id}/payment")
async def update_extract_payment(
    extract_id: str,
    payment: ExtractPaymentRequest,
    current_user: User = Depends(get_current_user)
):
    """تحديث حالة صرف المستخلص - لبيت الخبرة (admin) فقط"""
    # التحقق من أن المستخدم هو admin فقط
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="هذه الميزة متاحة لبيت الخبرة فقط")
    
    extract = await db.extracts.find_one({"id": extract_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    
    if not extract:
        raise HTTPException(status_code=404, detail="المستخلص غير موجود")
    
    update_data = {
        "is_paid": payment.is_paid,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # تحديث المبلغ الكلي إذا تم إرساله
    if payment.total_amount is not None:
        update_data["total_amount"] = payment.total_amount
    
    if payment.is_paid:
        update_data["paid_amount"] = payment.paid_amount
        # حساب المبلغ المتبقي
        total = payment.total_amount or extract.get('total_amount') or 0
        paid = payment.paid_amount or 0
        update_data["remaining_amount"] = max(0, total - paid)
        
        # استخدام تاريخ الصرف المحدد من المستخدم أو التاريخ الحالي
        if payment.paid_date:
            update_data["paid_at"] = payment.paid_date
        else:
            update_data["paid_at"] = datetime.now(timezone.utc).isoformat()
        update_data["paid_by"] = current_user.id
        update_data["paid_by_name"] = current_user.full_name
        update_data["payment_notes"] = payment.payment_notes
    else:
        # إلغاء الصرف
        update_data["paid_amount"] = None
        update_data["remaining_amount"] = None
        update_data["paid_at"] = None
        update_data["paid_by"] = None
        update_data["paid_by_name"] = None
        update_data["payment_notes"] = None
    
    await db.extracts.update_one({"id": extract_id}, {"$set": update_data})
    
    # جلب البيانات المحدثة
    updated = await db.extracts.find_one({"id": extract_id}, {"_id": 0})
    
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    if isinstance(updated.get('updated_at'), str):
        updated['updated_at'] = datetime.fromisoformat(updated['updated_at'])
    if isinstance(updated.get('paid_at'), str):
        updated['paid_at'] = datetime.fromisoformat(updated['paid_at'])
    
    return ExtractResponse(**updated)


@api_router.get("/reports/export/pdf")
async def export_reports_pdf(
    search: Optional[str] = Query(None),
    governorate: Optional[str] = Query(None),
    project: Optional[str] = Query(None),
    contractor: Optional[str] = Query(None),
    report_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    license_status: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    start_date_from: Optional[str] = Query(None),
    start_date_to: Optional[str] = Query(None),
    created_by: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from arabic_reshaper import reshape
    from bidi.algorithm import get_display
    import io
    
    # تسجيل الخط العربي
    arabic_font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'NotoSansArabic-Regular.ttf')
    pdfmetrics.registerFont(TTFont('Arabic', arabic_font_path))
    # تسجيل خط يدعم اللاتينية للأعمدة مثل CCP
    noto_sans_path = os.path.join(os.path.dirname(__file__), 'fonts', 'NotoSans-Regular.ttf')
    if os.path.exists(noto_sans_path):
        pdfmetrics.registerFont(TTFont('NotoSans', noto_sans_path))
    
    query = {"is_deleted": {"$ne": True}}
    
    # التصفية الهرمية: عرض تقارير المستخدم + المستخدمين الفرعيين
    if current_user.role != "admin":
        # تصفية حسب صلاحيات المشاريع إذا كانت محددة
        if len(current_user.projects) > 0:
            query["project"] = {"$in": current_user.projects}
        
        # تصفية حسب صلاحيات المحافظات إذا كانت محددة
        if len(current_user.governorates) > 0:
            query["governorate"] = {"$in": current_user.governorates}
        
        # ملاحظة: تم إلغاء قيود "restricted_users" على التصدير — أي مستخدم يستطيع تصدير
        # جميع البلاغات في مشاريعه ومحافظاته بغض النظر عن المُنشئ.
    
    if search:
        query["$or"] = [
            {"report_number": {"$regex": search, "$options": "i"}},
            {"license_number": {"$regex": search, "$options": "i"}}
        ]
    
    if governorate:
        query["governorate"] = governorate
    
    if project:
        query["project"] = project
    if contractor:
        query["contractor"] = contractor
    
    if report_type:
        query["report_type"] = report_type
    
    if status:
        query["status"] = status
    
    # فلتر حسب المستخدم (للـ Admin ومستوى 2 ومحمود هارون ومدحت)
    if created_by:
        # البحث بـ username أو user_id
        if "$and" in query:
            # إذا كان هناك $and موجود، نضيف شرط جديد
            query["$and"].append({
                "$or": [
                    {"created_by": created_by},
                    {"created_by": {"$regex": created_by, "$options": "i"}}
                ]
            })
        else:
            # إذا لم يكن هناك $and، نستخدم created_by مباشرة
            query["created_by"] = created_by
    
    # فلترة حسب حالة الرخصة أو الحالة
    if license_status == 'status_fixed':
        query["status"] = "تم الإصلاح"
    elif license_status == 'status_asphalt':
        query["status"] = {"$in": [_ASPHALT_CANONICAL, "بانتظار الأسفلت"]}
    elif license_status == 'status_in_progress':
        query["wfm_closed"] = {"$ne": True}
    elif license_status == 'status_wfm_closed':
        query["wfm_closed"] = True
    elif license_status == 'review_pending':
        query["review_status"] = {"$in": ["بانتظار المراجعة", "قيد المراجعة", None]}
    elif license_status == 'license_issued':
        query["license_number"] = {"$regex": "[0-9]"}
    elif license_status == 'license_not_issued':
        query["$or"] = [
            {"license_number": {"$exists": False}},
            {"license_number": None},
            {"license_number": {"$not": {"$regex": "[0-9]"}}}
        ]
    elif license_status and license_status.startswith('custom_'):
        custom_status_name = license_status[len('custom_'):]
        query["status"] = custom_status_name
    
    # فلترة بتاريخ استلام البلاغ (created_at) - يدعم string و datetime
    if date_from or date_to:
        from datetime import datetime as dt, timedelta
        
        try:
            if date_from and date_to:
                # كلا التاريخين موجودان
                date_from_obj = dt.fromisoformat(date_from)
                date_to_obj = dt.fromisoformat(date_to)
                next_day = date_to_obj + timedelta(days=1)
                
                # فلتر يدعم string و datetime
                date_filter = {
                    "$or": [
                        {"created_at": {"$gte": f"{date_from}T00:00:00", "$lt": next_day.strftime("%Y-%m-%dT00:00:00")}},
                        {"created_at": {"$gte": date_from_obj, "$lt": next_day}}
                    ]
                }
            elif date_from:
                date_from_obj = dt.fromisoformat(date_from)
                date_filter = {
                    "$or": [
                        {"created_at": {"$gte": f"{date_from}T00:00:00"}},
                        {"created_at": {"$gte": date_from_obj}}
                    ]
                }
            elif date_to:
                date_to_obj = dt.fromisoformat(date_to)
                next_day = date_to_obj + timedelta(days=1)
                date_filter = {
                    "$or": [
                        {"created_at": {"$lt": next_day.strftime("%Y-%m-%dT00:00:00")}},
                        {"created_at": {"$lt": next_day}}
                    ]
                }
            
            # إضافة فلتر التاريخ للـ query
            if "$and" in query:
                query["$and"].append(date_filter)
            elif "$or" in query:
                query["$and"] = [{"$or": query.pop("$or")}, date_filter]
            else:
                if "$or" in date_filter:
                    query["$and"] = [date_filter]
                else:
                    query.update(date_filter)
        except Exception as e:
            print(f"Date filter error: {e}")

    # فلترة بتاريخ مباشرة البلاغ (start_date)
    if start_date_from or start_date_to:
        from datetime import datetime as dt, timedelta
        try:
            if start_date_from and start_date_to:
                s_from = dt.fromisoformat(start_date_from)
                s_to = dt.fromisoformat(start_date_to)
                s_next = s_to + timedelta(days=1)
                sdate_filter = {
                    "$or": [
                        {"start_date": {"$gte": f"{start_date_from}T00:00:00", "$lt": s_next.strftime("%Y-%m-%dT00:00:00")}},
                        {"start_date": {"$gte": s_from, "$lt": s_next}}
                    ]
                }
            elif start_date_from:
                s_from = dt.fromisoformat(start_date_from)
                sdate_filter = {
                    "$or": [
                        {"start_date": {"$gte": f"{start_date_from}T00:00:00"}},
                        {"start_date": {"$gte": s_from}}
                    ]
                }
            else:
                s_to = dt.fromisoformat(start_date_to)
                s_next = s_to + timedelta(days=1)
                sdate_filter = {
                    "$or": [
                        {"start_date": {"$lt": s_next.strftime("%Y-%m-%dT00:00:00")}},
                        {"start_date": {"$lt": s_next}}
                    ]
                }
            
            if "$and" in query:
                query["$and"].append(sdate_filter)
            elif "$or" in query:
                query["$and"] = [{"$or": query.pop("$or")}, sdate_filter]
            else:
                query["$and"] = [sdate_filter]
        except Exception as e:
            print(f"Start date filter error: {e}")
    
    # جلب البلاغات بدون حد (None) لدعم أي عدد من البلاغات
    # ⚡ استثناء حقل الصور الكامل لتسريع الاستعلام - نجلب فقط عدد الصور
    projection = {
        "_id": 0,
        "images": 0  # استثناء الصور الكاملة
    }
    reports = await db.reports.find(query, projection).sort("created_at", -1).to_list(None)
    
    # جلب عدد الصور لكل بلاغ بشكل منفصل وسريع
    report_ids = [r.get('id') for r in reports if r.get('id')]
    images_counts = {}
    if report_ids:
        pipeline = [
            {"$match": {"id": {"$in": report_ids}}},
            {"$project": {"id": 1, "images_count": {"$size": {"$ifNull": ["$images", []]}}}}
        ]
        counts_cursor = db.reports.aggregate(pipeline)
        async for doc in counts_cursor:
            images_counts[doc.get('id')] = doc.get('images_count', 0)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=15, leftMargin=15, topMargin=15, bottomMargin=15)
    elements = []
    
    # إضافة الشعارات في أعلى الصفحة
    from reportlab.platypus import Image, Spacer
    from reportlab.lib.units import mm
    
    # مسارات الشعارات (ديناميكية من إعدادات المنصة)
    branding = await db.platform_settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    bayt_logo_path = await _resolve_logo_path(branding.get("company_logo_url"), default_filename="bayt-alkhibra-logo.png")
    nwc_logo_path = await _resolve_logo_path(branding.get("partner_logo_url"), default_filename="nwc-logo.png")
    
    # إنشاء جدول للشعارات (شركة المياه يمين - بيت الخبرة يسار)
    logo_data = []
    logo_row = []
    
    # شعار بيت الخبرة (يسار) - تصغير الشعارات
    if bayt_logo_path and os.path.exists(bayt_logo_path):
        bayt_logo = Image(bayt_logo_path, width=40*mm, height=20*mm)
        logo_row.append(bayt_logo)
    else:
        logo_row.append('')
    
    # مسافة فارغة في المنتصف
    logo_row.append('')
    
    # شعار شركة المياه الوطنية (يمين) - تصغير الشعارات
    if nwc_logo_path and os.path.exists(nwc_logo_path):
        nwc_logo = Image(nwc_logo_path, width=40*mm, height=20*mm)
        logo_row.append(nwc_logo)
    else:
        logo_row.append('')
    
    logo_data.append(logo_row)
    
    # إنشاء جدول الشعارات بعرض الصفحة
    page_width = landscape(A4)[0] - 60  # عرض الصفحة ناقص الهوامش
    logo_table = Table(logo_data, colWidths=[page_width*0.3, page_width*0.4, page_width*0.3])
    logo_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),    # بيت الخبرة يسار
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),   # شركة المياه يمين
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(logo_table)
    elements.append(Spacer(1, 5*mm))  # تقليل المسافة بعد الشعارات
    
    # دالة لاختصار اسم المشروع
    def shorten_project_name(project_name):
        if not project_name:
            return ''
        if 'الغربية' in project_name:
            return 'مشروع المحافظات الغربية'
        elif 'الشمالية' in project_name:
            return 'مشروع المحافظات الشمالية'
        elif 'الجنوبية' in project_name:
            return 'مشروع المحافظات الجنوبية'
        return project_name
    
    # دالة للحصول على اسم المشروع الكامل
    def get_full_project_name(proj):
        return proj if proj else ""
    
    # إنشاء العنوان الديناميكي
    title_style = ParagraphStyle(
        'ArabicTitle',
        fontName='Arabic',
        fontSize=16,  # تصغير حجم العنوان
        alignment=TA_CENTER,
        textColor=colors.HexColor('#366092'),
        spaceBefore=0,
        spaceAfter=8,  # تقليل المسافة بعد العنوان
        leading=20
    )
    
    # إنشاء عنوان ديناميكي حسب الفلاتر والشهر
    # استخراج الشهر من date_from إذا كان موجوداً
    month_text = ""
    if date_from:
        try:
            date_obj = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            month_num = date_obj.month
            # أسماء الأشهر بالعربية
            months_ar = ["", "يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو", 
                        "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]
            month_text = f" - شهر {months_ar[month_num]}"
        except:
            pass
    
    if governorate and project:
        short_project = shorten_project_name(project)
        title_text = reshape(f"بلاغات {governorate} - {short_project}{month_text}")
    elif governorate:
        title_text = reshape(f"بلاغات محافظة {governorate}{month_text}")
    elif project:
        short_project = shorten_project_name(project)
        title_text = reshape(f"بلاغات {short_project}{month_text}")
    elif month_text:
        title_text = reshape(f"البلاغات{month_text}")
    else:
        title_text = reshape("تقرير البلاغات")
    
    title_text = get_display(title_text)
    title = Paragraph(title_text, title_style)
    elements.append(title)
    
    # إضافة اسم المشروع الكامل - تقليل المسافات
    project_style = ParagraphStyle(
        'ProjectStyle',
        fontName='Arabic',
        fontSize=10,  # تصغير حجم الخط
        alignment=TA_CENTER,
        textColor=colors.HexColor('#333333'),
        spaceBefore=0,
        spaceAfter=5,  # تقليل المسافة
        leading=14
    )
    
    # استخراج المشروع من البلاغات أو الفلتر
    project_for_display = project
    if not project_for_display and reports:
        # الحصول على المشروع الأكثر تكراراً
        proj_counts = {}
        for r in reports:
            p = r.get('project', '')
            if p:
                proj_counts[p] = proj_counts.get(p, 0) + 1
        if proj_counts:
            project_for_display = max(proj_counts, key=proj_counts.get)
    
    full_project_display = get_full_project_name(project_for_display)
    if full_project_display:
        project_text = reshape(full_project_display)
        project_text = get_display(project_text)
        project_para = Paragraph(project_text, project_style)
        elements.append(project_para)
    
    # إضافة م / محمود هارون فوق الجدول
    author_style = ParagraphStyle(
        'AuthorStyle2',
        fontName='Arabic',
        fontSize=8,
        alignment=TA_RIGHT,
        textColor=colors.HexColor('#555555'),
        spaceBefore=2,
        spaceAfter=5,
        leading=14
    )
    author_text = reshape("تنفيذ م-محمود محمد هارون مدير النظام وتحليل البيانات")
    author_text = get_display(author_text)
    author_para = Paragraph(author_text, author_style)
    elements.append(author_para)
    
    # إضافة مسافة صغيرة قبل الجدول
    elements.append(Spacer(1, 3*mm))
    
    # إعداد البيانات - 11 عمود (مع العمق والقطر)
    headers = [
        "رقم", "المحافظة", "المشروع", "رقم البلاغ", "رقم الرخصة", 
        "الحالة", "خط العرض", "خط الطول", "نوع البلاغ", "المقاول", "تاريخ الإنشاء"
    ]
    
    # عكس ترتيب الأعمدة لتبدأ من اليمين (RTL)
    headers_reversed = headers[::-1]
    
    # ستايلات خلايا الجدول لضمان التنسيق والتفاف النص والترميز الصحيح
    header_style = ParagraphStyle(
        'HeaderStyle2',
        fontName='Arabic',
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.whitesmoke,
        leading=10
    )
    
    cell_style = ParagraphStyle(
        'CellStyle2',
        fontName='Arabic',
        fontSize=8,
        alignment=TA_CENTER,
        leading=10,
        wordWrap='RTL'
    )
    
    latin_cell_style = ParagraphStyle(
        'LatinCellStyle2',
        fontName='NotoSans' if os.path.exists(noto_sans_path) else 'Helvetica',
        fontSize=8,
        alignment=TA_CENTER,
        leading=10
    )
    
    # معالجة العناوين - الطريقة الأصلية البسيطة التي كانت تعمل
    processed_headers = []
    for header in headers_reversed:
        processed_headers.append(Paragraph(arabic_text(header), header_style))
    
    data = [processed_headers]
    
    # معالجة البيانات
    for idx, report in enumerate(reports, 1):
        created_at = report.get('created_at')
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at)
        elif isinstance(created_at, datetime):
            pass
        else:
            created_at = None
        
        # تكوين الإحداثيات
        lat = report.get('latitude', '')
        lng = report.get('longitude', '')
        coordinates = f"{lat}, {lng}" if lat and lng else ''
        
        # إضافة CCB- أو CCP- حسب المشروع
        report_num = report.get('report_number', '')
        if report_num:
            report_num = str(report_num).replace('CCP-', '').replace('CCB-', '')
            proj_name = report.get('project', '')
            if 'الغربي' in proj_name:
                report_num = f"CCB-{report_num}"
            else:
                report_num = f"CCP-{report_num}"
        
        # جلب قيم العمق والقطر
        depth_val = report.get('depth_meters', '')
        diameter_val = report.get('diameter_mm', '')
        depth_str = str(depth_val) if depth_val else '-'
        diameter_str = str(diameter_val) if diameter_val else '-'
        
        # ترتيب الصف من اليمين لليسار مباشرة (بدون عكس)
        # معالجة رقم الرخصة - إذا كان فارغ نعرض "لم يتم إصدار رخصة"
        license_num = report.get('license_number') or ''
        license_num = str(license_num).strip()
        if not license_num:
            license_num = 'لم يتم إصدار رخصة'
        
        row = [
            created_at.strftime('%Y-%m-%d') if created_at else '',  # تاريخ الإنشاء - يمين
            report.get('contractor', ''),  # المقاول
            diameter_str,  # القطر بالملليمتر
            depth_str,  # العمق بالمتر
            report.get('report_type', ''),  # نوع البلاغ
            report.get('status', ''),  # الحالة
            "مغلقة بواسطة الاستشاري" if report.get('wfm_closed') else "قيد المعالجة", # حالة الرخصة
            license_num,  # رقم الرخصة
            report_num,  # رقم البلاغ مع CCP- (بدون عكس)
            shorten_project_name(report.get('project', '')),  # المشروع
            report.get('governorate', ''),  # المحافظة
            str(idx)  # رقم - يسار
        ]
        
        processed_row = [
            Paragraph(created_at.strftime('%Y-%m-%d') if created_at else '', latin_cell_style),
            Paragraph(arabic_text(report.get('contractor', '')), cell_style),
            Paragraph(diameter_str, latin_cell_style),
            Paragraph(depth_str, latin_cell_style),
            Paragraph(arabic_text(report.get('report_type', '')), cell_style),
            Paragraph(arabic_text(report.get('status', '')), cell_style),
            Paragraph(arabic_text("مغلقة بواسطة الاستشاري") if report.get('wfm_closed') else arabic_text("قيد المعالجة"), cell_style),
            Paragraph(arabic_text(license_num), cell_style),
            Paragraph(report_num, latin_cell_style),
            Paragraph(arabic_text(shorten_project_name(report.get('project', ''))), cell_style),
            Paragraph(arabic_text(report.get('governorate', '')), cell_style),
            Paragraph(str(idx), latin_cell_style)
        ]
        
        data.append(processed_row)
    
    # إنشاء الجدول - 11 عمود متوازن ضمن عرض الصفحة landscape A4 (802 points)
    # العروض بالـ points (معكوسة من اليمين لليسار): تاريخ، مقاول، قطر، عمق، نوع، حالة، رقم رخصة، رقم بلاغ، مشروع، محافظة، رقم
    # تعديل: توسيع الحالة (70→85)، تضييق رقم الرخصة (90→75)
    # المجموع: 65+75+50+50+65+85+75+85+105+60+28 = 743 points (ضمن 802)
    # إنشاء الجدول - 12 عمود متوازن مع توسيع خانة رقم البلاغ والحالة
    # العروض: تاريخ(50), مقاول(70), قطر(35), عمق(35), نوع(50), حالة(125), حالة رخصة(85), رقم رخصة(50), رقم بلاغ(110), مشروع(120), محافظة(35), رقم(20)
    col_widths = [50, 70, 50, 55, 55, 125, 95, 110, 140, 35, 20]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    # العمود 7 من اليسار (index 7) هو رقم البلاغ (CCB-)
    # والعمود 0 هو التاريخ (يحتاج خط لاتيني أيضاً)
    table.setStyle(TableStyle([
        # تنسيق الرأس
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        
        # تنسيق الجسم
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # ========== استخراج اسم المراقب من البلاغات ==========
    supervisor_name = ""
    if reports:
        # جلب أسماء المستخدمين الذين أنشأوا البلاغات
        created_by_ids = {}
        for r in reports:
            # أولاً نحاول الحصول على الاسم مباشرة
            creator_name = r.get('created_by_name', '')
            creator_id = r.get('created_by', '')
            
            if creator_name and creator_name != creator_id:
                # الاسم موجود ومختلف عن الـ ID
                created_by_ids[creator_name] = created_by_ids.get(creator_name, 0) + 1
            elif creator_id:
                # نحتاج لجلب الاسم من قاعدة البيانات
                created_by_ids[creator_id] = created_by_ids.get(creator_id, 0) + 1
        
        # أخذ الـ ID/الاسم الأكثر تكراراً
        if created_by_ids:
            most_common = max(created_by_ids, key=created_by_ids.get)
            
            # إذا كان UUID، نجلب الاسم من قاعدة البيانات
            if len(most_common) == 36 and '-' in most_common:  # UUID format
                user_doc = await db.users.find_one({"id": most_common}, {"_id": 0, "full_name": 1, "username": 1})
                if user_doc:
                    supervisor_name = user_doc.get('full_name') or user_doc.get('username', '')
            else:
                supervisor_name = most_common
    
    # ========== إضافة قسم التوقيعات مباشرة تحت الجدول ==========
    elements.append(Spacer(1, 10*mm))
    
    # دالة مساعدة للنص العربي في التوقيعات - معالجة محسنة
    def sig_arabic(text):
        import html
        if not text: return ""
        if not isinstance(text, str): text = str(text)
        if '\\u' in text:
            try: text = text.encode('utf-8').decode('unicode-escape')
            except: pass
        text = html.unescape(text)
        return get_display(reshape(text))
    
    # بيانات التوقيع مع اسم المراقب
    platform_settings = await db.platform_settings.find_one({"key": "platform_name"}, {"_id": 0}) or {}
    platform_name = platform_settings.get("value", "مكتب بيت الخبرة للاستشارات الهندسية")

    company_name = branding.get("company_name")
    if not company_name:
        company_name = platform_name
        
    partner_company = branding.get("partner_company_name")
    if not partner_company:
        partner_company = "شركة المياه الوطنية"
        
    sig_data = [
        [
            sig_arabic(partner_company),
            "",
            "",
            sig_arabic(company_name)
        ],
        [
            sig_arabic("الاسم: ........................"),
            "",
            "",
            sig_arabic(f"الاسم: {supervisor_name}") if supervisor_name else sig_arabic("الاسم: ........................")
        ],
        [
            sig_arabic("التوقيع: ........................"),
            "",
            "",
            sig_arabic("التوقيع: ........................")
        ]
    ]
    
    sig_table = Table(sig_data, colWidths=[page_width*0.30, page_width*0.20, page_width*0.20, page_width*0.30])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    
    # استخدام KeepTogether لضمان بقاء التوقيع في صفحة واحدة
    from reportlab.platypus import KeepTogether
    elements.append(KeepTogether([Spacer(1, 5*mm), sig_table]))
    
    doc.build(elements)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reports.pdf"}
    )


# ============= CARS MANAGEMENT (سيارات المشاريع) =============

@api_router.get("/cars")
async def get_cars(
    project: Optional[str] = None,
    page: int = 1,
    limit: int = 10,
    current_user: User = Depends(get_current_user)
):
    """
    جلب قائمة السيارات مع الترقيم - نظام الصلاحيات:
    - المستوى 1 (Admin/بيت الخبرة): يرى جميع السيارات
    - من لديه صلاحية cars_manage: يرى جميع السيارات (يمكنه الإدارة)
    - المستوى 2 (Manager): يرى سيارات فريقه فقط
    - المستوى 3 (User): يرى فقط السيارة المسلمة له شخصياً
    """
    query = {}
    
    # فلترة حسب المشروع
    if project:
        query["project"] = project
    
    # التحقق من الصلاحيات
    is_admin = current_user.role == "admin"
    is_level2 = current_user.can_create_subusers and not is_admin
    has_cars_permission = "cars" in (current_user.permissions or [])
    has_manage_permission = "cars_manage" in (current_user.permissions or [])
    
    if is_admin or has_manage_permission:
        # المستوى 1 (بيت الخبرة) أو من لديه صلاحية الإدارة: يرى جميع السيارات
        pass
    elif is_level2:
        # المستوى 2 (Manager): يرى سيارات فريقه + سياراته
        team_users = await db.users.find(
            {"created_by": current_user.id},
            {"_id": 0, "id": 1}
        ).to_list(1000)
        team_user_ids = [u["id"] for u in team_users]
        team_user_ids.append(current_user.id)
        query["assigned_user_id"] = {"$in": team_user_ids}
    elif has_cars_permission:
        # المستوى 3 مع صلاحية عرض السيارات: يرى سياراته فقط
        query["assigned_user_id"] = current_user.id
    else:
        # المستوى 3 العادي: يرى سياراته فقط
        query["assigned_user_id"] = current_user.id
    
    # Get total count
    total_count = await db.cars.count_documents(query)
    total_pages = (total_count + limit - 1) // limit if limit > 0 else 1
    
    # Apply pagination
    skip = (page - 1) * limit
    cars = await db.cars.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "cars": cars,
        "total_count": total_count,
        "total_pages": total_pages,
        "current_page": page,
        "limit": limit
    }


@api_router.get("/cars/all-users")
async def get_all_users_for_cars(current_user: User = Depends(get_current_user)):
    """جلب جميع المستخدمين من جميع المشاريع - لمن لديه صلاحية cars_manage"""
    has_manage_permission = "cars_manage" in (current_user.permissions or [])
    is_admin = current_user.role == "admin"
    
    if not is_admin and not has_manage_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    # جلب جميع المستخدمين غير الأدمن
    users = await db.users.find(
        {"role": {"$ne": "admin"}},
        {"_id": 0, "id": 1, "username": 1, "full_name": 1, "projects": 1, "assigned_projects": 1}
    ).to_list(1000)
    
    return users


@api_router.get("/cars/{car_id}")
async def get_car(car_id: str, current_user: User = Depends(get_current_user)):
    """جلب سيارة واحدة"""
    car = await db.cars.find_one({"id": car_id}, {"_id": 0})
    if not car:
        raise HTTPException(status_code=404, detail="السيارة غير موجودة")
    
    is_admin = current_user.role == "admin"
    has_cars_permission = "cars" in (current_user.permissions or [])
    is_assigned = car.get("assigned_user_id") == current_user.id
    
    if not is_admin and not has_cars_permission and not is_assigned:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    return car


@api_router.post("/cars")
async def create_car(car: CarCreate, current_user: User = Depends(get_current_user)):
    """
    إضافة/تسليم سيارة جديدة - نظام الصلاحيات:
    - المستوى 1 (Admin/بيت الخبرة): يمكنه تسليم السيارات لأي شخص
    - من لديه صلاحية cars_manage: يمكنه تسليم السيارات لأي شخص من أي مشروع
    - المستوى 2 (Manager) بدون صلاحية: يمكنه تسليم لفريقه فقط
    - المستوى 3: لا يمكنه تسليم سيارات
    """
    is_admin = current_user.role == "admin"
    has_manage_permission = "cars_manage" in (current_user.permissions or [])
    is_level2 = current_user.can_create_subusers and not is_admin
    
    # التحقق من صلاحية التسليم
    if is_admin or has_manage_permission:
        # المستوى 1 أو من لديه صلاحية cars_manage: يمكنه تسليم لأي شخص
        pass
    elif is_level2:
        # المستوى 2 بدون صلاحية: يمكنه تسليم لفريقه فقط
        team_users = await db.users.find(
            {"created_by": current_user.id},
            {"_id": 0, "id": 1}
        ).to_list(1000)
        team_user_ids = [u["id"] for u in team_users]
        team_user_ids.append(current_user.id)
        
        if car.assigned_user_id not in team_user_ids:
            raise HTTPException(status_code=403, detail="يمكنك تسليم سيارات فقط لأعضاء فريقك")
    else:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية لتسليم سيارة")
    
    car_doc = {
        "id": str(uuid.uuid4()),
        "project": car.project,
        "assigned_user_id": car.assigned_user_id,
        "assigned_user_name": car.assigned_user_name,
        "car_type": car.car_type,
        "plate_number": car.plate_number,
        "model": car.model,
        "authorization_start": car.authorization_start,
        "authorization_end": car.authorization_end,
        "color": car.color,
        "notes": car.notes or "",
        "kilometers": car.kilometers or "",
        "image": "",
        "created_by": current_user.id,
        "created_by_name": current_user.full_name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.cars.insert_one(car_doc)
    car_doc.pop("_id", None)
    return car_doc


@api_router.put("/cars/{car_id}")
async def update_car(
    car_id: str, 
    car_update: CarUpdate, 
    current_user: User = Depends(get_current_user)
):
    """تحديث سيارة - نفس صلاحيات التسليم"""
    car = await db.cars.find_one({"id": car_id}, {"_id": 0})
    if not car:
        raise HTTPException(status_code=404, detail="السيارة غير موجودة")
    
    is_admin = current_user.role == "admin"
    has_manage_permission = "cars_manage" in (current_user.permissions or [])
    is_level2 = current_user.can_create_subusers and not is_admin
    
    if is_admin or has_manage_permission:
        # Admin أو من لديه صلاحية: يمكنه التعديل
        pass
    elif is_level2:
        # تحقق أن السيارة مسلمة لأحد أفراد فريقه
        team_users = await db.users.find(
            {"created_by": current_user.id},
            {"_id": 0, "id": 1}
        ).to_list(1000)
        team_user_ids = [u["id"] for u in team_users]
        team_user_ids.append(current_user.id)
        
        if car.get("assigned_user_id") not in team_user_ids:
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية لتعديل هذه السيارة")
    else:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية للتعديل")
    
    update_data = {k: v for k, v in car_update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.cars.update_one({"id": car_id}, {"$set": update_data})
    return {"message": "تم تحديث السيارة بنجاح"}


@api_router.delete("/cars/{car_id}")
async def delete_car(car_id: str, current_user: User = Depends(get_current_user)):
    """حذف سيارة - نفس صلاحيات التسليم"""
    car = await db.cars.find_one({"id": car_id}, {"_id": 0})
    if not car:
        raise HTTPException(status_code=404, detail="السيارة غير موجودة")
    
    is_admin = current_user.role == "admin"
    has_manage_permission = "cars_manage" in (current_user.permissions or [])
    is_level2 = current_user.can_create_subusers and not is_admin
    
    if is_admin or has_manage_permission:
        # Admin أو من لديه صلاحية: يمكنه الحذف
        pass
    elif is_level2:
        # تحقق أن السيارة مسلمة لأحد أفراد فريقه
        team_users = await db.users.find(
            {"created_by": current_user.id},
            {"_id": 0, "id": 1}
        ).to_list(1000)
        team_user_ids = [u["id"] for u in team_users]
        team_user_ids.append(current_user.id)
        
        if car.get("assigned_user_id") not in team_user_ids:
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية لحذف هذه السيارة")
    else:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية للحذف")
    
    await db.cars.delete_one({"id": car_id})
    return {"message": "تم حذف السيارة بنجاح"}


@api_router.post("/cars/{car_id}/image")
async def upload_car_image(
    car_id: str,
    image: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """رفع صورة السيارة - نفس صلاحيات التعديل"""
    car = await db.cars.find_one({"id": car_id}, {"_id": 0})
    if not car:
        raise HTTPException(status_code=404, detail="السيارة غير موجودة")
    
    is_admin = current_user.role == "admin"
    is_level2 = current_user.can_create_subusers and not is_admin
    has_manage_permission = "cars_manage" in (current_user.permissions or [])
    
    if is_admin:
        pass
    elif is_level2 and has_manage_permission:
        # تحقق أن السيارة مسلمة لأحد أفراد فريقه
        team_users = await db.users.find(
            {"created_by": current_user.id},
            {"_id": 0, "id": 1}
        ).to_list(1000)
        team_user_ids = [u["id"] for u in team_users]
        team_user_ids.append(current_user.id)
        
        if car.get("assigned_user_id") not in team_user_ids:
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    else:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    content = await image.read()
    import asyncio as _asyncio
    loop = _asyncio.get_event_loop()
    content = await loop.run_in_executor(thread_pool, compress_image, content)
    image_base64 = await loop.run_in_executor(
        thread_pool,
        lambda: _store_image_bytes(content, category="cars", filename=image.filename, content_type=image.content_type)
    )
    
    await db.cars.update_one(
        {"id": car_id},
        {"$set": {"image": image_base64, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "تم رفع الصورة بنجاح"}


@api_router.get("/cars/project/{project_name}/users")
async def get_project_users(project_name: str, current_user: User = Depends(get_current_user)):
    """جلب مستخدمي مشروع معين - يشمل المستخدمين المعينين للمشروع + جميع المستخدمين بدون مشاريع"""
    # جلب المستخدمين المعينين لهذا المشروع أو بدون مشاريع محددة
    # مع استبعاد المستخدمين الذين تم حذفهم من قائمة السيارات
    users = await db.users.find(
        {
            "$and": [
                # استبعاد المستخدمين المحذوفين من هذا المشروع
                {"$or": [
                    {"excluded_from_cars_projects": {"$exists": False}},
                    {"excluded_from_cars_projects": {"$ne": project_name}}
                ]},
                # شرط المشروع
                {"$or": [
                    {"projects": project_name},
                    {"assigned_projects": project_name},
                    # المستخدمين بدون مشاريع (يمكن تسليمهم سيارات)
                    {"$and": [
                        {"$or": [{"projects": {"$exists": False}}, {"projects": {"$size": 0}}, {"projects": None}]},
                        {"$or": [{"assigned_projects": {"$exists": False}}, {"assigned_projects": {"$size": 0}}, {"assigned_projects": None}]},
                        {"role": {"$ne": "admin"}}
                    ]}
                ]}
            ]
        },
        {"_id": 0, "id": 1, "username": 1, "full_name": 1}
    ).to_list(500)
    
    return users


@api_router.delete("/cars/project-users/{user_id}")
async def remove_user_from_cars_list(user_id: str, project: str = Query(...), current_user: User = Depends(get_current_user)):
    """إزالة مستخدم من قائمة مستلمي السيارات (إزالته من المشروع)"""
    if current_user.role != "admin" and not current_user.can_create_subusers:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    # إضافة المشروع لقائمة المشاريع المستبعدة للمستخدم
    result = await db.users.update_one(
        {"id": user_id},
        {
            "$pull": {"projects": project, "assigned_projects": project},
            "$addToSet": {"excluded_from_cars_projects": project}
        }
    )
    
    # إذا لم يتم تعديل شيء (المستخدم ليس له مشاريع)، نضيفه للقائمة المستبعدة فقط
    if result.modified_count == 0:
        await db.users.update_one(
            {"id": user_id},
            {"$addToSet": {"excluded_from_cars_projects": project}}
        )
    
    return {"message": "تم إزالة المستخدم من القائمة"}


# ============= WATER CONNECTIONS (توصيلات المياه) =============

@api_router.get("/water-connections")
async def get_water_connections(
    project: Optional[str] = None,
    governorate: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    request_number: Optional[str] = None,
    ccb_report_number: Optional[str] = None,
    customer_name: Optional[str] = None,
    search: Optional[str] = None,
    exact: Optional[bool] = False,
    page: int = 1,
    limit: int = 10,
    current_user: User = Depends(get_current_user)
):
    """جلب قائمة توصيلات المياه مع الترقيم - مفلترة حسب مشاريع المستخدم"""
    is_admin = current_user.role == "admin"
    # صلاحية توصيلات المياه: إما عامة أو في أي مشروع من مشاريعه
    has_permission = user_has_any_project_permission(current_user, "water_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    query = {"is_deleted": {"$ne": True}}
    
    # ⚡ فلترة حسب المشاريع المسموحة (لغير الأدمن)
    allowed_projects = None
    if not is_admin:
        allowed_projects = get_projects_with_permission(current_user, "water_connections")
        if not allowed_projects:
            return {"connections": [], "total_count": 0, "total_pages": 0, "current_page": page, "limit": limit}
        query["project"] = {"$in": allowed_projects}
    
    # ⚡ فلترة مرنة حسب اسم المشروع (للجميع)
    if project:
        keywords = [k for k in project.replace('-', ' ').split() if len(k) > 1]
        if keywords:
            p_regex = ".*".join(keywords).replace('أ', '[أا]').replace('إ', '[إا]').replace('ا', '[اأإ]')
            regex_query = {"$regex": p_regex, "$options": "i"}
            
            if not is_admin:
                # دمج البحث مع الصلاحيات المتاحة
                query["$and"] = [
                    {"project": regex_query},
                    {"project": {"$in": allowed_projects}}
                ]
                if "project" in query: del query["project"]
            else:
                query["project"] = regex_query
    
    if status:
        query["request_status"] = status

    if governorate:
        gov_p = normalize_arabic_regex(governorate.strip())
        gov_filter = {"$or": [{"governorate": {"$regex": gov_p, "$options": "i"}}, {"area": {"$regex": gov_p, "$options": "i"}}]}
        if "$or" in query or "$and" in query:
            if "$and" not in query: query["$and"] = []
            if "$or" in query: query["$and"].append({"$or": query.pop("$or")})
            query["$and"].append(gov_filter)
        else:
            query.update(gov_filter)

    if request_number:
        query["request_number"] = {"$regex": request_number, "$options": "i"}
    
    if ccb_report_number:
        query["ccb_report_number"] = {"$regex": ccb_report_number, "$options": "i"}
        
    if customer_name:
        query["customer_name"] = {"$regex": customer_name, "$options": "i"}

    if search:
        if exact:
            query["$or"] = query.get("$or", []) + [
                {"id": search},
                {"request_number": search},
                {"ccb_report_number": search}
            ]
        else:
            query["$or"] = query.get("$or", []) + [
                {"request_number": {"$regex": search, "$options": "i"}},
                {"ccb_report_number": {"$regex": search, "$options": "i"}},
                {"customer_name": {"$regex": search, "$options": "i"}},
                {"phone_number": {"$regex": search, "$options": "i"}},
                {"area": {"$regex": search, "$options": "i"}},
                {"contractor": {"$regex": search, "$options": "i"}}
            ]

    if date_from or date_to:
        date_query = {}
        if date_from: date_query["$gte"] = date_from
        if date_to: date_query["$lte"] = date_to
        query["work_order_date"] = date_query
    
    # ⚡ فلترة حسب المحافظات والتسلسل الهرمي (لغير الأدمن)
    if not is_admin:
        hierarchy_filter = await get_hierarchy_filter(current_user)
        user_governorates = current_user.governorates if hasattr(current_user, 'governorates') and current_user.governorates else []
        has_all_govs = any(g in ["الكل", "جميع المحافظات", "كل المحافظات"] for g in user_governorates)
        
        if not has_all_govs:
            if len(user_governorates) > 0:
                gov_patterns = []
                for g in user_governorates:
                    p = normalize_arabic_regex(g)
                    gov_patterns.append(p)
                gov_regex = f"({'|'.join(gov_patterns)})"
                
                # تطبيق الفلترة الصارمة: يجب أن يكون من إنتاج المستخدم أو تابعيه AND ضمن المحافظات المسندة
                query.update(hierarchy_filter)
                
                gov_or = [
                    {"governorate": {"$regex": gov_regex, "$options": "i"}},
                    {"area": {"$regex": gov_regex, "$options": "i"}}
                ]
                
                if "$and" in query:
                    query["$and"].append({"$or": gov_or})
                else:
                    query["$or"] = gov_or
            else:
                # لا توجد محافظات محددة - يرى فقط ما أنشأه هو أو تابعوه
                query.update(hierarchy_filter)
    
    # Get total count
    total_count = await db.water_connections.count_documents(query)
    total_pages = (total_count + limit - 1) // limit if limit > 0 else 1
    
    # Apply pagination - استبعاد الصور لتسريع الاستجابة
    skip = (page - 1) * limit
    projection = {"_id": 0, "images": 0}  # استبعاد الصور من القائمة
    connections = await db.water_connections.find(query, projection).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # إضافة عدد الصور لكل توصيلة (بدلاً من الصور نفسها)
    for conn in connections:
        images_count = await db.water_connections.find_one({"id": conn["id"]}, {"images": 1})
        conn["images_count"] = len(images_count.get("images", [])) if images_count else 0
    
    return {
        "connections": connections,
        "total_count": total_count,
        "total_pages": total_pages,
        "current_page": page,
        "limit": limit
    }


@api_router.get("/water-connections/{connection_id}")
async def get_water_connection(connection_id: str, current_user: User = Depends(get_current_user)):
    """جلب توصيلة مياه واحدة"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "water_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    connection = await db.water_connections.find_one({"id": connection_id}, {"_id": 0})
    if not connection:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة")
    
    return connection


@api_router.get("/water-connections/{connection_id}/images")
async def get_water_connection_images(connection_id: str, current_user: User = Depends(get_current_user)):
    """جلب صور توصيلة مياه فقط - للتحميل عند الطلب"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "water_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    connection = await db.water_connections.find_one({"id": connection_id}, {"_id": 0, "images": 1})
    if not connection:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة")
    
    return {"images": connection.get("images", [])}


@api_router.post("/water-connections")
async def create_water_connection(data: WaterConnectionCreate, current_user: User = Depends(get_current_user)):
    """إضافة توصيلة مياه جديدة"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "water_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    # التحقق من عدم تكرار الأرقام المهمة في نفس المشروع
    data_dict = data.dict()
    duplicate_checks = []
    
    # التحقق من رقم الطلب
    if data_dict.get('request_number'):
        existing = await db.water_connections.find_one({
            "project": data_dict['project'],
            "request_number": data_dict['request_number']
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"رقم الطلب {data_dict['request_number']} موجود مسبقاً في هذا المشروع")
    
    # التحقق من رقم الحصر/التقييد
    if data_dict.get('restriction_number'):
        existing = await db.water_connections.find_one({
            "project": data_dict['project'],
            "restriction_number": data_dict['restriction_number']
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"رقم الحصر {data_dict['restriction_number']} موجود مسبقاً في هذا المشروع")
    
    # التحقق من رقم البلاغ CCB
    if data_dict.get('ccb_report_number'):
        existing = await db.water_connections.find_one({
            "project": data_dict['project'],
            "ccb_report_number": data_dict['ccb_report_number']
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"رقم بلاغ CCP {data_dict['ccb_report_number']} موجود مسبقاً في هذا المشروع")
    
    # التحقق من رقم التصريح
    if data_dict.get('permit_number'):
        existing = await db.water_connections.find_one({
            "project": data_dict['project'],
            "permit_number": data_dict['permit_number']
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"رقم التصريح {data_dict['permit_number']} موجود مسبقاً في هذا المشروع")
    
    doc = {
        "id": str(uuid.uuid4()),
        **data_dict,
        "created_by": current_user.id,
        "created_by_name": current_user.full_name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # ضمان وجود المحافظة في كلا الحقلين للتوافق
    if doc.get('governorate') and not doc.get('area'):
        doc['area'] = doc['governorate']
    elif doc.get('area') and not doc.get('governorate'):
        doc['governorate'] = doc['area']
    
    await db.water_connections.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/water-connections/{connection_id}")
async def update_water_connection(connection_id: str, data: dict, current_user: User = Depends(get_current_user)):
    """تحديث توصيلة مياه"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "water_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    connection = await db.water_connections.find_one({"id": connection_id}, {"_id": 0})
    if not connection:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة")
    
    # معالجة الصور ورفعها لـ Cloudinary إذا كانت موجودة في التحديث
    if "images" in data and data["images"]:
        data["images"] = await process_images_for_storage(data["images"], category="water_connections")
    
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.water_connections.update_one({"id": connection_id}, {"$set": data})
    return {"message": "تم تحديث التوصيلة بنجاح"}


@api_router.delete("/water-connections/{connection_id}")
async def delete_water_connection(connection_id: str, current_user: User = Depends(get_current_user)):
    """حذف توصيلة مياه"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "water_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    result = await db.water_connections.update_one(
        {"id": connection_id},
        {"$set": {
            "is_deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.id
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة")
    
    return {"message": "تم حذف التوصيلة بنجاح"}


# ============= إحصائيات التوصيلات للـ Dashboard =============

@api_router.get("/connections-stats")
async def get_connections_stats(
    project: Optional[str] = None,
    month: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    print(f"DEBUG: get_connections_stats called with project='{project}', month='{month}'")
    """جلب إحصائيات توصيلات المياه والصرف الصحي للـ Dashboard مع دعم فلترة الشهر"""
    try:
        # فلتر حسب المشروع
        water_filter = {"is_deleted": {"$ne": True}}
        sewage_filter = {"is_deleted": {"$ne": True}}
        
        # فلترة حسب المشروع المسموحة (لغير الأدمن)
        allowed_projects = None
        if current_user.role != "admin":
            allowed_projects = getattr(current_user, 'projects', [])
            if not allowed_projects:
                return {"water": {"total": 0}, "sewage": {"total": 0}}
            
            # التحقق من الصلاحية للمشروع المختار
            if project:
                # التحقق من الصلاحية للمشروع المختار بمرونة
                has_permission = False
                for up in allowed_projects:
                    up_keywords = [k for k in up.replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'أعمال', 'إصلاح']]
                    proj_keywords = [k for k in project.replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'أعمال', 'إصلاح']]
                    is_match = any(k in project for k in up_keywords) or any(k in up for k in proj_keywords)
                    
                    if is_match:
                        has_permission = True
                        break
                if not has_permission:
                    return {"water": {"total": 0}, "sewage": {"total": 0}}
                
                water_filter["project"] = get_flexible_project_query(project)
                sewage_filter["project"] = get_flexible_project_query(project)
            else:
                flex_proj = get_flexible_in_query(allowed_projects, "project")
                water_filter.update(flex_proj)
                sewage_filter.update(flex_proj)
            
            # فلترة حسب المحافظات والتسلسل الهرمي
            user_governorates = current_user.governorates if hasattr(current_user, 'governorates') and current_user.governorates else []
            
            # تطبيق الفلترة الهرمية الشاملة (Recursive)
            hierarchy_filter = await get_hierarchy_filter(current_user)
            
            if len(user_governorates) > 0:
                gov_patterns = []
                for g in user_governorates:
                    p = normalize_arabic_regex(g)
                    gov_patterns.append(p)
                gov_regex = f"({'|'.join(gov_patterns)})"
                
                for f in [water_filter, sewage_filter]:
                    f['$or'] = [
                        hierarchy_filter,
                        {'area': {'$regex': gov_regex, '$options': 'i'}} # التوصيلات تستخدم area
                    ]
            else:
                water_filter.update(hierarchy_filter)
                sewage_filter.update(hierarchy_filter)
        elif project:
            water_filter["project"] = get_flexible_project_query(project)
            sewage_filter["project"] = get_flexible_project_query(project)
        
        # تطبيق فلتر الشهر الشامل (Universal Date Filter)
        if month:
            from datetime import datetime as dt, timezone
            
            # month format: "2024-01"
            year, month_num = month.split('-')
            
            # 1. تحضير قيم الفلترة
            month_regex = f"^{month}" # يبدأ بـ YYYY-MM
            
            # 2. تحضير نطاق التاريخ لكائنات datetime
            date_from_obj = dt(int(year), int(month_num), 1, 0, 0, 0, tzinfo=timezone.utc)
            if int(month_num) == 12:
                date_to_obj = dt(int(year) + 1, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
            else:
                date_to_obj = dt(int(year), int(month_num) + 1, 1, 0, 0, 0, tzinfo=timezone.utc)
            
            # بناء فلتر التاريخ الذي يدعم string و datetime
            # نفحص عدة حقول تاريخ لضمان الشمولية
            date_fields = ["created_at", "added_at", "work_order_date"]
            
            date_filters = []
            for field in date_fields:
                date_filters.append({
                    "$or": [
                        {field: {"$regex": month_regex}},
                        {field: {"$gte": date_from_obj, "$lt": date_to_obj}}
                    ]
                })
            
            # إضافة فلتر التاريخ للـ query الأساسية
            month_query = {"$or": date_filters}
            
            if "$and" not in water_filter:
                water_filter = {"$and": [water_filter, month_query]}
                sewage_filter = {"$and": [sewage_filter, month_query]}
            else:
                water_filter["$and"].append(month_query)
                sewage_filter["$and"].append(month_query)
        
        # إحصائيات توصيلات المياه
        water_total = await db.water_connections.count_documents(water_filter)
        water_new = await db.water_connections.count_documents({**water_filter, "request_status": "جديد"}) if "$and" not in water_filter else await db.water_connections.count_documents({"$and": [water_filter, {"request_status": "جديد"}]})
        water_in_progress = await db.water_connections.count_documents({**water_filter, "request_status": "قيد التنفيذ"}) if "$and" not in water_filter else await db.water_connections.count_documents({"$and": [water_filter, {"request_status": "قيد التنفيذ"}]})
        water_completed = await db.water_connections.count_documents({**water_filter, "request_status": "مكتمل"}) if "$and" not in water_filter else await db.water_connections.count_documents({"$and": [water_filter, {"request_status": "مكتمل"}]})
        water_cancelled = await db.water_connections.count_documents({**water_filter, "request_status": "ملغي"}) if "$and" not in water_filter else await db.water_connections.count_documents({"$and": [water_filter, {"request_status": "ملغي"}]})
        
        # إحصائيات توصيلات الصرف الصحي
        sewage_total = await db.sewage_connections.count_documents(sewage_filter)
        sewage_new = await db.sewage_connections.count_documents({**sewage_filter, "request_status": "جديد"}) if "$and" not in sewage_filter else await db.sewage_connections.count_documents({"$and": [sewage_filter, {"request_status": "جديد"}]})
        sewage_in_progress = await db.sewage_connections.count_documents({**sewage_filter, "request_status": "قيد التنفيذ"}) if "$and" not in sewage_filter else await db.sewage_connections.count_documents({"$and": [sewage_filter, {"request_status": "قيد التنفيذ"}]})
        sewage_completed = await db.sewage_connections.count_documents({**sewage_filter, "request_status": "مكتمل"}) if "$and" not in sewage_filter else await db.sewage_connections.count_documents({"$and": [sewage_filter, {"request_status": "مكتمل"}]})
        sewage_cancelled = await db.sewage_connections.count_documents({**sewage_filter, "request_status": "ملغي"}) if "$and" not in sewage_filter else await db.sewage_connections.count_documents({"$and": [sewage_filter, {"request_status": "ملغي"}]})
        
        return {
            "water": {
                "total": water_total,
                "new": water_new,
                "in_progress": water_in_progress,
                "completed": water_completed,
                "cancelled": water_cancelled
            },
            "sewage": {
                "total": sewage_total,
                "new": sewage_new,
                "in_progress": sewage_in_progress,
                "completed": sewage_completed,
                "cancelled": sewage_cancelled
            },
            "grand_total": water_total + sewage_total
        }
    except Exception as e:
        import traceback
        logging.error(f"Error in get_connections_stats: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


# ============= SEWAGE CONNECTIONS (توصيلات الصرف الصحي) =============

@api_router.get("/sewage-connections")
async def get_sewage_connections(
    project: Optional[str] = None,
    governorate: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    request_number: Optional[str] = None,
    ccb_report_number: Optional[str] = None,
    customer_name: Optional[str] = None,
    search: Optional[str] = None,
    exact: Optional[bool] = False,
    page: int = 1,
    limit: int = 10,
    current_user: User = Depends(get_current_user)
):
    """جلب قائمة توصيلات الصرف الصحي مع الترقيم - مفلترة حسب مشاريع المستخدم"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "sewage_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    query = {"is_deleted": {"$ne": True}}
    
    # ⚡ فلترة حسب المشاريع المسموحة (لغير الأدمن)
    allowed_projects = None
    if not is_admin:
        allowed_projects = get_projects_with_permission(current_user, "sewage_connections")
        if not allowed_projects:
            return {"connections": [], "total_count": 0, "total_pages": 0, "current_page": page, "limit": limit}
        query["project"] = {"$in": allowed_projects}
    
    # ⚡ فلترة مرنة حسب اسم المشروع (للجميع)
    if project:
        keywords = [k for k in project.replace('-', ' ').split() if len(k) > 1]
        if keywords:
            p_regex = ".*".join(keywords).replace('أ', '[أا]').replace('إ', '[إا]').replace('ا', '[اأإ]')
            regex_query = {"$regex": p_regex, "$options": "i"}
            
            if not is_admin:
                # دمج البحث مع الصلاحيات المتاحة
                query["$and"] = [
                    {"project": regex_query},
                    {"project": {"$in": allowed_projects}}
                ]
                if "project" in query: del query["project"]
            else:
                query["project"] = regex_query
    
    if status:
        query["request_status"] = status
    
    if governorate:
        gov_p = normalize_arabic_regex(governorate.strip())
        gov_filter = {"$or": [{"governorate": {"$regex": gov_p, "$options": "i"}}, {"area": {"$regex": gov_p, "$options": "i"}}]}
        if "$or" in query or "$and" in query:
            if "$and" not in query: query["$and"] = []
            if "$or" in query: query["$and"].append({"$or": query.pop("$or")})
            query["$and"].append(gov_filter)
        else:
            query.update(gov_filter)

    if request_number:
        query["request_number"] = {"$regex": request_number, "$options": "i"}
    
    if ccb_report_number:
        query["ccb_report_number"] = {"$regex": ccb_report_number, "$options": "i"}
        
    if customer_name:
        query["customer_name"] = {"$regex": customer_name, "$options": "i"}

    if search:
        if exact:
            search_filter = {"$or": [
                {"id": search},
                {"request_number": search},
                {"ccb_report_number": search}
            ]}
        else:
            search_filter = {"$or": [
                {"request_number": {"$regex": search, "$options": "i"}},
                {"ccb_report_number": {"$regex": search, "$options": "i"}},
                {"customer_name": {"$regex": search, "$options": "i"}},
                {"phone_number": {"$regex": search, "$options": "i"}},
                {"area": {"$regex": search, "$options": "i"}},
                {"contractor": {"$regex": search, "$options": "i"}}
            ]}
        if "$and" in query: query["$and"].append(search_filter)
        elif "$or" in query:
            if "$and" not in query: query["$and"] = []
            query["$and"].append({"$or": query.pop("$or")})
            query["$and"].append(search_filter)
        else:
            query.update(search_filter)

    if date_from or date_to:
        date_query = {}
        if date_from: date_query["$gte"] = date_from
        if date_to: date_query["$lte"] = date_to
        query["work_order_date"] = date_query
    
    # ⚡ فلترة حسب المحافظات والتسلسل الهرمي (لغير الأدمن)
    if not is_admin:
        hierarchy_filter = await get_hierarchy_filter(current_user)
        user_governorates = current_user.governorates if hasattr(current_user, 'governorates') and current_user.governorates else []
        has_all_govs = any(g in ["الكل", "جميع المحافظات", "كل المحافظات"] for g in user_governorates)
        
        if not has_all_govs:
            if len(user_governorates) > 0:
                gov_patterns = []
                for g in user_governorates:
                    p = normalize_arabic_regex(g)
                    gov_patterns.append(p)
                gov_regex = f"({'|'.join(gov_patterns)})"
                
                # تطبيق الفلترة الصارمة: يجب أن يكون من إنتاج المستخدم أو تابعيه AND ضمن المحافظات المسندة
                query.update(hierarchy_filter)
                
                gov_or = [
                    {"governorate": {"$regex": gov_regex, "$options": "i"}},
                    {"area": {"$regex": gov_regex, "$options": "i"}}
                ]
                
                if "$and" in query:
                    query["$and"].append({"$or": gov_or})
                else:
                    query["$or"] = gov_or
            else:
                if "$and" in query:
                    query["$and"].append(hierarchy_filter)
                else:
                    query.update(hierarchy_filter)
    
    # Get total count
    total_count = await db.sewage_connections.count_documents(query)
    total_pages = (total_count + limit - 1) // limit if limit > 0 else 1
    
    # Apply pagination - استبعاد الصور لتسريع الاستجابة
    skip = (page - 1) * limit
    projection = {"_id": 0, "images": 0}  # استبعاد الصور من القائمة
    connections = await db.sewage_connections.find(query, projection).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # إضافة عدد الصور لكل توصيلة (بدلاً من الصور نفسها)
    for conn in connections:
        images_count = await db.sewage_connections.find_one({"id": conn["id"]}, {"images": 1})
        conn["images_count"] = len(images_count.get("images", [])) if images_count else 0
    
    return {
        "connections": connections,
        "total_count": total_count,
        "total_pages": total_pages,
        "current_page": page,
        "limit": limit
    }


@api_router.get("/sewage-connections/{connection_id}")
async def get_sewage_connection(connection_id: str, current_user: User = Depends(get_current_user)):
    """جلب توصيلة صرف صحي واحدة"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "sewage_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    connection = await db.sewage_connections.find_one({"id": connection_id}, {"_id": 0})
    if not connection:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة")
    
    return connection


@api_router.get("/sewage-connections/{connection_id}/images")
async def get_sewage_connection_images(connection_id: str, current_user: User = Depends(get_current_user)):
    """جلب صور توصيلة صرف صحي فقط - للتحميل عند الطلب"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "sewage_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    connection = await db.sewage_connections.find_one({"id": connection_id}, {"_id": 0, "images": 1})
    if not connection:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة")
    
    return {"images": connection.get("images", [])}


@api_router.post("/sewage-connections")
async def create_sewage_connection(data: SewageConnectionCreate, current_user: User = Depends(get_current_user)):
    """إضافة توصيلة صرف صحي جديدة"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "sewage_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    # التحقق من عدم تكرار الأرقام المهمة في نفس المشروع
    data_dict = data.dict()
    
    # التحقق من رقم الطلب
    if data_dict.get('request_number'):
        existing = await db.sewage_connections.find_one({
            "project": data_dict['project'],
            "request_number": data_dict['request_number']
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"رقم الطلب {data_dict['request_number']} موجود مسبقاً في هذا المشروع")
    
    # التحقق من رقم الحصر/التقييد
    if data_dict.get('restriction_number'):
        existing = await db.sewage_connections.find_one({
            "project": data_dict['project'],
            "restriction_number": data_dict['restriction_number']
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"رقم الحصر {data_dict['restriction_number']} موجود مسبقاً في هذا المشروع")
    
    # التحقق من رقم البلاغ CCB
    if data_dict.get('ccb_report_number'):
        existing = await db.sewage_connections.find_one({
            "project": data_dict['project'],
            "ccb_report_number": data_dict['ccb_report_number']
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"رقم بلاغ CCP {data_dict['ccb_report_number']} موجود مسبقاً في هذا المشروع")
    
    # التحقق من رقم التصريح
    if data_dict.get('permit_number'):
        existing = await db.sewage_connections.find_one({
            "project": data_dict['project'],
            "permit_number": data_dict['permit_number']
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"رقم التصريح {data_dict['permit_number']} موجود مسبقاً في هذا المشروع")
    
    doc = {
        "id": str(uuid.uuid4()),
        **data_dict,
        "created_by": current_user.id,
        "created_by_name": current_user.full_name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # ضمان وجود المحافظة في كلا الحقلين للتوافق
    if doc.get('governorate') and not doc.get('area'):
        doc['area'] = doc['governorate']
    elif doc.get('area') and not doc.get('governorate'):
        doc['governorate'] = doc['area']
    
    await db.sewage_connections.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/sewage-connections/{connection_id}")
async def update_sewage_connection(connection_id: str, data: dict, current_user: User = Depends(get_current_user)):
    """تحديث توصيلة صرف صحي"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "sewage_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    connection = await db.sewage_connections.find_one({"id": connection_id}, {"_id": 0})
    if not connection:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة")
    
    # معالجة الصور ورفعها لـ Cloudinary إذا كانت موجودة في التحديث
    if "images" in data and data["images"]:
        data["images"] = await process_images_for_storage(data["images"], category="sewage_connections")
    
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.sewage_connections.update_one({"id": connection_id}, {"$set": data})
    return {"message": "تم تحديث التوصيلة بنجاح"}


@api_router.delete("/sewage-connections/{connection_id}")
async def delete_sewage_connection(connection_id: str, current_user: User = Depends(get_current_user)):
    """حذف توصيلة صرف صحي"""
    is_admin = current_user.role == "admin"
    has_permission = user_has_any_project_permission(current_user, "sewage_connections")
    
    if not is_admin and not has_permission:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    result = await db.sewage_connections.update_one(
        {"id": connection_id},
        {"$set": {
            "is_deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.id
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="التوصيلة غير موجودة")
    
    return {"message": "تم حذف التوصيلة بنجاح"}


# ============= IMPORT WATER CONNECTIONS =============

@api_router.post("/water-connections/import-excel")
async def import_water_connections_from_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """استيراد توصيلات المياه من ملف Excel"""
    # التحقق من الصلاحية: admin أو صلاحية water_connections_import
    if not user_has_any_project_permission(current_user, 'water_connections_import'):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية استيراد توصيلات المياه")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="يجب رفع ملف Excel (.xlsx أو .xls)")
    
    try:
        import pandas as pd
        from io import BytesIO
        
        content = await file.read()
        content_buffer = BytesIO(content)
        
        # قراءة الملف وتحديد صف الأسماء
        df_raw = pd.read_excel(content_buffer, header=None)
        header_row = None
        for i in range(min(10, len(df_raw))):
            row_values = [str(v) for v in df_raw.iloc[i].tolist()]
            row_str = ' '.join(row_values)
            if 'رقم الحساب' in row_str or 'رقم الطلب' in row_str or 'المشروع' in row_str:
                header_row = i
                break
        
        if header_row is not None:
            content_buffer.seek(0)
            df = pd.read_excel(content_buffer, header=header_row)
        else:
            content_buffer.seek(0)
            df = pd.read_excel(content_buffer)
        
        # إزالة الأعمدة الفارغة
        df = df.dropna(axis=1, how='all')
        df = df.dropna(how='all')
        
        # تنظيف أسماء الأعمدة
        df.columns = [str(col).replace(' *', '').replace('*', '').strip() for col in df.columns]
        
        # الحصول على المشروع الافتراضي
        default_project = current_user.projects[0] if current_user.projects else ''
        
        imported = 0
        skipped = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                def safe_str(col_names, default=''):
                    if isinstance(col_names, str):
                        col_names = [col_names]
                    for col in col_names:
                        for df_col in df.columns:
                            if col in df_col:
                                val = row[df_col]
                                if pd.notna(val) and str(val).strip().lower() not in ['nan', 'none', '']:
                                    return str(val).strip()
                    return default
                
                # قراءة رقم الحساب أو رقم الطلب كمعرف
                account_num = safe_str(['رقم الحساب', 'account_number', 'حساب'], '')
                request_num = safe_str(['رقم الطلب', 'request_number', 'طلب'], '')
                
                if not account_num and not request_num:
                    continue
                
                # الحصول على المشروع
                project = safe_str(['المشروع', 'project'], default_project)
                
                # التحقق من عدم التكرار في نفس المشروع
                existing = await db.water_connections.find_one({
                    "$or": [
                        {"account_number": account_num, "project": project} if account_num else {"_id": None},
                        {"request_number": request_num, "project": project} if request_num else {"_id": None}
                    ],
                    "is_deleted": {"$ne": True}
                })
                
                if existing:
                    skipped += 1
                    continue
                
                # تحضير البيانات
                connection_data = {
                    "id": str(uuid.uuid4()),
                    "project": project,
                    "contractor": safe_str(['المقاول', 'contractor'], ''),
                    "account_number": account_num,
                    "request_number": request_num,
                    "restriction_number": safe_str(['رقم التقييد', 'رقم الحصر', 'restriction'], ''),
                    "ccb_report_number": safe_str(['رقم بلاغ CCB', 'رقم بلاغ CCP', 'ccb', 'بلاغ'], ''),
                    "customer_name": safe_str(['اسم العميل', 'عميل', 'customer'], ''),
                    "phone_number": safe_str(['رقم الجوال', 'جوال', 'phone'], ''),
                    "area": safe_str(['المنطقة', 'منطقة', 'area'], ''),
                    "work_order_date": safe_str(['تاريخ أمر العمل', 'أمر العمل', 'تاريخ أمر الشغل'], ''),
                    "diameter": safe_str(['القطر', 'diameter'], ''),
                    "connection_length": safe_str(['طول الماسورة', 'طول التوصيلة', 'طول'], ''),
                    "notes": safe_str(['ملاحظات', 'notes'], ''),
                    "latitude": safe_str(['خط العرض', 'latitude'], ''),
                    "longitude": safe_str(['خط الطول', 'longitude'], ''),
                    "commissioning_date": safe_str(['تاريخ التشغيل', 'تشغيل'], ''),
                    "permit_number": safe_str(['رقم التصريح', 'تصريح'], ''),
                    "request_status": safe_str(['حالة الطلب', 'الحالة', 'status'], 'جديد'),
                    "images": [],
                    "is_deleted": {"$ne": True},
                    "created_by": current_user.id,
                    "created_by_name": current_user.full_name or current_user.username,
                    "created_at": datetime.now(timezone.utc)
                }
                
                await db.water_connections.insert_one(connection_data)
                imported += 1
                
            except Exception as e:
                errors.append(f"سطر {idx + 2}: {str(e)}")
        
        return {
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "total": len(df),
            "errors": errors[:10],
            "message": f"تم استيراد {imported} توصيلة مياه بنجاح" + (f"، تم تخطي {skipped} موجود مسبقاً" if skipped > 0 else "")
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"خطأ في معالجة الملف: {str(e)}")


@api_router.post("/sewage-connections/import-excel")
async def import_sewage_connections_from_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """استيراد توصيلات الصرف الصحي من ملف Excel"""
    # التحقق من الصلاحية: admin أو صلاحية sewage_connections_import
    if not user_has_any_project_permission(current_user, 'sewage_connections_import'):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية استيراد توصيلات الصرف الصحي")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="يجب رفع ملف Excel (.xlsx أو .xls)")
    
    try:
        import pandas as pd
        from io import BytesIO
        
        content = await file.read()
        content_buffer = BytesIO(content)
        
        # قراءة الملف وتحديد صف الأسماء
        df_raw = pd.read_excel(content_buffer, header=None)
        header_row = None
        for i in range(min(10, len(df_raw))):
            row_values = [str(v) for v in df_raw.iloc[i].tolist()]
            row_str = ' '.join(row_values)
            if 'رقم الحساب' in row_str or 'رقم الطلب' in row_str or 'المشروع' in row_str:
                header_row = i
                break
        
        if header_row is not None:
            content_buffer.seek(0)
            df = pd.read_excel(content_buffer, header=header_row)
        else:
            content_buffer.seek(0)
            df = pd.read_excel(content_buffer)
        
        # إزالة الأعمدة الفارغة
        df = df.dropna(axis=1, how='all')
        df = df.dropna(how='all')
        
        # تنظيف أسماء الأعمدة
        df.columns = [str(col).replace(' *', '').replace('*', '').strip() for col in df.columns]
        
        # الحصول على المشروع الافتراضي
        default_project = current_user.projects[0] if current_user.projects else ''
        
        imported = 0
        skipped = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                def safe_str(col_names, default=''):
                    if isinstance(col_names, str):
                        col_names = [col_names]
                    for col in col_names:
                        for df_col in df.columns:
                            if col in df_col:
                                val = row[df_col]
                                if pd.notna(val) and str(val).strip().lower() not in ['nan', 'none', '']:
                                    return str(val).strip()
                    return default
                
                # قراءة رقم الحساب أو رقم الطلب كمعرف
                account_num = safe_str(['رقم الحساب', 'account_number', 'حساب'], '')
                request_num = safe_str(['رقم الطلب', 'request_number', 'طلب'], '')
                
                if not account_num and not request_num:
                    continue
                
                # الحصول على المشروع
                project = safe_str(['المشروع', 'project'], default_project)
                
                # التحقق من عدم التكرار في نفس المشروع
                existing = await db.sewage_connections.find_one({
                    "$or": [
                        {"account_number": account_num, "project": project} if account_num else {"_id": None},
                        {"request_number": request_num, "project": project} if request_num else {"_id": None}
                    ],
                    "is_deleted": {"$ne": True}
                })
                
                if existing:
                    skipped += 1
                    continue
                
                # تحضير البيانات
                connection_data = {
                    "id": str(uuid.uuid4()),
                    "project": project,
                    "contractors": [safe_str(['المقاول', 'contractor', 'المقاولين'], '')] if safe_str(['المقاول', 'contractor', 'المقاولين'], '') else [],
                    "account_number": account_num,
                    "request_number": request_num,
                    "restriction_number": safe_str(['رقم التقييد', 'رقم الحصر', 'restriction'], ''),
                    "ccb_report_number": safe_str(['رقم بلاغ CCB', 'رقم بلاغ CCP', 'ccb', 'بلاغ'], ''),
                    "customer_name": safe_str(['اسم العميل', 'عميل', 'customer'], ''),
                    "customer_number": safe_str(['رقم العميل', 'customer_number'], ''),
                    "phone_number": safe_str(['رقم الجوال', 'جوال', 'phone'], ''),
                    "area": safe_str(['المنطقة', 'منطقة', 'area'], ''),
                    "work_order_date": safe_str(['تاريخ أمر العمل', 'أمر العمل', 'تاريخ أمر الشغل'], ''),
                    "diameter": safe_str(['القطر', 'diameter'], ''),
                    "actual_length": safe_str(['العمق', 'طول على الطبيعة', 'actual_length', 'depth', 'manhole_depth'], ''),
                    "network_line_length": safe_str(['طول خط الشبكة', 'network_line_length', 'طول الشبكة'], ''),
                    "notes": safe_str(['ملاحظات', 'notes'], ''),
                    "latitude": safe_str(['خط العرض', 'latitude'], ''),
                    "longitude": safe_str(['خط الطول', 'longitude'], ''),
                    "commissioning_date": safe_str(['تاريخ التشغيل', 'تشغيل'], ''),
                    "permit": safe_str(['رقم التصريح', 'التصريح', 'permit'], ''),
                    "request_status": safe_str(['حالة الطلب', 'الحالة', 'status'], 'جديد'),
                    "images": [],
                    "is_deleted": {"$ne": True},
                    "created_by": current_user.id,
                    "created_by_name": current_user.full_name or current_user.username,
                    "created_at": datetime.now(timezone.utc)
                }
                
                await db.sewage_connections.insert_one(connection_data)
                imported += 1
                
            except Exception as e:
                errors.append(f"سطر {idx + 2}: {str(e)}")
        
        return {
            "success": True,
            "imported": imported,
            "skipped": skipped,
            "total": len(df),
            "errors": errors[:10],
            "message": f"تم استيراد {imported} توصيلة صرف صحي بنجاح" + (f"، تم تخطي {skipped} موجود مسبقاً" if skipped > 0 else "")
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"خطأ في معالجة الملف: {str(e)}")


# ============= EXPORT CONNECTIONS =============

class ExportRequest(BaseModel):
    connections: List[dict]
    project_name: str


@api_router.post("/water-connections/export/excel")
async def export_water_connections_excel(data: ExportRequest, current_user: User = Depends(get_current_user)):
    """تصدير توصيلات المياه إلى Excel - جميع البيانات"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "توصيلات المياه"
    ws.sheet_view.rightToLeft = True
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # جميع الحقول شاملة المحافظة والملاحظات وكل التفاصيل المدخلة
    headers = [
        '#', 'المشروع', 'المحافظة', 'المقاول', 'رقم الحساب', 'رقم الطلب', 'رقم الحصر', 'رقم البلاغ CCB',
        'اسم العميل', 'رقم الجوال', 'المنطقة / الحي', 'تاريخ أمر الشغل', 'القطر', 'طول التوصيلة',
        'رقم التصريح', 'تاريخ التعميد', 'تاريخ النشر', 'تاريخ الإصدار', 'تاريخ التنفيذ المتوقع',
        'نوع التوصيلة', 'عدد الوصلات', 'طول التوصيلة بدون الإضافي', 'طول الوصلات بدون الرئيسي',
        'قطر خط الشبكة 63', 'طول خط الشبكة', 'قطر خط الشبكة 16', 'رقم العداد', 'نوع العداد',
        'إزالة وتركيب العداد', 'تاريخ التنفيذ', 'تاريخ الإغلاق', 'الحالة', 'تاريخ الإلغاء', 'سبب الإلغاء', 'ملاحظات',
        'مراقب الاستشاري', 'تاريخ الإضافة', 'خط العرض', 'خط الطول'
    ]
    
    last_col = get_column_letter(len(headers))
    
    branding = await db.platform_settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    platform_settings = await db.platform_settings.find_one({"key": "platform_name"}, {"_id": 0}) or {}
    platform_name = platform_settings.get("value", "مكتب بيت الخبرة للاستشارات الهندسية")
    
    partner_company = branding.get("partner_company_name") or "شركة المياه الوطنية"
    company_name = branding.get("company_name") or platform_name

    # Title rows with company info
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = partner_company
    ws['A1'].font = Font(bold=True, size=18, color="0066CC")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = company_name
    ws['A2'].font = Font(bold=True, size=14, color="006600")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'] = f"تقرير توصيلات المياه - {data.project_name}"
    ws['A3'].font = Font(bold=True, size=16)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'A4:{last_col}4')
    ws['A4'] = f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A4'].font = Font(size=10, italic=True)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    # Header row
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Data rows
    for idx, conn in enumerate(data.connections, 1):
        row = idx + 6
        row_data = [
            idx,
            conn.get('project', ''),
            conn.get('governorate', ''),
            conn.get('contractor', ''),
            conn.get('account_number', ''),
            conn.get('request_number', ''),
            conn.get('restriction_number', ''),
            conn.get('ccb_report_number', ''),
            conn.get('customer_name', ''),
            conn.get('phone_number', ''),
            conn.get('area', ''),
            conn.get('work_order_date', ''),
            conn.get('diameter', ''),
            conn.get('connection_length', ''),
            conn.get('permit_number', ''),
            conn.get('commissioning_date', ''),
            conn.get('publication_date', ''),
            conn.get('issue_date', ''),
            conn.get('expected_execution_date', ''),
            conn.get('connection_type', ''),
            conn.get('connections_count', ''),
            conn.get('connection_length_without_extra', ''),
            conn.get('connections_length_without_main', ''),
            conn.get('network_diameter_63', ''),
            conn.get('network_line_length', ''),
            conn.get('network_diameter_16', ''),
            conn.get('meter_number', ''),
            conn.get('meter_type', ''),
            conn.get('meter_removal_installation', ''),
            conn.get('execution_date', ''),
            conn.get('system_closing_date', ''),
            conn.get('request_status', ''),
            conn.get('cancellation_date', ''),
            conn.get('cancellation_reason', ''),
            conn.get('notes', ''),
            conn.get('created_by_name', ''),
            conn.get('created_at', ''),
            conn.get('latitude', ''),
            conn.get('longitude', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
    
    # Summary row
    summary_row = len(data.connections) + 8
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    ws[f'A{summary_row}'] = f"إجمالي التوصيلات: {len(data.connections)}"
    ws[f'A{summary_row}'].font = Font(bold=True, size=12)
    
    # Adjust column widths dynamically
    col_widths = [5, 25, 15, 15, 12, 12, 12, 12, 20, 12, 15, 12, 8, 12, 12, 12, 12, 12, 12, 15, 10, 15, 15, 12, 12, 12, 12, 12, 15, 12, 12, 10, 12, 15, 20, 15, 15, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=water_connections.xlsx"}
    )


@api_router.post("/water-connections/export/pdf")
async def export_water_connections_pdf(data: ExportRequest, current_user: User = Depends(get_current_user)):
    """تصدير توصيلات المياه إلى PDF مع دعم العربية والشعارات"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from io import BytesIO
    from bidi.algorithm import get_display
    import arabic_reshaper
    
    # تسجيل الخط العربي
    font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'NotoSansArabic-Regular.ttf')
    try:
        pdfmetrics.registerFont(TTFont('Arabic', font_path))
    except:
        try:
            pdfmetrics.registerFont(TTFont('Arabic', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        except:
            pass
    
    def arabic_text(text):
        if not text:
            return ""
        import html
        try:
            if not isinstance(text, str):
                text = str(text)
            # فك تشفير أي نصوص مشفرة أو Unicode Escapes (مثل \u0627)
            if '\\u' in text:
                try: text = text.encode('utf-8').decode('unicode-escape')
                except: pass
            
            # فك تشفير HTML Entities (مثل &amp;)
            text = html.unescape(text)
            
            reshaped = arabic_reshaper.reshape(text)
            return get_display(reshaped)
        except:
            return str(text)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=40)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles with Arabic font
    title_style = ParagraphStyle('ArabicTitle', parent=styles['Title'], fontName='Arabic', fontSize=16, alignment=TA_CENTER, textColor=colors.HexColor('#0066CC'))
    subtitle_style = ParagraphStyle('ArabicSubtitle', parent=styles['Normal'], fontName='Arabic', fontSize=12, alignment=TA_CENTER, textColor=colors.HexColor('#006600'))
    normal_style = ParagraphStyle('ArabicNormal', parent=styles['Normal'], fontName='Arabic', fontSize=10, alignment=TA_CENTER)
    signature_style = ParagraphStyle('Signature', parent=styles['Normal'], fontName='Arabic', fontSize=9, alignment=TA_RIGHT, textColor=colors.HexColor('#333333'))
    
    branding = await db.platform_settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    platform_settings = await db.platform_settings.find_one({"key": "platform_name"}, {"_id": 0}) or {}
    platform_name = platform_settings.get("value", "مكتب بيت الخبرة للاستشارات الهندسية")
    
    partner_company = branding.get("partner_company_name") or "شركة المياه الوطنية"
    company_name = branding.get("company_name") or platform_name

    # Header with company names
    elements.append(Paragraph(arabic_text(partner_company), title_style))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(arabic_text("National Water Company" if partner_company == "شركة المياه الوطنية" else ""), normal_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(arabic_text(company_name), subtitle_style))
    elements.append(Spacer(1, 15))
    elements.append(Paragraph(arabic_text(f"تقرير توصيلات المياه - {data.project_name}"), title_style))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(arabic_text(f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d')}"), normal_style))
    elements.append(Spacer(1, 15))
    
    # Table headers - الحقول الرئيسية
    headers = [arabic_text(h) for h in ['م', 'رقم الطلب', 'رقم CCB', 'العميل', 'الجوال', 'المقاول', 'المنطقة', 'القطر', 'الطول', 'الحالة', 'تاريخ التنفيذ']]
    table_data = [headers]
    
    for idx, conn in enumerate(data.connections, 1):
        row = [
            str(idx),
            str(conn.get('request_number', '') or ''),
            str(conn.get('ccb_report_number', '') or ''),
            arabic_text(conn.get('customer_name', '') or ''),
            str(conn.get('phone_number', '') or ''),
            arabic_text(conn.get('contractor', '') or ''),
            arabic_text(conn.get('area', '') or ''),
            str(conn.get('diameter', '') or ''),
            str(conn.get('connection_length', '') or ''),
            arabic_text(conn.get('request_status', '') or ''),
            str(conn.get('execution_date', '') or '')
        ]
        table_data.append(row)
    
    col_widths = [20, 55, 55, 80, 55, 70, 70, 35, 35, 50, 55]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Summary
    elements.append(Paragraph(arabic_text(f"إجمالي التوصيلات: {len(data.connections)}"), normal_style))
    elements.append(Spacer(1, 30))
    
    # Electronic signature section - تحسين التوقيع الإلكتروني
    # على اليسار: بيت الخبرة | على اليمين: شركة المياه الوطنية
    signature_table_data = [
        [arabic_text(partner_company), '', arabic_text(company_name)],
        [arabic_text("National Water Company" if partner_company == "شركة المياه الوطنية" else ""), '', arabic_text("Bayt Al-Khibra Engineering" if company_name == "مكتب بيت الخبرة للاستشارات الهندسية" else "")],
        ['', '', ''],
        [arabic_text("اسم المعتمد: ________________"), '', arabic_text(f"اسم المسؤول: {current_user.full_name}")],
        [arabic_text("التوقيع: ________________"), '', arabic_text("التوقيع: ________________")],
        [arabic_text(f"التاريخ: {datetime.now().strftime('%Y-%m-%d')}"), '', arabic_text(f"التاريخ: {datetime.now().strftime('%Y-%m-%d')}")]
    ]
    sig_table = Table(signature_table_data, colWidths=[220, 120, 220])
    sig_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('BACKGROUND', (0, 0), (0, 1), colors.HexColor('#E0F0FF')),
        ('BACKGROUND', (2, 0), (2, 1), colors.HexColor('#E0FFE0')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, 1), 10),
        ('FONTSIZE', (0, 2), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 1), (0, 1), 1, colors.HexColor('#0066CC')),
        ('LINEBELOW', (2, 1), (2, 1), 1, colors.HexColor('#006600')),
    ]))
    elements.append(sig_table)
    elements.append(Spacer(1, 15))
    
    # Footer
    elements.append(Paragraph(arabic_text(f"{partner_company} - {company_name}"), signature_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=water_connections.pdf"}
    )


@api_router.post("/sewage-connections/export/excel")
async def export_sewage_connections_excel(data: ExportRequest, current_user: User = Depends(get_current_user)):
    """تصدير توصيلات الصرف الصحي إلى Excel - جميع البيانات"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "توصيلات الصرف الصحي"
    ws.sheet_view.rightToLeft = True
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # جميع الحقول شاملة المحافظة والملاحظات ورقم الجوال وكل التفاصيل المدخلة
    headers = [
        '#', 'المشروع', 'المحافظة', 'المقاولين', 'رقم الطلب', 'رقم الحساب', 'رقم الحصر', 'رقم البلاغ CCB',
        'اسم العميل', 'رقم الجوال', 'رقم العميل', 'المنطقة / الحي', 'تاريخ أمر الشغل', 'القطر', 'رقم العداد',
        'التصريح', 'تاريخ التعميد', 'تاريخ النشر', 'تاريخ الإصدار', 'تاريخ التنفيذ المتوقع',
        'نوع الربط', 'تركيب فتح التهوية', 'تركيب غرفة تفتيش', 'Back Drop',
        'طول على الطبيعة', 'طول خط الشبكة', 'تكسير بيارة', 'هجمة', 'إزالة توصيلة',
        'تاريخ التنفيذ', 'تاريخ الإغلاق', 'الحالة', 'سبب الإلغاء', 'ملاحظات',
        'مراقب الاستشاري', 'تاريخ الإضافة', 'خط العرض', 'خط الطول'
    ]
    
    last_col = get_column_letter(len(headers))
    
    branding = await db.platform_settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    platform_settings = await db.platform_settings.find_one({"key": "platform_name"}, {"_id": 0}) or {}
    platform_name = platform_settings.get("value", "مكتب بيت الخبرة للاستشارات الهندسية")
    
    partner_company = branding.get("partner_company_name") or "شركة المياه الوطنية"
    company_name = branding.get("company_name") or platform_name

    # Title rows with company info
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = partner_company
    ws['A1'].font = Font(bold=True, size=18, color="228B22")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = company_name
    ws['A2'].font = Font(bold=True, size=14, color="006600")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'] = f"تقرير توصيلات الصرف الصحي - {data.project_name}"
    ws['A3'].font = Font(bold=True, size=16)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells(f'A4:{last_col}4')
    ws['A4'] = f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A4'].font = Font(size=10, italic=True)
    ws['A4'].alignment = Alignment(horizontal='center')
    
    header_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    for idx, conn in enumerate(data.connections, 1):
        row = idx + 6
        contractors = ', '.join(conn.get('contractors', [])) if isinstance(conn.get('contractors'), list) else str(conn.get('contractors', ''))
        
        row_data = [
            idx,
            conn.get('project', ''),
            conn.get('governorate', ''),
            contractors,
            conn.get('request_number', ''),
            conn.get('account_number', ''),
            conn.get('restriction_number', ''),
            conn.get('ccb_report_number', ''),
            conn.get('customer_name', ''),
            conn.get('phone_number', ''),
            conn.get('customer_number', ''),
            conn.get('area', ''),
            conn.get('work_order_date', ''),
            conn.get('diameter', ''),
            conn.get('meter_number', ''),
            conn.get('permit', ''),
            conn.get('commissioning_date', ''),
            conn.get('publication_date', ''),
            conn.get('issue_date', ''),
            conn.get('expected_execution_date', ''),
            conn.get('connection_type', ''),
            'نعم' if conn.get('ventilation_installation') else 'لا',
            'نعم' if conn.get('inspection_room_installation') else 'لا',
            'نعم' if conn.get('back_drop') else 'لا',
            conn.get('actual_length', ''),
            conn.get('network_line_length', ''),
            'نعم' if conn.get('cesspool_breaking') else 'لا',
            'نعم' if conn.get('attack') else 'لا',
            'نعم' if conn.get('connection_removal') else 'لا',
            conn.get('execution_date', ''),
            conn.get('system_closing_date', ''),
            conn.get('request_status', ''),
            conn.get('cancellation_reason', ''),
            conn.get('notes', ''),
            conn.get('created_by_name', ''),
            conn.get('created_at', ''),
            conn.get('latitude', ''),
            conn.get('longitude', '')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
    
    # Summary row
    summary_row = len(data.connections) + 8
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    ws[f'A{summary_row}'] = f"إجمالي التوصيلات: {len(data.connections)}"
    ws[f'A{summary_row}'].font = Font(bold=True, size=12)
    
    # Adjust column widths dynamically
    col_widths = [5, 25, 15, 20, 12, 12, 12, 12, 20, 12, 12, 15, 12, 8, 12, 12, 12, 12, 12, 12, 15, 10, 10, 10, 12, 12, 10, 8, 10, 12, 12, 10, 15, 20, 15, 15, 12, 12]
    for i, width in enumerate(col_widths, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=sewage_connections.xlsx"}
    )


@api_router.post("/sewage-connections/export/pdf")
async def export_sewage_connections_pdf(data: ExportRequest, current_user: User = Depends(get_current_user)):
    """تصدير توصيلات الصرف الصحي إلى PDF مع دعم العربية والشعارات"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from io import BytesIO
    from bidi.algorithm import get_display
    import arabic_reshaper
    
    # تسجيل الخط العربي
    font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'NotoSansArabic-Regular.ttf')
    try:
        pdfmetrics.registerFont(TTFont('Arabic', font_path))
    except:
        try:
            pdfmetrics.registerFont(TTFont('Arabic', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        except:
            pass
    
    def arabic_text(text):
        if not text:
            return ""
        import html
        try:
            if not isinstance(text, str):
                text = str(text)
            # فك تشفير أي نصوص مشفرة أو Unicode Escapes (مثل \u0627)
            if '\\u' in text:
                try: text = text.encode('utf-8').decode('unicode-escape')
                except: pass
            
            # فك تشفير HTML Entities (مثل &amp;)
            text = html.unescape(text)
            
            reshaped = arabic_reshaper.reshape(text)
            return get_display(reshaped)
        except:
            return str(text)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=40)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles with Arabic font
    title_style = ParagraphStyle('ArabicTitle', parent=styles['Title'], fontName='Arabic', fontSize=16, alignment=TA_CENTER, textColor=colors.HexColor('#228B22'))
    subtitle_style = ParagraphStyle('ArabicSubtitle', parent=styles['Normal'], fontName='Arabic', fontSize=12, alignment=TA_CENTER, textColor=colors.HexColor('#006600'))
    normal_style = ParagraphStyle('ArabicNormal', parent=styles['Normal'], fontName='Arabic', fontSize=10, alignment=TA_CENTER)
    signature_style = ParagraphStyle('Signature', parent=styles['Normal'], fontName='Arabic', fontSize=9, alignment=TA_RIGHT, textColor=colors.HexColor('#333333'))
    
    branding = await db.platform_settings.find_one({"key": "branding"}, {"_id": 0}) or {}
    platform_settings = await db.platform_settings.find_one({"key": "platform_name"}, {"_id": 0}) or {}
    platform_name = platform_settings.get("value", "مكتب بيت الخبرة للاستشارات الهندسية")
    
    partner_company = branding.get("partner_company_name") or "شركة المياه الوطنية"
    company_name = branding.get("company_name") or platform_name

    # Header with company names
    elements.append(Paragraph(arabic_text(partner_company), title_style))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(arabic_text("National Water Company" if partner_company == "شركة المياه الوطنية" else ""), normal_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(arabic_text(company_name), subtitle_style))
    elements.append(Spacer(1, 15))
    elements.append(Paragraph(arabic_text(f"تقرير توصيلات الصرف الصحي - {data.project_name}"), title_style))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(arabic_text(f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d')}"), normal_style))
    elements.append(Spacer(1, 15))
    
    # Table headers
    headers = [arabic_text(h) for h in ['م', 'رقم الطلب', 'رقم CCB', 'العميل', 'رقم العميل', 'المقاولين', 'المنطقة', 'نوع الربط', 'الحالة', 'تاريخ التنفيذ']]
    table_data = [headers]
    
    for idx, conn in enumerate(data.connections, 1):
        contractors = ', '.join(conn.get('contractors', [])) if isinstance(conn.get('contractors'), list) else str(conn.get('contractors', ''))
        row = [
            str(idx),
            str(conn.get('request_number', '') or ''),
            str(conn.get('ccb_report_number', '') or ''),
            arabic_text(conn.get('customer_name', '') or ''),
            str(conn.get('customer_number', '') or ''),
            arabic_text(contractors),
            arabic_text(conn.get('area', '') or ''),
            arabic_text(conn.get('connection_type', '') or ''),
            arabic_text(conn.get('request_status', '') or ''),
            str(conn.get('execution_date', '') or '')
        ]
        table_data.append(row)
    
    col_widths = [20, 55, 55, 75, 50, 75, 65, 55, 50, 55]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#228B22')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Summary
    elements.append(Paragraph(arabic_text(f"إجمالي التوصيلات: {len(data.connections)}"), normal_style))
    elements.append(Spacer(1, 30))
    
    # Electronic signature section - تحسين التوقيع الإلكتروني
    # على اليسار: بيت الخبرة | على اليمين: شركة المياه الوطنية
    signature_table_data = [
        [arabic_text(partner_company), '', arabic_text(company_name)],
        [arabic_text("National Water Company" if partner_company == "شركة المياه الوطنية" else ""), '', arabic_text("Bayt Al-Khibra Engineering" if company_name == "مكتب بيت الخبرة للاستشارات الهندسية" else "")],
        ['', '', ''],
        [arabic_text("اسم المعتمد: ________________"), '', arabic_text(f"اسم المسؤول: {current_user.full_name}")],
        [arabic_text("التوقيع: ________________"), '', arabic_text("التوقيع: ________________")],
        [arabic_text(f"التاريخ: {datetime.now().strftime('%Y-%m-%d')}"), '', arabic_text(f"التاريخ: {datetime.now().strftime('%Y-%m-%d')}")]
    ]
    sig_table = Table(signature_table_data, colWidths=[220, 120, 220])
    sig_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('BACKGROUND', (0, 0), (0, 1), colors.HexColor('#E0F0FF')),
        ('BACKGROUND', (2, 0), (2, 1), colors.HexColor('#E0FFE0')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, 1), 10),
        ('FONTSIZE', (0, 2), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 1), (0, 1), 1, colors.HexColor('#0066CC')),
        ('LINEBELOW', (2, 1), (2, 1), 1, colors.HexColor('#006600')),
    ]))
    elements.append(sig_table)
    elements.append(Spacer(1, 15))
    
    # Footer
    elements.append(Paragraph(arabic_text(f"{partner_company} - {company_name}"), signature_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=sewage_connections.pdf"}
    )


# ============= DATA MIGRATION =============

@api_router.post("/admin/migrate-data")
async def migrate_old_data(current_user: User = Depends(get_current_admin_user)):
    """تحديث البيانات القديمة لإضافة حقل المشروع"""
    
    # تحديث جميع البلاغات التي ليس لديها حقل project
    result = await db.reports.update_many(
        {"project": {"$exists": False}},
        {"$set": {"project": "مشروع إصلاح أعمال المحافظات الغربية - القطاع الأوسط"}}
    )
    
    # تحديث جميع المستخدمين الذين ليس لديهم حقل projects
    user_result = await db.users.update_many(
        {"projects": {"$exists": False}},
        {"$set": {"projects": []}}
    )
    
    return {
        "message": "تم تحديث البيانات بنجاح",
        "reports_updated": result.modified_count,
        "users_updated": user_result.modified_count
    }


# ============= AI ANALYSIS =============

@api_router.post("/reports/analyze")
async def analyze_reports(current_user: User = Depends(get_current_user)):
    query = {"is_deleted": {"$ne": True}}
    
    # فلترة حسب صلاحيات المحافظات (إلا إذا كان admin بدون محافظات محددة)
    if current_user.role != "admin" or len(current_user.governorates) > 0:
        if len(current_user.governorates) > 0:
            query.update(get_flexible_in_query(current_user.governorates, "governorate"))
        
        # تطبيق الفلترة الهرمية للتحليل
        hierarchy_filter = await get_hierarchy_filter(current_user)
        query.update(hierarchy_filter)
        
        # تطبيق فلترة المشاريع بمرونة
        if current_user.projects:
            query.update(get_flexible_in_query(current_user.projects, "project"))
    
    reports = await db.reports.find(query, {"_id": 0}).to_list(1000)
    
    if not reports:
        return {"analysis": "لا توجد بلاغات لتحليلها"}
    
    report_summary = f"عدد البلاغات: {len(reports)}\n\n"
    
    status_count = {}
    for report in reports:
        status = report['status']
        status_count[status] = status_count.get(status, 0) + 1
    
    report_summary += "حسب الحالة:\n"
    for status, count in status_count.items():
        report_summary += f"- {status}: {count}\n"
    
    type_count = {}
    for report in reports:
        rtype = report['report_type']
        type_count[rtype] = type_count.get(rtype, 0) + 1
    
    report_summary += "\nحسب النوع:\n"
    for rtype, count in type_count.items():
        report_summary += f"- {rtype}: {count}\n"
    
    gov_count = {}
    for report in reports:
        gov = report['governorate']
        gov_count[gov] = gov_count.get(gov, 0) + 1
    
    report_summary += "\nحسب المحافظة:\n"
    for gov, count in gov_count.items():
        report_summary += f"- {gov}: {count}\n"
    
    contractor_count = {}
    for report in reports:
        contractor = report['contractor']
        contractor_count[contractor] = contractor_count.get(contractor, 0) + 1
    
    report_summary += "\nحسب المقاول:\n"
    for contractor, count in contractor_count.items():
        report_summary += f"- {contractor}: {count}\n"
    
    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=str(uuid.uuid4()),
            system_message="أنت مساعد ذكي متخصص في تحليل بيانات البلاغات. قدم تحليلاً شاملاً ونصائح عملية بناءً على البيانات المقدمة."
        )
        chat.with_model("openai", "gpt-4o-mini")
        
        user_message = UserMessage(
            text=f"قم بتحليل البيانات التالية للبلاغات وقدم رؤى وتوصيات:\n\n{report_summary}\n\nأريد تحليلاً يشمل: الأنماط، المشاكل المحتملة، التوصيات لتحسين الأداء."
        )
        
        response = await chat.send_message(user_message)
        
        return {
            "analysis": response,
            "summary": report_summary
        }
    except Exception as e:
        return {
            "analysis": "حدث خطأ في التحليل الذكي",
            "summary": report_summary,
            "error": str(e)
        }


@api_router.get("/")
async def root():
    return {"message": "نظام إدارة البلاغات المستلمة على WFM - API"}


@api_router.get("/health")
async def health_check():
    return {"status": "healthy"}


# ============= تنبيهات السيارات =============
@api_router.get("/fleet-cars/alerts")
async def get_fleet_car_alerts(current_user: User = Depends(get_current_user)):
    """جلب تنبيهات انتهاء الاستمارة/الفحص/التفويض"""
    user_perms = current_user.permissions or []
    has_perm = "cars_manage" in user_perms or "fleet_maintenance" in user_perms or "cars" in user_perms
    if not has_perm and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    alerts = []
    today = datetime.now(timezone.utc).date()
    one_month = timedelta(days=30)
    one_week = timedelta(days=7)
    
    cars = await db.fleet_cars.find({}, {"_id": 0}).to_list(1000)
    
    for car in cars:
        car_info = f"{car.get('car_type', '')} - {car.get('plate_number', '')}"
        
        # فحص الاستمارة
        if car.get('registration_end'):
            try:
                reg_end = datetime.fromisoformat(car['registration_end'].replace('Z', '+00:00')).date() if 'T' in car['registration_end'] else datetime.strptime(car['registration_end'], '%Y-%m-%d').date()
                diff = (reg_end - today).days
                if diff < 0:
                    alerts.append({
                        "type": "registration",
                        "car_id": car.get('id'),
                        "car_info": car_info,
                        "message": f"انتهت الاستمارة منذ {abs(diff)} يوم",
                        "date": car['registration_end'],
                        "priority": "expired",
                        "icon": "📋"
                    })
                elif diff <= 7:
                    alerts.append({
                        "type": "registration",
                        "car_id": car.get('id'),
                        "car_info": car_info,
                        "message": f"تنتهي الاستمارة خلال {diff} يوم",
                        "date": car['registration_end'],
                        "priority": "urgent",
                        "icon": "📋"
                    })
                elif diff <= 30:
                    alerts.append({
                        "type": "registration",
                        "car_id": car.get('id'),
                        "car_info": car_info,
                        "message": f"تنتهي الاستمارة خلال {diff} يوم",
                        "date": car['registration_end'],
                        "priority": "warning",
                        "icon": "📋"
                    })
            except: pass
        
        # فحص الفحص الدوري
        if car.get('inspection_end'):
            try:
                insp_end = datetime.fromisoformat(car['inspection_end'].replace('Z', '+00:00')).date() if 'T' in car['inspection_end'] else datetime.strptime(car['inspection_end'], '%Y-%m-%d').date()
                diff = (insp_end - today).days
                if diff < 0:
                    alerts.append({
                        "type": "inspection",
                        "car_id": car.get('id'),
                        "car_info": car_info,
                        "message": f"انتهى الفحص الدوري منذ {abs(diff)} يوم",
                        "date": car['inspection_end'],
                        "priority": "expired",
                        "icon": "🔧"
                    })
                elif diff <= 7:
                    alerts.append({
                        "type": "inspection",
                        "car_id": car.get('id'),
                        "car_info": car_info,
                        "message": f"ينتهي الفحص الدوري خلال {diff} يوم",
                        "date": car['inspection_end'],
                        "priority": "urgent",
                        "icon": "🔧"
                    })
                elif diff <= 30:
                    alerts.append({
                        "type": "inspection",
                        "car_id": car.get('id'),
                        "car_info": car_info,
                        "message": f"ينتهي الفحص الدوري خلال {diff} يوم",
                        "date": car['inspection_end'],
                        "priority": "warning",
                        "icon": "🔧"
                    })
            except: pass
        
        # فحص التفويض
        if car.get('authorization_end'):
            try:
                auth_end = datetime.fromisoformat(car['authorization_end'].replace('Z', '+00:00')).date() if 'T' in car['authorization_end'] else datetime.strptime(car['authorization_end'], '%Y-%m-%d').date()
                diff = (auth_end - today).days
                if diff < 0:
                    alerts.append({
                        "type": "authorization",
                        "car_id": car.get('id'),
                        "car_info": car_info,
                        "message": f"انتهى التفويض منذ {abs(diff)} يوم",
                        "date": car['authorization_end'],
                        "priority": "expired",
                        "icon": "📝"
                    })
                elif diff <= 7:
                    alerts.append({
                        "type": "authorization",
                        "car_id": car.get('id'),
                        "car_info": car_info,
                        "message": f"ينتهي التفويض خلال {diff} يوم",
                        "date": car['authorization_end'],
                        "priority": "urgent",
                        "icon": "📝"
                    })
                elif diff <= 30:
                    alerts.append({
                        "type": "authorization",
                        "car_id": car.get('id'),
                        "car_info": car_info,
                        "message": f"ينتهي التفويض خلال {diff} يوم",
                        "date": car['authorization_end'],
                        "priority": "warning",
                        "icon": "📝"
                    })
            except: pass
    
    # ترتيب التنبيهات حسب الأولوية
    priority_order = {"expired": 0, "urgent": 1, "warning": 2}
    alerts.sort(key=lambda x: priority_order.get(x['priority'], 3))
    
    return alerts


# ============= CHAT WEBSOCKET & ENDPOINTS =============

# WebSocket endpoint للدردشة الفورية
@app.websocket("/api/ws/chat/{user_id}")
async def websocket_endpoint(websocket: WebSocket, user_id: str, token: str = None):
    # التحقق من التوكن
    if not token:
        await websocket.close(code=1008, reason="Missing authentication token")
        return
    
    try:
        # التحقق من صحة التوكن
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        token_user_id = payload.get("sub")
        
        # التأكد من أن المستخدم يتصل بـ WebSocket الخاص به فقط
        if token_user_id != user_id:
            await websocket.close(code=1008, reason="Unauthorized")
            return
            
    except JWTError:
        await websocket.close(code=1008, reason="Invalid token")
        return
    
    await manager.connect(websocket, user_id)
    try:
        while True:
            data = await websocket.receive_json()
            # إرسال الرسالة للمستقبل
            if data.get('type') == 'message':
                await manager.send_personal_message(data, data['receiver_id'])
            elif data.get('type') == 'delivered':
                await manager.send_personal_message(data, data['sender_id'])
            elif data.get('type') == 'read':
                await manager.send_personal_message(data, data['sender_id'])
    except WebSocketDisconnect:
        manager.disconnect(user_id)


# إرسال رسالة جديدة
@api_router.post("/chat/messages")
async def send_message(
    receiver_id: str = Form(...),
    message: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user)
):
    file_url = None
    file_type = None
    
    # رفع الملف إن وجد
    if file:
        file_ext = file.filename.split('.')[-1].lower()
        file_type = 'image' if file_ext in ['jpg', 'jpeg', 'png', 'gif'] else 'audio'
        file_name = f"chat_{current_user.id}_{int(time.time())}.{file_ext}"
        file_path = Path(f"/app/backend/uploads/chat/{file_name}")
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(file_path, "wb") as f:
            f.write(await file.read())
        
        file_url = f"/api/uploads/chat/{file_name}"
    
    # حفظ الرسالة في قاعدة البيانات
    msg = Message(
        sender_id=current_user.id,
        receiver_id=receiver_id,
        message=message,
        file_url=file_url,
        file_type=file_type
    )
    
    doc = msg.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.messages.insert_one(doc)
    
    # إرسال عبر WebSocket للمستقبل
    await manager.send_personal_message({
        'type': 'new_message',
        'message': doc
    }, receiver_id)
    
    return {"success": True, "message_id": msg.id}


# جلب المحادثات (قائمة المستخدمين الذين تواصلت معهم)
@api_router.get("/chat/conversations")
async def get_conversations(current_user: User = Depends(get_current_user)):
    # جلب جميع المستخدمين الذين أرسلت لهم أو استقبلت منهم رسائل
    messages = await db.messages.find({
        "$or": [
            {"sender_id": current_user.id},
            {"receiver_id": current_user.id}
        ]
    }, {"_id": 0}).to_list(10000)
    
    # استخراج user_ids الفريدة
    user_ids = set()
    for msg in messages:
        if msg['sender_id'] != current_user.id:
            user_ids.add(msg['sender_id'])
        if msg['receiver_id'] != current_user.id:
            user_ids.add(msg['receiver_id'])
    
    # جلب بيانات المستخدمين
    users = []
    for user_id in user_ids:
        user = await db.users.find_one({"id": user_id}, {"_id": 0, "hashed_password": 0})
        if user:
            # جلب آخر رسالة
            last_msg = await db.messages.find_one(
                {
                    "$or": [
                        {"sender_id": current_user.id, "receiver_id": user_id},
                        {"sender_id": user_id, "receiver_id": current_user.id}
                    ]
                },
                {"_id": 0},
                sort=[("created_at", -1)]
            )
            
            # عد الرسائل غير المقروءة
            unread_count = await db.messages.count_documents({
                "sender_id": user_id,
                "receiver_id": current_user.id,
                "is_read": False
            })
            
            users.append({
                **user,
                "last_message": last_msg,
                "unread_count": unread_count
            })
    
    # ترتيب حسب آخر رسالة
    users.sort(key=lambda x: x['last_message']['created_at'] if x.get('last_message') else '', reverse=True)
    
    return users


# جلب الرسائل مع مستخدم معين
@api_router.get("/chat/messages/{other_user_id}")
async def get_messages(other_user_id: str, current_user: User = Depends(get_current_user)):
    messages = await db.messages.find({
        "$or": [
            {"sender_id": current_user.id, "receiver_id": other_user_id},
            {"sender_id": other_user_id, "receiver_id": current_user.id}
        ]
    }, {"_id": 0}).sort("created_at", 1).to_list(1000)
    
    # تحديث الرسائل كـ مقروءة
    await db.messages.update_many(
        {"sender_id": other_user_id, "receiver_id": current_user.id, "is_read": False},
        {"$set": {"is_read": True, "read_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # إرسال إشعار بالقراءة
    await manager.send_personal_message({
        'type': 'messages_read',
        'reader_id': current_user.id
    }, other_user_id)
    
    return messages


# تعديل رسالة (خلال 10 دقائق)
@api_router.put("/chat/messages/{message_id}")
async def edit_message(
    message_id: str,
    new_message: str = Form(...),
    current_user: User = Depends(get_current_user)
):
    msg = await db.messages.find_one({"id": message_id, "sender_id": current_user.id}, {"_id": 0})
    
    if not msg:
        raise HTTPException(404, "الرسالة غير موجودة")
    
    # التحقق من مرور 10 دقائق
    created_at = datetime.fromisoformat(msg['created_at'])
    now = datetime.now(timezone.utc)
    if (now - created_at).total_seconds() > 600:  # 10 دقائق
        raise HTTPException(400, "انتهت مدة التعديل (10 دقائق)")
    
    # تحديث الرسالة
    await db.messages.update_one(
        {"id": message_id},
        {"$set": {
            "message": new_message,
            "is_edited": True,
            "edited_at": now.isoformat()
        }}
    )
    
    # إرسال التحديث عبر WebSocket
    await manager.send_personal_message({
        'type': 'message_edited',
        'message_id': message_id,
        'new_message': new_message
    }, msg['receiver_id'])
    
    return {"success": True}


# حذف رسالة
@api_router.delete("/chat/messages/{message_id}")
async def delete_message(message_id: str, current_user: User = Depends(get_current_user)):
    result = await db.messages.delete_one({"id": message_id, "sender_id": current_user['id']})
    
    if result.deleted_count == 0:
        raise HTTPException(404, "الرسالة غير موجودة")
    
    return {"success": True}


# جلب قائمة جميع المستخدمين لبدء محادثة جديدة (حسب المشروع)
@api_router.get("/chat/users")
async def get_all_users(current_user: User = Depends(get_current_user)):
    # جلب مشاريع المستخدم الحالي
    user_projects = current_user.projects or []
    
    # إذا كان admin بدون مشاريع محددة، يرى الجميع
    if current_user.role == 'admin' and not user_projects:
        users = await db.users.find(
            {"id": {"$ne": current_user.id}},
            {"_id": 0, "hashed_password": 0}
        ).to_list(1000)
    else:
        # فلترة المستخدمين الذين لديهم مشروع مشترك
        users = await db.users.find(
            {
                "id": {"$ne": current_user.id},
                "$or": [
                    {"projects": {"$in": user_projects}},
                    {"projects": {"$size": 0}}  # المستخدمين بدون مشاريع محددة
                ]
            },
            {"_id": 0, "hashed_password": 0}
        ).to_list(1000)
    
    # إضافة معلومات المشاريع المشتركة لكل مستخدم
    for user in users:
        user_shared_projects = []
        if user.get('projects'):
            user_shared_projects = list(set(user['projects']) & set(user_projects))
        user['shared_projects'] = user_shared_projects
    
    return users


# ============= نظام دردشة بسيط =============

# جلب الرسائل مع مستخدم (آخر 100 رسالة فقط للأداء)
@api_router.get("/chat/simple/messages/{other_user_id}")
async def get_simple_messages(other_user_id: str, current_user: User = Depends(get_current_user)):
    messages = await db.simple_messages.find({
        "$or": [
            {"sender_id": current_user.id, "receiver_id": other_user_id},
            {"sender_id": other_user_id, "receiver_id": current_user.id}
        ]
    }, {"_id": 0}).sort("created_at", -1).limit(100).to_list(100)
    
    # عكس الترتيب لعرضها من الأقدم للأحدث
    messages.reverse()
    
    # تحديث الرسائل كمقروءة
    await db.simple_messages.update_many(
        {"sender_id": other_user_id, "receiver_id": current_user.id, "is_read": False},
        {"$set": {"is_read": True}}
    )
    
    return messages


# إرسال رسالة
@api_router.post("/chat/simple/send")
async def send_simple_message(
    request: dict,
    current_user: User = Depends(get_current_user)
):
    receiver_id = request.get("receiver_id")
    message = request.get("message")
    
    if not receiver_id or not message:
        raise HTTPException(status_code=400, detail="receiver_id and message are required")
    
    msg = {
        "id": str(uuid.uuid4()),
        "sender_id": current_user.id,
        "receiver_id": receiver_id,
        "message": message,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "is_read": False
    }
    
    await db.simple_messages.insert_one(msg)
    
    return {"success": True, "message_id": msg["id"]}


# عدد الرسائل غير المقروءة
@api_router.get("/chat/unread-count")
async def get_unread_count(current_user: User = Depends(get_current_user)):
    count = await db.simple_messages.count_documents({
        "receiver_id": current_user.id,
        "is_read": False
    })
    return {"count": count}


@api_router.get("/health")
async def health_check():
    return {"status": "healthy"}

# ============= SUPPORT MESSAGES =============

@api_router.post("/support/messages", response_model=SupportMessageResponse)
async def create_support_message(message_data: SupportMessageCreate):
    """إنشاء رسالة دعم جديدة"""
    try:
        message = SupportMessage(
            name=message_data.name,
            email=message_data.email,
            message=message_data.message
        )
        await db.support_messages.insert_one(message.model_dump())
        return SupportMessageResponse(**message.model_dump())
    except Exception as e:
        logger.error(f"Error creating support message: {str(e)}")
        raise HTTPException(status_code=500, detail="حدث خطأ في إرسال الرسالة")


@api_router.get("/support/messages", response_model=List[SupportMessageResponse])
async def get_support_messages(
    status: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    """جلب رسائل الدعم - حسب الصلاحيات"""
    try:
        user_permissions = current_user.permissions or []
        has_support_perm = ("support_messages" in user_permissions) or user_has_any_project_permission(current_user, "support_messages")
        
        if current_user.role != "admin" and not has_support_perm:
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        
        query = {}
        if status:
            query["status"] = status
        
        messages = await db.support_messages.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
        for msg in messages:
            if isinstance(msg.get('created_at'), str):
                msg['created_at'] = datetime.fromisoformat(msg['created_at'])
        return [SupportMessageResponse(**msg) for msg in messages]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail="حدث خطأ")


@api_router.get("/support/messages/count")
async def get_support_messages_count(current_user: User = Depends(get_current_user)):
    """عدد الرسائل الجديدة"""
    try:
        user_permissions = current_user.permissions or []
        has_support_perm = ("support_messages" in user_permissions) or user_has_any_project_permission(current_user, "support_messages")
        if current_user.role != "admin" and not has_support_perm:
            return {"count": 0}
        
        query = {"status": "جديدة"}
        count = await db.support_messages.count_documents(query)
        return {"count": count}
    except:
        return {"count": 0}


@api_router.get("/support/ticket-status")
async def get_public_ticket_status(email: str = Query(...)):
    """التحقق من حالة تذكرة الدعم بالإيميل - بدون تسجيل دخول"""
    try:
        tickets = await db.support_messages.find(
            {"email": email},
            {"_id": 0, "id": 1, "status": 1, "message": 1, "created_at": 1}
        ).sort("created_at", -1).to_list(10)
        
        if not tickets:
            return {"has_tickets": False, "tickets": []}
        
        return {
            "has_tickets": True,
            "tickets": tickets
        }
    except:
        return {"has_tickets": False, "tickets": []}


@api_router.patch("/support/messages/{message_id}/status")
async def update_support_message_status(
    message_id: str,
    status: str = Query(...),
    current_user: User = Depends(get_current_user)
):
    """تحديث حالة الرسالة"""
    try:
        user_permissions = current_user.permissions or []
        has_support_perm = ("support_messages" in user_permissions) or user_has_any_project_permission(current_user, "support_messages")
        if current_user.role != "admin" and not has_support_perm:
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        update_data = {"status": status}
        if status == "تم الحل":
            update_data["resolved_by"] = current_user.username
            update_data["resolved_at"] = datetime.now(timezone.utc)
        result = await db.support_messages.update_one({"id": message_id}, {"$set": update_data})
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="غير موجودة")
        return {"message": "تم التحديث"}
    except HTTPException:
        raise
    except:
        raise HTTPException(status_code=500, detail="خطأ")

@api_router.delete("/support/messages/{message_id}")
async def delete_support_message(
    message_id: str,
    current_user: User = Depends(get_current_user)
):
    """حذف رسالة دعم"""
    try:
        user_permissions = current_user.permissions or []
        has_support_perm = ("support_messages" in user_permissions) or user_has_any_project_permission(current_user, "support_messages")
        if current_user.role != "admin" and not has_support_perm:
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        result = await db.support_messages.delete_one({"id": message_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="الرسالة غير موجودة")
        return {"message": "تم حذف الرسالة بنجاح"}
    except HTTPException:
        raise
    except:
        raise HTTPException(status_code=500, detail="خطأ في حذف الرسالة")


# ============= PLATFORM SETTINGS ENDPOINTS =============

@api_router.get("/settings/platform")
async def get_platform_settings():
    """جلب إعدادات المنصة (اسم المنصة) - متاح للجميع"""
    try:
        settings = await db.platform_settings.find_one({"key": "platform_name"}, {"_id": 0})
        theme_settings = await db.platform_settings.find_one({"key": "platform_theme"}, {"_id": 0})
        branding = await db.platform_settings.find_one({"key": "branding"}, {"_id": 0}) or {}
        if not settings:
            settings = {"value": "بيت الخبرة", "logo_url": ""}
        
        # القيم الافتراضية للنصوص
        default_company_description = "شركة بيت الخبرة للاستشارات الهندسية هي شركة سعودية متخصصة في تقديم الاستشارات والدراسات الهندسية وإدارة المشاريع، تسعى لتقديم حلول متكاملة وفق أعلى معايير الجودة العالمية."
        default_login_pm_description = "يدير فريق المشاريع المحترف المهندس أحمد عبيدات - مدير عام المشاريع، الذي يتميز بخبرته الواسعة في قيادة وتنسيق المشاريع المعقدة، وضمان تنفيذها وفق أعلى معايير الجودة."
        default_login_pc_description = "يعمل إلى جانبه الأستاذ أحمد حافظ، منسق المشاريع، وهو متخصص في إدارة وتنسيق جميع جوانب المشروع بدقة واحترافية لضمان سير العمل بسلاسة وفعالية."
        default_login_team_description = "يتكون فريق بيت الخبرة من مجموعة من المهندسين والاستشاريين المؤهلين، الذين يسعون دائمًا لتقديم حلول مبتكرة ومتكاملة في جميع مراحل المشروع، مع الحرص على التواصل المستمر مع شركائنا لضمان نجاح كل مشروع."
        
        return {
            "platform_name": settings.get("value", "بيت الخبرة"), 
            "logo_url": settings.get("logo_url", ""),
            "theme": theme_settings.get("theme", "blue") if theme_settings else "blue",
            "dark_mode": theme_settings.get("dark_mode", False) if theme_settings else False,
            # معلومات العلامة التجارية (قابلة للتحرير من الإعدادات)
            "company_name": branding.get("company_name", "شركة بيت الخبرة للإستشارات الهندسية"),
            "company_description": branding.get("company_description") or default_company_description,
            "project_manager_name": branding.get("project_manager_name", "المهندس أحمد عبيدات"),
            "project_manager_title": branding.get("project_manager_title", "مدير عام المشاريع"),
            "project_coordinator_name": branding.get("project_coordinator_name", "الأستاذ أحمد حافظ"),
            "project_coordinator_title": branding.get("project_coordinator_title", "منسق المشاريع"),
            "consultant_name": branding.get("consultant_name", "مكتب بيت الخبرة للاستشارات الهندسية"),
            "partner_company_name": branding.get("partner_company_name", "شركة المياة الوطنية"),
            "company_logo_url": branding.get("company_logo_url", "/bayt-alkhibra-logo.png"),
            "partner_logo_url": branding.get("partner_logo_url", "/nwc-logo.png"),
            # النصوص الثلاثة في صفحة الدخول - قابلة للتحرير بشكل منفصل
            "login_pm_description": branding.get("login_pm_description") or default_login_pm_description,
            "login_pc_description": branding.get("login_pc_description") or default_login_pc_description,
            "login_team_description": branding.get("login_team_description") or default_login_team_description,
            "login_description": branding.get("login_description") or default_login_team_description,  # للتوافق مع النسخة القديمة
            "copyright_text": branding.get("copyright_text", "جميع الحقوق محفوظة"),
            "dashboard_title": branding.get("dashboard_title", "نظام إدارة البلاغات والمشاريع - WFM"),
            "footer_year": branding.get("footer_year", "2026"),
            "login_footer_description": branding.get("login_footer_description", "نظام إدارة البلاغات المستلمة من WFM لربط المكاتب الاستشارية مع شركة المياه الوطنية."),
            "internal_footer_description": branding.get("internal_footer_description", "نظام إدارة البلاغات المستلمة من WFM"),
            "global_announcement": branding.get("global_announcement", ""),
            "show_announcement": branding.get("show_announcement", False),
            "flash_announcement": branding.get("flash_announcement", True),
            "vision_logo_url": branding.get("vision_logo_url", ""),
            "occasion_watermark": branding.get("occasion_watermark", "none"),
            "custom_occasion_text_ar": branding.get("custom_occasion_text_ar", ""),
            "custom_occasion_text_en": branding.get("custom_occasion_text_en", ""),
            "occasions_list": branding.get("occasions_list", []),
        }
    except Exception:
        return {"platform_name": "بيت الخبرة", "logo_url": "", "theme": "blue", "dark_mode": False}


class BrandingUpdate(BaseModel):
    company_name: Optional[str] = None
    company_description: Optional[str] = None
    project_manager_name: Optional[str] = None
    project_manager_title: Optional[str] = None
    project_coordinator_name: Optional[str] = None
    project_coordinator_title: Optional[str] = None
    consultant_name: Optional[str] = None
    partner_company_name: Optional[str] = None
    company_logo_url: Optional[str] = None
    partner_logo_url: Optional[str] = None
    login_description: Optional[str] = None
    login_pm_description: Optional[str] = None
    login_pc_description: Optional[str] = None
    login_team_description: Optional[str] = None
    copyright_text: Optional[str] = None
    dashboard_title: Optional[str] = None
    footer_year: Optional[str] = None
    login_footer_description: Optional[str] = None
    internal_footer_description: Optional[str] = None
    global_announcement: Optional[str] = None
    show_announcement: Optional[bool] = None
    flash_announcement: Optional[bool] = None
    vision_logo_url: Optional[str] = None
    occasion_watermark: Optional[str] = None
    custom_occasion_text_ar: Optional[str] = None
    custom_occasion_text_en: Optional[str] = None
    occasions_list: Optional[list] = None


@api_router.put("/settings/branding")
async def update_branding_settings(
    data: BrandingUpdate,
    current_user: User = Depends(get_current_user)
):
    """تحديث إعدادات العلامة التجارية (الأدمن فقط)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="هذه الصلاحية متاحة للأدمن فقط")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user.id}
    payload = data.model_dump(exclude_none=True)
    update_data.update(payload)
    
    await db.platform_settings.update_one(
        {"key": "branding"},
        {"$set": update_data},
        upsert=True
    )
    return {"success": True, "message": "تم تحديث إعدادات العلامة التجارية بنجاح"}


@api_router.put("/settings/platform")
async def update_platform_settings(
    platform_name: str = None,
    logo_url: str = None,
    current_user: User = Depends(get_current_user)
):
    """تحديث إعدادات المنصة (الأدمن فقط)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="هذه الصلاحية متاحة للأدمن فقط")
    
    try:
        update_data = {"updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user.id}
        if platform_name is not None:
            update_data["value"] = platform_name
        if logo_url is not None:
            update_data["logo_url"] = logo_url
        
        await db.platform_settings.update_one(
            {"key": "platform_name"},
            {"$set": update_data},
            upsert=True
        )
        return {"success": True, "message": "تم تحديث الإعدادات بنجاح"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل في تحديث الإعدادات: {str(e)}")


class ThemeUpdate(BaseModel):
    theme: str
    dark_mode: bool = False

class PersonalThemeUpdate(BaseModel):
    personal_theme: Optional[str] = None  # None يعني استخدام الثيم الافتراضي

@api_router.put("/settings/personal-theme")
async def update_personal_theme(
    theme_data: PersonalThemeUpdate,
    current_user: User = Depends(get_current_user)
):
    """تحديث الثيم الشخصي للمستخدم"""
    # التحقق من صحة الثيم
    valid_themes = ['blue', 'green', 'gray', 'gray-light', 'purple', 'teal', 'amber', 'rose', 'slate', None]
    if theme_data.personal_theme not in valid_themes:
        raise HTTPException(status_code=400, detail="ثيم غير صالح")
    
    try:
        await db.users.update_one(
            {"id": current_user.id},
            {"$set": {"personal_theme": theme_data.personal_theme}}
        )
        return {"success": True, "message": "تم تحديث الثيم الشخصي بنجاح", "personal_theme": theme_data.personal_theme}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل في تحديث الثيم: {str(e)}")

@api_router.put("/settings/theme")
async def update_platform_theme(
    theme_data: ThemeUpdate,
    current_user: User = Depends(get_current_user)
):
    """تحديث ألوان المنصة (الأدمن فقط)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="هذه الصلاحية متاحة للأدمن فقط")
    
    # التحقق من صحة الثيم
    valid_themes = ['blue', 'green', 'gray', 'gray-light', 'purple', 'teal', 'amber', 'rose', 'slate']
    if theme_data.theme not in valid_themes:
        raise HTTPException(status_code=400, detail="ثيم غير صالح")
    
    try:
        await db.platform_settings.update_one(
            {"key": "platform_theme"},
            {"$set": {
                "theme": theme_data.theme,
                "dark_mode": theme_data.dark_mode,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user.id
            }},
            upsert=True
        )
        return {"success": True, "message": "تم تحديث ألوان المنصة بنجاح"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل في تحديث الألوان: {str(e)}")


# ============= نظام مشاريع الإيصال =============

class ConnectionProjectCreate(BaseModel):
    name: str
    description: Optional[str] = ""

class ConnectionProjectUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None

@api_router.get("/connection-projects")
async def get_connection_projects(current_user: User = Depends(get_current_user)):
    """جلب جميع مشاريع الإيصال"""
    try:
        if current_user.role == "admin" or not current_user.projects:
            projects = await db.connection_projects.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
        else:
            query = get_loose_in_query(current_user.projects, "name")
            projects = await db.connection_projects.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
        return projects
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل في جلب المشاريع: {str(e)}")

@api_router.post("/connection-projects")
async def create_connection_project(
    project: ConnectionProjectCreate,
    current_user: User = Depends(get_current_user)
):
    """إنشاء مشروع إيصال جديد"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="هذه الصلاحية متاحة للأدمن فقط")
    
    # التحقق من عدم وجود مشروع بنفس الاسم
    existing = await db.connection_projects.find_one({"name": project.name})
    if existing:
        raise HTTPException(status_code=400, detail="يوجد مشروع بنفس الاسم")
    
    project_id = str(uuid.uuid4())
    project_data = {
        "id": project_id,
        "name": project.name,
        "description": project.description,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id
    }
    
    await db.connection_projects.insert_one(project_data)
    return {"success": True, "message": "تم إنشاء المشروع بنجاح", "project": {k: v for k, v in project_data.items() if k != '_id'}}

@api_router.put("/connection-projects/{project_id}")
async def update_connection_project(
    project_id: str,
    project: ConnectionProjectUpdate,
    current_user: User = Depends(get_current_user)
):
    """تعديل مشروع إيصال"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="هذه الصلاحية متاحة للأدمن فقط")
    
    existing = await db.connection_projects.find_one({"id": project_id})
    if not existing:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    update_data = {}
    if project.name is not None and project.name != existing["name"]:
        old_name = existing["name"]
        new_name = project.name
        # التحقق من عدم وجود مشروع آخر بنفس الاسم
        other = await db.connection_projects.find_one({"name": new_name, "id": {"$ne": project_id}})
        if other:
            raise HTTPException(status_code=400, detail="يوجد مشروع آخر بنفس الاسم")
        update_data["name"] = new_name
        
        # تحديث المشروع في جميع المستخدمين
        await db.users.update_many({"projects": old_name}, {"$set": {"projects.$": new_name}})
        
        # تحديث project_permissions للمستخدمين
        users_with_perms = await db.users.find({f"project_permissions.{old_name}": {"$exists": True}}).to_list(1000)
        for u in users_with_perms:
            perms = u.get("project_permissions", {}).get(old_name)
            if perms is not None:
                await db.users.update_one(
                    {"id": u["id"]},
                    {"$set": {f"project_permissions.{new_name}": perms}, "$unset": {f"project_permissions.{old_name}": ""}}
                )
                
        # تحديث المشروع في التوصيلات
        await db.water_connections.update_many({"project": old_name}, {"$set": {"project": new_name}})
        await db.sewage_connections.update_many({"project": old_name}, {"$set": {"project": new_name}})
        
        # تحديث المشروع في الفواتير والمستخلصات وطلبات الموظفين والمقاولين
        await db.contractors.update_many({"project": old_name}, {"$set": {"project": new_name}})
        await db.invoices.update_many({"project": old_name}, {"$set": {"project": new_name}})
        await db.extracts.update_many({"project": old_name}, {"$set": {"project": new_name}})
        await db.employee_requests.update_many({"project": old_name}, {"$set": {"project": new_name}})
        
    if project.description is not None:
        update_data["description"] = project.description
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user.id
    
    await db.connection_projects.update_one({"id": project_id}, {"$set": update_data})
    return {"success": True, "message": "تم تحديث المشروع بنجاح"}

@api_router.delete("/connection-projects/{project_id}")
async def delete_connection_project(
    project_id: str,
    current_user: User = Depends(get_current_user)
):
    """حذف مشروع إيصال"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="هذه الصلاحية متاحة للأدمن فقط")
    
    existing = await db.connection_projects.find_one({"id": project_id})
    if not existing:
        raise HTTPException(status_code=404, detail="المشروع غير موجود")
    
    project_name = existing.get("name")
    
    await db.connection_projects.delete_one({"id": project_id})
    
    # حذف الارتباطات من جميع الأماكن
    if project_name:
        await db.project_governorates.delete_many({"project": project_name})
        await db.deleted_governorates.delete_many({"project": project_name})
        await db.users.update_many(
            {},
            {"$pull": {"projects": project_name}}
        )
        
    return {"success": True, "message": "تم حذف المشروع بنجاح"}

@api_router.get("/user-connection-projects/{user_id}")
async def get_user_connection_projects(
    user_id: str,
    current_user: User = Depends(get_current_user)
):
    """جلب مشاريع الإيصال المسموح بها للمستخدم"""
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # إذا كان admin يرى كل المشاريع
    if user.get("role") == "admin":
        projects = await db.connection_projects.find({}, {"_id": 0}).to_list(100)
        return projects
    
    # جلب صلاحيات التوصيلات للمستخدم
    connection_permissions = user.get("connection_permissions", {})
    allowed_projects = []
    
    for project_id, perms in connection_permissions.items():
        if perms.get("water_connections") or perms.get("sewage_connections"):
            project = await db.connection_projects.find_one({"id": project_id}, {"_id": 0})
            if project:
                project["permissions"] = perms
                allowed_projects.append(project)
    
    return allowed_projects


# ============= نظام أسطول السيارات والصيانة =============

@api_router.get("/fleet-cars/export/excel")
async def export_fleet_cars_excel(current_user: User = Depends(get_current_user)):
    """تصدير سجل السيارات إلى Excel"""
    if "cars_manage" not in (current_user.permissions or []) and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    cars = await db.fleet_cars.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    # تحضير البيانات
    data = []
    for car in cars:
        data.append({
            'رقم اللوحة': car.get('plate_number', ''),
            'نوع السيارة': car.get('car_type', ''),
            'الموديل': car.get('model', ''),
            'اللون': car.get('color', ''),
            'المالك': car.get('owner_name', ''),
            'الشركة': car.get('company', ''),
            'المشروع': car.get('project_name', ''),
            'المستخدم الحالي': car.get('current_user_name', ''),
            'بداية الاستمارة': car.get('registration_start', ''),
            'نهاية الاستمارة': car.get('registration_end', ''),
            'بداية الفحص': car.get('inspection_start', ''),
            'نهاية الفحص': car.get('inspection_end', ''),
            'بداية التفويض': car.get('authorization_start', ''),
            'نهاية التفويض': car.get('authorization_end', ''),
            'ملاحظات': car.get('notes', '')
        })
    
    # إنشاء Excel
    import pandas as pd
    from io import BytesIO
    
    df = pd.DataFrame(data)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='سجل السيارات')
        
        # تنسيق العمود
        worksheet = writer.sheets['سجل السيارات']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=fleet_cars_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@api_router.get("/fleet-cars")
async def get_fleet_cars(current_user: User = Depends(get_current_user)):
    """جلب جميع سيارات الأسطول"""
    if "cars_manage" not in (current_user.permissions or []) and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    cars = await db.fleet_cars.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return cars

@api_router.post("/fleet-cars")
async def create_fleet_car(car: FleetCarCreate, current_user: User = Depends(get_current_user)):
    """إضافة سيارة جديدة للأسطول"""
    if "cars_manage" not in (current_user.permissions or []) and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    car_doc = {
        "id": str(uuid.uuid4()),
        **car.dict(),
        "image": "",
        "created_by": current_user.id,
        "created_by_name": current_user.full_name,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.fleet_cars.insert_one(car_doc)
    car_doc.pop("_id", None)
    return car_doc

@api_router.put("/fleet-cars/{car_id}")
async def update_fleet_car(car_id: str, car: FleetCarUpdate, current_user: User = Depends(get_current_user)):
    """تحديث سيارة في الأسطول"""
    if "cars_manage" not in (current_user.permissions or []) and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    # جلب السيارة الحالية للتحقق من تغيير المستخدم
    existing_car = await db.fleet_cars.find_one({"id": car_id}, {"_id": 0})
    if not existing_car:
        raise HTTPException(status_code=404, detail="السيارة غير موجودة")
    
    update_data = {k: v for k, v in car.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # التحقق من تغيير المستخدم وتسجيله في السجل (باستخدام الاسم)
    old_user_name = existing_car.get("current_user_name", "")
    new_user_name = update_data.get("current_user_name", old_user_name)
    
    if new_user_name and new_user_name != old_user_name:
        # إغلاق السجل القديم (إذا وجد)
        if old_user_name:
            await db.car_user_history.update_one(
                {"fleet_car_id": car_id, "user_name": old_user_name, "returned_date": None},
                {"$set": {"returned_date": datetime.now(timezone.utc).isoformat()}}
            )
        
        # إنشاء سجل جديد للمستخدم الجديد
        history_doc = {
            "id": str(uuid.uuid4()),
            "fleet_car_id": car_id,
            "user_name": new_user_name,
            "assigned_date": datetime.now(timezone.utc).isoformat(),
            "returned_date": None,
            "assigned_by": current_user.id,
            "assigned_by_name": current_user.full_name,
            "notes": ""
        }
        await db.car_user_history.insert_one(history_doc)
    
    await db.fleet_cars.update_one({"id": car_id}, {"$set": update_data})
    return {"message": "تم التحديث"}

@api_router.get("/fleet-cars/{car_id}/history")
async def get_car_user_history(car_id: str, current_user: User = Depends(get_current_user)):
    """جلب سجل تسليم السيارة"""
    user_perms = current_user.permissions or []
    has_perm = "cars_manage" in user_perms or "fleet_maintenance" in user_perms
    if not has_perm and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    history = await db.car_user_history.find({"fleet_car_id": car_id}, {"_id": 0}).sort("assigned_date", -1).to_list(100)
    return history

@api_router.delete("/fleet-cars/{car_id}")
async def delete_fleet_car(car_id: str, current_user: User = Depends(get_current_user)):
    """حذف سيارة من الأسطول"""
    if "cars_manage" not in (current_user.permissions or []) and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    await db.fleet_cars.delete_one({"id": car_id})
    await db.maintenance_records.delete_many({"fleet_car_id": car_id})
    await db.car_user_history.delete_many({"fleet_car_id": car_id})
    return {"message": "تم الحذف"}

@api_router.post("/fleet-cars/{car_id}/image")
async def upload_fleet_car_image(car_id: str, file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """رفع صورة لسيارة الأسطول"""
    user_perms = current_user.permissions or []
    has_perm = "cars_manage" in user_perms or "fleet_maintenance" in user_perms
    if not has_perm and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    contents = await file.read()
    img = Image.open(io.BytesIO(contents))
    img.thumbnail((800, 600), Image.LANCZOS)
    buffer = io.BytesIO()
    img.save(buffer, format="JPEG", quality=60, optimize=True)
    image_url = _store_image_bytes(buffer.getvalue(), category="fleet-cars", filename=file.filename, content_type="image/jpeg")
    
    await db.fleet_cars.update_one({"id": car_id}, {"$set": {"image": image_url}})
    return {"message": "تم رفع الصورة", "image": image_url}

# سجلات الصيانة
@api_router.get("/fleet-cars/{car_id}/maintenance")
async def get_maintenance_records(car_id: str, current_user: User = Depends(get_current_user)):
    """جلب سجلات صيانة سيارة"""
    user_perms = current_user.permissions or []
    has_perm = "cars_manage" in user_perms or "fleet_maintenance" in user_perms
    if not has_perm and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    records = await db.maintenance_records.find({"fleet_car_id": car_id}, {"_id": 0}).sort("date", -1).to_list(500)
    return records

@api_router.post("/fleet-cars/{car_id}/maintenance")
async def create_maintenance_record(car_id: str, record: MaintenanceRecordCreate, current_user: User = Depends(get_current_user)):
    """إضافة سجل صيانة"""
    if "cars_manage" not in (current_user.permissions or []) and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    record_doc = {
        "id": str(uuid.uuid4()),
        "fleet_car_id": car_id,
        **record.dict(),
        "images": [],
        "created_by": current_user.id,
        "created_by_name": current_user.full_name,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.maintenance_records.insert_one(record_doc)
    record_doc.pop("_id", None)
    return record_doc

class MaintenanceRecordUpdate(BaseModel):
    """تحديث سجل الصيانة - جميع الحقول اختيارية"""
    maintenance_type: Optional[str] = None
    description: Optional[str] = None
    cost: Optional[float] = None
    workshop: Optional[str] = None
    date: Optional[str] = None
    notes: Optional[str] = None


@api_router.put("/maintenance/{record_id}")
async def update_maintenance_record(
    record_id: str,
    maintenance_type: Optional[str] = None,
    description: Optional[str] = None,
    cost: Optional[str] = None,  # نستخدم str لتجنب خطأ التحويل
    workshop: Optional[str] = None,
    date: Optional[str] = None,
    notes: Optional[str] = None,
    body: Optional[MaintenanceRecordUpdate] = Body(None),
    current_user: User = Depends(get_current_user)
):
    """تحديث سجل صيانة - يدعم Query Params و JSON Body"""
    if "cars_manage" not in (current_user.permissions or []) and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    # دعم كل من query params و JSON body
    data = body.dict(exclude_none=True) if body else {}
    if maintenance_type: data["maintenance_type"] = maintenance_type
    if description: data["description"] = description
    if cost not in (None, ""):
        try:
            data["cost"] = float(cost)
        except (ValueError, TypeError):
            pass
    if workshop: data["workshop"] = workshop
    if date: data["date"] = date
    if notes: data["notes"] = notes
    
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.maintenance_records.update_one({"id": record_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="سجل الصيانة غير موجود")
    return {"message": "تم التحديث بنجاح"}

@api_router.delete("/maintenance/{record_id}")
async def delete_maintenance_record(record_id: str, current_user: User = Depends(get_current_user)):
    """حذف سجل صيانة"""
    if "cars_manage" not in (current_user.permissions or []) and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    await db.maintenance_records.delete_one({"id": record_id})
    return {"message": "تم الحذف"}

@api_router.post("/maintenance/{record_id}/images")
async def upload_maintenance_images(record_id: str, files: List[UploadFile] = File(...), current_user: User = Depends(get_current_user)):
    """رفع صور للصيانة (فواتير، صدمات، إلخ)"""
    user_perms = current_user.permissions or []
    has_perm = "cars_manage" in user_perms or "fleet_maintenance" in user_perms
    if not has_perm and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    # التحقق من وجود السجل
    record = await db.maintenance_records.find_one({"id": record_id})
    if not record:
        raise HTTPException(status_code=404, detail="سجل الصيانة غير موجود")
    
    images = []
    for file in files[:10]:
        try:
            contents = await file.read()
            img = Image.open(io.BytesIO(contents))
            img.thumbnail((1200, 900), Image.LANCZOS)
            buffer = io.BytesIO()
            img.save(buffer, format="JPEG", quality=50, optimize=True)
            url = _store_image_bytes(buffer.getvalue(), category="maintenance", filename=file.filename, content_type="image/jpeg")
            images.append(url)
        except Exception as e:
            logger.error(f"Error processing image: {str(e)}")
            continue
    
    if images:
        await db.maintenance_records.update_one({"id": record_id}, {"$push": {"images": {"$each": images}}})
        return {"message": f"تم رفع {len(images)} صورة", "count": len(images)}
    else:
        raise HTTPException(status_code=400, detail="لم يتم رفع أي صور")

@api_router.delete("/maintenance/{record_id}/images/{image_index}")
async def delete_maintenance_image(record_id: str, image_index: int, current_user: User = Depends(get_current_user)):
    """حذف صورة من سجل الصيانة"""
    if "cars_manage" not in (current_user.permissions or []) and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    record = await db.maintenance_records.find_one({"id": record_id})
    if record and "images" in record and 0 <= image_index < len(record["images"]):
        record["images"].pop(image_index)
        await db.maintenance_records.update_one({"id": record_id}, {"$set": {"images": record["images"]}})
    return {"message": "تم حذف الصورة"}


# ============= OBJECT STORAGE ENDPOINTS =============
@api_router.post("/storage/upload")
async def upload_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """رفع ملف إلى التخزين السحابي مع ضغط الصور"""
    data = await file.read()
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "bin"
    content_type = file.content_type or "application/octet-stream"
    
    if ext == "bin" or ext == "":
        if "pdf" in content_type.lower():
            ext = "pdf"
        elif "jpeg" in content_type.lower() or "jpg" in content_type.lower():
            ext = "jpg"
        elif "png" in content_type.lower():
            ext = "png"
    
    # ضغط الصور إذا كانت أكبر من 30 كيلوبايت
    if ext in ("jpg", "jpeg", "png", "webp") and len(data) > 30000:
        try:
            from PIL import Image
            img = Image.open(io.BytesIO(data))
            if img.mode == 'RGBA':
                img = img.convert('RGB')
            # تصغير الأبعاد
            max_dim = 800
            if max(img.size) > max_dim:
                img.thumbnail((max_dim, max_dim), Image.LANCZOS)
            # ضغط
            output = io.BytesIO()
            quality = 60
            img.save(output, format='JPEG', quality=quality, optimize=True)
            while output.tell() > 30000 and quality > 10:
                quality -= 10
                output = io.BytesIO()
                img.save(output, format='JPEG', quality=quality, optimize=True)
            data = output.getvalue()
            ext = 'jpg'
            content_type = 'image/jpeg'
        except Exception as e:
            logging.warning(f"Image compression failed: {e}")
            
    # ضغط ملفات البوربوينت
    if ext in ("pptx", "ppt") and len(data) > 50000:
        try:
            import zipfile
            import io
            
            def try_compress_pptx(data_bytes, img_quality=60, max_dim=1200):
                in_buffer = io.BytesIO(data_bytes)
                out_buffer = io.BytesIO()
                with zipfile.ZipFile(in_buffer, 'r') as in_zip:
                    with zipfile.ZipFile(out_buffer, 'w', compression=zipfile.ZIP_DEFLATED, compresslevel=9) as out_zip:
                        for item in in_zip.infolist():
                            content = in_zip.read(item.filename)
                            if item.filename.startswith("ppt/media/") and item.filename.lower().endswith((".png", ".jpg", ".jpeg")):
                                try:
                                    from PIL import Image
                                    img = Image.open(io.BytesIO(content))
                                    if max(img.size) > max_dim:
                                        img.thumbnail((max_dim, max_dim), Image.LANCZOS)
                                    img_out = io.BytesIO()
                                    if item.filename.lower().endswith(".png"):
                                        img.save(img_out, format="PNG", optimize=True)
                                    else:
                                        img.save(img_out, format="JPEG", quality=img_quality, optimize=True)
                                    optimized_content = img_out.getvalue()
                                    if len(optimized_content) < len(content):
                                        content = optimized_content
                                except Exception as img_err:
                                    logging.warning(f"Error optimizing embedded image {item.filename}: {img_err}")
                            out_zip.writestr(item, content)
                return out_buffer.getvalue()

            # First attempt: quality 60, max_dim 1200
            compressed_data = try_compress_pptx(data, img_quality=60, max_dim=1200)
            
            # If still > 2MB (2,097,152 bytes), try more aggressive settings: quality 40, max_dim 800
            if len(compressed_data) > 2097152:
                compressed_data = try_compress_pptx(data, img_quality=40, max_dim=800)
                
            # If still > 2MB, try even more aggressive: quality 30, max_dim 640
            if len(compressed_data) > 2097152:
                compressed_data = try_compress_pptx(data, img_quality=30, max_dim=640)

            if len(compressed_data) < len(data):
                data = compressed_data
        except Exception as e:
            logging.warning(f"PowerPoint compression failed: {e}")
            
    # ضغط ملفات PDF
    if ext == "pdf" and len(data) > 50000:
        try:
            import fitz  # PyMuPDF
            
            def try_compress_pdf(data_bytes):
                doc = fitz.open("pdf", data_bytes)
                # compress with PyMuPDF
                new_bytes = doc.write(garbage=4, deflate=True, clean=True)
                doc.close()
                return new_bytes
                
            compressed_data = try_compress_pdf(data)
            if len(compressed_data) < len(data):
                data = compressed_data
        except Exception as e:
            logging.warning(f"PDF compression failed: {e}")
    
    path = f"sery17/uploads/{current_user.id}/{uuid.uuid4()}.{ext}"
    
    try:
        result = put_object(path, data, content_type)
        file_record = {
            "id": str(uuid.uuid4()),
            "storage_path": result["path"],
            "original_filename": file.filename,
            "content_type": content_type,
            "size": len(data),
            "uploaded_by": current_user.id,
            "is_deleted": {"$ne": True},
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.uploaded_files.insert_one(file_record)
        file_record.pop("_id", None)
        return {"file_id": file_record["id"], "storage_path": result["path"], "size": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل رفع الملف: {str(e)}")

@api_router.get("/storage/files/{file_path:path}")
async def download_file(
    file_path: str,
    download: int = Query(0),
    auth: str = Query(None),
    authorization: str = None
):
    """تحميل أو عرض ملف من التخزين السحابي"""
    from urllib.parse import unquote
    decoded_path = unquote(file_path)
    
    # Normalize potential single-slash after protocol caused by path parsing
    if decoded_path.startswith("http:/") and not decoded_path.startswith("http://"):
        decoded_path = decoded_path.replace("http:/", "http://", 1)
    elif decoded_path.startswith("https:/") and not decoded_path.startswith("https://"):
        decoded_path = decoded_path.replace("https:/", "https://", 1)
        
    try:
        record = await db.uploaded_files.find_one({"storage_path": decoded_path, "is_deleted": {"$ne": True}}, {"_id": 0})
        if not record:
            normalized_path = decoded_path.replace("http://", "").replace("https://", "")
            record = await db.uploaded_files.find_one({
                "$or": [
                    {"storage_path": decoded_path},
                    {"storage_path": {"$regex": normalized_path}}
                ]
            })
            
        ct = "application/octet-stream"
        filename = "file"
        if record:
            ct = record.get("content_type", ct)
            filename = record.get("original_filename", "file")
            
        data, fetched_ct = _get_object(decoded_path)
        if not record:
            ct = fetched_ct
            
        headers = {}
        from urllib.parse import quote
        safe_filename = quote(filename)
        
        if download == 1:
            headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{safe_filename}"
        else:
            headers["Content-Disposition"] = f"inline; filename*=UTF-8''{safe_filename}"
            
        return Response(content=data, media_type=ct, headers=headers)
    except Exception as e:
        logging.error(f"Download/view file error for {decoded_path}: {e}")
        raise HTTPException(status_code=500, detail=f"فشل تحميل أو عرض الملف: {str(e)}")


# ============= HR MANAGEMENT ENDPOINTS =============
# شؤون الموظفين - الموظفين والعقود والرواتب

class HREmployee(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    employee_number: Optional[str] = None
    nationality: Optional[str] = None
    company: Optional[str] = None
    project: Optional[str] = None
    id_number: Optional[str] = None
    id_expiry: Optional[str] = None
    insurance_expiry: Optional[str] = None
    religion: Optional[str] = None
    status: str = "على رأس العمل"
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class HRContract(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_name: str
    nationality: Optional[str] = None
    company: Optional[str] = None
    project: Optional[str] = None
    contract_duration: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class HRSalary(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_name: str
    project: Optional[str] = None
    company: Optional[str] = None
    basic_salary: float = 0
    housing_allowance: float = 0
    transport_allowance: float = 0
    other_allowances: float = 0
    overtime: float = 0
    bonus: float = 0
    deduction_type: Optional[str] = None
    deduction_amount: float = 0
    month: int
    year: int
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# الموظفين

class AdvanceCustody(BaseModel):
    id: Optional[str] = None
    employee_name: str
    employee_number: Optional[str] = None
    project: Optional[str] = None
    company: Optional[str] = None
    type: str
    amount: Optional[float] = 0
    paid_amount: Optional[float] = 0
    remaining_amount: Optional[float] = 0
    item_description: Optional[str] = None
    date: str
    status: str
    action_date: Optional[str] = None
    notes: Optional[str] = None
    payment_history: Optional[list] = []

@api_router.get("/hr/employees")
async def get_hr_employees(current_user: User = Depends(get_current_user)):
    employees = await db.hr_employees.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return employees

@api_router.post("/hr/employees")
async def create_hr_employee(employee: HREmployee, current_user: User = Depends(get_current_user)):
    employee_dict = employee.model_dump()
    employee_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.hr_employees.insert_one(employee_dict)
    employee_dict.pop("_id", None)
    return employee_dict

@api_router.put("/hr/employees/{employee_id}")
async def update_hr_employee(employee_id: str, employee_data: dict, current_user: User = Depends(get_current_user)):
    employee_data.pop("id", None)
    employee_data.pop("_id", None)
    await db.hr_employees.update_one({"id": employee_id}, {"$set": employee_data})
    return {"message": "تم تحديث الموظف"}

@api_router.delete("/hr/employees/{employee_id}")
async def delete_hr_employee(employee_id: str, current_user: User = Depends(get_current_user)):
    await db.hr_employees.delete_one({"id": employee_id})
    return {"message": "تم حذف الموظف"}

# العقود
@api_router.get("/hr/contracts")
async def get_hr_contracts(current_user: User = Depends(get_current_user)):
    contracts = await db.hr_contracts.find({}, {"_id": 0}).sort("employee_name", 1).to_list(1000)
    return contracts

@api_router.post("/hr/contracts")
async def create_hr_contract(contract: HRContract, current_user: User = Depends(get_current_user)):
    contract_dict = contract.model_dump()
    contract_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.hr_contracts.insert_one(contract_dict)
    contract_dict.pop("_id", None)
    return contract_dict

@api_router.put("/hr/contracts/{contract_id}")
async def update_hr_contract(contract_id: str, contract_data: dict, current_user: User = Depends(get_current_user)):
    contract_data.pop("id", None)
    contract_data.pop("_id", None)
    await db.hr_contracts.update_one({"id": contract_id}, {"$set": contract_data})
    return {"message": "تم تحديث العقد"}

@api_router.delete("/hr/contracts/{contract_id}")
async def delete_hr_contract(contract_id: str, current_user: User = Depends(get_current_user)):
    await db.hr_contracts.delete_one({"id": contract_id})
    return {"message": "تم حذف العقد"}

# الرواتب
@api_router.get("/hr/salaries")
async def get_hr_salaries(current_user: User = Depends(get_current_user)):
    salaries = await db.hr_salaries.find({}, {"_id": 0}).sort([("year", -1), ("month", -1)]).to_list(1000)
    return salaries

@api_router.post("/hr/salaries")
async def create_hr_salary(salary: HRSalary, current_user: User = Depends(get_current_user)):
    salary_dict = salary.model_dump()
    salary_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.hr_salaries.insert_one(salary_dict)
    salary_dict.pop("_id", None)
    return salary_dict

@api_router.put("/hr/salaries/{salary_id}")
async def update_hr_salary(salary_id: str, salary_data: dict, current_user: User = Depends(get_current_user)):
    salary_data.pop("id", None)
    salary_data.pop("_id", None)
    await db.hr_salaries.update_one({"id": salary_id}, {"$set": salary_data})
    return {"message": "تم تحديث الراتب"}

@api_router.delete("/hr/salaries/{salary_id}")
async def delete_hr_salary(salary_id: str, current_user: User = Depends(get_current_user)):
    await db.hr_salaries.delete_one({"id": salary_id})
    return {"message": "تم حذف الراتب"}

# الشركات (من الموظفين)
@api_router.get("/hr/companies")
async def get_hr_companies(current_user: User = Depends(get_current_user)):
    companies = await db.hr_employees.distinct("company")
    return [{"name": c} for c in companies if c]

# تصدير HR بيانات Excel
@api_router.get("/hr/export/excel")
async def export_hr_excel(type: str = "employees", current_user: User = Depends(get_current_user)):
    """تصدير بيانات شؤون الموظفين إلى Excel"""
    import pandas as pd
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    if type == "employees":
        data = await db.hr_employees.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
        columns_map = {
            'name': 'الاسم', 'nationality': 'الجنسية', 'company': 'الشركة',
            'project': 'المشروع', 'id_number': 'رقم الإقامة', 'id_expiry': 'انتهاء الإقامة',
            'insurance_expiry': 'انتهاء التأمين', 'religion': 'الديانة', 'status': 'الحالة'
        }
        filename = "employees"
    elif type == "contracts":
        data = await db.hr_contracts.find({}, {"_id": 0}).sort("employee_name", 1).to_list(1000)
        columns_map = {
            'employee_name': 'الاسم', 'nationality': 'الجنسية', 'company': 'الشركة',
            'project': 'المشروع', 'contract_duration': 'مدة العقد',
            'start_date': 'تاريخ البداية', 'end_date': 'تاريخ الانتهاء', 'notes': 'ملاحظات'
        }
        filename = "contracts"
    elif type == "salaries":
        data = await db.hr_salaries.find({}, {"_id": 0}).sort([("year", -1), ("month", -1)]).to_list(1000)
        columns_map = {
            'employee_name': 'الاسم', 'project': 'المشروع', 'company': 'الشركة',
            'basic_salary': 'الراتب الأساسي', 'housing_allowance': 'بدل السكن',
            'transport_allowance': 'بدل التنقل', 'other_allowances': 'بدلات أخرى',
            'deduction_type': 'نوع الخصم', 'deduction_amount': 'مبلغ الخصم',
            'month': 'الشهر', 'year': 'السنة', 'notes': 'ملاحظات'
        }
        filename = "salaries"
    else:
        return {"error": "نوع غير صحيح"}
    
    # Clean data
    clean_data = []
    for item in data:
        row = {}
        for key, label in columns_map.items():
            val = item.get(key, '')
            row[label] = val if val else ''
        clean_data.append(row)
    
    df = pd.DataFrame(clean_data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )

# إشعارات شؤون الموظفين

@api_router.get("/hr/advances-custodies")
async def get_advances_custodies(current_user: User = Depends(get_current_user)):
    user_permissions = current_user.permissions or []
    has_hr_perm = ("hr_management" in user_permissions) or user_has_any_project_permission(current_user, "hr_management")
    if current_user.role != "admin" and not has_hr_perm:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
    
    docs = await db.hr_advances_custodies.find({}, {"_id": 0}).to_list(1000)
    return docs

@api_router.post("/hr/advances-custodies")
async def create_advance_custody(data: AdvanceCustody, current_user: User = Depends(get_current_user)):
    user_permissions = current_user.permissions or []
    has_hr_perm = ("hr_management" in user_permissions) or user_has_any_project_permission(current_user, "hr_management")
    if current_user.role != "admin" and not has_hr_perm:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        
    doc = data.model_dump(exclude_none=True)
    if 'id' not in doc or not doc['id']:
        import uuid
        doc['id'] = str(uuid.uuid4())
    doc["created_by"] = current_user.username
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.hr_advances_custodies.insert_one(doc)
    return {"message": "تم الإضافة بنجاح", "id": doc['id']}

@api_router.put("/hr/advances-custodies/{item_id}")
async def update_advance_custody(item_id: str, data: AdvanceCustody, current_user: User = Depends(get_current_user)):
    user_permissions = current_user.permissions or []
    has_hr_perm = ("hr_management" in user_permissions) or user_has_any_project_permission(current_user, "hr_management")
    if current_user.role != "admin" and not has_hr_perm:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        
    update_data = data.model_dump(exclude={"id"}, exclude_none=True)
    await db.hr_advances_custodies.update_one({"id": item_id}, {"$set": update_data})
    return {"message": "تم التعديل بنجاح"}

@api_router.delete("/hr/advances-custodies/{item_id}")
async def delete_advance_custody(item_id: str, current_user: User = Depends(get_current_user)):
    user_permissions = current_user.permissions or []
    has_hr_perm = ("hr_management" in user_permissions) or user_has_any_project_permission(current_user, "hr_management")
    if current_user.role != "admin" and not has_hr_perm:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية")
        
    await db.hr_advances_custodies.delete_one({"id": item_id})
    return {"message": "تم الحذف بنجاح"}

@api_router.get("/hr/alerts")
async def get_hr_alerts(current_user: User = Depends(get_current_user)):
    """جلب إشعارات انتهاء الإقامة والتأمين والعقود"""
    from datetime import timedelta
    alerts = []
    today = datetime.now(timezone.utc).date()
    
    # إشعارات الموظفين - إقامة وتأمين (20 يوم)
    employees = await db.hr_employees.find({}, {"_id": 0}).to_list(1000)
    for emp in employees:
        # انتهاء الإقامة
        if emp.get('id_expiry'):
            try:
                expiry = datetime.strptime(emp['id_expiry'], '%Y-%m-%d').date()
                days_left = (expiry - today).days
                if 0 <= days_left <= 20:
                    alerts.append({
                        "type": "id_expiry",
                        "employee_id": emp.get('id', ''),
                        "employee_name": emp.get('name', ''),
                        "message": f"إقامة {emp.get('name', '')} تنتهي خلال {days_left} يوم ({emp['id_expiry']})",
                        "days_left": days_left,
                        "priority": "urgent" if days_left <= 5 else "warning",
                        "expiry_date": emp['id_expiry']
                    })
                elif days_left < 0:
                    alerts.append({
                        "type": "id_expiry",
                        "employee_id": emp.get('id', ''),
                        "employee_name": emp.get('name', ''),
                        "message": f"إقامة {emp.get('name', '')} منتهية منذ {abs(days_left)} يوم!",
                        "days_left": days_left,
                        "priority": "expired",
                        "expiry_date": emp['id_expiry']
                    })
            except (ValueError, TypeError):
                pass
        
        # انتهاء التأمين
        if emp.get('insurance_expiry'):
            try:
                expiry = datetime.strptime(emp['insurance_expiry'], '%Y-%m-%d').date()
                days_left = (expiry - today).days
                if 0 <= days_left <= 20:
                    alerts.append({
                        "type": "insurance_expiry",
                        "employee_id": emp.get('id', ''),
                        "employee_name": emp.get('name', ''),
                        "message": f"تأمين {emp.get('name', '')} ينتهي خلال {days_left} يوم ({emp['insurance_expiry']})",
                        "days_left": days_left,
                        "priority": "urgent" if days_left <= 5 else "warning",
                        "expiry_date": emp['insurance_expiry']
                    })
                elif days_left < 0:
                    alerts.append({
                        "type": "insurance_expiry",
                        "employee_id": emp.get('id', ''),
                        "employee_name": emp.get('name', ''),
                        "message": f"تأمين {emp.get('name', '')} منتهي منذ {abs(days_left)} يوم!",
                        "days_left": days_left,
                        "priority": "expired",
                        "expiry_date": emp['insurance_expiry']
                    })
            except (ValueError, TypeError):
                pass
    
    # إشعارات العقود - 70 يوم
    contracts = await db.hr_contracts.find({}, {"_id": 0}).to_list(1000)
    for contract in contracts:
        if contract.get('end_date'):
            try:
                expiry = datetime.strptime(contract['end_date'], '%Y-%m-%d').date()
                days_left = (expiry - today).days
                if 0 <= days_left <= 70:
                    alerts.append({
                        "type": "contract_expiry",
                        "employee_id": contract.get('id', ''),
                        "employee_name": contract.get('employee_name', ''),
                        "message": f"عقد {contract.get('employee_name', '')} ينتهي خلال {days_left} يوم ({contract['end_date']})",
                        "days_left": days_left,
                        "priority": "urgent" if days_left <= 20 else "warning",
                        "expiry_date": contract['end_date']
                    })
                elif days_left < 0:
                    alerts.append({
                        "type": "contract_expiry",
                        "employee_id": contract.get('id', ''),
                        "employee_name": contract.get('employee_name', ''),
                        "message": f"عقد {contract.get('employee_name', '')} منتهي منذ {abs(days_left)} يوم!",
                        "days_left": days_left,
                        "priority": "expired",
                        "expiry_date": contract['end_date']
                    })
            except (ValueError, TypeError):
                pass
    
    # ترتيب حسب الأولوية
    priority_order = {"expired": 0, "urgent": 1, "warning": 2}
    alerts.sort(key=lambda x: (priority_order.get(x['priority'], 3), x.get('days_left', 999)))
    
    return {"alerts": alerts, "count": len(alerts)}


# ============= TRASH EMPTY ALL ENDPOINT =============
@api_router.delete("/trash/empty-all")
async def empty_all_trash(current_user: User = Depends(get_current_user)):
    user_perms = current_user.permissions or []
    has_trash = current_user.role == "admin" or "trash" in user_perms or user_has_any_project_permission(current_user, "trash")
    if not has_trash and getattr(current_user, "username", "") != "Eng Mahmoud Haroun":
        raise HTTPException(status_code=403, detail="Forbidden")
    
    # Delete from all collections
    await db.reports.delete_many({"is_deleted": True})
    await db.safety_reports.delete_many({"is_deleted": True})
    await db.quality_reports.delete_many({"is_deleted": True})
    await db.business_reports.delete_many({"is_deleted": True})
    await db.water_connections.delete_many({"is_deleted": True})
    await db.sewage_connections.delete_many({"is_deleted": True})
    await db.extracts.delete_many({"is_deleted": True})
    
    return {"message": "All trash emptied"}


# ============= UNIFIED DELETED ITEMS (سجل المحذوفات) =============

@api_router.get("/trash/count")
async def get_trash_count(current_user: User = Depends(get_current_user)):
    user_permissions = current_user.permissions or []
    has_trash_perm = current_user.role == "admin" or "trash" in user_permissions or user_has_any_project_permission(current_user, "trash")
    if not has_trash_perm:
        return {"count": 0}
        
    base_query = {"is_deleted": True, "deleted_by": {"$exists": True}}
    if current_user.role != "admin":
        base_query["deleted_by"] = current_user.id
        
    c_inv = await db.invoices.count_documents(base_query)
    c_req = await db.employee_requests.count_documents(base_query)
    c_rep = await db.reports.count_documents(base_query)
    c_ext = await db.extracts.count_documents(base_query)
    
    water_query = {**base_query}
    sewage_query = {**base_query}
    safety_query = {**base_query}
    quality_query = {**base_query}
    business_query = {**base_query}
    
    if current_user.role != "admin":
        water_query["project"] = {"$in": get_projects_with_permission(current_user, "water_connections")}
        sewage_query["project"] = {"$in": get_projects_with_permission(current_user, "sewage_connections")}
        safety_query["project"] = {"$in": get_projects_with_permission(current_user, "safety_reports")}
        quality_query["project"] = {"$in": get_projects_with_permission(current_user, "quality_reports")}
        business_query["project"] = {"$in": get_projects_with_permission(current_user, "business_reports")}
        
    c_w = await db.water_connections.count_documents(water_query)
    c_s = await db.sewage_connections.count_documents(sewage_query)
    c_sr = await db.safety_reports.count_documents(safety_query)
    c_qr = await db.quality_reports.count_documents(quality_query)
    c_br = await db.business_reports.count_documents(business_query)
    
    total = c_inv + c_req + c_rep + c_ext + c_w + c_s + c_sr + c_qr + c_br
    return {"count": total}


@api_router.get("/deleted-items")
async def get_deleted_items(
    item_type: Optional[str] = None,
    page: int = 1,
    limit: int = 30,
    current_user: User = Depends(get_current_user)
):
    """جلب العناصر المحذوفة حديثاً فقط (آخر 30 يوم)"""
    user_permissions = current_user.permissions or []
    has_trash_perm = current_user.role == "admin" or "trash" in user_permissions or user_has_any_project_permission(current_user, "trash")
    if not has_trash_perm:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية الوصول لسجل المحذوفات")
    
    # فقط العناصر المحذوفة حديثاً (التي تحتوي على deleted_by)
    base_query = {"is_deleted": True, "deleted_by": {"$exists": True}}
    if current_user.role != "admin":
        base_query["deleted_by"] = current_user.id
    
    items = []
    skip = (page - 1) * limit
    total_count = 0
    
    # إذا كان نوع العنصر محدداً، يمكننا جلب تلك المجموعة فقط بـ skip و limit
    if item_type:
        collection_map = {
            "invoice": (db.invoices, "invoice", "فاتورة"),
            "employee_request": (db.employee_requests, "employee_request", "طلب موظف"),
            "report": (db.reports, "report", "بلاغ"),
            "extract": (db.extracts, "extract", "مستخلص"),
            "water_connection": (db.water_connections, "water_connection", "توصيلة مياه"),
            "sewage_connection": (db.sewage_connections, "sewage_connection", "توصيلة صرف صحي"),
            "safety_report": (db.safety_reports, "safety_report", "تقرير سلامة"),
            "quality_report": (db.quality_reports, "quality_report", "تقرير جودة"),
            "business_report": (db.business_reports, "business_report", "تقرير أعمال")
        }
        
        target = collection_map.get(item_type)
        if target:
            coll, t_name, t_label = target
            q = {**base_query}
            if t_name == "water_connection" and current_user.role != "admin":
                allowed_projects = get_projects_with_permission(current_user, "water_connections")
                q["project"] = {"$in": allowed_projects}
                q["deleted_by"] = current_user.id
            elif t_name == "sewage_connection" and current_user.role != "admin":
                allowed_projects = get_projects_with_permission(current_user, "sewage_connections")
                q["project"] = {"$in": allowed_projects}
                q["deleted_by"] = current_user.id
                
            total_count = await coll.count_documents(q)
            
            projection = {"_id": 0}
            if t_name == "report":
                projection = {"_id": 0, "id": 1, "report_number": 1, "project": 1, "status": 1, "deleted_at": 1, "deleted_by": 1}
                
            docs = await coll.find(q, projection).sort("deleted_at", -1).skip(skip).limit(limit).to_list(limit)
            
            for doc in docs:
                deleted_by_name = "غير معروف"
                if doc.get("deleted_by"):
                    user_doc = await db.users.find_one({"id": doc["deleted_by"]}, {"_id": 0, "full_name": 1})
                    deleted_by_name = user_doc.get("full_name") if user_doc else "غير معروف"
                doc["deleted_by_name"] = deleted_by_name
                items.append({"type": t_name, "type_label": t_label, "data": doc})
    else:
        # إذا لم يكن النوع محدداً (الكل)، ندمج المجموعات حتى skip + limit ثم نقص
        max_fetch = skip + limit
        
        # 1. الفواتير
        invs = await db.invoices.find(base_query, {"_id": 0}).sort("deleted_at", -1).limit(max_fetch).to_list(max_fetch)
        for inv in invs:
            items.append({"type": "invoice", "type_label": "فاتورة", "data": inv})
            
        # 2. طلبات الموظفين
        reqs = await db.employee_requests.find(base_query, {"_id": 0}).sort("deleted_at", -1).limit(max_fetch).to_list(max_fetch)
        for req in reqs:
            items.append({"type": "employee_request", "type_label": "طلب موظف", "data": req})
            
        # 3. البلاغات
        reps = await db.reports.find(base_query, {"_id": 0, "id": 1, "report_number": 1, "project": 1, "status": 1, "deleted_at": 1, "deleted_by": 1}).sort("deleted_at", -1).limit(max_fetch).to_list(max_fetch)
        for rep in reps:
            items.append({"type": "report", "type_label": "بلاغ", "data": rep})
            
        # 4. المستخلصات
        exts = await db.extracts.find(base_query, {"_id": 0}).sort("deleted_at", -1).limit(max_fetch).to_list(max_fetch)
        for ext in exts:
            items.append({"type": "extract", "type_label": "مستخلص", "data": ext})
            
        # 5. توصيلات المياه
        water_query = {**base_query}
        if current_user.role != "admin":
            allowed_projects = get_projects_with_permission(current_user, "water_connections")
            water_query["project"] = {"$in": allowed_projects}
            water_query["deleted_by"] = current_user.id
        wconns = await db.water_connections.find(water_query, {"_id": 0}).sort("deleted_at", -1).limit(max_fetch).to_list(max_fetch)
        for wc in wconns:
            items.append({"type": "water_connection", "type_label": "توصيلة مياه", "data": wc})
            
        # 6. توصيلات الصرف
        sewage_query = {**base_query}
        if current_user.role != "admin":
            allowed_projects = get_projects_with_permission(current_user, "sewage_connections")
            sewage_query["project"] = {"$in": allowed_projects}
            sewage_query["deleted_by"] = current_user.id
        sconns = await db.sewage_connections.find(sewage_query, {"_id": 0}).sort("deleted_at", -1).limit(max_fetch).to_list(max_fetch)
        for sc in sconns:
            items.append({"type": "sewage_connection", "type_label": "توصيلة صرف صحي", "data": sc})
            
        # 7. تقارير السلامة
        safety_query = {**base_query}
        if current_user.role != "admin":
            allowed_projects = get_projects_with_permission(current_user, "safety_reports")
            safety_query["project"] = {"$in": allowed_projects}
        sreps = await db.safety_reports.find(safety_query, {"_id": 0}).sort("deleted_at", -1).limit(max_fetch).to_list(max_fetch)
        for sr in sreps:
            items.append({"type": "safety_report", "type_label": "تقرير سلامة", "data": sr})
            
        # 8. تقارير الجودة
        quality_query = {**base_query}
        if current_user.role != "admin":
            allowed_projects = get_projects_with_permission(current_user, "quality_reports")
            quality_query["project"] = {"$in": allowed_projects}
        qreps = await db.quality_reports.find(quality_query, {"_id": 0}).sort("deleted_at", -1).limit(max_fetch).to_list(max_fetch)
        for qr in qreps:
            items.append({"type": "quality_report", "type_label": "تقرير جودة", "data": qr})
            
        # 9. تقارير الأعمال
        business_query = {**base_query}
        if current_user.role != "admin":
            allowed_projects = get_projects_with_permission(current_user, "business_reports")
            business_query["project"] = {"$in": allowed_projects}
        breps = await db.business_reports.find(business_query, {"_id": 0}).sort("deleted_at", -1).limit(max_fetch).to_list(max_fetch)
        for br in breps:
            items.append({"type": "business_report", "type_label": "تقرير أعمال", "data": br})

        # الترتيب
        items.sort(key=lambda x: x["data"].get("deleted_at", ""), reverse=True)
        
        # حساب الإجمالي
        c_inv = await db.invoices.count_documents(base_query)
        c_req = await db.employee_requests.count_documents(base_query)
        c_rep = await db.reports.count_documents(base_query)
        c_ext = await db.extracts.count_documents(base_query)
        c_w = await db.water_connections.count_documents(water_query)
        c_s = await db.sewage_connections.count_documents(sewage_query)
        c_sr = await db.safety_reports.count_documents(safety_query)
        c_qr = await db.quality_reports.count_documents(quality_query)
        c_br = await db.business_reports.count_documents(business_query)
        total_count = c_inv + c_req + c_rep + c_ext + c_w + c_s + c_sr + c_qr + c_br
        
        # القص للصفحة الحالية
        items = items[skip:skip+limit]
        
        # جلب أسماء الموظفين الذين قاموا بالحذف
        for item in items:
            doc = item["data"]
            deleted_by_name = "غير معروف"
            if doc.get("deleted_by"):
                user_doc = await db.users.find_one({"id": doc["deleted_by"]}, {"_id": 0, "full_name": 1})
                deleted_by_name = user_doc.get("full_name") if user_doc else "غير معروف"
            doc["deleted_by_name"] = deleted_by_name

    import math
    pages_count = math.ceil(total_count / limit) if limit > 0 else 1
    
    return {
        "items": items,
        "count": total_count,
        "total": total_count,
        "page": page,
        "pages": pages_count
    }


@api_router.post("/deleted-items/{item_type}/{item_id}/restore")
async def restore_deleted_item(item_type: str, item_id: str, current_user: User = Depends(get_current_user)):
    """استعادة عنصر محذوف"""
    user_permissions = current_user.permissions or []
    has_trash_perm = current_user.role == "admin" or "trash" in user_permissions or user_has_any_project_permission(current_user, "trash")
    if not has_trash_perm:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية الاستعادة")
    
    collection_map = {
        "invoice": db.invoices, 
        "employee_request": db.employee_requests, 
        "report": db.reports,
        "extract": db.extracts,
        "water_connection": db.water_connections,
        "sewage_connection": db.sewage_connections,
        "safety_report": db.safety_reports,
        "quality_report": db.quality_reports,
        "business_report": db.business_reports
    }
    collection = collection_map.get(item_type)
    if collection is None:
        raise HTTPException(status_code=400, detail="نوع غير صحيح")
    
    result = await collection.update_one({"id": item_id}, {"$unset": {"is_deleted": "", "deleted_by": "", "deleted_at": ""}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="العنصر غير موجود")
    return {"message": "تمت الاستعادة بنجاح"}


@api_router.delete("/deleted-items/{item_type}/{item_id}/permanent")
async def permanently_delete_item(item_type: str, item_id: str, current_user: User = Depends(get_current_user)):
    """حذف نهائي"""
    if current_user.role != "admin":
        user_permissions = current_user.permissions or []
        has_trash_perm = current_user.role == "admin" or "trash" in user_permissions or user_has_any_project_permission(current_user, "trash")
        if not has_trash_perm:
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية الحذف النهائي")
    
    collection_map = {
        "invoice": db.invoices, 
        "employee_request": db.employee_requests, 
        "report": db.reports,
        "extract": db.extracts,
        "water_connection": db.water_connections,
        "sewage_connection": db.sewage_connections,
        "safety_report": db.safety_reports,
        "quality_report": db.quality_reports,
        "business_report": db.business_reports
    }
    collection = collection_map.get(item_type)
    if collection is None:
        raise HTTPException(status_code=400, detail="نوع غير صحيح")
    
    result = await collection.delete_one({"id": item_id, "is_deleted": True})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="العنصر غير موجود")
    return {"message": "تم الحذف النهائي"}


@api_router.post("/deleted-items/bulk-permanent-delete")
async def bulk_permanent_delete_items(data: dict, current_user: User = Depends(get_current_user)):
    """حذف جماعي نهائي من سلة المحذوفات"""
    if current_user.role != "admin":
        user_permissions = current_user.permissions or []
        has_trash_perm = current_user.role == "admin" or "trash" in user_permissions or user_has_any_project_permission(current_user, "trash")
        if not has_trash_perm:
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية الحذف الجماعي")
    
    items = data.get("items", [])
    if not items:
        return {"message": "لا توجد عناصر للحذف", "deleted_count": 0}
    
    collection_map = {
        "invoice": db.invoices,
        "employee_request": db.employee_requests,
        "report": db.reports,
        "extract": db.extracts,
        "water_connection": db.water_connections,
        "sewage_connection": db.sewage_connections,
        "safety_report": db.safety_reports,
        "quality_report": db.quality_reports,
        "business_report": db.business_reports
    }
    
    deleted_count = 0
    for item in items:
        item_type = item.get("type")
        item_id = item.get("id")
        collection = collection_map.get(item_type)
        if collection is not None and item_id:
            try:
                result = await collection.delete_one({"id": item_id, "is_deleted": True})
                if result.deleted_count > 0:
                    deleted_count += 1
            except Exception as e:
                logger.error(f"Error bulk deleting item {item_id}: {str(e)}")
                
    return {"message": f"تم حذف {deleted_count} عنصر نهائياً", "deleted_count": deleted_count}



@api_router.put("/wfm/toggle/{report_id}")
async def toggle_report_wfm(report_id: str, current_user: User = Depends(get_current_user)):
    """تبديل حالة إغلاق WFM للبلاغ"""
    # التحقق من الصلاحيات
    user_govs = current_user.governorates if hasattr(current_user, 'governorates') and current_user.governorates else []
    has_all_govs = any(g.strip() in ["الكل", "جميع المحافظات", "كل المحافظات", "الكل "] for g in user_govs)
    is_admin = current_user.role == "admin"
    is_level2 = current_user.can_create_subusers
    user_perms = current_user.permissions or []
    has_reports_edit = "reports_edit" in user_perms or user_has_any_project_permission(current_user, "reports_edit")
    
    if not (is_admin or is_level2 or has_all_govs or has_reports_edit):
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية لتعديل حالة WFM")
        
    try:
        # نستخدم $ne: True بدلاً من False لالتقاط البلاغات التي ليس لها حقل is_deleted أو قيمته None
        report = await db.reports.find_one({"id": report_id, "is_deleted": {"$ne": True}})
        if not report:
            print(f"ERROR: Report {report_id} not found or deleted")
            raise HTTPException(status_code=404, detail="البلاغ غير موجود")
            
        new_status = not report.get("wfm_closed", False)
        update_fields = {
            "wfm_closed": new_status, 
            "updated_at": datetime.now(timezone.utc)
        }
        if new_status:
            closer_name = getattr(current_user, 'full_name', None) or getattr(current_user, 'username', '')
            update_fields["wfm_closed_by"] = closer_name
        else:
            update_fields["wfm_closed_by"] = None
        
        result = await db.reports.update_one(
            {"id": report_id}, 
            {"$set": update_fields}
        )
        print(f"DEBUG: Update WFM for {report_id} to {new_status}. Modified count: {result.modified_count}")
        
        msg = "تم اغلاق الرخصة علي منصة البنية التحتية" if new_status else "تم فتح معالجة الرخصة"
        return {"message": msg, "wfm_closed": new_status}
    except HTTPException:
        # أعد رمي استثناءات HTTP (403, 404) كما هي بدون تحويلها إلى 500
        raise
    except Exception as e:
        import traceback
        print(f"FATAL ERROR in toggle_report_wfm for {report_id}:")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"خطأ داخلي في السيرفر: {str(e)}")


# Include router moved to the end of the file

# GZip Compression
from starlette.middleware.gzip import GZipMiddleware
app.add_middleware(GZipMiddleware, minimum_size=1000)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_db():
    """إنشاء الـ indexes لتحسين الأداء + تهيئة Object Storage"""
    # تهيئة Emergent Object Storage (تُستدعى مرة واحدة)
    try:
        from storage import init_storage as _init_storage
        _init_storage()
    except Exception as e:
        logger.error(f"Storage init error: {e}")
    
    async def create_index_safe(collection, keys, **kwargs):
        try:
            await collection.create_index(keys, **kwargs)
            logger.info(f"Index created: {kwargs.get('name', 'unnamed')}")
        except Exception as e:
            error_msg = str(e)
            if 'IndexOptionsConflict' in error_msg or 'already exists' in error_msg.lower():
                logger.info(f"Index already exists (skipped): {kwargs.get('name', 'unnamed')}")
            else:
                logger.error(f"Error creating index {kwargs.get('name', 'unnamed')}: {error_msg}")
                raise
    try:
        await create_index_safe(db.reports, "id", name="idx_id")
        await create_index_safe(db.reports, "report_number", name="idx_report_number")
        await create_index_safe(db.reports, "license_number", name="idx_license_number")
        await create_index_safe(db.reports, "governorate", name="idx_governorate")
        await create_index_safe(db.reports, "project", name="idx_project")
        await create_index_safe(db.reports, "contractor", name="idx_contractor")
        await create_index_safe(db.reports, "status", name="idx_status")
        await create_index_safe(db.reports, "report_type", name="idx_report_type")
        await create_index_safe(db.reports, "is_deleted", name="idx_is_deleted")
        await create_index_safe(db.reports, "created_at", name="idx_created_at")
        await create_index_safe(db.reports, "created_by", name="idx_created_by")
        await create_index_safe(db.reports, [("is_deleted", 1), ("project", 1), ("governorate", 1), ("created_at", -1)], name="idx_reports_main_query")
        await create_index_safe(db.reports, [("is_deleted", 1), ("contractor", 1), ("status", 1), ("created_at", -1)], name="idx_reports_contractor_status")
        await create_index_safe(db.reports, [("is_deleted", 1), ("report_type", 1), ("created_at", -1)], name="idx_reports_type")
        await create_index_safe(db.reports, [("is_deleted", 1), ("report_number", 1)], name="idx_reports_search_number")
        await create_index_safe(db.reports, [("is_deleted", 1), ("license_number", 1)], name="idx_reports_search_license")
        await create_index_safe(db.reports, [("is_deleted", 1), ("created_at", -1), ("governorate", 1), ("project", 1)], name="idx_reports_48h")
        await create_index_safe(db.reports, [("is_deleted", 1), ("deleted_at", -1)], name="idx_reports_trash")
        await create_index_safe(db.users, "id", name="idx_users_id")
        await create_index_safe(db.users, "username", unique=True, name="idx_users_username_unique")
        await create_index_safe(db.users, "email", name="idx_users_email")
        await create_index_safe(db.users, "created_by", name="idx_users_created_by")
        await create_index_safe(db.users, "role", name="idx_users_role")
        await create_index_safe(db.contractors, "id", name="idx_contractors_id")
        await create_index_safe(db.contractors, [("name", 1), ("project", 1)], unique=True, name="idx_contractors_name_project")
        await create_index_safe(db.contractors, "project", name="idx_contractors_project")
        await create_index_safe(db.extracts, "id", name="idx_extracts_id")
        await create_index_safe(db.extracts, [("project", 1), ("month", 1), ("year", 1)], name="idx_extracts_project_month_year")
        await create_index_safe(db.extracts, "status", name="idx_extracts_status")
        await create_index_safe(db.extracts, "created_at", name="idx_extracts_created_at")
        logger.info("Database indexes created successfully")
    except Exception as e:
        logger.error(f"Error creating indexes: {str(e)}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


# ========== Safety Reports API ==========
@api_router.get("/safety-reports")
async def get_safety_reports(
    project: Optional[str] = None,
    governorate: Optional[str] = None,
    count_only: Optional[bool] = False,
    status_filter: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "safety_reports" not in user_perms:
        return []
    query = {"is_deleted": {"$ne": True}}
    

        
    and_clauses = []
    
    if user_doc.get("role") != "admin":
        user_govs = user_doc.get("governorates", [])
        user_projs = user_doc.get("projects", [])
        
        # Apply governorate restriction
        if governorate:
            gov_query = get_flexible_in_query([governorate], "governorate")
            if gov_query:
                and_clauses.append(gov_query)
        elif user_govs:
            gov_query = get_flexible_in_query(user_govs, "governorate")
            if gov_query:
                and_clauses.append(gov_query)
                
        # Apply project restriction
        if project:
            proj_query = get_flexible_in_query([project], "project")
            if proj_query:
                and_clauses.append(proj_query)
        elif user_projs:
            proj_query = get_loose_in_query(user_projs, "project")
            if proj_query:
                and_clauses.append(proj_query)
    else:
        # Admin - no permission restrictions
        if project:
            proj_query = get_flexible_in_query([project], "project")
            if proj_query:
                and_clauses.append(proj_query)
        if governorate:
            gov_query = get_flexible_in_query([governorate], "governorate")
            if gov_query:
                and_clauses.append(gov_query)
                
    if and_clauses:
        query["$and"] = and_clauses
    
    if status_filter:
        if status_filter == 'قيد المراجعة':
            query['$or'] = [{'status': 'قيد المراجعة'}, {'status': {'$exists': False}}, {'status': None}]
        else:
            query['status'] = status_filter
            
    if count_only:
        c = await db.safety_reports.count_documents(query)
        return {"count": c}
        
    records = await db.safety_reports.find(query, {"_id": 0, "image": 0, "images": 0}).sort("date", -1).to_list(100)
    for r in records:
        if not r.get("status"):
            r["status"] = "قيد المراجعة"
    return records

@api_router.get("/safety-reports/{report_id}")
async def get_safety_report(report_id: str, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "safety_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
    record = await db.safety_reports.find_one({"id": report_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Not found")
    if not record.get("status"):
        record["status"] = "قيد المراجعة"
    return record


@api_router.post("/safety-reports")
async def create_safety_report(request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions") or {}
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "safety_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية إضافة تقارير السلامة")
    body = await request.json()
    record = {
        "id": str(uuid.uuid4()),
        "date": body.get("date", ""),
        "project": body.get("project", ""),
        "governorate": body.get("governorate", ""),
        "notes": body.get("notes", ""),
        "image": body.get("image", ""),
        "images": body.get("images", []),
        "status": "قيد المراجعة",
        "created_by": user_doc.get("username", ""),
        "created_at": datetime.utcnow().isoformat()
    }
    await db.safety_reports.insert_one(record)
    record.pop("_id", None)
    return record


@api_router.put("/safety-reports/{report_id}")
async def update_safety_report(report_id: str, request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions") or {}
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "safety_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية تعديل تقارير السلامة")
    body = await request.json()
    update_data = {k: v for k, v in body.items() if k in ["date", "project", "governorate", "notes", "consultant_note", "consultant_reply", "image", "images", "report_note_processed", "consultant_note_processed", "status"]}
    
    if "status" in body:
        new_status = body.get("status")
        is_admin = user_doc.get("role") == "admin"
        is_manager = user_doc.get("can_create_subusers", False)
        
        existing = await db.safety_reports.find_one({"id": report_id})
        if not existing:
            raise HTTPException(status_code=404, detail="التقرير غير موجود")
            
        report_project = existing.get("project", "")
        user_projects = user_doc.get("projects", [])
        
        project_matched = False
        def normalize_str(s):
            if not s: return ""
            return s.strip().lower().replace(" ", "").replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").replace("ة", "ه").replace("ى", "ي")
            
        report_proj_norm = normalize_str(report_project)
        for p in user_projects:
            if normalize_str(p) == report_proj_norm:
                project_matched = True
                break
                
        if is_admin or (is_manager and project_matched):
            update_data["status"] = new_status
        else:
            raise HTTPException(status_code=403, detail="لا تملك صلاحية تعديل حالة هذا التقرير")
            
    update_data["updated_at"] = datetime.utcnow().isoformat()
    await db.safety_reports.update_one({"id": report_id}, {"$set": update_data})
    return {"message": "تم التحديث بنجاح"}


@api_router.delete("/safety-reports/{report_id}")
async def delete_safety_report(report_id: str, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "safety_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية حذف تقارير السلامة")
    result = await db.safety_reports.update_one(
        {"id": report_id},
        {"$set": {
            "is_deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.id
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="غير موجود")
    return {"message": "تم الحذف بنجاح"}


# ========== Quality Reports API ==========
@api_router.get("/quality-reports")
async def get_quality_reports(
    project: Optional[str] = None,
    governorate: Optional[str] = None,
    count_only: Optional[bool] = False,
    status_filter: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "quality_reports" not in user_perms:
        return []
    query = {"is_deleted": {"$ne": True}}
    

        
    and_clauses = []
    
    if user_doc.get("role") != "admin":
        user_govs = user_doc.get("governorates", [])
        user_projs = user_doc.get("projects", [])
        
        # Apply governorate restriction
        if governorate:
            gov_query = get_flexible_in_query([governorate], "governorate")
            if gov_query:
                and_clauses.append(gov_query)
        elif user_govs:
            gov_query = get_flexible_in_query(user_govs, "governorate")
            if gov_query:
                and_clauses.append(gov_query)
                
        # Apply project restriction
        if project:
            proj_query = get_flexible_in_query([project], "project")
            if proj_query:
                and_clauses.append(proj_query)
        elif user_projs:
            proj_query = get_loose_in_query(user_projs, "project")
            if proj_query:
                and_clauses.append(proj_query)
    else:
        # Admin - no permission restrictions
        if project:
            proj_query = get_flexible_in_query([project], "project")
            if proj_query:
                and_clauses.append(proj_query)
        if governorate:
            gov_query = get_flexible_in_query([governorate], "governorate")
            if gov_query:
                and_clauses.append(gov_query)
                
    if and_clauses:
        query["$and"] = and_clauses
    
    if status_filter:
        if status_filter == 'قيد المراجعة':
            query['$or'] = [{'status': 'قيد المراجعة'}, {'status': {'$exists': False}}, {'status': None}]
        else:
            query['status'] = status_filter
            
    if count_only:
        c = await db.quality_reports.count_documents(query)
        return {"count": c}
        
    records = await db.quality_reports.find(query, {"_id": 0, "image": 0, "images": 0}).sort("date", -1).to_list(100)
    for r in records:
        if not r.get("status"):
            r["status"] = "قيد المراجعة"
    return records

@api_router.get("/quality-reports/{report_id}")
async def get_quality_report(report_id: str, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "quality_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
    record = await db.quality_reports.find_one({"id": report_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Not found")
    if not record.get("status"):
        record["status"] = "قيد المراجعة"
    return record


@api_router.post("/quality-reports")
async def create_quality_report(request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "quality_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية إضافة تقارير الجودة")
    body = await request.json()
    # رفع الصورة إلى Cloudinary إذا كانت base64
    image_val = body.get("image", "")
    images_val = body.get("images", [])
    if images_val:
        processed_list = []
        for img in images_val:
            if img.startswith("data:"):
                try:
                    p = await process_images_for_storage([img], category="quality_reports")
                    processed_list.append(p[0] if p else img)
                except:
                    processed_list.append(img)
            else:
                processed_list.append(img)
        images_val = processed_list
    if image_val and isinstance(image_val, str) and image_val.startswith("data:"):
        try:
            processed = await process_images_for_storage([image_val], category="quality_reports")
            image_val = processed[0] if processed else image_val
        except Exception as e:
            logging.error(f"Quality report image upload failed: {e}")
    record = {
        "id": str(uuid.uuid4()),
        "date": body.get("date", ""),
        "project": body.get("project", ""),
        "governorate": body.get("governorate", ""),
        "notes": body.get("notes", ""),
        "image": image_val,
        "images": images_val,
        "status": "قيد المراجعة",
        "created_by": user_doc.get("username", ""),
        "created_at": datetime.utcnow().isoformat()
    }
    await db.quality_reports.insert_one(record)
    record.pop("_id", None)
    return record


@api_router.put("/quality-reports/{report_id}")
async def update_quality_report(report_id: str, request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "quality_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية تعديل تقارير الجودة")
    body = await request.json()
    update_data = {k: v for k, v in body.items() if k in ["date", "project", "governorate", "notes", "consultant_note", "consultant_reply", "report_note_processed", "consultant_note_processed", "image", "images"]}
    
    if "status" in body:
        new_status = body.get("status")
        is_admin = user_doc.get("role") == "admin"
        is_manager = user_doc.get("can_create_subusers", False)
        
        existing = await db.quality_reports.find_one({"id": report_id})
        if not existing:
            raise HTTPException(status_code=404, detail="التقرير غير موجود")
            
        report_project = existing.get("project", "")
        user_projects = user_doc.get("projects", [])
        
        project_matched = False
        def normalize_str(s):
            if not s: return ""
            return s.strip().lower().replace(" ", "").replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").replace("ة", "ه").replace("ى", "ي")
            
        report_proj_norm = normalize_str(report_project)
        for p in user_projects:
            if normalize_str(p) == report_proj_norm:
                project_matched = True
                break
                
        if is_admin or (is_manager and project_matched):
            update_data["status"] = new_status
        else:
            raise HTTPException(status_code=403, detail="لا تملك صلاحية تعديل حالة هذا التقرير")
            
    update_data["updated_at"] = datetime.utcnow().isoformat()
    await db.quality_reports.update_one({"id": report_id}, {"$set": update_data})
    return {"message": "تم التحديث بنجاح"}


@api_router.delete("/quality-reports/{report_id}")
async def delete_quality_report(report_id: str, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "quality_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية حذف تقارير الجودة")
    result = await db.quality_reports.update_one(
        {"id": report_id},
        {"$set": {
            "is_deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.id
        }}
    )
    if result.modified_count == 0:
        pass # Not found or already deleted, return success anyway
    return {"message": "تم الحذف بنجاح"}



# ========== Warehouse Visits API ==========
@api_router.get("/warehouse-visits")
async def get_warehouse_visits(count_only: Optional[bool] = False,
    status_filter: Optional[str] = None,
    current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "quality_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    query = {"is_deleted": {"$ne": True}}
    
    if user_doc.get("role") != "admin":
        user_projs = user_doc.get("projects", [])
        if user_projs:
            proj_query = get_loose_in_query(user_projs, "project")
            if proj_query:
                query.update(proj_query)
                
    records = await db.warehouse_visits.find(query, {"_id": 0}).sort("date", -1).to_list(100)
    return records

@api_router.post("/warehouse-visits")
async def create_warehouse_visit(request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "quality_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    data = await request.json()
    # رفع الصورة إلى Cloudinary إذا كانت base64
    image_val = data.get("image", "")
    images_val = data.get("images", [])
    if images_val:
        processed_list = []
        for img in images_val:
            if img.startswith("data:"):
                try:
                    p = await process_images_for_storage([img], category="warehouse_visits")
                    processed_list.append(p[0] if p else img)
                except:
                    processed_list.append(img)
            else:
                processed_list.append(img)
        images_val = processed_list
    if image_val and isinstance(image_val, str) and image_val.startswith("data:"):
        try:
            processed = await process_images_for_storage([image_val], category="warehouse_visits")
            image_val = processed[0] if processed else image_val
        except Exception as e:
            logging.error(f"Warehouse visit image upload failed: {e}")
    record = {
        "id": str(uuid.uuid4()),
        "date": data.get("date", ""),
        "project": data.get("project", ""),
        "governorate": data.get("governorate", ""),
        "notes": data.get("notes", ""),
        "image": image_val,
        "images": images_val,
        "created_by": user_doc.get("username", ""),
        "created_at": datetime.utcnow().isoformat(),
        "status": "قيد المراجعة"
    }
    await db.warehouse_visits.insert_one(record)
    record.pop("_id", None)
    return record

@api_router.put("/warehouse-visits/{visit_id}")
async def update_warehouse_visit(visit_id: str, request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "quality_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    update_data = await request.json()
    
    # Preserve original metadata
    existing = await db.warehouse_visits.find_one({"id": visit_id})
    if existing:
        if "created_by" in existing:
            update_data["created_by"] = existing["created_by"]
        if "created_at" in existing:
            update_data["created_at"] = existing["created_at"]
            
    await db.warehouse_visits.update_one({"id": visit_id}, {"$set": update_data})
    return {"message": "Updated successfully"}

@api_router.delete("/warehouse-visits/{visit_id}")
async def delete_warehouse_visit(visit_id: str, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "quality_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    result = await db.warehouse_visits.update_one(
        {"id": visit_id},
        {"$set": {
            "is_deleted": True,
            "deleted_at": datetime.utcnow().isoformat(),
            "deleted_by": user_doc.get("id")
        }}
    )
    return {"message": "Deleted successfully"}

# ========== Business Reports API ==========
@api_router.get("/business-reports")
async def get_business_reports(
    project: Optional[str] = None,
    governorate: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    count_only: Optional[bool] = False,
    status_filter: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "business_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
    query = {"is_deleted": {"$ne": True}}
    

        
    and_clauses = []
    
    if user_doc.get("role") != "admin":
        user_govs = user_doc.get("governorates", [])
        user_projs = user_doc.get("projects", [])
        
        # Apply governorate restriction
        if governorate:
            gov_query = get_flexible_in_query([governorate], "governorate")
            if gov_query:
                and_clauses.append({"$or": [
                    gov_query,
                    {"governorate": {"$in": ["جميع المحافظات", "الكل", "كل المحافظات"]}}
                ]})
        elif user_govs:
            gov_query = get_flexible_in_query(user_govs, "governorate")
            if gov_query:
                and_clauses.append(gov_query)
                
        # Apply project restriction
        if project:
            proj_query = get_flexible_in_query([project], "project")
            if proj_query:
                and_clauses.append(proj_query)
        elif user_projs:
            proj_query = get_loose_in_query(user_projs, "project")
            if proj_query:
                and_clauses.append(proj_query)
    else:
        # Admin - no permission restrictions
        if project:
            proj_query = get_flexible_in_query([project], "project")
            if proj_query:
                and_clauses.append(proj_query)
        if governorate:
            gov_query = get_flexible_in_query([governorate], "governorate")
            if gov_query:
                and_clauses.append({"$or": [
                    gov_query,
                    {"governorate": {"$in": ["جميع المحافظات", "الكل", "كل المحافظات"]}}
                ]})
                
    if and_clauses:
        query["$and"] = and_clauses
                
    if date_from:
        query["date_from"] = {"$gte": date_from}
    if date_to:
        query["date_to"] = {"$lte": date_to}
        
    records = await db.business_reports.find(query, {"_id": 0, "files": 0, "file_url": 0}).sort("date_from", -1).to_list(100)
    for r in records:
        if not r.get("status"):
            r["status"] = "قيد المراجعة"
    return records

@api_router.get("/business-reports/{report_id}")
async def get_business_report(report_id: str, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "business_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
    record = await db.business_reports.find_one({"id": report_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Not found")
    if not record.get("status"):
        record["status"] = "قيد المراجعة"
    return record


@api_router.post("/business-reports")
async def create_business_report(request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "business_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
    body = await request.json()
    record = {
        "id": str(uuid.uuid4()),
        "date_from": body.get("date_from", ""),
        "date_to": body.get("date_to", ""),
        "project": body.get("project", ""),
        "governorate": body.get("governorate", ""),
        "notes": body.get("notes", ""),
        "file_url": body.get("file_url", ""),
        "file_name": body.get("file_name", ""),
        "files": body.get("files", []),
        "status": "قيد المراجعة",
        "created_by": user_doc.get("username", ""),
        "created_at": datetime.utcnow().isoformat()
    }
    await db.business_reports.insert_one(record)
    record.pop("_id", None)
    return record


@api_router.put("/business-reports/{report_id}")
async def update_business_report(report_id: str, request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "business_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
    body = await request.json()
    update_data = {k: v for k, v in body.items() if k in ["date_from", "date_to", "project", "governorate", "notes", "file_url", "file_name", "files"]}

    if "status" in body:
        new_status = body.get("status")
        is_admin = user_doc.get("role") == "admin"
        is_manager = user_doc.get("can_create_subusers", False)
        existing = await db.business_reports.find_one({"id": report_id})
        if not existing:
            raise HTTPException(status_code=404, detail="التقرير غير موجود")
        report_project = existing.get("project", "")
        user_projects = user_doc.get("projects", [])
        project_matched = False
        def normalize_str(s):
            if not s: return ""
            return s.strip().lower().replace(" ", "").replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").replace("ة", "ه").replace("ى", "ي")
        report_proj_norm = normalize_str(report_project)
        for p in user_projects:
            if normalize_str(p) == report_proj_norm:
                project_matched = True
                break
        if is_admin or (is_manager and project_matched):
            update_data["status"] = new_status
        else:
            raise HTTPException(status_code=403, detail="لا تملك صلاحية تعديل حالة هذا التقرير")

    update_data["updated_at"] = datetime.utcnow().isoformat()
    await db.business_reports.update_one({"id": report_id}, {"$set": update_data})
    return {"message": "Success"}


@api_router.delete("/business-reports/{report_id}")
async def delete_business_report(report_id: str, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "business_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
    result = await db.business_reports.update_one(
        {"id": report_id},
        {"$set": {
            "is_deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.id
        }}
    )
    if result.modified_count == 0:
        pass # Not found or already deleted, return success anyway
    return {"message": "Success"}


# ========== Trash Endpoints for Reports ==========
@api_router.get("/safety-reports-trash")
async def get_safety_reports_trash(
    page: int = 1,
    limit: int = 15,
    current_user: User = Depends(get_current_user)
):
    user_perms = current_user.permissions or []
    has_trash = current_user.role == "admin" or "trash" in user_perms or user_has_any_project_permission(current_user, "trash")
    if not has_trash: raise HTTPException(status_code=403, detail="Forbidden")
    query = {"is_deleted": True}
    if current_user.role != "admin":
        query["deleted_by"] = current_user.id
    total_count = await db.safety_reports.count_documents(query)
    skip = (page - 1) * limit
    reports = await db.safety_reports.find(query, {"_id": 0}).sort("deleted_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"items": reports, "total": total_count}

@api_router.post("/safety-reports-trash/{report_id}/restore")
async def restore_safety_reports_trash(report_id: str, current_user: User = Depends(get_current_user)):
    user_perms = current_user.permissions or []
    if current_user.role != "admin" and "trash" not in user_perms and not user_has_any_project_permission(current_user, "trash"):
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.safety_reports.update_one({"id": report_id}, {"$unset": {"is_deleted": "", "deleted_by": "", "deleted_at": ""}})
    return {"message": "Restored"}

@api_router.delete("/safety-reports-trash/{report_id}/permanent")
async def permanent_safety_reports_trash(report_id: str, current_user: User = Depends(get_current_user)):
    user_perms = current_user.permissions or []
    if current_user.role != "admin" and "trash" not in user_perms and not user_has_any_project_permission(current_user, "trash"):
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.safety_reports.delete_one({"id": report_id, "is_deleted": True})
    return {"message": "Deleted"}


# ========== Work Permits API ==========
@api_router.get("/work-permits")
async def get_work_permits(
    project: Optional[str] = None,
    governorate: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "work_permits" not in user_perms and "safety_reports" not in user_perms:
        return []
    query = {"is_deleted": {"$ne": True}}
    

        
    and_clauses = []
    if user_doc.get("role") != "admin":
        user_govs = user_doc.get("governorates", [])
        user_projs = user_doc.get("projects", [])
        if governorate:
            gov_query = get_flexible_in_query([governorate], "governorate")
            if gov_query:
                and_clauses.append(gov_query)
        elif user_govs:
            gov_query = get_flexible_in_query(user_govs, "governorate")
            if gov_query:
                and_clauses.append(gov_query)
        if project:
            proj_query = get_flexible_in_query([project], "project")
            if proj_query:
                and_clauses.append(proj_query)
        elif user_projs:
            proj_query = get_loose_in_query(user_projs, "project")
            if proj_query:
                and_clauses.append(proj_query)
    else:
        if project:
            proj_query = get_flexible_in_query([project], "project")
            if proj_query:
                and_clauses.append(proj_query)
        if governorate:
            gov_query = get_flexible_in_query([governorate], "governorate")
            if gov_query:
                and_clauses.append(gov_query)
    if and_clauses:
        query["$and"] = and_clauses
    records = await db.work_permits.find(query, {"_id": 0, "image": 0}).sort("date", -1).to_list(100)
    for r in records:
        if not r.get("status"):
            r["status"] = "قيد المراجعة"
    return records


@api_router.get("/work-permits/{permit_id}")
async def get_work_permit(permit_id: str, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "work_permits" not in user_perms and "safety_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
    record = await db.work_permits.find_one({"id": permit_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Not found")
    if not record.get("status"):
        record["status"] = "قيد المراجعة"
    return record


@api_router.post("/work-permits")
async def create_work_permit(request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions") or {}
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "work_permits" not in user_perms and "safety_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية إضافة تصاريح العمل")
    body = await request.json()
    record = {
        "id": str(uuid.uuid4()),
        "date": body.get("date", ""),
        "project": body.get("project", ""),
        "governorate": body.get("governorate", ""),
        "notes": body.get("notes", ""),
        "image": body.get("image", ""),
        "images": body.get("images", []),
        "status": "قيد المراجعة",
        "created_by": user_doc.get("username", ""),
        "created_at": datetime.utcnow().isoformat()
    }
    await db.work_permits.insert_one(record)
    record.pop("_id", None)
    return record


@api_router.put("/work-permits/{permit_id}")
async def update_work_permit(permit_id: str, request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions") or {}
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "work_permits" not in user_perms and "safety_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية تعديل تصاريح العمل")
    body = await request.json()
    update_data = {k: v for k, v in body.items() if k in ["date", "project", "governorate", "notes", "consultant_note", "consultant_reply", "report_note_processed", "consultant_note_processed", "image", "images", "status"]}
    update_data["updated_at"] = datetime.utcnow().isoformat()
    await db.work_permits.update_one({"id": permit_id}, {"$set": update_data})
    return {"message": "تم التحديث بنجاح"}


@api_router.delete("/work-permits/{permit_id}")
async def delete_work_permit(permit_id: str, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "work_permits" not in user_perms and "safety_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية حذف تصاريح العمل")
    result = await db.work_permits.update_one(
        {"id": permit_id},
        {"$set": {
            "is_deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": user_doc.get("username", "")
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="غير موجود")
    return {"message": "تم الحذف بنجاح"}


# ============= نقاط نهاية المخالفات =============

@api_router.get("/violations")
async def get_violations(
    type: Optional[str] = "safety",
    project: Optional[str] = None,
    governorate: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    # صلاحية الوصول: safety_reports أو business_reports
    if user_doc.get("role") != "admin" and "safety_reports" not in user_perms and "business_reports" not in user_perms:
        return []
    query = {"is_deleted": {"$ne": True}}
    if type == "safety":
        query["type"] = {"$in": ["safety", None, ""]}
    else:
        query["type"] = type
        


    and_clauses = []
    if user_doc.get("role") != "admin":
        user_govs = user_doc.get("governorates", [])
        user_projs = user_doc.get("projects", [])
        if governorate:
            gq = get_flexible_in_query([governorate], "governorate")
            if gq: and_clauses.append(gq)
        elif user_govs:
            gq = get_flexible_in_query(user_govs, "governorate")
            if gq: and_clauses.append(gq)
        if project:
            pq = get_flexible_in_query([project], "project")
            if pq: and_clauses.append(pq)
        elif user_projs:
            pq = get_loose_in_query(user_projs, "project")
            if pq: and_clauses.append(pq)
    else:
        if project:
            pq = get_flexible_in_query([project], "project")
            if pq: and_clauses.append(pq)
        if governorate:
            gq = get_flexible_in_query([governorate], "governorate")
            if gq: and_clauses.append(gq)
    if and_clauses:
        query["$and"] = and_clauses
    records = await db.violations.find(query, {"_id": 0, "images": 0}).sort("date", -1).to_list(100)
    return records

@api_router.get("/violations/{violation_id}")
async def get_violation(violation_id: str, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "safety_reports" not in user_perms and "business_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="Forbidden")
    record = await db.violations.find_one({"id": violation_id, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="Not found")
    return record


@api_router.post("/violations")
async def create_violation(request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions") or {}
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "safety_reports" not in user_perms and "business_reports" not in user_perms:
        raise HTTPException(status_code=403, detail="لا تملك صلاحية إضافة مخالفات")
    body = await request.json()
    record = {
        "id": str(uuid.uuid4()),
        "date": body.get("date", ""),
        "project": body.get("project", ""),
        "governorate": body.get("governorate", ""),
        "violation_type": body.get("violation_type", ""),
        "type": body.get("type", "safety"),
        "notes": body.get("notes", ""),
        "images": body.get("images", []),
        "created_by": user_doc.get("username", ""),
        "created_at": datetime.utcnow().isoformat()
    }
    await db.violations.insert_one(record)
    record.pop("_id", None)
    return record


@api_router.put("/violations/{violation_id}")
async def update_violation(violation_id: str, request: Request, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions") or {}
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "safety_reports_edit" not in user_perms and "business_reports_edit" not in user_perms:
        raise HTTPException(status_code=403, detail="غير مصرح بالتعديل")
    body = await request.json()
    update_data = {k: v for k, v in body.items() if k not in ["id", "_id", "created_by", "created_at"]}
    await db.violations.update_one({"id": violation_id}, {"$set": update_data})
    updated = await db.violations.find_one({"id": violation_id}, {"_id": 0})
    return updated or {}


@api_router.delete("/violations/{violation_id}")
async def delete_violation(violation_id: str, current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions") or {}
    for plist in pp.values():
        user_perms.update(plist or [])
    if user_doc.get("role") != "admin" and "safety_reports_delete" not in user_perms and "business_reports_delete" not in user_perms:
        raise HTTPException(status_code=403, detail="غير مصرح بالحذف")
    await db.violations.delete_one({"id": violation_id})
    return {"message": "تم الحذف"}


@api_router.get("/quality-reports-trash")
async def get_quality_reports_trash(
    page: int = 1,
    limit: int = 15,
    current_user: User = Depends(get_current_user)
):
    user_perms = current_user.permissions or []
    has_trash = current_user.role == "admin" or "trash" in user_perms or user_has_any_project_permission(current_user, "trash")
    if not has_trash: raise HTTPException(status_code=403, detail="Forbidden")
    query = {"is_deleted": True}
    if current_user.role != "admin":
        query["deleted_by"] = current_user.id
    total_count_qr = await db.quality_reports.count_documents(query)
    total_count_wv = await db.warehouse_visits.count_documents(query)
    total_count = total_count_qr + total_count_wv
    skip = (page - 1) * limit
    
    qr_reports = await db.quality_reports.find(query, {"_id": 0}).to_list(1000)
    wv_reports = await db.warehouse_visits.find(query, {"_id": 0}).to_list(1000)
    
    for r in qr_reports: r['trash_type'] = 'quality'
    for r in wv_reports: r['trash_type'] = 'warehouse_visit'
    
    all_reports = qr_reports + wv_reports
    all_reports.sort(key=lambda x: x.get('deleted_at', ''), reverse=True)
    
    reports = all_reports[skip:skip+limit]
    return {"items": reports, "total": total_count}

@api_router.post("/quality-reports-trash/{report_id}/restore")
async def restore_quality_reports_trash(report_id: str, current_user: User = Depends(get_current_user)):
    user_perms = current_user.permissions or []
    if current_user.role != "admin" and "trash" not in user_perms and not user_has_any_project_permission(current_user, "trash"):
        raise HTTPException(status_code=403, detail="Forbidden")
    res1 = await db.quality_reports.update_one({"id": report_id}, {"$unset": {"is_deleted": "", "deleted_by": "", "deleted_at": ""}})
    res2 = await db.warehouse_visits.update_one({"id": report_id}, {"$unset": {"is_deleted": "", "deleted_by": "", "deleted_at": ""}})
    if res1.modified_count == 0 and res2.modified_count == 0:
        raise HTTPException(status_code=404, detail="Report not found in trash")
    return {"message": "Restored"}

@api_router.delete("/quality-reports-trash/{report_id}/permanent")
async def permanent_quality_reports_trash(report_id: str, current_user: User = Depends(get_current_user)):
    user_perms = current_user.permissions or []
    if current_user.role != "admin" and "trash" not in user_perms and not user_has_any_project_permission(current_user, "trash"):
        raise HTTPException(status_code=403, detail="Forbidden")
    res1 = await db.quality_reports.delete_one({"id": report_id, "is_deleted": True})
    res2 = await db.warehouse_visits.delete_one({"id": report_id, "is_deleted": True})
    if res1.deleted_count == 0 and res2.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Report not found in trash")
    return {"message": "Deleted"}

@api_router.get("/business-reports-trash")
async def get_business_reports_trash(
    page: int = 1,
    limit: int = 15,
    current_user: User = Depends(get_current_user)
):
    user_perms = current_user.permissions or []
    has_trash = current_user.role == "admin" or "trash" in user_perms or user_has_any_project_permission(current_user, "trash")
    if not has_trash: raise HTTPException(status_code=403, detail="Forbidden")
    query = {"is_deleted": True}
    if current_user.role != "admin":
        query["deleted_by"] = current_user.id
    total_count = await db.business_reports.count_documents(query)
    skip = (page - 1) * limit
    reports = await db.business_reports.find(query, {"_id": 0}).sort("deleted_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"items": reports, "total": total_count}

@api_router.post("/business-reports-trash/{report_id}/restore")
async def restore_business_reports_trash(report_id: str, current_user: User = Depends(get_current_user)):
    user_perms = current_user.permissions or []
    if current_user.role != "admin" and "trash" not in user_perms and not user_has_any_project_permission(current_user, "trash"):
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.business_reports.update_one({"id": report_id}, {"$unset": {"is_deleted": "", "deleted_by": "", "deleted_at": ""}})
    return {"message": "Restored"}

@api_router.delete("/business-reports-trash/{report_id}/permanent")
async def permanent_business_reports_trash(report_id: str, current_user: User = Depends(get_current_user)):
    user_perms = current_user.permissions or []
    if current_user.role != "admin" and "trash" not in user_perms and not user_has_any_project_permission(current_user, "trash"):
        raise HTTPException(status_code=403, detail="Forbidden")
    await db.business_reports.delete_one({"id": report_id, "is_deleted": True})
    return {"message": "Deleted"}


# Include router - MUST be after ALL endpoint definitions
class ConsultantNoteUpdate(BaseModel):
    consultant_note: str

@api_router.put("/reports/{report_id}/consultant_note")
async def update_consultant_note(report_id: str, payload: ConsultantNoteUpdate, current_user: User = Depends(get_current_user)):
    report = await db.reports.find_one({"id": report_id})
    if not report: raise HTTPException(status_code=404, detail="Report not found")
    
    if not has_project_permission(current_user.model_dump() if hasattr(current_user, 'model_dump') else current_user, report.get("project"), "consultant_notes"):
        raise HTTPException(status_code=403, detail="Not authorized to add consultant notes")
    
    from datetime import datetime, timezone
    update_data = {"consultant_note": payload.consultant_note}
    if not report.get("consultant_note_date") and payload.consultant_note:
        update_data["consultant_note_date"] = datetime.now(timezone.utc).isoformat()
        
    if payload.consultant_note != report.get("consultant_note", ""):
        update_data["consultant_note_processed"] = False
        update_data["consultant_note_processed_date"] = ""
        
    await db.reports.update_one({"id": report_id}, {"$set": update_data})
    
    if payload.consultant_note:
        import uuid
        from datetime import datetime, timezone
        target_users = []
        
        # 1. Admins
        admins = await db.users.find({"role": "admin", "is_deleted": {"$ne": True}}).to_list(None)
        for admin in admins:
            if admin.get("id") != current_user.id:
                target_users.append(admin)
                
        # 2. Level 2 & 3 users with access to this project & governorate who have 'consultant_notes' permission
        users = await db.users.find({"role": {"$ne": "admin"}, "is_deleted": {"$ne": True}}).to_list(None)
        
        def user_has_access(u):
            if u.get("id") == current_user.id: return False
            perms = u.get("permissions", [])
            pp = u.get("project_permissions", {})
            has_perm = "consultant_notes" in perms or any("consultant_notes" in p for p in pp.values())
            if not has_perm: return False
            
            u_projs = u.get("projects", [])
            report_proj = report.get("project", "")
            has_proj = False
            for p in u_projs:
                p_keys = [k for k in p.replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'في', 'محافظة']]
                r_keys = [k for k in report_proj.replace('-', ' ').split() if len(k) > 2 and k not in ['مشروع', 'في', 'محافظة']]
                if report_proj == p or any(k in report_proj for k in p_keys) or any(k in p for k in r_keys):
                    has_proj = True
                    break
            if not has_proj: return False
            
            u_govs = u.get("governorates", [])
            report_gov = report.get("governorate", "")
            has_all_govs = any(g.strip() in ["الكل", "جميع المحافظات", "كل المحافظات", "الكل "] for g in u_govs)
            
            has_gov = False
            if has_all_govs:
                has_gov = True
            else:
                for g in u_govs:
                    if g == report_gov or g in report_gov or report_gov in g:
                        has_gov = True
                        break
                        
            if not has_gov: return False
            
            return True

        for u in users:
            if user_has_access(u):
                target_users.append(u)
                
        # 3. Ensure creator gets it too if not already in list
        creator_id = report.get("created_by")
        creator = await db.users.find_one({"$or": [{"id": creator_id}, {"username": creator_id}]})
        if creator and creator.get("id") != current_user.id:
            if not any(t.get("id") == creator.get("id") for t in target_users):
                target_users.append(creator)
                
        # Send messages
        for t_user in target_users:
            message_id = str(uuid.uuid4())
            new_msg = {
                "id": message_id,
                "sender_id": current_user.id,
                "receiver_id": t_user.get("id"),
                "message": f"تمت إضافة/تعديل ملاحظة الاستشاري على البلاغ {report.get('report_number', '')}:\n{payload.consultant_note}",
                "created_at": datetime.now(timezone.utc),
                "is_read": False,
                "is_delivered": False,
                "is_edited": False
            }
            await db.messages.insert_one(new_msg)
            
    return {"message": "Success", "consultant_note": payload.consultant_note}



# ============= CHAT ENDPOINTS =============



@api_router.put("/chat/v2/groups/{group_id}")
async def update_chat_group(group_id: str, payload: dict, current_user: User = Depends(get_current_user)):
    group = await db.chat_groups.find_one({"id": group_id})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
        
    # Only creator can update
    if group.get("created_by") != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this group")
        
    update_data = {}
    if payload.get("name"):
        update_data["name"] = payload.get("name")
    if payload.get("members") is not None:
        members = payload.get("members")
        if current_user.id not in members:
            members.append(current_user.id)
        update_data["members"] = members
    if payload.get("avatar"):
        update_data["avatar"] = payload.get("avatar")
        
    if update_data:
        await db.chat_groups.update_one({"id": group_id}, {"$set": update_data})
        
    return {"message": "Group updated"}


@api_router.delete("/chat/v2/groups/{group_id}")
async def delete_chat_group(group_id: str, current_user: User = Depends(get_current_user)):
    group = await db.chat_groups.find_one({"id": group_id})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
        
    # Only creator can delete
    if group.get("created_by") != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this group")
        
    # Delete the group
    await db.chat_groups.delete_one({"id": group_id})
    
    # Optionally delete all messages in this group
    await db.chat_messages.delete_many({"receiver_id": group_id})
        
    return {"message": "Group deleted successfully"}

@api_router.post("/chat/v2/groups")
async def create_chat_group(payload: dict, current_user: User = Depends(get_current_user)):
    if not payload.get("name") or not payload.get("members"):
        raise HTTPException(status_code=400, detail="Name and members are required")
    
    group_id = f"group_{str(uuid.uuid4())[:8]}"
    members = payload.get("members")
    if current_user.id not in members:
        members.append(current_user.id)
        
    new_group = {
        "id": group_id,
        "name": payload.get("name"),
        "created_by": current_user.id,
        "members": members,
        "created_at": datetime.utcnow(),
        "is_group": True,
        "avatar": payload.get("avatar")
    }
    await db.chat_groups.insert_one(new_group)
    new_group["_id"] = str(new_group["_id"])
    new_group["created_at"] = new_group["created_at"].isoformat()
    return {"message": "Group created", "group": new_group}

@api_router.get("/chat/v2/contacts")
async def get_chat_contacts(current_user: User = Depends(get_current_user)):
    contacts = []
    
    if current_user.role == 'admin':
        # Admin (Level 1) sees ONLY Level 2 users (can_create_subusers = True, created by admin or None/missing)
        users = await db.users.find({
            "can_create_subusers": True,
            "role": {"$ne": "admin"},
            "$or": [
                {"created_by": current_user.id},
                {"created_by": None},
                {"created_by": {"$exists": False}}
            ]
        }, {"id": 1, "username": 1, "full_name": 1, "profile_picture": 1, "role": 1, "can_create_subusers": 1}).to_list(1000)
        contacts.extend(users)
        
    elif getattr(current_user, 'can_create_subusers', False):
        # Level 2 (Manager) sees Admin (Level 1)
        admins = await db.users.find({"role": "admin"}, {"id": 1, "username": 1, "full_name": 1, "profile_picture": 1, "role": 1, "can_create_subusers": 1}).to_list(100)
        contacts.extend(admins)
        
        # Level 2 sees ALL subusers created by them (Level 3)
        subusers = await db.users.find({
            "created_by": current_user.id,
            "role": {"$ne": "admin"}
        }, {"id": 1, "username": 1, "full_name": 1, "profile_picture": 1, "role": 1, "can_create_subusers": 1}).to_list(1000)
        contacts.extend(subusers)
        
    else:
        # Level 3 (Subordinate) sees ONLY their Level 2 creator (manager) if they share a project
        creator_id = getattr(current_user, 'created_by', None)
        if creator_id:
            manager = await db.users.find_one({
                "id": creator_id,
                "can_create_subusers": True
            }, {"id": 1, "username": 1, "full_name": 1, "profile_picture": 1, "role": 1, "can_create_subusers": 1, "projects": 1})
            if manager:
                user_projects = getattr(current_user, 'projects', []) or []
                manager_projects = manager.get('projects', []) or []
                if user_projects and manager_projects:
                    shared = set(user_projects).intersection(set(manager_projects))
                    if shared:
                        contacts.append(manager)
                else:
                    contacts.append(manager)
            
    # إضافة الاستثناءات (صلاحيات الدردشة الخاصة الممنوحة من الأدمن)
    # 1. المستخدمون الذين أضافهم هذا المستخدم في قائمته المسموحة
    my_allowed_ids = getattr(current_user, 'allowed_chat_users', []) or []
    if my_allowed_ids:
        explicit_users = await db.users.find(
            {"id": {"$in": my_allowed_ids}},
            {"id": 1, "username": 1, "full_name": 1, "profile_picture": 1, "role": 1, "can_create_subusers": 1}
        ).to_list(100)
        contacts.extend(explicit_users)
        
    # 2. المستخدمون الذين لديهم هذا المستخدم في قائمتهم المسموحة (تبادل الظهور)
    users_who_allow_me = await db.users.find(
        {"allowed_chat_users": current_user.id},
        {"id": 1, "username": 1, "full_name": 1, "profile_picture": 1, "role": 1, "can_create_subusers": 1}
    ).to_list(100)
    contacts.extend(users_who_allow_me)

    seen = set()
    unique_contacts = []
    for c in contacts:
        if c["id"] not in seen and c["id"] != current_user.id:
            seen.add(c["id"])
            
            # Count unread messages from this contact
            unread_count = await db.chat_messages.count_documents({
                "sender_id": c["id"],
                "receiver_id": current_user.id,
                "is_read": False,
                "is_deleted": False
            })
            c["unread_count"] = unread_count
            
            # Get last message
            last_msg = await db.chat_messages.find_one(
                {
                    "$or": [
                        {"sender_id": current_user.id, "receiver_id": c["id"]},
                        {"sender_id": c["id"], "receiver_id": current_user.id}
                    ],
                    "is_deleted": False
                },
                sort=[("created_at", -1)]
            )
            if last_msg:
                c["last_message"] = last_msg.get("text") if last_msg.get("text") else "صورة 📷"
                c["last_message_time"] = last_msg.get("created_at")
            
            unique_contacts.append(c)
            

    # Fetch groups
    groups = await db.chat_groups.find({"members": current_user.id}).to_list(100)
    for g in groups:
        unread_count = await db.chat_messages.count_documents({
            "receiver_id": g["id"],
            "sender_id": {"$ne": current_user.id},
            "read_by": {"$ne": current_user.id},
            "is_deleted": False
        })
        g["unread_count"] = unread_count
        g["full_name"] = g.get("name")
        g["is_group"] = True
        
        last_msg = await db.chat_messages.find_one({
            "receiver_id": g["id"],
            "is_deleted": False
        }, sort=[("created_at", -1)])
        
        if last_msg:
            sender = await db.users.find_one({"id": last_msg["sender_id"]}, {"full_name": 1, "username": 1})
            sender_name = sender.get("full_name") or sender.get("username") if sender else "Unknown"
            g["last_message"] = f"{sender_name}: {last_msg.get('text', 'مرفق')}"
            g["last_message_time"] = last_msg.get("created_at")
        else:
            g["last_message"] = "بدأت المجموعة"
            g["last_message_time"] = g.get("created_at")
                
        unique_contacts.append(g)

    unique_contacts.sort(key=lambda x: x.get("last_message_time") or datetime.min, reverse=True)
            
    return [{
        "id": c["id"],
        "name": c.get("full_name") or c.get("username"),
        "avatar": c.get("avatar") or c.get("profile_picture"),
        "last_message": c.get("last_message"),
        "last_message_time": c.get("last_message_time").isoformat() if isinstance(c.get("last_message_time"), datetime) else c.get("last_message_time"),
        "unread_count": c.get("unread_count", 0),
        "is_group": c.get("is_group", False),
        "members": c.get("members", []),
        "created_by": c.get("created_by")
    } for c in unique_contacts]
class LinkUsersRequest(BaseModel):
    user_ids: List[str]

@api_router.post("/chat/v2/link-users")
async def link_chat_users(req: LinkUsersRequest, current_user: User = Depends(get_current_user)):
    """ربط شخصين ببعضهما لتبادل الظهور في الدردشة"""
    if len(req.user_ids) != 2:
        raise HTTPException(status_code=400, detail="يجب تحديد مستخدمين اثنين للربط")
    
    is_admin = current_user.role == "admin"
    is_level2 = getattr(current_user, "can_create_subusers", False)
    
    if not (is_admin or is_level2):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
        
    u1_id, u2_id = req.user_ids
    
    u1 = await db.users.find_one({"id": u1_id})
    u2 = await db.users.find_one({"id": u2_id})
    
    if not u1 or not u2:
        raise HTTPException(status_code=404, detail="أحد المستخدمين غير موجود")
        
    if not is_admin:
        def can_manage(u):
            return u["id"] == current_user.id or u.get("created_by") == current_user.id
            
        if not (can_manage(u1) and can_manage(u2)):
            raise HTTPException(status_code=403, detail="يمكنك فقط ربط المستخدمين التابعين لفريقك")
            
    if u2_id in u1.get("allowed_chat_users", []):
        raise HTTPException(status_code=400, detail="المستخدمان مربوطان مسبقاً")
            
    await db.users.update_one(
        {"id": u1_id},
        {"$addToSet": {"allowed_chat_users": u2_id}}
    )
    await db.users.update_one(
        {"id": u2_id},
        {"$addToSet": {"allowed_chat_users": u1_id}}
    )
    
    return {"message": "تم ربط المستخدمين بنجاح للدردشة المباشرة"}

@api_router.post("/chat/v2/unlink-users")
async def unlink_chat_users(req: LinkUsersRequest, current_user: User = Depends(get_current_user)):
    """إلغاء ربط شخصين في الدردشة"""
    if len(req.user_ids) != 2:
        raise HTTPException(status_code=400, detail="يجب تحديد مستخدمين اثنين لإلغاء الربط")
    
    is_admin = current_user.role == "admin"
    is_level2 = getattr(current_user, "can_create_subusers", False)
    
    if not (is_admin or is_level2):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
        
    u1_id, u2_id = req.user_ids
    
    u1 = await db.users.find_one({"id": u1_id})
    u2 = await db.users.find_one({"id": u2_id})
    
    if not u1 or not u2:
        raise HTTPException(status_code=404, detail="أحد المستخدمين غير موجود")
        
    if not is_admin:
        def can_manage(u):
            return u["id"] == current_user.id or u.get("created_by") == current_user.id
            
        if not (can_manage(u1) and can_manage(u2)):
            raise HTTPException(status_code=403, detail="يمكنك فقط إلغاء ربط المستخدمين التابعين لفريقك")
            
    if u2_id not in u1.get("allowed_chat_users", []):
        raise HTTPException(status_code=400, detail="المستخدمان غير مربوطين مسبقاً")
        
    await db.users.update_one(
        {"id": u1_id},
        {"$pull": {"allowed_chat_users": u2_id}}
    )
    await db.users.update_one(
        {"id": u2_id},
        {"$pull": {"allowed_chat_users": u1_id}}
    )
    
    return {"message": "تم إلغاء ربط المستخدمين بنجاح"}

@api_router.get("/chat/v2/links")
async def get_chat_links(current_user: User = Depends(get_current_user)):
    """جلب قائمة الارتباطات الحالية للمستخدمين (لمديري الفرق والأدمن)"""
    is_admin = current_user.role == "admin"
    is_level2 = getattr(current_user, "can_create_subusers", False)
    
    if not (is_admin or is_level2):
        return []
        
    users_with_links = await db.users.find(
        {"allowed_chat_users": {"$exists": True, "$not": {"$size": 0}}}
    ).to_list(1000)
    
    links = []
    seen_pairs = set()
    
    for u in users_with_links:
        if not is_admin:
            if u["id"] != current_user.id and u.get("created_by") != current_user.id:
                continue
                
        for linked_id in u.get("allowed_chat_users", []):
            pair_key = tuple(sorted([u["id"], linked_id]))
            if pair_key not in seen_pairs:
                seen_pairs.add(pair_key)
                
                linked_user = await db.users.find_one({"id": linked_id})
                if linked_user:
                    links.append({
                        "user1_id": u["id"],
                        "user1_name": u.get("full_name") or u.get("username"),
                        "user2_id": linked_id,
                        "user2_name": linked_user.get("full_name") or linked_user.get("username")
                    })
                    
    return links

@api_router.get("/chat/v2/unread-count")
async def get_chat_unread_count(current_user: User = Depends(get_current_user)):
    count = await db.chat_messages.count_documents({
        "receiver_id": current_user.id,
        "is_read": False,
        "is_deleted": False
    })
    
    # حساب إشعارات المجموعات غير المقروءة
    groups = await db.chat_groups.find({"members": current_user.id}).to_list(100)
    group_ids = [g["id"] for g in groups]
    if group_ids:
        group_count = await db.chat_messages.count_documents({
            "receiver_id": {"$in": group_ids},
            "sender_id": {"$ne": current_user.id},
            "read_by": {"$ne": current_user.id},
            "is_deleted": False
        })
        count += group_count
        
    return {"unread_count": count}

@api_router.get("/chat/v2/messages/{contact_id}")
async def get_chat_messages(contact_id: str, current_user: User = Depends(get_current_user)):
    # Mark messages from this contact to current_user as read
    await db.chat_messages.update_many(
        {"sender_id": contact_id, "receiver_id": current_user.id, "is_read": False},
        {"$set": {"is_read": True}}
    )
    
    if contact_id.startswith("group_"):
        await db.chat_messages.update_many(
            {"receiver_id": contact_id, "sender_id": {"$ne": current_user.id}},
            {"$addToSet": {"read_by": current_user.id}}
        )
        messages = await db.chat_messages.find(
            {
                "receiver_id": contact_id,
                "is_deleted": False,
                "cleared_by": {"$ne": current_user.id}
            }
        ).sort("created_at", 1).to_list(500)
    else:
        messages = await db.chat_messages.find(
            {
                "$or": [
                    {"sender_id": current_user.id, "receiver_id": contact_id},
                    {"sender_id": contact_id, "receiver_id": current_user.id}
                ],
                "is_deleted": False,
                "cleared_by": {"$ne": current_user.id}
            }
        ).sort("created_at", 1).to_list(500)
    
    for msg in messages:
        msg["_id"] = str(msg["_id"])
        if isinstance(msg.get("created_at"), datetime):
            msg["created_at"] = msg["created_at"].isoformat()
    return messages

@api_router.post("/chat/v2/messages")
async def send_chat_message(msg_in: dict, current_user: User = Depends(get_current_user)):
    if not msg_in.get("receiver_id"):
        raise HTTPException(status_code=400, detail="Receiver ID is required")
    if not msg_in.get("text") and not msg_in.get("image_url"):
        raise HTTPException(status_code=400, detail="Text or image is required")
        
    new_msg = {
        "id": str(uuid.uuid4()),
        "sender_id": current_user.id,
        "receiver_id": msg_in.get("receiver_id"),
        "text": msg_in.get("text"),
        "image_url": msg_in.get("image_url"),
        "created_at": datetime.utcnow(),
        "is_deleted": False,
        "is_read": False,
        "read_by": [],
        "is_edited": False
    }
    await db.chat_messages.insert_one(new_msg)
    new_msg["_id"] = str(new_msg["_id"])
    new_msg["created_at"] = new_msg["created_at"].isoformat()
    return new_msg

@api_router.put("/chat/v2/messages/{message_id}")
async def edit_chat_message(message_id: str, payload: dict, current_user: User = Depends(get_current_user)):
    msg = await db.chat_messages.find_one({"id": message_id})
    if not msg:
        raise HTTPException(status_code=404, detail="Message not found")
        
    if msg["sender_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="لا يمكنك تعديل رسالة لم ترسلها")
        
    # Check 15 minutes limit
    created_at = msg.get("created_at")
    if isinstance(created_at, str):
        try:
            created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        except:
            created_at = datetime.utcnow()
            
    if created_at.tzinfo is not None:
        created_at = created_at.replace(tzinfo=None)
        
    time_diff = datetime.utcnow() - created_at
    if time_diff.total_seconds() > 20 * 60:
        raise HTTPException(status_code=403, detail="لا يمكنك تعديل الرسالة بعد مرور 20 دقيقة")
        
    text = payload.get("text")
    if not text or not text.strip():
        raise HTTPException(status_code=400, detail="Text cannot be empty")
        
    await db.chat_messages.update_one(
        {"id": message_id},
        {"$set": {"text": text.strip(), "is_edited": True}}
    )
    return {"message": "Message edited successfully", "text": text.strip()}

@api_router.delete("/chat/v2/messages/{message_id}")
async def delete_chat_message(message_id: str, for_everyone: bool = True, current_user: User = Depends(get_current_user)):
    msg = await db.chat_messages.find_one({"id": message_id})
    if not msg:
        raise HTTPException(status_code=404, detail="Message not found")
        
    if for_everyone:
        if msg["sender_id"] != current_user.id:
            raise HTTPException(status_code=403, detail="لا يمكنك حذف رسالة لم ترسلها للجميع")
            
        # Check 20 mins limit
        created_at = msg.get("created_at")
        if isinstance(created_at, str):
            try:
                created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
            except:
                created_at = datetime.utcnow()
                
        if created_at.tzinfo is not None:
            created_at = created_at.replace(tzinfo=None)
            
        time_diff = datetime.utcnow() - created_at
        if time_diff.total_seconds() > 20 * 60:
            raise HTTPException(status_code=403, detail="لا يمكنك حذف الرسالة بعد مرور 20 دقيقة")
            
        await db.chat_messages.update_one({"id": message_id}, {"$set": {"is_deleted": True}})
    else:
        # Delete for me only (hide from my view)
        await db.chat_messages.update_one({"id": message_id}, {"$addToSet": {"cleared_by": current_user.id}})
        
    return {"message": "Message deleted successfully", "for_everyone": for_everyone}

@api_router.delete("/chat/v2/conversation/{contact_id}")
async def delete_chat_conversation(contact_id: str, current_user: User = Depends(get_current_user)):
    # يحذف المحادثة من شاشة المستخدم الحالي فقط (مثل نظام الواتساب) عبر إضافته لقائمة cleared_by
    result = await db.chat_messages.update_many(
        {
            "$or": [
                {"sender_id": current_user.id, "receiver_id": contact_id},
                {"sender_id": contact_id, "receiver_id": current_user.id}
            ]
        },
        {"$addToSet": {"cleared_by": current_user.id}}
    )
    return {"message": "Conversation cleared for user successfully", "deleted_count": result.modified_count}
@api_router.get("/dashboard/badges")
async def get_dashboard_badges(current_user: User = Depends(get_current_user)):
    user_doc = current_user if isinstance(current_user, dict) else current_user.model_dump() if hasattr(current_user, 'model_dump') else current_user.dict()
    user_perms = set(user_doc.get("permissions", []))
    pp = user_doc.get("project_permissions", {})
    for plist in pp.values():
        user_perms.update(plist or [])
    role = getattr(current_user, "role", user_doc.get("role"))
    
    user_govs = user_doc.get("governorates", [])
    user_projs = user_doc.get("projects", [])
    
    common_and = []
    if role != "admin":
        if user_govs:
            gov_q = get_flexible_in_query(user_govs, "governorate")
            if gov_q: common_and.append(gov_q)
        if user_projs:
            proj_q = get_loose_in_query(user_projs, "project")
            if proj_q: common_and.append(proj_q)
            
    def get_query(extra_filter):
        q = {"is_deleted": {"$ne": True}}
        if extra_filter: q.update(extra_filter)
        if common_and: q["$and"] = common_and.copy()
        return q

    badges = {
        "safety": 0, "quality": 0, "warehouse": 0, "business": 0,
        "safety_notes": 0, "quality_notes": 0, "work_permits_notes": 0, "violations_notes": 0,
        "work_permits": 0, "violations": 0, "report_notes": 0, "consultant": 0
    }
    
    import asyncio
    tasks = []
    keys = []
    
    is_level_3 = role != "admin" and not user_doc.get("can_create_subusers", False)
    
    def get_note_query():
        if is_level_3:
            return {"consultant_note": {"$ne": "", "$exists": True}, "report_note_processed": {"$ne": True}}
        else:
            return {"consultant_reply": {"$ne": "", "$exists": True}, "consultant_note_processed": {"$ne": True}}
            
    note_q = get_note_query()

    if role == "admin" or "safety_reports" in user_perms:
        tasks.append(db.safety_reports.count_documents(get_query({'$or': [{'status': 'قيد المراجعة'}, {'status': {'$exists': False}}, {'status': None}]})))
        keys.append("safety")
        tasks.append(db.safety_reports.count_documents(get_query(note_q)))
        keys.append("safety_notes")
        
    if role == "admin" or "quality_reports" in user_perms:
        tasks.append(db.quality_reports.count_documents(get_query({"status": "قيد المراجعة"})))
        keys.append("quality")
        tasks.append(db.quality_reports.count_documents(get_query(note_q)))
        keys.append("quality_notes")
        
        tasks.append(db.warehouse_visits.count_documents(get_query({"status": "قيد المراجعة"})))
        keys.append("warehouse")
        
    if role == "admin" or "business_reports" in user_perms:
        tasks.append(db.business_reports.count_documents(get_query({"status": "قيد المراجعة"})))
        keys.append("business")
        
    if role == "admin" or "work_permits" in user_perms:
        tasks.append(db.work_permits.count_documents(get_query({"status": "قيد المراجعة"})))
        keys.append("work_permits")
        tasks.append(db.work_permits.count_documents(get_query(note_q)))
        keys.append("work_permits_notes")
        
    if role == "admin" or "violations" in user_perms:
        tasks.append(db.violations.count_documents(get_query({"status": "قيد المعالجة"})))
        keys.append("violations")
        tasks.append(db.violations.count_documents(get_query(note_q)))
        keys.append("violations_notes")
        
    if role == "admin" or "report_notes" in user_perms:
        tasks.append(db.reports.count_documents(get_query({"notes": {"$ne": "", "$exists": True}, "report_note_processed": {"$ne": True}})))
        keys.append("report_notes")
        
    if role == "admin" or "consultant_notes" in user_perms or "owner_notes" in user_perms:
        tasks.append(db.reports.count_documents(get_query({"consultant_note": {"$ne": "", "$exists": True}, "consultant_note_processed": {"$ne": True}})))
        keys.append("consultant")
        
    results = await asyncio.gather(*tasks) if tasks else []
    for k, v in zip(keys, results):
        badges[k] = v
        
    return badges


import fitz  # PyMuPDF
import base64

@api_router.post("/compress-pdf")
async def compress_pdf(data: dict, current_user: User = Depends(get_current_user)):
    pdf_base64 = data.get("pdf", "")
    if not pdf_base64.startswith("data:application/pdf"):
        return {"pdf": pdf_base64}
        
    try:
        header, encoded = pdf_base64.split(",", 1)
        pdf_bytes = base64.b64decode(encoded)
        
        # Compress using PyMuPDF
        doc = fitz.open("pdf", pdf_bytes)
        
        # Basic optimization
        new_bytes = doc.write(garbage=4, deflate=True, clean=True)
        doc.close()
        
        new_base64 = header + "," + base64.b64encode(new_bytes).decode('utf-8')
        return {"pdf": new_base64}
    except Exception as e:
        print("PDF compression error:", str(e))
        return {"pdf": pdf_base64}




# ============= MEETINGS MODULE =============
import uuid
import shutil
import os
from datetime import datetime

UPLOADS_DIR = "uploads"
os.makedirs(UPLOADS_DIR, exist_ok=True)

@api_router.get("/meetings")
async def get_meetings(
    skip: int = 0,
    limit: int = 10,
    date: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user)
):
    query = {}
    if date:
        query["date"] = date
    if type and type != "الكل":
        query["type"] = type
    if search:
        query["$or"] = [
            {"title": {"$regex": search, "$options": "i"}},
            {"contractor": {"$regex": search, "$options": "i"}},
            {"consultant": {"$regex": search, "$options": "i"}},
            {"project": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    if current_user.role != "admin":
        has_meetings = user_has_any_project_permission(current_user, "meetings")
        has_meetings_add = user_has_any_project_permission(current_user, "meetings_add")
        if not has_meetings and not has_meetings_add:
            raise HTTPException(status_code=403, detail="غير مصرح")
        
    total = await db.meetings.count_documents(query)
    meetings = await db.meetings.find(query).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Remove _id from dicts
    for m in meetings:
        if "_id" in m:
            del m["_id"]
            
    return {"total": total, "meetings": meetings}

@api_router.post("/meetings")
async def create_meeting(
    title: str = Form(...),
    type: str = Form(...),
    date: str = Form(...),
    contractor: str = Form(""),
    consultant: str = Form(""),
    project: str = Form(""),
    governorate: str = Form(""),
    description: str = Form(""),
    pdfs: List[UploadFile] = File(default=[]),
    images: List[UploadFile] = File(default=[]),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin" and not user_has_any_project_permission(current_user, "meetings_add"):
        raise HTTPException(status_code=403, detail="غير مصرح")

    meeting_id = str(uuid.uuid4())
    pdf_paths = []
    image_paths = []

    # Process PDFs
    for pdf in (pdfs or []):
        if pdf and pdf.filename:
            try:
                pdf_bytes = await pdf.read()
                pdf_bytes_to_upload = pdf_bytes
                if len(pdf_bytes) > 150 * 1024:
                    try:
                        import fitz
                        doc = fitz.open("pdf", pdf_bytes)
                        pdf_bytes_to_upload = doc.tobytes(garbage=3, deflate=True)
                        doc.close()
                    except Exception as e:
                        print("PDF compression error:", e)
                
                secure_url = _upload_image(pdf_bytes_to_upload, category="meetings", ext="pdf", content_type="application/pdf")
                pdf_paths.append(secure_url)
            except Exception as e:
                print("PDF upload error to Cloudinary:", e)

    # Process Images
    for img in (images or []):
        if img and img.filename:
            try:
                img_bytes = await img.read()
                secure_url = _upload_image(img_bytes, category="meetings", ext="jpg", content_type="image/jpeg")
                image_paths.append(secure_url)
            except Exception as e:
                print("Image upload error to Cloudinary:", e)

    new_meeting = {
        "id": meeting_id,
        "title": title,
        "type": type,
        "date": date,
        "contractor": contractor,
        "consultant": consultant,
        "project": project,
        "governorate": governorate,
        "description": description,
        "pdf_paths": pdf_paths,
        "images": image_paths,
        "created_by": current_user.username,
        "created_at": datetime.utcnow().isoformat(),
        "updated_at": datetime.utcnow().isoformat()
    }

    await db.meetings.insert_one(new_meeting.copy())
    return {"message": "تم إضافة الاجتماع بنجاح", "meeting": new_meeting}

@api_router.delete("/meetings/{meeting_id}")
async def delete_meeting(meeting_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin" and not user_has_any_project_permission(current_user, "meetings"):
        raise HTTPException(status_code=403, detail="غير مصرح")
        
    meeting = await db.meetings.find_one({"id": meeting_id})
    if not meeting:
        raise HTTPException(status_code=404, detail="الاجتماع غير موجود")
        
    # Delete files
    for pdf_path in meeting.get("pdf_paths", []):
        try:
            os.remove(os.path.join(UPLOADS_DIR, os.path.basename(pdf_path)))
        except: pass
        
    for img_path in meeting.get("images", []):
        try:
            os.remove(os.path.join(UPLOADS_DIR, os.path.basename(img_path)))
        except: pass

    await db.meetings.delete_one({"id": meeting_id})
    return {"message": "تم حذف الاجتماع بنجاح"}

@api_router.put("/meetings/{meeting_id}")
async def update_meeting(
    meeting_id: str,
    title: str = Form(...),
    type: str = Form(...),
    date: str = Form(...),
    contractor: str = Form(""),
    consultant: str = Form(""),
    project: str = Form(""),
    governorate: str = Form(""),
    description: str = Form(""),
    existing_images: str = Form("[]"),
    existing_pdfs: str = Form("[]"),
    pdfs: List[UploadFile] = File(default=[]),
    images: List[UploadFile] = File(default=[]),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin" and not user_has_any_project_permission(current_user, "meetings"):
        raise HTTPException(status_code=403, detail="غير مصرح")
        
    meeting = await db.meetings.find_one({"id": meeting_id})
    if not meeting:
        raise HTTPException(status_code=404, detail="الاجتماع غير موجود")

    import json
    try:
        kept_images = json.loads(existing_images)
    except:
        kept_images = []
        
    try:
        kept_pdfs = json.loads(existing_pdfs)
    except:
        kept_pdfs = []

    # delete removed files
    for old_pdf in meeting.get("pdf_paths", []):
        if old_pdf not in kept_pdfs:
            try: os.remove(os.path.join(UPLOADS_DIR, os.path.basename(old_pdf)))
            except: pass
        
    for old_img in meeting.get("images", []):
        if old_img not in kept_images:
            try: os.remove(os.path.join(UPLOADS_DIR, os.path.basename(old_img)))
            except: pass

    final_pdfs = kept_pdfs
    for pdf in (pdfs or []):
        if pdf and pdf.filename:
            try:
                pdf_bytes = await pdf.read()
                pdf_bytes_to_upload = pdf_bytes
                if len(pdf_bytes) > 150 * 1024:
                    try:
                        import fitz
                        doc = fitz.open("pdf", pdf_bytes)
                        pdf_bytes_to_upload = doc.tobytes(garbage=3, deflate=True)
                        doc.close()
                    except Exception as e:
                        print("PDF compression error:", e)
                
                secure_url = _upload_image(pdf_bytes_to_upload, category="meetings", ext="pdf", content_type="application/pdf")
                final_pdfs.append(secure_url)
            except Exception as e:
                print("PDF upload error to Cloudinary:", e)

    final_images = kept_images
    for img in (images or []):
        if img and img.filename:
            try:
                img_bytes = await img.read()
                secure_url = _upload_image(img_bytes, category="meetings", ext="jpg", content_type="image/jpeg")
                final_images.append(secure_url)
            except Exception as e:
                print("Image upload error to Cloudinary:", e)

    update_data = {
        "title": title,
        "type": type,
        "date": date,
        "contractor": contractor,
        "consultant": consultant,
        "project": project,
        "governorate": governorate,
        "description": description,
        "pdf_paths": final_pdfs,
        "images": final_images,
        "updated_at": datetime.utcnow().isoformat()
    }

    await db.meetings.update_one({"id": meeting_id}, {"$set": update_data})
    
    meeting.update(update_data)
    if "_id" in meeting:
        del meeting["_id"]
    return {"message": "تم تعديل الاجتماع بنجاح", "meeting": meeting}


app.include_router(api_router)
